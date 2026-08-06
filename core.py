"""
core.py — Moteur comptable (sans interface graphique).

Toute la logique métier vit ici, indépendamment de Tkinter, pour rester
testable en ligne de commande. main.py ne fait qu'appeler ces fonctions.
"""
import json
import os
import sys
import sqlite3
from datetime import date, datetime, timedelta


def _resource_dir():
    """Dossier des ressources bundlées : gère le cas exécutable PyInstaller."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Comptes SYSCOHADA "spéciaux" utilisés par les calculs automatiques
# (repris de la maquette Excel d'origine).
# ---------------------------------------------------------------------------
COMPTES_STOCK = ["310000", "320000", "331000", "360000"]
# Préfixes (3 chiffres) des mêmes comptes, utilisés uniquement pour agréger le
# total des stocks au Bilan (capture aussi d'éventuels sous-comptes de stock
# détaillés) — le suivi détaillé (onglet Stocks) reste lui scopé aux 4 comptes
# maîtres ci-dessus.
COMPTES_STOCK_PREFIXES = ["31", "32", "33", "36"]


def account_racine(code):
    """Racine (compte de rattachement) d'un compte : 1 chiffre pour les classes
    1,2,3,5,6,7,8,9 ; 2 chiffres pour la classe 4 (40 à 49), qui se subdivise
    par nature de tiers (40=Fournisseurs, 41=Clients, 42=Personnel,
    43=Organismes sociaux, 44=État, 45=Organismes internationaux,
    46=Associés/Groupe, 47=Débiteurs/créditeurs divers, 48=Régularisations,
    49=Dépréciations)."""
    code = str(code)
    if not code:
        return ""
    return code[:2] if code[0] == "4" else code[:1]


RACINE_FOURNISSEURS = "40"
RACINE_CLIENTS = "41"

# ---------------------------------------------------------------------------
# Facturation — mapping compte de vente (classe 70) -> impact sur les stocks.
# Un compte de vente lié à des marchandises (classe 31) ou des produits finis
# (classe 36) déclenche une sortie de stock automatique à la validation de la
# facture ; un compte de service (ex. 706000) n'impacte aucun stock.
# Le rattachement se fait par PRÉFIXE (3 chiffres) pour couvrir aussi les
# sous-comptes détaillés (ex. 701100, 701900... tous rattachés au préfixe 701).
# ---------------------------------------------------------------------------
VENTE_STOCK_MAPPING = {
    "701": ("marchandise", "310000", "603100"),   # Ventes marchandises -> stock 31, coût 603100
    "702": ("produit_fini", "360000", "736000"),  # Ventes produits finis -> stock 36, coût 736000
}
COMPTE_TVA_VENTES = "443100"  # État, T.V.A. facturée sur ventes
TVA_TAUX_DEFAUT = 18.0

# Achats (classe 6) -> impact sur les stocks : un achat de marchandises ou de
# matières premières augmente le stock correspondant (par préfixe, ex. 602101
# "Achat clinker" est bien rattaché au préfixe 602) ; un achat de service
# (ex. 622000) n'impacte aucun stock.
ACHAT_STOCK_MAPPING = {
    "601": ("marchandise", "310000", "603100"),        # Achats marchandises -> stock 31
    "602": ("matiere_premiere", "320000", "603200"),   # Achats matières premières -> stock 32
}
RETENUE_TAUX_DEFAUT = 0.0
COMPTE_RETENUE_DEFAUT = "447800"  # État, autres impôts et contributions (retenues à la source)


def _match_stock_mapping(compte, mapping):
    """Retrouve le mapping stock applicable à un compte, par préfixe de 3
    chiffres (ex. 602101 correspond au préfixe 602)."""
    if not compte or len(compte) < 3:
        return None
    return mapping.get(compte[:3])

RACINE_LABELS = {
    "1": "Comptes de ressources durables",
    "2": "Comptes d'actif immobilisé",
    "3": "Comptes de stocks",
    "40": "Fournisseurs et comptes rattachés",
    "41": "Clients et comptes rattachés",
    "42": "Personnel",
    "43": "Organismes sociaux",
    "44": "État et collectivités publiques",
    "45": "Organismes internationaux",
    "46": "Associés et groupe",
    "47": "Débiteurs et créditeurs divers",
    "48": "Comptes de régularisation",
    "49": "Dépréciations et provisions sur tiers",
    "5": "Comptes de trésorerie",
    "6": "Comptes de charges",
    "7": "Comptes de produits",
    "8": "Comptes des autres charges et produits (HAO)",
    "9": "Comptes analytiques/engagements",
}

COMPTES_TRESORERIE = ["521", "531", "570", "585"]
COMPTES_CAPITAL = ["101", "118", "121"]
COMPTE_SUBVENTIONS = "141"
COMPTE_PROVISIONS = "191"
COMPTES_DETTES_FIN = ["162", "165"]
COMPTES_PRODUITS_EXPL = ["701", "702", "705", "706", "736"]
COMPTE_SUBV_EXPL = "710"
COMPTE_AUTRES_PRODUITS = "758"
COMPTES_ACHATS = ["601", "602", "604", "605", "603"]
COMPTES_TRANSPORT = ["610", "614"]
COMPTES_SERVICES_EXT = ["622", "624", "625", "626", "627", "628",
                         "631", "632", "633"]
COMPTES_IMPOTS = ["641", "645"]
COMPTE_AUTRES_CHARGES = "651"
COMPTES_PERSONNEL = ["661", "663", "664"]
COMPTES_DOTATIONS = ["681", "691"]
COMPTES_PRODUITS_FIN = ["771", "776"]
COMPTES_CHARGES_FIN = ["671", "676"]

# ---------------------------------------------------------------------------
# Liasse fiscale — codes SYSCOHADA "système normal" (BILAN / RESULTAT)
# NB : les totaux (AD, AI, AZ, BK, BT, BZ, CP, DD, DP, DT, DZ) sont fiables
# (dérivés directement de la partie double). Le détail par ligne (AE..AN,
# CA/CH/CJ, DA/DJ/DK/DM/DR) est une répartition indicative par plage de
# comptes — à vérifier avec votre expert-comptable avant tout dépôt officiel.
# ---------------------------------------------------------------------------
RANGES_INCORP = {"AE": (211000, 211999), "AF": (212000, 214999),
                  "AG": (215000, 216999), "AH": (217000, 219999)}
RANGE_AMORT_INCORP = (281000, 281999)
RANGES_CORP = {"AJ": [(220000, 229999)], "AK": [(230000, 233999)],
               "AL": [(234000, 239999)],
               "AM": [(240000, 244999), (246000, 249999)],
               "AN": [(245000, 245999)]}
RANGE_AMORT_CORP = (282000, 297999)
RANGE_AVANCES_IMMO = (250000, 252999)
RANGE_TITRES_PARTICIPATION = (260000, 268999)
RANGE_AUTRES_IMMO_FIN = (270000, 278999)

RANGES_CAPITAUX = {"CA": [(101000, 104999)], "CD": [(105000, 105999)],
                    "CF_CG": [(110000, 118999)], "CH": [(120000, 129999)],
                    "CL": [(140000, 148999)], "CM": [(150000, 158999)]}
RANGE_DETTES_FIN = (160000, 168999)
RANGE_DETTES_LOCATION = (170000, 178999)
RANGE_PROVISIONS_RC = (190000, 198999)

RANGE_STOCKS = (300000, 399999)
RANGE_AVANCES_FOURN = (409000, 409999)
RANGE_CLIENTS = (411000, 419999)
RANGE_FOURNISSEURS = (401000, 408999)
RANGE_DETTES_FISC_SOC = (420000, 449999)
RANGE_AUTRES_DETTES = (450000, 499999)


def _in_ranges(code_int, ranges):
    if isinstance(ranges, tuple):
        ranges = [ranges]
    return any(lo <= code_int <= hi for lo, hi in ranges)


def _sum_range(balance, ranges, classe=None):
    total = 0.0
    for b in balance:
        code_int = int(b["code"])
        if classe and b["classe"] != classe:
            continue
        if _in_ranges(code_int, ranges):
            total += b["solde_cloture"]
    return total


def compute_liasse_bilan(conn, stock_initial=0.0, exercice=None):
    """Bilan au format SYSCOHADA système normal (codes officiels)."""
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    bilan_simple = compute_bilan(conn, stock_initial=stock_initial, exercice=exercice)

    # --- Détail indicatif Immobilisations incorporelles ---
    incorp_brut = {k: _sum_range(balance, [rng]) for k, rng in RANGES_INCORP.items()}
    total_incorp_brut = sum(incorp_brut.values())
    amort_incorp_total = -_sum_range(balance, [RANGE_AMORT_INCORP])  # positif
    incorp_net = {}
    for k, brut in incorp_brut.items():
        part = (brut / total_incorp_brut * amort_incorp_total) if total_incorp_brut else 0
        incorp_net[k] = brut - part

    # --- Détail indicatif Immobilisations corporelles ---
    corp_brut = {k: _sum_range(balance, rngs) for k, rngs in RANGES_CORP.items()}
    total_corp_brut = sum(corp_brut.values())
    amort_corp_total = -_sum_range(balance, [RANGE_AMORT_CORP])
    corp_net = {}
    for k, brut in corp_brut.items():
        part = (brut / total_corp_brut * amort_corp_total) if total_corp_brut else 0
        corp_net[k] = brut - part

    avances_immo = _sum_range(balance, [RANGE_AVANCES_IMMO])
    titres_participation = _sum_range(balance, [RANGE_TITRES_PARTICIPATION])
    autres_immo_fin = _sum_range(balance, [RANGE_AUTRES_IMMO_FIN])

    # --- Détail indicatif Capitaux propres ---
    capitaux_detail = {k: -_sum_range(balance, rngs) for k, rngs in RANGES_CAPITAUX.items()}
    dettes_financieres = -_sum_range(balance, [RANGE_DETTES_FIN])
    dettes_location = -_sum_range(balance, [RANGE_DETTES_LOCATION])
    provisions_rc = -_sum_range(balance, [RANGE_PROVISIONS_RC])

    # --- Détail indicatif Passif circulant ---
    fournisseurs = -_sum_range(balance, [RANGE_FOURNISSEURS])
    avances_fourn = -_sum_range(balance, [RANGE_AVANCES_FOURN])
    dettes_fisc_soc = -_sum_range(balance, [RANGE_DETTES_FISC_SOC])
    autres_dettes = -_sum_range(balance, [RANGE_AUTRES_DETTES])

    # --- Détail indicatif Actif circulant ---
    avances_versees = _sum_range(balance, [RANGE_AVANCES_FOURN])
    clients = _sum_range(balance, [RANGE_CLIENTS])

    return {
        "totaux": bilan_simple,
        "actif_detail": {
            **{k: {"brut": incorp_brut[k], "net": incorp_net[k]} for k in incorp_brut},
            **{k: {"brut": corp_brut[k], "net": corp_net[k]} for k in corp_brut},
            "AP": {"brut": avances_immo, "net": avances_immo},
            "AR": {"brut": titres_participation, "net": titres_participation},
            "AS": {"brut": autres_immo_fin, "net": autres_immo_fin},
        },
        "actif_circulant_detail": {
            "BH": avances_versees, "BI": clients,
        },
        "passif_detail": {
            **capitaux_detail,
            "DA": dettes_financieres, "DB": dettes_location, "DC": provisions_rc,
            "DJ": fournisseurs, "DH_avances": avances_fourn,
            "DK": dettes_fisc_soc, "DM": autres_dettes,
        },
    }


def compute_liasse_resultat(conn, exercice=None):
    """Compte de résultat au format SYSCOHADA système normal (codes officiels)."""
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)

    def net_produit(codes):
        d, c = _sum_accounts(balance, codes)
        return c - d

    def net_charge(codes):
        d, c = _sum_accounts(balance, codes)
        return d - c

    ta = net_produit(["701"])
    ra = net_charge(["601"])
    ra_stock = net_charge(["603100"])  # variation de stock de marchandises (préfixe précis pour ne pas doubler avec 603200)
    xa = ta - ra - ra_stock  # marge commerciale

    tb = net_produit(["702"])
    tc = net_produit(["705", "706"])
    td = 0.0
    xb = ta + tb + tc + td

    stock_d, stock_c = _sum_accounts(balance, ["360"])
    te = stock_d - stock_c
    th = net_produit(["758"])
    tg = net_produit(["710"])

    rc = net_charge(["602", "603200"])
    re = net_charge(["604", "605"])
    rg = net_charge(["610", "614"])
    rh = net_charge(["622", "624", "625", "626", "627", "628",
                      "631", "632", "633"])
    ri = net_charge(["641", "645"])
    rj = net_charge(["651"])
    xc = xb + (-ra) + (-ra_stock) + te + tg + th + (-rc) + (-re) + (-rg) + (-rh) + (-ri) + (-rj)

    rk = net_charge(["661", "663", "664"])
    xd = xc - rk

    rl = net_charge(["681", "691"])
    xe = xd - rl

    tk = net_produit(["771", "776"])
    rm = net_charge(["671", "676"])
    xf = tk - rm
    xg = xe + xf

    xh = 0.0  # Résultat HAO — non tracé dans cette application
    rq = 0.0  # Participation des travailleurs — non tracée
    rs = 0.0  # Impôt sur le résultat — non tracé (IS à calculer/saisir séparément)
    xi = xg + xh + rq + rs

    return {
        "TA": ta, "RA": ra, "RA_STOCK": ra_stock, "XA": xa,
        "TB": tb, "TC": tc, "TD": td, "XB": xb,
        "TE": te, "TG": tg, "TH": th,
        "RC": rc, "RE": re, "RG": rg, "RH": rh, "RI": ri, "RJ": rj, "XC": xc,
        "RK": rk, "XD": xd,
        "RL": rl, "XE": xe,
        "TK": tk, "RM": rm, "XF": xf, "XG": xg,
        "XH": xh, "RQ": rq, "RS": rs, "XI": xi,
    }


def default_db_path():
    """Emplacement du fichier de données, à côté de l'exécutable."""
    base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    folder = os.path.join(base, "SaisieComptable")
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "comptabilite.db")


def get_connection(db_path=None):
    db_path = db_path or default_db_path()
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    init_db(conn)
    return conn


def init_db(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            classe TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            piece TEXT,
            journal TEXT,
            compte TEXT NOT NULL,
            tiers TEXT,
            libelle TEXT,
            debit REAL NOT NULL DEFAULT 0,
            credit REAL NOT NULL DEFAULT 0,
            flux_code TEXT,
            analytic_code TEXT,
            budget_code TEXT,
            donor_code TEXT,
            quantite REAL NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS opening_balances (
            code TEXT NOT NULL,
            exercice TEXT NOT NULL,
            solde REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (code, exercice)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS exercices (
            exercice TEXT PRIMARY KEY,
            cloture INTEGER NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS analytic_codes (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS budget_codes (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            montant REAL NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS donor_codes (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS fournisseurs (
            code TEXT PRIMARY KEY,
            raison_sociale TEXT NOT NULL,
            contact TEXT,
            telephone TEXT,
            adresse TEXT,
            delai_paiement_jours INTEGER NOT NULL DEFAULT 30,
            delai_livraison_jours INTEGER NOT NULL DEFAULT 15
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS commandes_fournisseurs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fournisseur_code TEXT NOT NULL,
            piece TEXT,
            libelle TEXT,
            montant REAL NOT NULL DEFAULT 0,
            date_commande TEXT NOT NULL,
            date_livraison_prevue TEXT,
            date_livraison_reelle TEXT,
            date_echeance_paiement TEXT,
            date_paiement_reel TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS clients (
            code TEXT PRIMARY KEY,
            raison_sociale TEXT NOT NULL,
            contact TEXT,
            telephone TEXT,
            adresse TEXT,
            delai_paiement_jours INTEGER NOT NULL DEFAULT 30
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS factures_clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_code TEXT NOT NULL,
            piece TEXT,
            libelle TEXT,
            montant REAL NOT NULL DEFAULT 0,
            date_facture TEXT NOT NULL,
            date_echeance_paiement TEXT,
            date_paiement_reel TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS produits_finis (
            code TEXT PRIMARY KEY,
            nom TEXT NOT NULL,
            description TEXT,
            quantite_produite REAL NOT NULL DEFAULT 1,
            marge_pourcentage REAL NOT NULL DEFAULT 30,
            compte_stock TEXT NOT NULL DEFAULT '360000'
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS recette_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produit_code TEXT NOT NULL,
            type_ligne TEXT NOT NULL,
            libelle TEXT NOT NULL,
            compte TEXT,
            quantite REAL NOT NULL DEFAULT 0,
            cout_unitaire REAL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS factures_vente (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            date_facture TEXT NOT NULL,
            client_code TEXT NOT NULL,
            entete TEXT,
            pied_page TEXT,
            tva_taux REAL NOT NULL DEFAULT 0,
            statut TEXT NOT NULL DEFAULT 'brouillon',
            piece TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS facture_vente_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            facture_id INTEGER NOT NULL,
            compte_vente TEXT NOT NULL,
            libelle TEXT NOT NULL,
            quantite REAL NOT NULL DEFAULT 0,
            prix_unitaire REAL NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS factures_achat (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            date_facture TEXT NOT NULL,
            fournisseur_code TEXT NOT NULL,
            entete TEXT,
            pied_page TEXT,
            retenue_taux REAL NOT NULL DEFAULT 0,
            retenue_compte TEXT NOT NULL DEFAULT '447800',
            statut TEXT NOT NULL DEFAULT 'brouillon',
            piece TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS facture_achat_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            facture_id INTEGER NOT NULL,
            compte_achat TEXT NOT NULL,
            libelle TEXT NOT NULL,
            quantite REAL NOT NULL DEFAULT 0,
            prix_unitaire REAL NOT NULL DEFAULT 0
        )
    """)
    conn.commit()
    _migrate(conn)
    if conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0] == 0:
        load_plan_comptable(conn)
    ensure_racine_accounts(conn)


def _migrate(conn):
    """Ajoute les colonnes/tables manquantes si la base a été créée par une version antérieure."""
    cols = [r["name"] for r in conn.execute("PRAGMA table_info(entries)")]
    if "analytic_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN analytic_code TEXT")
    if "budget_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN budget_code TEXT")
    if "donor_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN donor_code TEXT")
    if "quantite" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN quantite REAL NOT NULL DEFAULT 0")
    if "fournisseur_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN fournisseur_code TEXT")
    if "client_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN client_code TEXT")

    pf_cols = [r["name"] for r in conn.execute("PRAGMA table_info(produits_finis)")]
    if pf_cols and "compte_stock" not in pf_cols:
        conn.execute("ALTER TABLE produits_finis ADD COLUMN compte_stock TEXT NOT NULL DEFAULT '360000'")

    # Migre l'ancien mécanisme "stock_initial_<compte>" (settings) vers opening_balances
    default_exercice = str(datetime.today().year)
    old_rows = conn.execute("SELECT key, value FROM settings WHERE key LIKE 'stock_initial_%'").fetchall()
    for row in old_rows:
        code = row["key"].replace("stock_initial_", "")
        try:
            val = float(row["value"])
        except (TypeError, ValueError):
            val = 0.0
        if val:
            conn.execute("INSERT OR REPLACE INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)",
                         (code, default_exercice, val))
        conn.execute("DELETE FROM settings WHERE key = ?", (row["key"],))

    # Migre l'ancienne table opening_balances (sans colonne exercice) vers le nouveau schéma
    ob_cols = [r["name"] for r in conn.execute("PRAGMA table_info(opening_balances)")]
    if "exercice" not in ob_cols:
        conn.execute("ALTER TABLE opening_balances RENAME TO opening_balances_old")
        conn.execute("""
            CREATE TABLE opening_balances (
                code TEXT NOT NULL, exercice TEXT NOT NULL, solde REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (code, exercice)
            )
        """)
        for row in conn.execute("SELECT code, solde FROM opening_balances_old"):
            conn.execute("INSERT OR REPLACE INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)",
                         (row["code"], default_exercice, row["solde"]))
        conn.execute("DROP TABLE opening_balances_old")

    if conn.execute("SELECT 1 FROM exercices WHERE exercice = ?", (default_exercice,)).fetchone() is None:
        # S'assure qu'au moins l'exercice courant existe dans la table
        conn.execute("INSERT OR IGNORE INTO exercices (exercice, cloture) VALUES (?, 0)", (default_exercice,))
    conn.commit()


def get_setting(conn, key, default=0.0):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return float(row["value"]) if row else default


def set_setting(conn, key, value):
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
    conn.commit()


def get_text_setting(conn, key, default=""):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


# ---------------------------------------------------------------------------
# Exercices comptables
# ---------------------------------------------------------------------------
def get_current_exercice(conn):
    ex = get_text_setting(conn, "exercice_courant", "")
    if ex:
        return ex
    ex = str(datetime.today().year)
    set_setting(conn, "exercice_courant", ex)
    conn.execute("INSERT OR IGNORE INTO exercices (exercice, cloture) VALUES (?, 0)", (ex,))
    conn.commit()
    return ex


def set_current_exercice(conn, exercice):
    conn.execute("INSERT OR IGNORE INTO exercices (exercice, cloture) VALUES (?, 0)", (exercice,))
    set_setting(conn, "exercice_courant", exercice)


def list_exercices(conn):
    """Tous les exercices connus : ceux créés explicitement + toutes les années
    présentes dans les écritures, triés."""
    years_in_entries = {r[0][:4] for r in conn.execute("SELECT DISTINCT date FROM entries") if r[0]}
    years_known = {r["exercice"] for r in conn.execute("SELECT exercice FROM exercices")}
    all_years = sorted(years_in_entries | years_known)
    cloture_map = {r["exercice"]: bool(r["cloture"]) for r in conn.execute("SELECT exercice, cloture FROM exercices")}
    return [{"exercice": y, "cloture": cloture_map.get(y, False)} for y in all_years]


def is_exercice_cloture(conn, exercice):
    row = conn.execute("SELECT cloture FROM exercices WHERE exercice = ?", (exercice,)).fetchone()
    return bool(row["cloture"]) if row else False


def _exercice_of_date(date_str):
    return (date_str or "")[:4]


def close_exercice(conn, exercice):
    """Clôture un exercice : calcule les soldes de clôture de tous les comptes de
    bilan (classes 1 à 5), les reporte comme soldes d'ouverture de l'exercice
    suivant, y intègre le résultat net de l'exercice clôturé (compte 121000 —
    report à nouveau), puis marque l'exercice comme clôturé."""
    if is_exercice_cloture(conn, exercice):
        raise ValueError(f"L'exercice {exercice} est déjà clôturé.")
    next_exercice = str(int(exercice) + 1)

    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    cr = compute_compte_resultat(conn, exercice=exercice)
    resultat_net = cr["resultat_net"]

    for b in balance:
        if b["classe"] not in ("1", "2", "3", "4", "5"):
            continue
        cloture = b["solde_cloture"]
        if b["code"] == "121000":
            cloture -= resultat_net  # intègre le résultat net dans le report à nouveau
        if cloture:
            conn.execute(
                "INSERT OR REPLACE INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)",
                (b["code"], next_exercice, cloture),
            )

    conn.execute("INSERT OR REPLACE INTO exercices (exercice, cloture) VALUES (?, 1)", (exercice,))
    conn.execute("INSERT OR IGNORE INTO exercices (exercice, cloture) VALUES (?, 0)", (next_exercice,))
    conn.commit()
    return next_exercice


# ---------------------------------------------------------------------------
# Soldes d'ouverture (report à nouveau) — un solde signé par compte, PAR
# EXERCICE. Balance de clôture = solde d'ouverture de l'exercice + mouvements
# de l'exercice (Débit - Crédit) enregistrés à des dates de cet exercice.
# ---------------------------------------------------------------------------
def get_opening_balance(conn, code, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    row = conn.execute("SELECT solde FROM opening_balances WHERE code = ? AND exercice = ?",
                        (code, exercice)).fetchone()
    return row["solde"] if row else 0.0


def set_opening_balance(conn, code, value, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    conn.execute("INSERT OR REPLACE INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)",
                 (code, exercice, value))
    conn.commit()


def list_opening_balances(conn, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    rows = conn.execute("""
        SELECT o.code, a.label, a.classe, o.solde
        FROM opening_balances o JOIN accounts a ON a.code = o.code
        WHERE o.solde != 0 AND o.exercice = ?
        ORDER BY o.code
    """, (exercice,)).fetchall()
    return [dict(r) for r in rows]


def total_opening_balance(conn, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    row = conn.execute("SELECT COALESCE(SUM(solde), 0) t FROM opening_balances WHERE exercice = ?",
                        (exercice,)).fetchone()
    return row["t"]


def export_opening_balances_xlsx(conn, path, exercice=None):
    """Exporte la balance d'ouverture (soldes d'ouverture) de l'exercice en .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    exercice = exercice or get_current_exercice(conn)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Soldes ouverture"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, label in enumerate(["N° Compte", "Libellé", "Solde (débit +, crédit -)"], start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    for r, row in enumerate(list_opening_balances(conn, exercice=exercice), start=2):
        ws.cell(row=r, column=1, value=row["code"])
        ws.cell(row=r, column=2, value=row["label"])
        ws.cell(row=r, column=3, value=row["solde"])
    for i, w in enumerate([14, 40, 20], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_opening_balances_xlsx(conn, path, exercice=None):
    """Importe une balance d'ouverture depuis un .xlsx et ÉCRASE les soldes
    d'ouverture existants pour cet exercice (les autres exercices ne sont
    pas affectés)."""
    import openpyxl

    exercice = exercice or get_current_exercice(conn)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]
    aliases = {"code": ["n° compte", "code", "compte"],
               "solde": ["solde (débit +, crédit -)", "solde", "solde débit", "montant"]}
    colmap = {}
    for key, alist in aliases.items():
        for i, h in enumerate(headers):
            if h in alist:
                colmap[key] = i
                break
    if "code" not in colmap or "solde" not in colmap:
        raise ValueError("Colonnes obligatoires introuvables (« N° Compte » et « Solde »).")

    rows = []
    warnings = []
    for r_idx, r in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in r]
        if all(v in (None, "") for v in values):
            continue
        code = str(values[colmap["code"]] or "").strip()
        if not code:
            continue
        try:
            solde = float(values[colmap["solde"]] or 0)
        except (TypeError, ValueError):
            warnings.append(f"Ligne {r_idx} : solde invalide pour le compte {code}, ignoré.")
            continue
        if not account_exists(conn, code):
            warnings.append(f"Ligne {r_idx} : compte « {code} » introuvable dans le Plan comptable — "
                             f"importé quand même.")
        rows.append((code, exercice, solde))

    if not rows:
        raise ValueError("Le fichier ne contient aucune ligne valide.")

    conn.execute("DELETE FROM opening_balances WHERE exercice = ?", (exercice,))
    conn.executemany("INSERT INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)", rows)
    conn.commit()
    return len(rows), warnings


def load_plan_comptable(conn, json_path=None):
    """Charge le plan comptable (bundlé avec l'application) dans la base."""
    if json_path is None:
        json_path = os.path.join(_resource_dir(), "plan_comptable.json")
    with open(json_path, encoding="utf-8") as f:
        accounts = json.load(f)
    conn.executemany(
        "INSERT OR REPLACE INTO accounts (code, label, classe) VALUES (?, ?, ?)",
        [(a["code"], a["label"], a["classe"]) for a in accounts],
    )
    conn.commit()
    ensure_racine_accounts(conn)


def ensure_racine_accounts(conn):
    """Insère les comptes racines (1 chiffre pour les classes 1,2,3,5,6,7,8,9 ;
    2 chiffres 40 à 49 pour la classe 4) s'ils n'existent pas déjà, sans écraser
    un compte que l'utilisateur aurait éventuellement créé avec ce même code.
    Grâce au tri alphabétique des codes (ex. '1' < '101000'), ces racines
    apparaissent en tête de chaque groupe dans les listes de comptes."""
    racines = []
    for c in ("1", "2", "3", "5", "6", "7", "8", "9"):
        racines.append((c, RACINE_LABELS.get(c, f"Classe {c}"), c))
    for r in range(40, 50):
        code = str(r)
        racines.append((code, RACINE_LABELS.get(code, f"Racine {code}"), "4"))
    for code, label, classe in racines:
        exists = conn.execute("SELECT 1 FROM accounts WHERE code = ?", (code,)).fetchone()
        if not exists:
            conn.execute("INSERT INTO accounts (code, label, classe) VALUES (?, ?, ?)",
                         (code, f"— {label} —", classe))
    conn.commit()


# ---------------------------------------------------------------------------
# Comptes
# ---------------------------------------------------------------------------
def search_accounts(conn, query, limit=50):
    query = (query or "").strip()
    if not query:
        rows = conn.execute("SELECT code, label, classe FROM accounts ORDER BY code LIMIT ?", (limit,))
    else:
        like = f"%{query}%"
        rows = conn.execute(
            "SELECT code, label, classe FROM accounts "
            "WHERE code LIKE ? OR label LIKE ? ORDER BY code LIMIT ?",
            (f"{query}%", like, limit),
        )
    return [dict(r) for r in rows]


def to_display_date(iso_str):
    """AAAA-MM-JJ (stockage) -> JJ/MM/AAAA (affichage)."""
    if not iso_str:
        return ""
    s = str(iso_str).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s  # déjà dans un autre format ou invalide : renvoyé tel quel


def to_iso_date(display_str):
    """JJ/MM/AAAA (saisie) -> AAAA-MM-JJ (stockage). Accepte aussi AAAA-MM-JJ en entrée."""
    s = (display_str or "").strip()
    if not s:
        return ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # format non reconnu : renvoyé tel quel (l'appelant peut valider)


def get_account_label(conn, code):
    row = conn.execute("SELECT label FROM accounts WHERE code = ?", (code,)).fetchone()
    return row["label"] if row else "Compte introuvable"


def account_exists(conn, code):
    return conn.execute("SELECT 1 FROM accounts WHERE code = ?", (code,)).fetchone() is not None


def add_account(conn, code, label, classe=None):
    code = str(code).strip()
    classe = classe or (code[0] if code else "")
    conn.execute("INSERT OR REPLACE INTO accounts (code, label, classe) VALUES (?, ?, ?)",
                 (code, label.strip(), classe))
    conn.commit()


def delete_account(conn, code):
    conn.execute("DELETE FROM accounts WHERE code = ?", (code,))
    conn.commit()


def export_plan_comptable_xlsx(conn, path):
    """Exporte tout le Plan comptable (code, libellé, classe) en .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan comptable"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, label in enumerate(["N° Compte", "Libellé", "Classe"], start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    for r, a in enumerate(conn.execute("SELECT code, label, classe FROM accounts ORDER BY code"), start=2):
        ws.cell(row=r, column=1, value=a["code"])
        ws.cell(row=r, column=2, value=a["label"])
        ws.cell(row=r, column=3, value=a["classe"])
    for i, w in enumerate([14, 45, 10], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_plan_comptable_xlsx(conn, path):
    """Importe un Plan comptable depuis un .xlsx et ÉCRASE l'ancien plan
    (toutes les fiches auxiliaires clients/fournisseurs et les écritures
    existantes ne sont PAS supprimées, mais leurs comptes ne seront plus
    reconnus s'ils ne figurent pas dans le nouveau plan)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]
    aliases = {"code": ["n° compte", "code", "compte"], "label": ["libellé", "libelle"],
               "classe": ["classe"]}
    colmap = {}
    for key, alist in aliases.items():
        for i, h in enumerate(headers):
            if h in alist:
                colmap[key] = i
                break
    if "code" not in colmap or "label" not in colmap:
        raise ValueError("Colonnes obligatoires introuvables (« N° Compte » et « Libellé »).")

    rows = []
    for r in ws.iter_rows(min_row=2):
        values = [c.value for c in r]
        if all(v in (None, "") for v in values):
            continue
        code = str(values[colmap["code"]] or "").strip()
        label = str(values[colmap["label"]] or "").strip()
        if not code or not label:
            continue
        classe = None
        if "classe" in colmap and colmap["classe"] < len(values) and values[colmap["classe"]]:
            classe = str(values[colmap["classe"]]).strip()
        if not classe:
            classe = code[:2] if code[:1] == "4" else code[:1]
        rows.append((code, label, classe))

    if not rows:
        raise ValueError("Le fichier ne contient aucun compte valide.")

    conn.execute("DELETE FROM accounts")
    conn.executemany("INSERT INTO accounts (code, label, classe) VALUES (?, ?, ?)", rows)
    conn.commit()
    ensure_racine_accounts(conn)
    return len(rows)


# ---------------------------------------------------------------------------
# Plans auxiliaires : analytique, budgétaire, bailleurs de fonds
# (même logique CRUD simple pour les 3, table dédiée chacun)
# ---------------------------------------------------------------------------
def _plan_list(conn, table, extra_cols=""):
    cols = "code, label" + (f", {extra_cols}" if extra_cols else "")
    rows = conn.execute(f"SELECT {cols} FROM {table} ORDER BY code").fetchall()
    return [dict(r) for r in rows]


def _plan_exists(conn, table, code):
    return conn.execute(f"SELECT 1 FROM {table} WHERE code = ?", (code,)).fetchone() is not None


def _plan_delete(conn, table, code):
    conn.execute(f"DELETE FROM {table} WHERE code = ?", (code,))
    conn.commit()


def list_analytic_codes(conn):
    return _plan_list(conn, "analytic_codes")


def analytic_code_exists(conn, code):
    return _plan_exists(conn, "analytic_codes", code)


def add_analytic_code(conn, code, label):
    conn.execute("INSERT OR REPLACE INTO analytic_codes (code, label) VALUES (?, ?)",
                 (code.strip(), label.strip()))
    conn.commit()


def delete_analytic_code(conn, code):
    _plan_delete(conn, "analytic_codes", code)


def list_budget_codes(conn):
    return _plan_list(conn, "budget_codes", extra_cols="montant")


def budget_code_exists(conn, code):
    return _plan_exists(conn, "budget_codes", code)


def add_budget_code(conn, code, label, montant=0):
    conn.execute("INSERT OR REPLACE INTO budget_codes (code, label, montant) VALUES (?, ?, ?)",
                 (code.strip(), label.strip(), montant or 0))
    conn.commit()


def delete_budget_code(conn, code):
    _plan_delete(conn, "budget_codes", code)


def list_donor_codes(conn):
    return _plan_list(conn, "donor_codes")


def donor_code_exists(conn, code):
    return _plan_exists(conn, "donor_codes", code)


def add_donor_code(conn, code, label):
    conn.execute("INSERT OR REPLACE INTO donor_codes (code, label) VALUES (?, ?)",
                 (code.strip(), label.strip()))
    conn.commit()


def delete_donor_code(conn, code):
    _plan_delete(conn, "donor_codes", code)


def _export_plan_generic_xlsx(conn, path, table, title, has_montant=False):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    headers = ["Code", "Libellé"] + (["Montant"] if has_montant else [])
    for i, label in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    cols = "code, label" + (", montant" if has_montant else "")
    for r, row in enumerate(conn.execute(f"SELECT {cols} FROM {table} ORDER BY code"), start=2):
        ws.cell(row=r, column=1, value=row["code"])
        ws.cell(row=r, column=2, value=row["label"])
        if has_montant:
            ws.cell(row=r, column=3, value=row["montant"])
    widths = [16, 40] + ([16] if has_montant else [])
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def _import_plan_generic_xlsx(conn, path, table, has_montant=False):
    """Importe un plan (code/libellé[/montant]) depuis un .xlsx et ÉCRASE
    l'ancien contenu de la table."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]
    aliases = {"code": ["code"], "label": ["libellé", "libelle"], "montant": ["montant"]}
    colmap = {}
    for key, alist in aliases.items():
        for i, h in enumerate(headers):
            if h in alist:
                colmap[key] = i
                break
    if "code" not in colmap or "label" not in colmap:
        raise ValueError("Colonnes obligatoires introuvables (« Code » et « Libellé »).")

    rows = []
    for r in ws.iter_rows(min_row=2):
        values = [c.value for c in r]
        if all(v in (None, "") for v in values):
            continue
        code = str(values[colmap["code"]] or "").strip()
        label = str(values[colmap["label"]] or "").strip()
        if not code or not label:
            continue
        montant = 0.0
        if has_montant and "montant" in colmap and colmap["montant"] < len(values):
            try:
                montant = float(values[colmap["montant"]] or 0)
            except (TypeError, ValueError):
                montant = 0.0
        rows.append((code, label, montant) if has_montant else (code, label))

    if not rows:
        raise ValueError("Le fichier ne contient aucune ligne valide.")

    conn.execute(f"DELETE FROM {table}")
    if has_montant:
        conn.executemany(f"INSERT INTO {table} (code, label, montant) VALUES (?, ?, ?)", rows)
    else:
        conn.executemany(f"INSERT INTO {table} (code, label) VALUES (?, ?)", rows)
    conn.commit()
    return len(rows)


def export_analytic_codes_xlsx(conn, path):
    return _export_plan_generic_xlsx(conn, path, "analytic_codes", "Plan analytique")


def import_analytic_codes_xlsx(conn, path):
    return _import_plan_generic_xlsx(conn, path, "analytic_codes")


def export_budget_codes_xlsx(conn, path):
    return _export_plan_generic_xlsx(conn, path, "budget_codes", "Plan budgétaire", has_montant=True)


def import_budget_codes_xlsx(conn, path):
    return _import_plan_generic_xlsx(conn, path, "budget_codes", has_montant=True)


def export_donor_codes_xlsx(conn, path):
    return _export_plan_generic_xlsx(conn, path, "donor_codes", "Plan bailleurs")


def import_donor_codes_xlsx(conn, path):
    return _import_plan_generic_xlsx(conn, path, "donor_codes")


# ---------------------------------------------------------------------------
# Équilibrage d'une pièce comptable
# ---------------------------------------------------------------------------
def get_piece_balance(conn, piece):
    """Retourne (total_debit, total_credit) pour toutes les lignes d'une pièce donnée."""
    row = conn.execute(
        "SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries WHERE piece = ?",
        (piece,),
    ).fetchone()
    return row["d"], row["c"]


# ---------------------------------------------------------------------------
# Fournisseurs (fiche auxiliaire)
# ---------------------------------------------------------------------------
def list_fournisseurs(conn, query=None):
    if query:
        like = f"%{query}%"
        rows = conn.execute(
            "SELECT * FROM fournisseurs WHERE code LIKE ? OR raison_sociale LIKE ? ORDER BY code",
            (f"{query}%", like),
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM fournisseurs ORDER BY code").fetchall()
    return [dict(r) for r in rows]


def fournisseur_exists(conn, code):
    return conn.execute("SELECT 1 FROM fournisseurs WHERE code = ?", (code,)).fetchone() is not None


def get_fournisseur(conn, code):
    row = conn.execute("SELECT * FROM fournisseurs WHERE code = ?", (code,)).fetchone()
    return dict(row) if row else None


def add_fournisseur(conn, code, raison_sociale, contact="", telephone="", adresse="",
                     delai_paiement_jours=30, delai_livraison_jours=15):
    conn.execute(
        """INSERT OR REPLACE INTO fournisseurs
           (code, raison_sociale, contact, telephone, adresse, delai_paiement_jours, delai_livraison_jours)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (code.strip(), raison_sociale.strip(), contact, telephone, adresse,
         int(delai_paiement_jours or 0), int(delai_livraison_jours or 0)),
    )
    conn.commit()


def delete_fournisseur(conn, code):
    conn.execute("DELETE FROM fournisseurs WHERE code = ?", (code,))
    conn.commit()


FOURNISSEUR_IMPORT_COLUMNS = [
    ("code", "Code fournisseur", ["code", "code fournisseur"]),
    ("raison_sociale", "Raison sociale", ["raison sociale", "nom", "dénomination"]),
    ("contact", "Contact", ["contact"]),
    ("telephone", "Téléphone", ["téléphone", "telephone", "tel"]),
    ("adresse", "Adresse", ["adresse"]),
    ("delai_paiement_jours", "Délai paiement (jours)", ["délai paiement (jours)", "delai paiement", "délai paiement"]),
    ("delai_livraison_jours", "Délai livraison (jours)", ["délai livraison (jours)", "delai livraison", "délai livraison"]),
]


def export_fournisseurs_template(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fournisseurs"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, (_, label, _) in enumerate(FOURNISSEUR_IMPORT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    example = ["FRS-0001", "Etablissements Dupont", "M. Dupont", "+226 70 00 00 00",
               "Ouagadougou", 30, 15]
    for i, val in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=val)
    for i, w in enumerate([14, 30, 18, 16, 26, 16, 16], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_fournisseurs_from_xlsx(conn, path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]

    colmap = {}
    for key, _, aliases in FOURNISSEUR_IMPORT_COLUMNS:
        for i, h in enumerate(headers):
            if h in aliases:
                colmap[key] = i
                break

    if "code" not in colmap or "raison_sociale" not in colmap:
        raise ValueError(
            "Colonnes obligatoires introuvables (« Code fournisseur » et « Raison sociale »). "
            "Utilisez le bouton « Télécharger un modèle »."
        )

    imported, warnings = 0, []

    def get(values, key, default=None):
        idx = colmap.get(key)
        if idx is None or idx >= len(values):
            return default
        return values[idx]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue
        code = str(get(values, "code") or "").strip()
        raison = str(get(values, "raison_sociale") or "").strip()
        if not code or not raison:
            warnings.append(f"Ligne {row_idx} : code ou raison sociale manquant, ligne ignorée.")
            continue
        try:
            dp = int(get(values, "delai_paiement_jours", 30) or 30)
        except (TypeError, ValueError):
            dp = 30
        try:
            dl = int(get(values, "delai_livraison_jours", 15) or 15)
        except (TypeError, ValueError):
            dl = 15
        add_fournisseur(conn, code, raison, str(get(values, "contact") or ""),
                         str(get(values, "telephone") or ""), str(get(values, "adresse") or ""),
                         dp, dl)
        imported += 1
    return imported, warnings


def compute_achats_par_fournisseur(conn, date_from=None, date_to=None):
    """Total Débit/Crédit/Solde par fournisseur, sur les seuls comptes fournisseurs
    (racine 40, tous les comptes 40xxxx) tagués avec le code fournisseur — le solde reflète ce qui reste
    dû (négatif = nous devons au fournisseur), sur une plage de dates optionnelle."""
    query = """
        SELECT e.fournisseur_code AS code,
               COALESCE(f.raison_sociale, e.fournisseur_code) AS raison_sociale,
               COALESCE(SUM(e.debit), 0) AS debit,
               COALESCE(SUM(e.credit), 0) AS credit
        FROM entries e
        LEFT JOIN fournisseurs f ON f.code = e.fournisseur_code
        WHERE e.fournisseur_code IS NOT NULL AND e.fournisseur_code != ''
          AND e.compte LIKE '40%'
    """
    params = []
    if date_from:
        query += " AND e.date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND e.date <= ?"
        params.append(date_to)
    query += " GROUP BY e.fournisseur_code, raison_sociale ORDER BY raison_sociale"
    rows = conn.execute(query, params).fetchall()
    result = []
    total_debit = total_credit = 0.0
    for r in rows:
        solde = r["debit"] - r["credit"]
        result.append({"code": r["code"], "raison_sociale": r["raison_sociale"],
                        "debit": r["debit"], "credit": r["credit"], "solde": solde})
        total_debit += r["debit"]
        total_credit += r["credit"]
    return result, total_debit, total_credit


# ---------------------------------------------------------------------------
# Contrats / commandes fournisseurs — suivi des délais de paiement et de
# livraison, avec détection des dépassements.
# ---------------------------------------------------------------------------
def add_commande(conn, fournisseur_code, piece, libelle, montant, date_commande,
                  date_livraison_prevue=None, date_paiement_prevue_override=None):
    fournisseur = get_fournisseur(conn, fournisseur_code)
    delai_paiement = fournisseur["delai_paiement_jours"] if fournisseur else 30
    delai_livraison = fournisseur["delai_livraison_jours"] if fournisseur else 15
    base = datetime.strptime(date_commande, "%Y-%m-%d")
    if not date_livraison_prevue:
        date_livraison_prevue = (base + timedelta(days=delai_livraison)).strftime("%Y-%m-%d")
    date_echeance_paiement = date_paiement_prevue_override or (
        base + timedelta(days=delai_paiement)).strftime("%Y-%m-%d")
    conn.execute(
        """INSERT INTO commandes_fournisseurs
           (fournisseur_code, piece, libelle, montant, date_commande, date_livraison_prevue,
            date_echeance_paiement)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (fournisseur_code, piece, libelle, montant, date_commande, date_livraison_prevue,
         date_echeance_paiement),
    )
    conn.commit()


def update_commande(conn, commande_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE commandes_fournisseurs SET {cols} WHERE id = ?", (*fields.values(), commande_id))
    conn.commit()


def delete_commande(conn, commande_id):
    conn.execute("DELETE FROM commandes_fournisseurs WHERE id = ?", (commande_id,))
    conn.commit()


def list_commandes(conn, fournisseur_code=None, date_from=None, date_to=None):
    query = """SELECT c.*, COALESCE(f.raison_sociale, c.fournisseur_code) AS raison_sociale
               FROM commandes_fournisseurs c LEFT JOIN fournisseurs f ON f.code = c.fournisseur_code
               WHERE 1=1"""
    params = []
    if fournisseur_code:
        query += " AND c.fournisseur_code = ?"
        params.append(fournisseur_code)
    if date_from:
        query += " AND c.date_commande >= ?"
        params.append(date_from)
    if date_to:
        query += " AND c.date_commande <= ?"
        params.append(date_to)
    query += " ORDER BY c.date_commande DESC, c.id DESC"
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    today = date.today().strftime("%Y-%m-%d")
    for r in rows:
        # Statut livraison
        if r["date_livraison_reelle"]:
            retard = (datetime.strptime(r["date_livraison_reelle"], "%Y-%m-%d")
                      - datetime.strptime(r["date_livraison_prevue"], "%Y-%m-%d")).days if r["date_livraison_prevue"] else 0
            r["statut_livraison"] = f"Livré (retard {retard} j)" if retard > 0 else "Livré à temps"
            r["depassement_livraison"] = retard > 0
        elif r["date_livraison_prevue"] and today > r["date_livraison_prevue"]:
            retard = (datetime.strptime(today, "%Y-%m-%d")
                      - datetime.strptime(r["date_livraison_prevue"], "%Y-%m-%d")).days
            r["statut_livraison"] = f"EN RETARD ({retard} j)"
            r["depassement_livraison"] = True
        else:
            r["statut_livraison"] = "En attente"
            r["depassement_livraison"] = False
        # Statut paiement
        if r["date_paiement_reel"]:
            retard = (datetime.strptime(r["date_paiement_reel"], "%Y-%m-%d")
                      - datetime.strptime(r["date_echeance_paiement"], "%Y-%m-%d")).days if r["date_echeance_paiement"] else 0
            r["statut_paiement"] = f"Payé (retard {retard} j)" if retard > 0 else "Payé à temps"
            r["depassement_paiement"] = retard > 0
        elif r["date_echeance_paiement"] and today > r["date_echeance_paiement"]:
            retard = (datetime.strptime(today, "%Y-%m-%d")
                      - datetime.strptime(r["date_echeance_paiement"], "%Y-%m-%d")).days
            r["statut_paiement"] = f"EN RETARD ({retard} j)"
            r["depassement_paiement"] = True
        else:
            r["statut_paiement"] = "En attente"
            r["depassement_paiement"] = False
    return rows


# ---------------------------------------------------------------------------
# Clients (fiche auxiliaire)
# ---------------------------------------------------------------------------
def list_clients(conn, query=None):
    if query:
        like = f"%{query}%"
        rows = conn.execute(
            "SELECT * FROM clients WHERE code LIKE ? OR raison_sociale LIKE ? ORDER BY code",
            (f"{query}%", like),
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM clients ORDER BY code").fetchall()
    return [dict(r) for r in rows]


def client_exists(conn, code):
    return conn.execute("SELECT 1 FROM clients WHERE code = ?", (code,)).fetchone() is not None


def get_client(conn, code):
    row = conn.execute("SELECT * FROM clients WHERE code = ?", (code,)).fetchone()
    return dict(row) if row else None


def add_client(conn, code, raison_sociale, contact="", telephone="", adresse="",
                delai_paiement_jours=30):
    conn.execute(
        """INSERT OR REPLACE INTO clients
           (code, raison_sociale, contact, telephone, adresse, delai_paiement_jours)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (code.strip(), raison_sociale.strip(), contact, telephone, adresse, int(delai_paiement_jours or 0)),
    )
    conn.commit()


def delete_client(conn, code):
    conn.execute("DELETE FROM clients WHERE code = ?", (code,))
    conn.commit()


CLIENT_IMPORT_COLUMNS = [
    ("code", "Code client", ["code", "code client"]),
    ("raison_sociale", "Raison sociale", ["raison sociale", "nom", "dénomination"]),
    ("contact", "Contact", ["contact"]),
    ("telephone", "Téléphone", ["téléphone", "telephone", "tel"]),
    ("adresse", "Adresse", ["adresse"]),
    ("delai_paiement_jours", "Délai paiement (jours)", ["délai paiement (jours)", "delai paiement", "délai paiement"]),
]


def export_clients_template(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, (_, label, _) in enumerate(CLIENT_IMPORT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    example = ["CLI-0001", "Société ABC", "Mme Traoré", "+226 70 11 11 11", "Ouagadougou", 30]
    for i, val in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=val)
    for i, w in enumerate([14, 30, 18, 16, 26, 18], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_clients_from_xlsx(conn, path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]

    colmap = {}
    for key, _, aliases in CLIENT_IMPORT_COLUMNS:
        for i, h in enumerate(headers):
            if h in aliases:
                colmap[key] = i
                break

    if "code" not in colmap or "raison_sociale" not in colmap:
        raise ValueError(
            "Colonnes obligatoires introuvables (« Code client » et « Raison sociale »). "
            "Utilisez le bouton « Télécharger un modèle »."
        )

    imported, warnings = 0, []

    def get(values, key, default=None):
        idx = colmap.get(key)
        if idx is None or idx >= len(values):
            return default
        return values[idx]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue
        code = str(get(values, "code") or "").strip()
        raison = str(get(values, "raison_sociale") or "").strip()
        if not code or not raison:
            warnings.append(f"Ligne {row_idx} : code ou raison sociale manquant, ligne ignorée.")
            continue
        try:
            dp = int(get(values, "delai_paiement_jours", 30) or 30)
        except (TypeError, ValueError):
            dp = 30
        add_client(conn, code, raison, str(get(values, "contact") or ""),
                   str(get(values, "telephone") or ""), str(get(values, "adresse") or ""), dp)
        imported += 1
    return imported, warnings


def compute_ventes_par_client(conn, date_from=None, date_to=None):
    """Total Débit/Crédit/Solde par client, sur les seuls comptes clients
    (racine 41, tous les comptes 41xxxx)
    tagués avec le code client — solde positif = montant restant dû par le client
    (à recouvrer), sur une plage de dates optionnelle."""
    query = """
        SELECT e.client_code AS code,
               COALESCE(c.raison_sociale, e.client_code) AS raison_sociale,
               COALESCE(SUM(e.debit), 0) AS debit,
               COALESCE(SUM(e.credit), 0) AS credit
        FROM entries e
        LEFT JOIN clients c ON c.code = e.client_code
        WHERE e.client_code IS NOT NULL AND e.client_code != ''
          AND e.compte LIKE '41%'
    """
    params = []
    if date_from:
        query += " AND e.date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND e.date <= ?"
        params.append(date_to)
    query += " GROUP BY e.client_code, raison_sociale ORDER BY raison_sociale"
    rows = conn.execute(query, params).fetchall()
    result = []
    total_debit = total_credit = 0.0
    for r in rows:
        solde = r["debit"] - r["credit"]
        result.append({"code": r["code"], "raison_sociale": r["raison_sociale"],
                        "debit": r["debit"], "credit": r["credit"], "solde": solde})
        total_debit += r["debit"]
        total_credit += r["credit"]
    return result, total_debit, total_credit


# ---------------------------------------------------------------------------
# Recouvrement — factures clients avec échéance et retard de paiement
# ---------------------------------------------------------------------------
def add_facture(conn, client_code, piece, libelle, montant, date_facture,
                 date_echeance_override=None):
    client = get_client(conn, client_code)
    delai_paiement = client["delai_paiement_jours"] if client else 30
    base = datetime.strptime(date_facture, "%Y-%m-%d")
    date_echeance = date_echeance_override or (base + timedelta(days=delai_paiement)).strftime("%Y-%m-%d")
    conn.execute(
        """INSERT INTO factures_clients
           (client_code, piece, libelle, montant, date_facture, date_echeance_paiement)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (client_code, piece, libelle, montant, date_facture, date_echeance),
    )
    conn.commit()


def update_facture(conn, facture_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE factures_clients SET {cols} WHERE id = ?", (*fields.values(), facture_id))
    conn.commit()


def delete_facture(conn, facture_id):
    conn.execute("DELETE FROM factures_clients WHERE id = ?", (facture_id,))
    conn.commit()


def list_factures(conn, client_code=None, date_from=None, date_to=None):
    query = """SELECT f.*, COALESCE(c.raison_sociale, f.client_code) AS raison_sociale
               FROM factures_clients f LEFT JOIN clients c ON c.code = f.client_code
               WHERE 1=1"""
    params = []
    if client_code:
        query += " AND f.client_code = ?"
        params.append(client_code)
    if date_from:
        query += " AND f.date_facture >= ?"
        params.append(date_from)
    if date_to:
        query += " AND f.date_facture <= ?"
        params.append(date_to)
    query += " ORDER BY f.date_facture DESC, f.id DESC"
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    today = date.today().strftime("%Y-%m-%d")
    for r in rows:
        if r["date_paiement_reel"]:
            retard = (datetime.strptime(r["date_paiement_reel"], "%Y-%m-%d")
                      - datetime.strptime(r["date_echeance_paiement"], "%Y-%m-%d")).days if r["date_echeance_paiement"] else 0
            r["statut_paiement"] = f"Payé (retard {retard} j)" if retard > 0 else "Payé à temps"
            r["depassement"] = retard > 0
        elif r["date_echeance_paiement"] and today > r["date_echeance_paiement"]:
            retard = (datetime.strptime(today, "%Y-%m-%d")
                      - datetime.strptime(r["date_echeance_paiement"], "%Y-%m-%d")).days
            r["statut_paiement"] = f"EN RETARD ({retard} j)"
            r["depassement"] = True
        else:
            r["statut_paiement"] = "En attente"
            r["depassement"] = False
    return rows


# ---------------------------------------------------------------------------
def _check_exercice_editable(conn, date_str):
    exercice = _exercice_of_date(date_str)
    if exercice and is_exercice_cloture(conn, exercice):
        raise ValueError(
            f"L'exercice {exercice} est clôturé : impossible d'ajouter, modifier ou supprimer "
            f"une écriture datée de cet exercice."
        )


def add_entry(conn, date_str, piece, journal, compte, tiers, libelle, debit, credit,
              flux_code="", analytic_code="", budget_code="", donor_code="", quantite=0,
              fournisseur_code="", client_code=""):
    _check_exercice_editable(conn, date_str)
    conn.execute(
        """INSERT INTO entries (date, piece, journal, compte, tiers, libelle, debit, credit,
                                 flux_code, analytic_code, budget_code, donor_code, quantite,
                                 fournisseur_code, client_code)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (date_str, piece, journal, compte, tiers, libelle, debit or 0, credit or 0,
         flux_code, analytic_code, budget_code, donor_code, quantite or 0, fournisseur_code, client_code),
    )
    conn.commit()


def add_balanced_entry(conn, date_str, piece, journal, compte_debit, compte_credit, montant,
                        tiers, libelle, analytic_code="", budget_code="", donor_code="", quantite=0,
                        fournisseur_code="", client_code=""):
    """Crée en une seule opération une écriture équilibrée par construction :
    une ligne au débit d'un compte, une ligne au crédit d'un autre, même montant.
    C'est le principe de la partie double — impossible de créer un déséquilibre
    en passant par cette fonction.

    Si une quantité est renseignée et que le compte débiteur est un compte
    d'achat lié à un stock (601x marchandises, 602x matières premières), une
    ENTRÉE de stock est automatiquement comptabilisée à sa suite. De même, si
    le compte créditeur est un compte de vente lié à un stock (701x
    marchandises, 702x produits finis), une SORTIE de stock est automatiquement
    comptabilisée (au coût unitaire moyen réel). Cela s'applique à toute
    écriture saisie directement dans l'onglet Saisie — pas seulement à celles
    créées via Facturation / Factures frs."""
    if montant <= 0:
        raise ValueError("Le montant doit être strictement positif.")
    if compte_debit == compte_credit:
        raise ValueError("Le compte débiteur et le compte créditeur doivent être différents.")
    _check_exercice_editable(conn, date_str)
    add_entry(conn, date_str, piece, journal, compte_debit, tiers, libelle, montant, 0,
              analytic_code=analytic_code, budget_code=budget_code, donor_code=donor_code,
              quantite=quantite, fournisseur_code=fournisseur_code, client_code=client_code)
    add_entry(conn, date_str, piece, journal, compte_credit, tiers, libelle, 0, montant,
              analytic_code=analytic_code, budget_code=budget_code, donor_code=donor_code,
              quantite=quantite, fournisseur_code=fournisseur_code, client_code=client_code)

    if quantite:
        achat_map = _match_stock_mapping(compte_debit, ACHAT_STOCK_MAPPING)
        if achat_map and compte_debit not in (achat_map[1], achat_map[2]):
            _, stock_compte, contre_compte = achat_map
            add_entry(conn, date_str, piece, journal, stock_compte, "", f"Entrée stock (auto) — {libelle}",
                      montant, 0, quantite=quantite)
            add_entry(conn, date_str, piece, journal, contre_compte, "", f"Entrée stock (auto) — {libelle}",
                      0, montant)

        vente_map = _match_stock_mapping(compte_credit, VENTE_STOCK_MAPPING)
        if vente_map and compte_credit not in (vente_map[1], vente_map[2]):
            _, stock_compte, cout_compte = vente_map
            stocks_by_code = {s["code"]: s for s in compute_stocks(conn, exercice=_exercice_of_date(date_str))}
            stock = stocks_by_code.get(stock_compte)
            cout_unitaire = stock["cout_unitaire_moyen"] if stock else None
            if cout_unitaire is not None:
                montant_sortie = quantite * cout_unitaire
                if montant_sortie > 0:
                    add_entry(conn, date_str, piece, journal, cout_compte, "", f"Sortie stock (auto) — {libelle}",
                              montant_sortie, 0)
                    add_entry(conn, date_str, piece, journal, stock_compte, "", f"Sortie stock (auto) — {libelle}",
                              0, montant_sortie, quantite=quantite)


def update_entry(conn, entry_id, **fields):
    if not fields:
        return
    row = conn.execute("SELECT date FROM entries WHERE id = ?", (entry_id,)).fetchone()
    if row:
        _check_exercice_editable(conn, row["date"])
    if "date" in fields:
        _check_exercice_editable(conn, fields["date"])
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE entries SET {cols} WHERE id = ?", (*fields.values(), entry_id))
    conn.commit()


def delete_entry(conn, entry_id):
    row = conn.execute("SELECT date FROM entries WHERE id = ?", (entry_id,)).fetchone()
    if row:
        _check_exercice_editable(conn, row["date"])
    conn.execute("DELETE FROM entries WHERE id = ?", (entry_id,))
    conn.commit()


def list_entries(conn, order_by="date", exercice=None):
    if exercice:
        date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
        rows = conn.execute(
            f"SELECT * FROM entries WHERE date >= ? AND date <= ? ORDER BY {order_by}, id",
            (date_from, date_to),
        ).fetchall()
    else:
        rows = conn.execute(f"SELECT * FROM entries ORDER BY {order_by}, id").fetchall()
    return [dict(r) for r in rows]


def totals_debit_credit(conn):
    row = conn.execute("SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries").fetchone()
    return row["d"], row["c"]


# ---------------------------------------------------------------------------
# Import massif d'écritures depuis un fichier .xlsx
# ---------------------------------------------------------------------------
IMPORT_COLUMNS = [
    ("date", "Date", ["date", "date piece", "date pièce"]),
    ("piece", "N° Pièce", ["pièce", "piece", "n° pièce", "n° piece", "numero piece", "num piece"]),
    ("journal", "Journal", ["journal"]),
    ("compte", "N° Compte", ["compte", "n° compte", "numero compte", "num compte"]),
    ("tiers", "Tiers", ["tiers"]),
    ("libelle", "Libellé", ["libellé", "libelle"]),
    ("debit", "Débit", ["débit", "debit"]),
    ("credit", "Crédit", ["crédit", "credit"]),
    ("quantite", "Quantité", ["quantité", "quantite", "qté", "qte"]),
    ("analytic_code", "Code analytique", ["code analytique", "analytique"]),
    ("budget_code", "Code budgétaire", ["code budgétaire", "code budgetaire", "budget"]),
    ("donor_code", "Code bailleur", ["code bailleur", "bailleur"]),
]


def export_import_template(path):
    """Génère un modèle .xlsx vierge (bon en-têtes) pour préparer un import massif."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ecritures"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, (_, label, _) in enumerate(IMPORT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    example = ["15/01/2024", "FA-0001", "AC", "601000", "Fournisseur X", "Achat marchandises",
               100000, 0, 10, "", "", ""]
    for i, val in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=val)
    example2 = ["15/01/2024", "FA-0001", "AC", "401000", "Fournisseur X", "Facture FA-0001",
                0, 100000, 0, "", "", ""]
    for i, val in enumerate(example2, start=1):
        ws.cell(row=3, column=i, value=val)
    for i, w in enumerate([12, 14, 10, 12, 20, 30, 14, 14, 10, 16, 16, 14], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_entries_from_xlsx(conn, path):
    """Importe en masse des écritures depuis un .xlsx. Reconnaît les en-têtes en
    français (Date, N° Compte, Débit, Crédit, etc. — voir IMPORT_COLUMNS) quel que
    soit leur ordre. Retourne (nb_importées, liste_avertissements)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]

    colmap = {}
    for key, _, aliases in IMPORT_COLUMNS:
        for i, h in enumerate(headers):
            if h in aliases:
                colmap[key] = i
                break

    if "date" not in colmap or "compte" not in colmap:
        raise ValueError(
            "Colonnes obligatoires introuvables dans le fichier (« Date » et « N° Compte »). "
            "Utilisez le bouton « Télécharger un modèle » pour obtenir les bons en-têtes."
        )

    valid_accounts = {r["code"] for r in conn.execute("SELECT code FROM accounts")}
    imported = 0
    warnings = []

    def get(values, key):
        idx = colmap.get(key)
        if idx is None or idx >= len(values):
            return None
        return values[idx]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue

        date_val = get(values, "date")
        if date_val in (None, ""):
            warnings.append(f"Ligne {row_idx} : date manquante, ligne ignorée.")
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        elif isinstance(date_val, date):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = to_iso_date(str(date_val))

        compte = get(values, "compte")
        compte = "" if compte is None else str(compte).strip()
        if compte.endswith(".0") and compte.replace(".0", "").isdigit():
            compte = compte[:-2]
        if not compte:
            warnings.append(f"Ligne {row_idx} : N° Compte manquant, ligne ignorée.")
            continue
        if compte not in valid_accounts:
            warnings.append(f"Ligne {row_idx} : compte '{compte}' absent du plan comptable (importée quand même).")

        def to_float(v, label):
            if v in (None, ""):
                return 0.0
            try:
                return float(v)
            except (TypeError, ValueError):
                warnings.append(f"Ligne {row_idx} : {label} invalide ('{v}'), remplacé par 0.")
                return 0.0

        debit = to_float(get(values, "debit"), "Débit")
        credit = to_float(get(values, "credit"), "Crédit")
        quantite = to_float(get(values, "quantite"), "Quantité")
        piece = get(values, "piece") or ""
        journal = get(values, "journal") or ""
        tiers = get(values, "tiers") or ""
        libelle = get(values, "libelle") or ""
        analytic_code = get(values, "analytic_code") or ""
        budget_code = get(values, "budget_code") or ""
        donor_code = get(values, "donor_code") or ""

        add_entry(conn, date_str, str(piece), str(journal), compte, str(tiers), str(libelle),
                  debit, credit, "", str(analytic_code), str(budget_code), str(donor_code), quantite)
        imported += 1

    return imported, warnings


# ---------------------------------------------------------------------------
# Balance
# ---------------------------------------------------------------------------
def compute_balance(conn, only_with_movement=True, include_zero_opening=True, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    rows = conn.execute("""
        SELECT a.code, a.label, a.classe,
               COALESCE(SUM(CASE WHEN e.date BETWEEN ? AND ? THEN e.debit ELSE 0 END), 0)  AS debit,
               COALESCE(SUM(CASE WHEN e.date BETWEEN ? AND ? THEN e.credit ELSE 0 END), 0) AS credit
        FROM accounts a
        LEFT JOIN entries e ON e.compte = a.code
        GROUP BY a.code, a.label, a.classe
        ORDER BY a.code
    """, (date_from, date_to, date_from, date_to)).fetchall()
    openings = {r["code"]: r["solde"] for r in conn.execute(
        "SELECT code, solde FROM opening_balances WHERE exercice = ?", (exercice,))}
    result = []
    for r in rows:
        debit, credit = r["debit"], r["credit"]
        ouverture = openings.get(r["code"], 0.0)
        if only_with_movement and debit == 0 and credit == 0 and ouverture == 0:
            continue
        solde_mouvement = debit - credit
        result.append({
            "code": r["code"], "label": r["label"], "classe": r["classe"],
            "debit": debit, "credit": credit, "solde": solde_mouvement,
            "solde_ouverture": ouverture,
            "solde_cloture": ouverture + solde_mouvement,
        })
    return result


def compute_balance_detaillee(conn, exercice=None):
    """Balance générale groupée par classe, avec un sous-total par classe et
    un total général — même structure que la Balance PDF de référence
    (N° compte, Libellé, Cumul Débit, Cumul Crédit, Solde Débit, Solde
    Crédit). Calculée à partir de la même compute_balance() que le Bilan,
    donc garantie cohérente avec lui."""
    balance = sorted(compute_balance(conn, only_with_movement=True, exercice=exercice),
                      key=lambda b: b["code"])
    classes = {}
    for b in balance:
        classes.setdefault(b["classe"], []).append(b)

    result_classes = []
    grand = {"cumul_debit": 0.0, "cumul_credit": 0.0, "solde_debit": 0.0, "solde_credit": 0.0}
    for classe in sorted(classes.keys()):
        lignes = []
        sous_total = {"cumul_debit": 0.0, "cumul_credit": 0.0, "solde_debit": 0.0, "solde_credit": 0.0}
        for b in classes[classe]:
            solde_cloture = b["solde_cloture"]
            solde_debit = solde_cloture if solde_cloture > 0 else 0.0
            solde_credit = -solde_cloture if solde_cloture < 0 else 0.0
            lignes.append({
                "code": b["code"], "label": b["label"], "solde_ouverture": b["solde_ouverture"],
                "cumul_debit": b["debit"], "cumul_credit": b["credit"],
                "solde_debit": solde_debit, "solde_credit": solde_credit,
            })
            sous_total["cumul_debit"] += b["debit"]
            sous_total["cumul_credit"] += b["credit"]
            sous_total["solde_debit"] += solde_debit
            sous_total["solde_credit"] += solde_credit
        result_classes.append({"classe": classe, "lignes": lignes, "sous_total": sous_total})
        for k in grand:
            grand[k] += sous_total[k]

    return {"classes": result_classes, "grand_total": grand}


def compute_tresorerie_detail(conn, exercice=None):
    """Détail de la trésorerie (classe 5) par compte réel — ex. chaque banque
    séparément (521110 WENDKUNI BANK, 521120 CORIS BANK...) — calculé à
    partir de la même compute_balance() que le Bilan et la Balance."""
    balance = compute_balance(conn, only_with_movement=True, exercice=exercice)
    lignes = [b for b in balance if b["classe"] == "5"]
    lignes.sort(key=lambda b: b["code"])
    total = sum(b["solde_cloture"] for b in lignes)
    return lignes, total


def _sum_accounts(balance, codes):
    """Somme Débit/Crédit pour tous les comptes dont le code COMMENCE PAR l'un
    des préfixes donnés (rétro-compatible : un préfixe de 6 chiffres ne
    matche que le compte exact ; un préfixe de 3 chiffres couvre aussi tous
    les sous-comptes détaillés, ex. « 602 » couvre 602000, 602101, 602102...)."""
    debit = credit = 0.0
    for b in balance:
        if any(b["code"].startswith(c) for c in codes):
            debit += b["debit"]
            credit += b["credit"]
    return debit, credit


def _sum_accounts_cloture(balance, codes):
    """Somme des soldes de CLÔTURE (ouverture + mouvements) pour tous les
    comptes dont le code commence par l'un des préfixes donnés."""
    return sum(b["solde_cloture"] for b in balance if any(b["code"].startswith(c) for c in codes))


def _sum_class(balance, classe, sign=None, field="solde_cloture"):
    total = 0
    for b in balance:
        if b["classe"] != classe:
            continue
        v = b[field]
        if sign == "pos" and v <= 0:
            continue
        if sign == "neg" and v >= 0:
            continue
        total += v
    return total


def _sum_racine(balance, racine, sign=None, field="solde_cloture"):
    """Somme les soldes des comptes dont la racine (cf. account_racine) correspond,
    avec un filtre de signe optionnel (utilisé pour les racines « fourre-tout »
    42 à 49, dont la nature actif/passif dépend du solde effectif)."""
    total = 0
    for b in balance:
        if account_racine(b["code"]) != racine:
            continue
        v = b[field]
        if sign == "pos" and v <= 0:
            continue
        if sign == "neg" and v >= 0:
            continue
        total += v
    return total


# ---------------------------------------------------------------------------
# Compte de résultat
# ---------------------------------------------------------------------------
def compute_compte_resultat(conn, exercice=None):
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)

    def net_produit(codes):
        d, c = _sum_accounts(balance, codes)
        return c - d

    def net_charge(codes):
        d, c = _sum_accounts(balance, codes)
        return d - c

    produits = {
        "Ventes (marchandises, produits finis, travaux, services)": net_produit(COMPTES_PRODUITS_EXPL),
        "Subventions d'exploitation": net_produit([COMPTE_SUBV_EXPL]),
        "Autres produits": net_produit([COMPTE_AUTRES_PRODUITS]),
    }
    total_produits = sum(produits.values())

    charges = {
        "Achats (marchandises et matières)": net_charge(COMPTES_ACHATS),
        "Transports": net_charge(COMPTES_TRANSPORT),
        "Services extérieurs": net_charge(COMPTES_SERVICES_EXT),
        "Impôts et taxes": net_charge(COMPTES_IMPOTS),
        "Autres charges": net_charge([COMPTE_AUTRES_CHARGES]),
        "Charges de personnel": net_charge(COMPTES_PERSONNEL),
        "Dotations aux amortissements et provisions": net_charge(COMPTES_DOTATIONS),
    }
    total_charges = sum(charges.values())

    resultat_exploitation = total_produits - total_charges

    produits_fin = net_produit(COMPTES_PRODUITS_FIN)
    charges_fin = net_charge(COMPTES_CHARGES_FIN)
    resultat_financier = produits_fin - charges_fin

    resultat_net = resultat_exploitation + resultat_financier

    return {
        "produits": produits, "total_produits": total_produits,
        "charges": charges, "total_charges": total_charges,
        "resultat_exploitation": resultat_exploitation,
        "produits_financiers": produits_fin, "charges_financieres": charges_fin,
        "resultat_financier": resultat_financier,
        "resultat_net": resultat_net,
    }


# ---------------------------------------------------------------------------
# Bilan
# ---------------------------------------------------------------------------
def compute_bilan(conn, stock_initial=0.0, exercice=None):
    """stock_initial : conservé pour compatibilité, normalement inutile désormais —
    utilisez la table des soldes d'ouverture (onglet « Soldes d'ouverture »)."""
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    cr = compute_compte_resultat(conn, exercice=exercice)

    immo_brutes = sum(b["solde_cloture"] for b in balance if b["classe"] == "2" and int(b["code"]) < 280000)
    amortissements = sum(b["solde_cloture"] for b in balance if b["classe"] == "2" and int(b["code"]) >= 280000)
    immo_nettes = immo_brutes + amortissements

    stocks = stock_initial + _sum_accounts_cloture(balance, COMPTES_STOCK_PREFIXES)

    # Comptes de tiers (classe 4) classés par racine plutôt que par simple signe :
    # racine 41 (Clients) toujours en créances, racine 40 (Fournisseurs) toujours
    # en dettes — les autres racines (42 Personnel à 49 Dépréciations) restent
    # classées par signe du solde, car leur nature actif/passif varie selon le cas.
    autres_racines_tiers = [str(r) for r in range(42, 50)]
    creances_clients = _sum_racine(balance, RACINE_CLIENTS)
    autres_creances = sum(_sum_racine(balance, r, sign="pos") for r in autres_racines_tiers)
    creances = creances_clients + autres_creances

    dettes_fournisseurs = -_sum_racine(balance, RACINE_FOURNISSEURS)
    autres_dettes_tiers = sum(-_sum_racine(balance, r, sign="neg") for r in autres_racines_tiers)
    dettes_circulantes = dettes_fournisseurs + autres_dettes_tiers

    treso_actif = _sum_class(balance, "5", sign="pos")
    total_actif = immo_nettes + stocks + creances + treso_actif

    capital = _sum_accounts_cloture(balance, COMPTES_CAPITAL) * -1
    subventions = _sum_accounts_cloture(balance, [COMPTE_SUBVENTIONS]) * -1
    provisions = _sum_accounts_cloture(balance, [COMPTE_PROVISIONS]) * -1
    resultat_net = cr["resultat_net"]
    dettes_financieres = _sum_accounts_cloture(balance, COMPTES_DETTES_FIN) * -1
    treso_passif = -_sum_class(balance, "5", sign="neg")
    total_passif = (capital + subventions + provisions + resultat_net
                     + dettes_financieres + dettes_circulantes + treso_passif)

    return {
        "actif": {
            "Immobilisations brutes": immo_brutes,
            "Amortissements (à déduire)": amortissements,
            "Immobilisations nettes": immo_nettes,
            "Stocks": stocks,
            "Créances et emplois assimilés": creances,
            "Trésorerie actif": treso_actif,
        },
        "total_actif": total_actif,
        "passif": {
            "Capital et réserves": capital,
            "Subventions d'investissement": subventions,
            "Provisions pour risques et charges": provisions,
            "Résultat net de l'exercice": resultat_net,
            "Dettes financières": dettes_financieres,
            "Dettes circulantes": dettes_circulantes,
            "Trésorerie passif": treso_passif,
        },
        "total_passif": total_passif,
        "ecart": total_actif - total_passif,
    }


def compute_grand_livre(conn, compte, tiers=None, date_from=None, date_to=None, exercice=None):
    """Détail chronologique des écritures d'un compte pour un exercice, avec
    solde cumulé démarrant au solde d'ouverture de l'exercice."""
    exercice = exercice or get_current_exercice(conn)
    if date_from is None:
        date_from = f"{exercice}-01-01"
    if date_to is None:
        date_to = f"{exercice}-12-31"
    query = "SELECT * FROM entries WHERE compte = ? AND date >= ? AND date <= ?"
    params = [compte, date_from, date_to]
    if tiers:
        query += " AND tiers LIKE ?"
        params.append(f"%{tiers}%")
    query += " ORDER BY date, id"
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    solde = get_opening_balance(conn, compte, exercice)
    for r in rows:
        solde += r["debit"] - r["credit"]
        r["solde_cumule"] = solde
    return rows


# ---------------------------------------------------------------------------
# Stocks
# ---------------------------------------------------------------------------
def compute_stocks(conn, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    by_code = {b["code"]: b for b in balance}
    result = []
    for code in COMPTES_STOCK:
        b = by_code.get(code, {"label": get_account_label(conn, code), "debit": 0.0, "credit": 0.0,
                                "solde_ouverture": 0.0})
        initial = get_opening_balance(conn, code, exercice)
        entrees, sorties = b["debit"], b["credit"]
        qte_row = conn.execute(
            """SELECT COALESCE(SUM(CASE WHEN debit > 0 THEN quantite ELSE 0 END), 0) AS qte_in,
                      COALESCE(SUM(CASE WHEN credit > 0 THEN quantite ELSE 0 END), 0) AS qte_out
               FROM entries WHERE compte = ? AND date >= ? AND date <= ?""",
            (code, date_from, date_to),
        ).fetchone()
        qte_entrees, qte_sorties = qte_row["qte_in"], qte_row["qte_out"]
        qte_initiale = get_setting(conn, f"stock_qte_initiale_{code}_{exercice}", 0.0)
        qte_finale = qte_initiale + qte_entrees - qte_sorties
        stock_final = initial + entrees - sorties
        cout_unitaire_moyen = (stock_final / qte_finale) if qte_finale else None
        result.append({
            "code": code, "label": b["label"], "stock_initial": initial,
            "entrees": entrees, "sorties": sorties,
            "stock_final": stock_final,
            "qte_initiale": qte_initiale, "qte_entrees": qte_entrees, "qte_sorties": qte_sorties,
            "qte_finale": qte_finale, "cout_unitaire_moyen": cout_unitaire_moyen,
        })
    return result


def compute_stocks_detail(conn, exercice=None, prefixes=None):
    """Détail du stock pour CHAQUE compte réel de la classe 3 (pas seulement les
    4 comptes centralisateurs) : tout compte 3xxxxx ayant un mouvement ou un
    solde d'ouverture sur l'exercice. `prefixes` (ex. ["31"], ["32"], ["36"])
    restreint optionnellement à une catégorie (marchandises/matières/produits
    finis)."""
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    balance = compute_balance(conn, only_with_movement=True, exercice=exercice)
    result = []
    for b in balance:
        if b["classe"] != "3":
            continue
        if prefixes and not any(b["code"].startswith(p) for p in prefixes):
            continue
        code = b["code"]
        initial = b["solde_ouverture"]
        entrees, sorties = b["debit"], b["credit"]
        qte_row = conn.execute(
            """SELECT COALESCE(SUM(CASE WHEN debit > 0 THEN quantite ELSE 0 END), 0) AS qte_in,
                      COALESCE(SUM(CASE WHEN credit > 0 THEN quantite ELSE 0 END), 0) AS qte_out
               FROM entries WHERE compte = ? AND date >= ? AND date <= ?""",
            (code, date_from, date_to),
        ).fetchone()
        qte_entrees, qte_sorties = qte_row["qte_in"], qte_row["qte_out"]
        qte_initiale = get_setting(conn, f"stock_qte_initiale_{code}_{exercice}", 0.0)
        qte_finale = qte_initiale + qte_entrees - qte_sorties
        stock_final = b["solde_cloture"]
        cout_unitaire_moyen = (stock_final / qte_finale) if qte_finale else None
        result.append({
            "code": code, "label": b["label"], "stock_initial": initial,
            "entrees": entrees, "sorties": sorties, "stock_final": stock_final,
            "qte_initiale": qte_initiale, "qte_entrees": qte_entrees, "qte_sorties": qte_sorties,
            "qte_finale": qte_finale, "cout_unitaire_moyen": cout_unitaire_moyen,
        })
    result.sort(key=lambda r: r["code"])
    return result


def set_stock_qte_initiale(conn, code, value, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    set_setting(conn, f"stock_qte_initiale_{code}_{exercice}", value)


def set_stock_initial(conn, code, value, exercice=None):
    set_opening_balance(conn, code, value, exercice=exercice)


def compute_mouvements_stocks(conn, exercice=None):
    """Détail chronologique de toutes les écritures sur les comptes de stock
    (classe 3), quelle que soit leur origine — saisie manuelle, ou générées
    automatiquement par la validation d'une Facture (vente) ou d'une Facture
    frs (achat). Pour chaque compte, les lignes sont triées par date et un
    cumul est tenu à la fois en VALEUR (solde du compte) et en QUANTITÉ
    (comme une fiche de stock), en partant du stock initial de l'exercice."""
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    stocks_synthese = {s["code"]: s for s in compute_stocks(conn, exercice=exercice)}

    result = []
    for code in COMPTES_STOCK:
        rows = conn.execute(
            """SELECT e.*, a.label AS compte_label FROM entries e
               JOIN accounts a ON a.code = e.compte
               WHERE e.compte = ? AND e.date >= ? AND e.date <= ?
               ORDER BY e.date, e.id""",
            (code, date_from, date_to),
        ).fetchall()
        synth = stocks_synthese.get(code, {})
        valeur_cumulee = synth.get("stock_initial", 0.0)
        qte_cumulee = synth.get("qte_initiale", 0.0)
        for r in rows:
            d = dict(r)
            libelle = d["libelle"] or ""
            if libelle.startswith("Entrée stock (auto) —") or libelle.startswith("Sortie stock (auto) —"):
                d["origine"] = "Saisie directe (auto)"
            elif libelle.startswith("Entrée stock —") or libelle.startswith("Sortie stock —"):
                d["origine"] = "Facturation" if libelle.startswith("Sortie stock —") else "Facture frs"
            else:
                d["origine"] = "Saisie manuelle"
            valeur_cumulee += (d["debit"] or 0) - (d["credit"] or 0)
            qte_cumulee += (d["quantite"] or 0) if (d["debit"] or 0) > 0 else -(d["quantite"] or 0)
            d["valeur_cumulee"] = valeur_cumulee
            d["qte_cumulee"] = qte_cumulee
            result.append(d)
    return result


# ---------------------------------------------------------------------------
# Production / coûts de fabrication (écritures taguées analytic_code = AN-FAB)
# ---------------------------------------------------------------------------
FLUX_FAB = "AN-FAB"
FAB_POSTES = [
    ("Matières premières et fournitures consommées", ["602", "604"]),
    ("Main-d'œuvre directe de production", ["661", "663", "664"]),
    ("Charges indirectes de fabrication", ["624", "625", "681"]),
]


def compute_production(conn, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)

    def net_produit(codes):
        d, c = _sum_accounts(balance, codes)
        return c - d

    ventes = net_produit(["702", "705", "706"])
    stock_d, stock_c = _sum_accounts(balance, ["360"])
    production_stockee = stock_d - stock_c
    valeur_production = ventes + production_stockee

    postes = []
    total_cout = 0.0
    for label, codes in FAB_POSTES:
        like_clause = " OR ".join("compte LIKE ?" for _ in codes)
        like_params = [f"{c}%" for c in codes]
        row = conn.execute(
            f"SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries "
            f"WHERE ({like_clause}) AND analytic_code = ? AND date >= ? AND date <= ?",
            (*like_params, FLUX_FAB, date_from, date_to),
        ).fetchone()
        montant = row["d"] - row["c"]
        postes.append({"label": label, "comptes": ", ".join(codes), "montant": montant})
        total_cout += montant

    return {
        "ventes": ventes,
        "production_stockee": production_stockee,
        "valeur_production": valeur_production,
        "postes_cout": postes,
        "cout_production": total_cout,
        "marge": valeur_production - total_cout,
    }


# ---------------------------------------------------------------------------
# Recettes de fabrication (nomenclature / BOM) — combine matières premières
# (coût réel issu des stocks comptables), main-d'œuvre et énergie pour
# calculer un coût de production, puis un prix de vente suggéré (+ marge).
# ---------------------------------------------------------------------------
LIGNE_TYPES = {
    "matiere": "Matière première (depuis un compte de stock)",
    "main_oeuvre": "Main-d'œuvre",
    "energie": "Énergie",
    "autre": "Autre charge de fabrication",
}


def add_produit_fini(conn, code, nom, description="", quantite_produite=1, marge_pourcentage=30,
                      compte_stock="360000"):
    conn.execute(
        """INSERT OR REPLACE INTO produits_finis
           (code, nom, description, quantite_produite, marge_pourcentage, compte_stock)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (code.strip(), nom.strip(), description, quantite_produite or 1, marge_pourcentage or 0,
         compte_stock or "360000"),
    )
    conn.commit()


def delete_produit_fini(conn, code):
    conn.execute("DELETE FROM recette_lignes WHERE produit_code = ?", (code,))
    conn.execute("DELETE FROM produits_finis WHERE code = ?", (code,))
    conn.commit()


def list_produits_finis(conn):
    return [dict(r) for r in conn.execute("SELECT * FROM produits_finis ORDER BY code").fetchall()]


def get_produit_fini(conn, code):
    row = conn.execute("SELECT * FROM produits_finis WHERE code = ?", (code,)).fetchone()
    return dict(row) if row else None


def add_recette_ligne(conn, produit_code, type_ligne, libelle, quantite, compte=None, cout_unitaire=None):
    conn.execute(
        """INSERT INTO recette_lignes (produit_code, type_ligne, libelle, compte, quantite, cout_unitaire)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (produit_code, type_ligne, libelle, compte, quantite or 0, cout_unitaire),
    )
    conn.commit()


def delete_recette_ligne(conn, ligne_id):
    conn.execute("DELETE FROM recette_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_recette_lignes(conn, produit_code):
    return [dict(r) for r in conn.execute(
        "SELECT * FROM recette_lignes WHERE produit_code = ? ORDER BY id", (produit_code,)
    ).fetchall()]


STOCK_VARIATION_PAR_PREFIXE = {
    "31": "603100",  # Variations des stocks de marchandises
    "32": "603200",  # Variations des stocks de matières premières
    "33": "603300",  # Variations des stocks d'autres approvisionnements
    "36": "736000",  # Variations des stocks de produits finis
}


def _compte_variation_stock(compte_stock):
    prefix = (compte_stock or "")[:2]
    return STOCK_VARIATION_PAR_PREFIXE.get(prefix, "603200")


def compute_cout_production(conn, produit_code, exercice=None):
    """Calcule le coût de production d'un produit fini à partir de sa recette :
    pour chaque ligne « matière première » liée à un compte de stock, le coût
    unitaire réel est repris automatiquement du coût unitaire moyen calculé
    dans l'onglet Stocks (valeur du stock / quantité) — sinon le coût unitaire
    saisi manuellement sur la ligne (main-d'œuvre, énergie, autre) est utilisé."""
    produit = get_produit_fini(conn, produit_code)
    if not produit:
        raise ValueError(f"Produit « {produit_code} » introuvable.")
    lignes = list_recette_lignes(conn, produit_code)
    stocks_by_code = {s["code"]: s for s in compute_stocks_detail(conn, exercice=exercice)}

    detail = []
    total = 0.0
    for l in lignes:
        cu = l["cout_unitaire"]
        source = "manuel"
        if l["type_ligne"] == "matiere" and l["compte"]:
            stock = stocks_by_code.get(l["compte"])
            if stock and stock["cout_unitaire_moyen"] is not None:
                cu = stock["cout_unitaire_moyen"]
                source = "stock (coût unitaire moyen)"
            elif cu is None:
                cu = 0.0
                source = "aucun coût connu — à saisir"
        elif cu is None:
            cu = 0.0
            source = "à saisir"
        montant = (l["quantite"] or 0) * (cu or 0)
        detail.append({**l, "cout_unitaire_utilise": cu, "source_cout": source, "montant": montant})
        total += montant

    qte_produite = produit["quantite_produite"] or 1
    cout_unitaire_produit = total / qte_produite if qte_produite else 0.0
    marge_pct = produit["marge_pourcentage"] or 0
    prix_vente_unitaire = cout_unitaire_produit * (1 + marge_pct / 100)
    prix_vente_total = prix_vente_unitaire * qte_produite

    return {
        "produit": produit,
        "lignes": detail,
        "cout_production_total": total,
        "quantite_produite": qte_produite,
        "cout_unitaire_produit": cout_unitaire_produit,
        "marge_pourcentage": marge_pct,
        "prix_vente_unitaire": prix_vente_unitaire,
        "prix_vente_total": prix_vente_total,
        "marge_unitaire": prix_vente_unitaire - cout_unitaire_produit,
    }


def valider_fabrication(conn, produit_code, date_str=None, piece=None, exercice=None):
    """Valide une fabrication à partir de sa recette :
    - impute comptablement la consommation de chaque matière première (le
      stock réel utilisé — ex. 321001 CLINKER — diminue en QUANTITÉ et en
      VALEUR, contrepartie en compte de variation de stock 603xxx) ;
    - place le produit fini dans son compte de stock (classe 36) en QUANTITÉ
      et en VALEUR, au coût de production + la marge paramétrée (compte
      736000 en contrepartie).
    Retourne (résultat de compute_cout_production, avertissements)."""
    resultat = compute_cout_production(conn, produit_code, exercice=exercice)
    produit = resultat["produit"]
    exercice = exercice or get_current_exercice(conn)
    if date_str is None:
        today = date.today()
        date_str = today.strftime("%Y-%m-%d") if str(today.year) == exercice else f"{exercice}-01-01"
    piece = piece or f"FAB-{produit_code}-{date_str}"
    warnings = []

    for l in resultat["lignes"]:
        if l["type_ligne"] != "matiere" or not l["compte"]:
            continue
        montant = l["montant"]
        qte = l["quantite"] or 0
        if montant <= 0 or qte <= 0:
            continue
        contre_compte = _compte_variation_stock(l["compte"])
        add_entry(conn, date_str, piece, "OD", contre_compte, "", f"Consommation fabrication — {l['libelle']}",
                  montant, 0)
        add_entry(conn, date_str, piece, "OD", l["compte"], "", f"Consommation fabrication — {l['libelle']}",
                  0, montant, quantite=qte)

    valeur_produit_fini = resultat["prix_vente_total"]
    qte_produite = resultat["quantite_produite"]
    if valeur_produit_fini > 0 and qte_produite > 0:
        compte_stock_pf = produit["compte_stock"]
        contre_compte_pf = _compte_variation_stock(compte_stock_pf)
        add_entry(conn, date_str, piece, "OD", compte_stock_pf, "", f"Production — {produit['nom']}",
                  valeur_produit_fini, 0, quantite=qte_produite)
        add_entry(conn, date_str, piece, "OD", contre_compte_pf, "", f"Production — {produit['nom']}",
                  0, valeur_produit_fini)
    else:
        warnings.append("Valeur ou quantité produite nulle — aucune entrée en stock de produit fini comptabilisée.")

    return resultat, warnings


# ---------------------------------------------------------------------------
# Facturation clients — présente une facture (entête + lignes + pied de page),
# et sa validation envoie automatiquement les écritures comptables dans la
# Saisie : Débit Client (411xxx) pour le TTC, Crédit compte(s) de vente (70x)
# pour le HT de chaque ligne, Crédit TVA (443100) pour la taxe, et pour les
# lignes liées à un stock (marchandises 31 ou produits finis 36), une sortie
# de stock automatique (Débit compte de coût / Crédit compte de stock).
# ---------------------------------------------------------------------------
def create_facture_vente(conn, numero, date_facture, client_code, entete="", pied_page="",
                          tva_taux=None):
    if tva_taux is None:
        tva_taux = get_setting(conn, "tva_taux_defaut", TVA_TAUX_DEFAUT)
    cur = conn.execute(
        """INSERT INTO factures_vente (numero, date_facture, client_code, entete, pied_page, tva_taux, statut)
           VALUES (?, ?, ?, ?, ?, ?, 'brouillon')""",
        (numero, date_facture, client_code, entete, pied_page, tva_taux),
    )
    conn.commit()
    return cur.lastrowid


def update_facture_vente(conn, facture_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE factures_vente SET {cols} WHERE id = ?", (*fields.values(), facture_id))
    conn.commit()


def delete_facture_vente(conn, facture_id):
    facture = get_facture_vente(conn, facture_id)
    if facture and facture["statut"] == "validee":
        raise ValueError("Impossible de supprimer une facture déjà validée (écritures envoyées en Saisie).")
    conn.execute("DELETE FROM facture_vente_lignes WHERE facture_id = ?", (facture_id,))
    conn.execute("DELETE FROM factures_vente WHERE id = ?", (facture_id,))
    conn.commit()


def get_facture_vente(conn, facture_id):
    row = conn.execute("SELECT * FROM factures_vente WHERE id = ?", (facture_id,)).fetchone()
    return dict(row) if row else None


def list_factures_vente(conn):
    rows = conn.execute("""
        SELECT f.*, COALESCE(c.raison_sociale, f.client_code) AS raison_sociale
        FROM factures_vente f LEFT JOIN clients c ON c.code = f.client_code
        ORDER BY f.date_facture DESC, f.id DESC
    """).fetchall()
    return [dict(r) for r in rows]


def add_ligne_facture_vente(conn, facture_id, compte_vente, libelle, quantite, prix_unitaire):
    conn.execute(
        """INSERT INTO facture_vente_lignes (facture_id, compte_vente, libelle, quantite, prix_unitaire)
           VALUES (?, ?, ?, ?, ?)""",
        (facture_id, compte_vente, libelle, quantite or 0, prix_unitaire or 0),
    )
    conn.commit()


def delete_ligne_facture_vente(conn, ligne_id):
    conn.execute("DELETE FROM facture_vente_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_lignes_facture_vente(conn, facture_id):
    rows = conn.execute(
        "SELECT * FROM facture_vente_lignes WHERE facture_id = ? ORDER BY id", (facture_id,)
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["montant_ht"] = (d["quantite"] or 0) * (d["prix_unitaire"] or 0)
        type_stock, stock_compte, cout_compte = _match_stock_mapping(d["compte_vente"], VENTE_STOCK_MAPPING) or (None, None, None)
        d["type_stock"] = type_stock
        d["stock_compte"] = stock_compte
        result.append(d)
    return result


def compute_facture_totals(conn, facture_id):
    facture = get_facture_vente(conn, facture_id)
    lignes = list_lignes_facture_vente(conn, facture_id)
    total_ht = sum(l["montant_ht"] for l in lignes)
    tva_taux = facture["tva_taux"] if facture else 0
    tva_montant = total_ht * (tva_taux or 0) / 100
    total_ttc = total_ht + tva_montant
    return {"total_ht": total_ht, "tva_taux": tva_taux, "tva_montant": tva_montant, "total_ttc": total_ttc}


def valider_facture_vente(conn, facture_id, exercice=None):
    """Envoie la facture en Saisie : une écriture équilibrée (Débit Client / Crédit
    ventes + TVA), plus une sortie de stock automatique pour chaque ligne liée à un
    compte de marchandises (31) ou de produits finis (36). Retourne la liste des
    avertissements (ex. coût unitaire de stock inconnu)."""
    facture = get_facture_vente(conn, facture_id)
    if not facture:
        raise ValueError("Facture introuvable.")
    if facture["statut"] == "validee":
        raise ValueError("Cette facture est déjà validée.")
    lignes = list_lignes_facture_vente(conn, facture_id)
    if not lignes:
        raise ValueError("La facture ne contient aucune ligne.")
    if not client_exists(conn, facture["client_code"]):
        raise ValueError(f"Le client « {facture['client_code']} » n'existe pas.")

    totals = compute_facture_totals(conn, facture_id)
    date_str = facture["date_facture"]
    piece = facture["numero"]
    warnings = []

    # Débit Client pour le TTC
    client = get_client(conn, facture["client_code"])
    tiers_label = client["raison_sociale"] if client else facture["client_code"]
    add_entry(conn, date_str, piece, "VE", "411000", tiers_label,
              facture["numero"], totals["total_ttc"], 0, client_code=facture["client_code"])

    # Crédit chaque compte de vente pour le HT de la ligne
    for l in lignes:
        add_entry(conn, date_str, piece, "VE", l["compte_vente"], "", l["libelle"],
                  0, l["montant_ht"], client_code=facture["client_code"], quantite=l["quantite"])

    # Crédit TVA facturée
    if totals["tva_montant"]:
        add_entry(conn, date_str, piece, "VE", COMPTE_TVA_VENTES, "", f"TVA {totals['tva_taux']:g}% facture {piece}",
                  0, totals["tva_montant"])

    # Sortie de stock automatique pour les lignes liées aux marchandises/produits finis
    stocks_by_code = {s["code"]: s for s in compute_stocks(conn, exercice=exercice)}
    for l in lignes:
        if not l["type_stock"]:
            continue
        _, stock_compte, cout_compte = _match_stock_mapping(l["compte_vente"], VENTE_STOCK_MAPPING)
        stock = stocks_by_code.get(stock_compte)
        cout_unitaire = stock["cout_unitaire_moyen"] if stock else None
        if cout_unitaire is None:
            warnings.append(
                f"Ligne « {l['libelle']} » : coût unitaire du stock {stock_compte} inconnu — "
                f"aucune sortie de stock comptabilisée pour cette ligne (renseignez un stock initial "
                f"ou des entrées avec quantité dans l'onglet Stocks)."
            )
            continue
        montant_sortie = (l["quantite"] or 0) * cout_unitaire
        if montant_sortie <= 0:
            continue
        add_entry(conn, date_str, piece, "VE", cout_compte, "", f"Sortie stock — {l['libelle']}",
                  montant_sortie, 0)
        add_entry(conn, date_str, piece, "VE", stock_compte, "", f"Sortie stock — {l['libelle']}",
                  0, montant_sortie, quantite=l["quantite"])

    update_facture_vente(conn, facture_id, statut="validee", piece=piece)
    return warnings


# ---------------------------------------------------------------------------
# Factures fournisseurs (achats) — présente une facture d'achat (entête +
# lignes + pied de page), et sa validation envoie automatiquement les
# écritures comptables dans la Saisie : Débit compte(s) d'achat (60x) pour le
# HT de chaque ligne, Crédit Fournisseur (401xxx) pour le net à payer, Crédit
# retenue à la source (44x, paramétrable) le cas échéant. Pour les lignes
# liées à un stock (marchandises 31 ou matières premières 32), une entrée de
# stock automatique est comptabilisée (le stock augmente).
# ---------------------------------------------------------------------------
def create_facture_achat(conn, numero, date_facture, fournisseur_code, entete="", pied_page="",
                          retenue_taux=None, retenue_compte=None):
    if retenue_taux is None:
        retenue_taux = get_setting(conn, "retenue_taux_defaut", RETENUE_TAUX_DEFAUT)
    if retenue_compte is None:
        retenue_compte = get_text_setting(conn, "retenue_compte_defaut", COMPTE_RETENUE_DEFAUT)
    cur = conn.execute(
        """INSERT INTO factures_achat
           (numero, date_facture, fournisseur_code, entete, pied_page, retenue_taux, retenue_compte, statut)
           VALUES (?, ?, ?, ?, ?, ?, ?, 'brouillon')""",
        (numero, date_facture, fournisseur_code, entete, pied_page, retenue_taux, retenue_compte),
    )
    conn.commit()
    return cur.lastrowid


def update_facture_achat(conn, facture_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE factures_achat SET {cols} WHERE id = ?", (*fields.values(), facture_id))
    conn.commit()


def delete_facture_achat(conn, facture_id):
    facture = get_facture_achat(conn, facture_id)
    if facture and facture["statut"] == "validee":
        raise ValueError("Impossible de supprimer une facture déjà validée (écritures envoyées en Saisie).")
    conn.execute("DELETE FROM facture_achat_lignes WHERE facture_id = ?", (facture_id,))
    conn.execute("DELETE FROM factures_achat WHERE id = ?", (facture_id,))
    conn.commit()


def get_facture_achat(conn, facture_id):
    row = conn.execute("SELECT * FROM factures_achat WHERE id = ?", (facture_id,)).fetchone()
    return dict(row) if row else None


def list_factures_achat(conn):
    rows = conn.execute("""
        SELECT f.*, COALESCE(fo.raison_sociale, f.fournisseur_code) AS raison_sociale
        FROM factures_achat f LEFT JOIN fournisseurs fo ON fo.code = f.fournisseur_code
        ORDER BY f.date_facture DESC, f.id DESC
    """).fetchall()
    return [dict(r) for r in rows]


def add_ligne_facture_achat(conn, facture_id, compte_achat, libelle, quantite, prix_unitaire):
    conn.execute(
        """INSERT INTO facture_achat_lignes (facture_id, compte_achat, libelle, quantite, prix_unitaire)
           VALUES (?, ?, ?, ?, ?)""",
        (facture_id, compte_achat, libelle, quantite or 0, prix_unitaire or 0),
    )
    conn.commit()


def delete_ligne_facture_achat(conn, ligne_id):
    conn.execute("DELETE FROM facture_achat_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_lignes_facture_achat(conn, facture_id):
    rows = conn.execute(
        "SELECT * FROM facture_achat_lignes WHERE facture_id = ? ORDER BY id", (facture_id,)
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["montant_ht"] = (d["quantite"] or 0) * (d["prix_unitaire"] or 0)
        type_stock, stock_compte, contre_compte = _match_stock_mapping(d["compte_achat"], ACHAT_STOCK_MAPPING) or (None, None, None)
        d["type_stock"] = type_stock
        d["stock_compte"] = stock_compte
        result.append(d)
    return result


def compute_facture_achat_totals(conn, facture_id):
    facture = get_facture_achat(conn, facture_id)
    lignes = list_lignes_facture_achat(conn, facture_id)
    total_ht = sum(l["montant_ht"] for l in lignes)
    retenue_taux = facture["retenue_taux"] if facture else 0
    retenue_montant = total_ht * (retenue_taux or 0) / 100
    net_a_payer = total_ht - retenue_montant
    return {"total_ht": total_ht, "retenue_taux": retenue_taux, "retenue_montant": retenue_montant,
            "net_a_payer": net_a_payer}


def valider_facture_achat(conn, facture_id, exercice=None):
    """Envoie la facture d'achat en Saisie : une écriture équilibrée (Débit achats /
    Crédit fournisseur + retenue), plus une entrée de stock automatique pour chaque
    ligne liée à un compte de marchandises (31) ou de matières premières (32).
    Retourne la liste des avertissements."""
    facture = get_facture_achat(conn, facture_id)
    if not facture:
        raise ValueError("Facture introuvable.")
    if facture["statut"] == "validee":
        raise ValueError("Cette facture est déjà validée.")
    lignes = list_lignes_facture_achat(conn, facture_id)
    if not lignes:
        raise ValueError("La facture ne contient aucune ligne.")
    if not fournisseur_exists(conn, facture["fournisseur_code"]):
        raise ValueError(f"Le fournisseur « {facture['fournisseur_code']} » n'existe pas.")

    totals = compute_facture_achat_totals(conn, facture_id)
    date_str = facture["date_facture"]
    piece = facture["numero"]
    warnings = []

    # Débit chaque compte d'achat pour le HT de la ligne
    for l in lignes:
        add_entry(conn, date_str, piece, "AC", l["compte_achat"], "", l["libelle"],
                  l["montant_ht"], 0, fournisseur_code=facture["fournisseur_code"], quantite=l["quantite"])

    # Crédit Fournisseur pour le net à payer (HT - retenue)
    fournisseur = get_fournisseur(conn, facture["fournisseur_code"])
    tiers_label = fournisseur["raison_sociale"] if fournisseur else facture["fournisseur_code"]
    add_entry(conn, date_str, piece, "AC", "401000", tiers_label,
              facture["numero"], 0, totals["net_a_payer"], fournisseur_code=facture["fournisseur_code"])

    # Crédit retenue fiscale à la source, si applicable
    if totals["retenue_montant"]:
        add_entry(conn, date_str, piece, "AC", facture["retenue_compte"], "",
                  f"Retenue {totals['retenue_taux']:g}% facture {piece}",
                  0, totals["retenue_montant"])

    # Entrée de stock automatique pour les lignes liées aux marchandises/matières premières
    for l in lignes:
        if not l["type_stock"]:
            continue
        _, stock_compte, contre_compte = _match_stock_mapping(l["compte_achat"], ACHAT_STOCK_MAPPING)
        montant_entree = l["montant_ht"]
        if montant_entree <= 0:
            continue
        add_entry(conn, date_str, piece, "AC", stock_compte, "", f"Entrée stock — {l['libelle']}",
                  montant_entree, 0, quantite=l["quantite"])
        add_entry(conn, date_str, piece, "AC", contre_compte, "", f"Entrée stock — {l['libelle']}",
                  0, montant_entree)

    update_facture_achat(conn, facture_id, statut="validee", piece=piece)
    return warnings


def compute_tft(conn, treso_ouverture=None, exercice=None):
    """treso_ouverture=None : dérivée automatiquement des soldes d'ouverture des
    comptes de trésorerie (521000/531000/570000/585000). Passez une valeur pour
    la forcer manuellement."""
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    if treso_ouverture is None:
        treso_ouverture = _sum_accounts_cloture(
            [dict(b, solde_cloture=b["solde_ouverture"]) for b in balance], COMPTES_TRESORERIE)
    treso_debit, treso_credit = _sum_accounts(balance, COMPTES_TRESORERIE)
    variation_totale = treso_debit - treso_credit

    def flux(code):
        like_clause = " OR ".join("compte LIKE ?" for _ in COMPTES_TRESORERIE)
        like_params = [f"{p}%" for p in COMPTES_TRESORERIE]
        rows = conn.execute(
            f"SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries "
            f"WHERE ({like_clause}) AND flux_code = ? AND date >= ? AND date <= ?",
            (*like_params, code, date_from, date_to),
        ).fetchone()
        return rows["d"] - rows["c"]

    exploitation = flux("FLUX-EXP")
    investissement = flux("FLUX-INV")
    financement = flux("FLUX-FIN")
    non_classes = variation_totale - (exploitation + investissement + financement)

    cloture = treso_ouverture + variation_totale
    return {
        "ouverture": treso_ouverture,
        "exploitation": exploitation,
        "investissement": investissement,
        "financement": financement,
        "non_classes": non_classes,
        "variation": variation_totale,
        "cloture": cloture,
    }


def compute_tft_officiel(conn, exercice=None):
    """TFT selon la disposition EXACTE du formulaire officiel SYSCOHADA
    (références ZA, FA à FE, ZB...), pour remplissage direct de la feuille
    TFT de la Liasse fiscale. Réutilise les mêmes briques que
    compute_tft_indirect / compute_situation_financiere — donc cohérent
    avec la Balance. Seule la partie confirmée visuellement (ZA, FA-FE) est
    actuellement mappée ; l'investissement et le financement restent à
    positionner une fois les numéros de ligne du modèle officiel confirmés."""
    exercice = exercice or get_current_exercice(conn)
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    cr = compute_liasse_resultat(conn, exercice=exercice)

    treso_ouverture = sum(b["solde_ouverture"] for b in balance if b["classe"] == "5")
    treso_cloture_reelle = sum(b["solde_cloture"] for b in balance if b["classe"] == "5")

    ebe = cr["XD"]
    revenus_financiers = cr["TK"]
    frais_financiers = -cr["RM"]
    fa_cafg = ebe + revenus_financiers + frais_financiers

    def _delta_prefixes(prefixes):
        ouv = sum(b["solde_ouverture"] for b in balance if any(b["code"].startswith(p) for p in prefixes))
        clo = sum(b["solde_cloture"] for b in balance if any(b["code"].startswith(p) for p in prefixes))
        return ouv, clo

    stock_ouv, stock_clo = _delta_prefixes(COMPTES_STOCK_PREFIXES)
    fc_variation_stocks = -(stock_clo - stock_ouv)

    racines_exploit = ["42", "43", "44", "45", "46"]
    creances_ouv = sum(b["solde_ouverture"] for b in balance if account_racine(b["code"]) == RACINE_CLIENTS)
    creances_clo = sum(b["solde_cloture"] for b in balance if account_racine(b["code"]) == RACINE_CLIENTS)
    for r in racines_exploit:
        creances_ouv += sum(b["solde_ouverture"] for b in balance
                             if account_racine(b["code"]) == r and b["solde_ouverture"] > 0)
        creances_clo += sum(b["solde_cloture"] for b in balance
                             if account_racine(b["code"]) == r and b["solde_cloture"] > 0)
    fd_variation_creances = -(creances_clo - creances_ouv)

    dettes_ouv = -sum(b["solde_ouverture"] for b in balance if account_racine(b["code"]) == RACINE_FOURNISSEURS)
    dettes_clo = -sum(b["solde_cloture"] for b in balance if account_racine(b["code"]) == RACINE_FOURNISSEURS)
    for r in racines_exploit:
        dettes_ouv += -sum(b["solde_ouverture"] for b in balance
                            if account_racine(b["code"]) == r and b["solde_ouverture"] < 0)
        dettes_clo += -sum(b["solde_cloture"] for b in balance
                            if account_racine(b["code"]) == r and b["solde_cloture"] < 0)
    fe_variation_passif = dettes_clo - dettes_ouv

    racines_hao = ["47", "48", "49"]
    hao_actif_ouv = sum(b["solde_ouverture"] for b in balance
                         if account_racine(b["code"]) in racines_hao and b["solde_ouverture"] > 0)
    hao_actif_clo = sum(b["solde_cloture"] for b in balance
                         if account_racine(b["code"]) in racines_hao and b["solde_cloture"] > 0)
    hao_passif_ouv = -sum(b["solde_ouverture"] for b in balance
                          if account_racine(b["code"]) in racines_hao and b["solde_ouverture"] < 0)
    hao_passif_clo = -sum(b["solde_cloture"] for b in balance
                          if account_racine(b["code"]) in racines_hao and b["solde_cloture"] < 0)
    fb_variation_hao = -((hao_actif_clo - hao_actif_ouv) - (hao_passif_clo - hao_passif_ouv))

    zb_flux_operationnel = fa_cafg + fb_variation_hao + fc_variation_stocks + fd_variation_creances + fe_variation_passif

    return {
        "ZA": treso_ouverture,
        "FA": fa_cafg, "FB": fb_variation_hao, "FC": fc_variation_stocks,
        "FD": fd_variation_creances, "FE": fe_variation_passif,
        "ZB": zb_flux_operationnel,
        "treso_cloture_reelle": treso_cloture_reelle,
    }


def compute_tft_indirect(conn, exercice=None):
    """TFT selon la méthode indirecte SYSCOHADA (avec CAFG), au même format
    que le modèle officiel : A) trésorerie d'ouverture, détermination de la
    CAFG, variations du BFR (stocks/créances/dettes circulantes), flux
    d'investissement (acquisitions/cessions d'immobilisations), flux de
    financement (capital, subventions, emprunts). Entièrement calculé à
    partir de compute_balance() et compute_liasse_resultat() — donc toujours
    cohérent avec la Balance et le Bilan. La ligne CONTRÔLE compare la
    trésorerie calculée à la trésorerie réelle de la Balance (classe 5) :
    tout écart signale un mouvement de trésorerie non correctement classé."""
    exercice = exercice or get_current_exercice(conn)
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    cr = compute_liasse_resultat(conn, exercice=exercice)

    # ---- Trésorerie (classe 5 entière, comme la Balance) ----
    treso_ouverture = sum(b["solde_ouverture"] for b in balance if b["classe"] == "5")
    treso_cloture_reelle = sum(b["solde_cloture"] for b in balance if b["classe"] == "5")

    # ---- CAFG (à partir des soldes déjà calculés pour le Compte de résultat) ----
    ebe = cr["XD"]  # Excédent brut d'exploitation = Valeur ajoutée - charges de personnel
    revenus_financiers = cr["TK"]      # produits financiers (771, 776)
    frais_financiers = -cr["RM"]       # charges financières (671, 676), en décaissement
    cafg = ebe + revenus_financiers + frais_financiers

    # ---- Variation du BFR (comparaison ouverture/clôture, cohérente avec le Bilan) ----
    def _delta_racines(prefixes):
        ouverture = sum(b["solde_ouverture"] for b in balance if any(b["code"].startswith(p) for p in prefixes))
        cloture = sum(b["solde_cloture"] for b in balance if any(b["code"].startswith(p) for p in prefixes))
        return ouverture, cloture

    stock_ouv, stock_clo = _delta_racines(COMPTES_STOCK_PREFIXES)
    variation_stocks = -(stock_clo - stock_ouv)  # une hausse de stock consomme de la trésorerie

    # Créances : racine 41 (toujours créances) + autres racines de tiers débitrices (42-49)
    autres_racines_tiers = [str(r) for r in range(42, 50)]
    creances_ouv = sum(b["solde_ouverture"] for b in balance if account_racine(b["code"]) == RACINE_CLIENTS)
    creances_clo = sum(b["solde_cloture"] for b in balance if account_racine(b["code"]) == RACINE_CLIENTS)
    for r in autres_racines_tiers:
        creances_ouv += sum(b["solde_ouverture"] for b in balance
                             if account_racine(b["code"]) == r and b["solde_ouverture"] > 0)
        creances_clo += sum(b["solde_cloture"] for b in balance
                             if account_racine(b["code"]) == r and b["solde_cloture"] > 0)
    variation_creances = -(creances_clo - creances_ouv)  # une hausse de créances consomme de la trésorerie

    # Dettes circulantes : racine 40 (toujours dettes) + autres racines créditrices (42-49)
    dettes_ouv = -sum(b["solde_ouverture"] for b in balance if account_racine(b["code"]) == RACINE_FOURNISSEURS)
    dettes_clo = -sum(b["solde_cloture"] for b in balance if account_racine(b["code"]) == RACINE_FOURNISSEURS)
    for r in autres_racines_tiers:
        dettes_ouv += -sum(b["solde_ouverture"] for b in balance
                            if account_racine(b["code"]) == r and b["solde_ouverture"] < 0)
        dettes_clo += -sum(b["solde_cloture"] for b in balance
                            if account_racine(b["code"]) == r and b["solde_cloture"] < 0)
    variation_dettes_circulantes = dettes_clo - dettes_ouv  # une hausse de dettes fournit de la trésorerie

    flux_operationnel = cafg + variation_stocks + variation_creances + variation_dettes_circulantes

    # ---- Flux d'investissement (acquisitions = débit de l'exercice sur les comptes d'immobilisations) ----
    def _debit_classe(prefixes):
        d, c = _sum_accounts(balance, prefixes)
        return d, c

    incorp_debit, incorp_credit = _debit_classe(["20", "21"])
    corp_debit, corp_credit = _debit_classe(["22", "23", "24"])
    fin_debit, fin_credit = _debit_classe(["26", "27"])
    acquisitions_incorp = -incorp_debit
    acquisitions_corp = -corp_debit
    acquisitions_fin = -fin_debit
    cessions_incorp = incorp_credit  # rare (compte 21 crédité lors d'une cession/sortie)
    cessions_corp = corp_credit
    cessions_fin = fin_credit
    flux_investissement = (acquisitions_incorp + acquisitions_corp + acquisitions_fin
                            + cessions_incorp + cessions_corp + cessions_fin)

    # ---- Flux de financement ----
    capital_debit, capital_credit = _debit_classe(["101", "104", "105"])
    augmentation_capital = capital_credit
    prelevements_capital = -capital_debit
    subv_debit, subv_credit = _debit_classe(["14"])
    subventions_recues = subv_credit
    dividendes_verses = 0.0  # non isolé dans le plan comptable par défaut
    flux_capitaux_propres = augmentation_capital + subventions_recues + prelevements_capital + dividendes_verses

    emprunts_debit, emprunts_credit = _debit_classe(["16", "17"])
    emprunts_nouveaux = emprunts_credit
    remboursements_emprunts = -emprunts_debit
    flux_capitaux_etrangers = emprunts_nouveaux + remboursements_emprunts

    flux_financement = flux_capitaux_propres + flux_capitaux_etrangers

    variation_treso_nette = flux_operationnel + flux_investissement + flux_financement
    treso_cloture_calculee = treso_ouverture + variation_treso_nette
    ecart = treso_cloture_calculee - treso_cloture_reelle

    return {
        "treso_ouverture": treso_ouverture,
        "ebe": ebe, "revenus_financiers": revenus_financiers, "frais_financiers": frais_financiers,
        "cafg": cafg,
        "variation_stocks": variation_stocks, "variation_creances": variation_creances,
        "variation_dettes_circulantes": variation_dettes_circulantes,
        "flux_operationnel": flux_operationnel,
        "acquisitions_incorp": acquisitions_incorp, "acquisitions_corp": acquisitions_corp,
        "acquisitions_fin": acquisitions_fin,
        "cessions_incorp": cessions_incorp, "cessions_corp": cessions_corp, "cessions_fin": cessions_fin,
        "flux_investissement": flux_investissement,
        "augmentation_capital": augmentation_capital, "subventions_recues": subventions_recues,
        "prelevements_capital": prelevements_capital, "dividendes_verses": dividendes_verses,
        "flux_capitaux_propres": flux_capitaux_propres,
        "emprunts_nouveaux": emprunts_nouveaux, "remboursements_emprunts": remboursements_emprunts,
        "flux_capitaux_etrangers": flux_capitaux_etrangers,
        "flux_financement": flux_financement,
        "variation_treso_nette": variation_treso_nette,
        "treso_cloture_calculee": treso_cloture_calculee,
        "treso_cloture_reelle": treso_cloture_reelle,
        "ecart": ecart,
    }


def compute_situation_financiere(conn, exercice=None):
    """Situation financière (FR - BFR - TN), présentée selon le modèle
    officiel : capacité d'autofinancement, ratios de rentabilité, puis
    analyse Fonds de Roulement / Besoin en Fonds de Roulement / Trésorerie
    Nette. Entièrement recalculé à partir de compute_bilan(),
    compute_liasse_resultat() et compute_tft_indirect() — donc toujours
    cohérent avec la Balance, le Bilan et le TFT."""
    exercice = exercice or get_current_exercice(conn)
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    b = compute_bilan(conn, exercice=exercice)
    cr = compute_liasse_resultat(conn, exercice=exercice)
    tft = compute_tft_indirect(conn, exercice=exercice)

    resultat_net = cr["XI"]
    resultat_exploitation = cr["XE"]
    cafg = tft["cafg"]
    dividendes_verses = tft["dividendes_verses"]
    autofinancement = cafg + dividendes_verses

    capitaux_propres_ressources = (b["passif"]["Capital et réserves"] + b["passif"]["Subventions d'investissement"]
                                    + b["passif"]["Provisions pour risques et charges"] + resultat_net)
    dettes_financieres = b["passif"]["Dettes financières"]
    ressources_stables = capitaux_propres_ressources + dettes_financieres
    actifs_immobilises = b["actif"]["Immobilisations nettes"]
    fonds_de_roulement = ressources_stables - actifs_immobilises

    racines_exploit = ["42", "43", "44", "45", "46"]

    def _somme_racine(racine, sign=None):
        total = 0.0
        for x in balance:
            if account_racine(x["code"]) != racine:
                continue
            v = x["solde_cloture"]
            if sign == "pos" and v <= 0:
                continue
            if sign == "neg" and v >= 0:
                continue
            total += v
        return total

    creances_exploit = _somme_racine(RACINE_CLIENTS)
    for r in racines_exploit:
        creances_exploit += _somme_racine(r, sign="pos")
    actif_circulant_exploitation = b["actif"]["Stocks"] + creances_exploit

    dettes_exploit = -_somme_racine(RACINE_FOURNISSEURS)
    for r in racines_exploit:
        dettes_exploit += -_somme_racine(r, sign="neg")
    passif_circulant_exploitation = dettes_exploit

    besoin_financement_exploitation = actif_circulant_exploitation - passif_circulant_exploitation

    racines_hao = ["47", "48", "49"]
    actif_circulant_hao = sum(_somme_racine(r, sign="pos") for r in racines_hao)
    passif_circulant_hao = sum(-_somme_racine(r, sign="neg") for r in racines_hao)
    besoin_financement_hao = actif_circulant_hao - passif_circulant_hao

    besoin_financement_global = besoin_financement_exploitation + besoin_financement_hao
    tresorerie_nette = fonds_de_roulement - besoin_financement_global

    treso_actif = b["actif"]["Trésorerie actif"]
    treso_passif = b["passif"]["Trésorerie passif"]
    treso_reelle = treso_actif - treso_passif
    controle_ecart = tresorerie_nette - treso_reelle

    rentabilite_economique = (resultat_exploitation / capitaux_propres_ressources * 100
                               ) if capitaux_propres_ressources else 0.0
    rentabilite_financiere = (resultat_net / capitaux_propres_ressources * 100
                               ) if capitaux_propres_ressources else 0.0

    endettement_financier_brut = dettes_financieres + treso_passif
    endettement_financier_net = endettement_financier_brut - treso_actif

    return {
        "resultat_net_comptable": resultat_net,
        "ebe": tft["ebe"], "revenus_financiers": tft["revenus_financiers"],
        "frais_financiers": tft["frais_financiers"], "cafg": cafg,
        "dividendes_verses": dividendes_verses, "autofinancement": autofinancement,
        "rentabilite_economique": rentabilite_economique, "rentabilite_financiere": rentabilite_financiere,
        "capitaux_propres_ressources": capitaux_propres_ressources,
        "dettes_financieres": dettes_financieres, "ressources_stables": ressources_stables,
        "actifs_immobilises": actifs_immobilises, "fonds_de_roulement": fonds_de_roulement,
        "actif_circulant_exploitation": actif_circulant_exploitation,
        "passif_circulant_exploitation": passif_circulant_exploitation,
        "besoin_financement_exploitation": besoin_financement_exploitation,
        "actif_circulant_hao": actif_circulant_hao, "passif_circulant_hao": passif_circulant_hao,
        "besoin_financement_hao": besoin_financement_hao,
        "besoin_financement_global": besoin_financement_global,
        "tresorerie_nette": tresorerie_nette, "controle_treso_reelle": treso_reelle,
        "controle_ecart": controle_ecart,
        "flux_operationnel": tft["flux_operationnel"], "flux_investissement": tft["flux_investissement"],
        "flux_financement": tft["flux_financement"], "variation_treso_nette": tft["variation_treso_nette"],
        "endettement_financier_brut": endettement_financier_brut,
        "treso_actif": treso_actif, "endettement_financier_net": endettement_financier_net,
    }


# ---------------------------------------------------------------------------
# Export de la liasse fiscale (.xlsx), mise en page SYSCOHADA système normal
# ---------------------------------------------------------------------------
COMPANY_FIELDS = {
    "societe_nom": "Dénomination sociale",
    "societe_sigle": "Sigle usuel",
    "societe_adresse": "Adresse",
    "societe_ifu": "N° IFU du contribuable",
    "societe_teledeclarant": "N° de télédéclarant (NES)",
    "exercice_clos_le": "Exercice clos le (JJ/MM/AAAA)",
}


def get_company_info(conn):
    return {k: conn.execute("SELECT value FROM settings WHERE key = ?", (k,)).fetchone()
            for k in COMPANY_FIELDS}


def get_company_value(conn, key, default=""):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_company_value(conn, key, value):
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()


def export_liasse_fiscale(conn, path, stock_initial=0.0, treso_ouverture=0.0):
    """Génère un classeur .xlsx : COUVERTURE, BILAN, RESULTAT, TFT
    (mise en page et codes SYSCOHADA système normal)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="999999")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    money_fmt = "#,##0"

    def company_row(ws, row=3):
        info = {k: get_company_value(conn, k) for k in COMPANY_FIELDS}
        ws.cell(row=row, column=1, value="Dénomination sociale :")
        ws.cell(row=row, column=3, value=info["societe_nom"])
        ws.cell(row=row + 1, column=1, value="Adresse :")
        ws.cell(row=row + 1, column=3, value=info["societe_adresse"])
        ws.cell(row=row + 2, column=1, value="N° IFU du contribuable :")
        ws.cell(row=row + 2, column=3, value=info["societe_ifu"])
        ws.cell(row=row + 2, column=6, value="Exercice clos le :")
        ws.cell(row=row + 2, column=7, value=info["exercice_clos_le"])
        ws.cell(row=row + 3, column=1, value="N° de télédéclarant (NES) :")
        ws.cell(row=row + 3, column=3, value=info["societe_teledeclarant"])
        for r in range(row, row + 4):
            ws.cell(row=r, column=1).font = bold

    # ---- COUVERTURE ----
    ws = wb.active
    ws.title = "COUVERTURE"
    ws["A1"] = "ÉTATS FINANCIERS — SYSTÈME COMPTABLE OHADA (SYSCOHADA), SYSTÈME NORMAL"
    ws["A1"].font = title_font
    company_row(ws, row=3)
    ws["A9"] = ("Généré automatiquement par l'application Saisie Comptable. Les totaux (AZ, BK, BT, BZ, "
                "CP, DD, DP, DT, DZ) sont calculés directement depuis vos écritures. Le détail par ligne "
                "(AE à AN, CA à CM, DA à DM) est une répartition indicative par plage de comptes — à faire "
                "vérifier par un expert-comptable avant tout dépôt officiel auprès de la DGI.")
    ws["A9"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A9:H9")
    ws.row_dimensions[9].height = 60
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 25

    # ---- BILAN ----
    liasse = compute_liasse_bilan(conn, stock_initial=stock_initial)
    bt = liasse["totaux"]
    ad_net = bt["actif"]["Immobilisations nettes"]
    stocks_net = bt["actif"]["Stocks"]
    creances_net = bt["actif"]["Créances et emplois assimilés"]
    treso_actif_net = bt["actif"]["Trésorerie actif"]
    total_actif = bt["total_actif"]

    ws = wb.create_sheet("BILAN")
    ws["A1"] = "BILAN — SYSTÈME NORMAL"
    ws["A1"].font = title_font
    company_row(ws, row=3)
    headers_row = 8
    ws.cell(row=headers_row, column=1, value="REF").font = bold
    ws.cell(row=headers_row, column=2, value="ACTIF").font = bold
    ws.cell(row=headers_row, column=3, value="BRUT").font = bold
    ws.cell(row=headers_row, column=4, value="AMORT/DEPREC").font = bold
    ws.cell(row=headers_row, column=5, value="NET").font = bold
    ws.cell(row=headers_row, column=7, value="REF").font = bold
    ws.cell(row=headers_row, column=8, value="PASSIF").font = bold
    ws.cell(row=headers_row, column=9, value="NET").font = bold
    for c in range(1, 10):
        ws.cell(row=headers_row, column=c).fill = header_fill

    ad = liasse["actif_detail"]
    actif_lines = [
        ("AE", "Frais de développement et de prospection", ad["AE"]),
        ("AF", "Brevets, licences, logiciels et droits similaires", ad["AF"]),
        ("AG", "Fonds commercial et droit au bail", ad["AG"]),
        ("AH", "Autres immobilisations incorporelles", ad["AH"]),
        ("AJ", "Terrains", ad["AJ"]),
        ("AK", "Bâtiments", ad["AK"]),
        ("AL", "Aménagements, agencements et installations", ad["AL"]),
        ("AM", "Matériel, mobilier et actifs biologiques", ad["AM"]),
        ("AN", "Matériel de transport", ad["AN"]),
        ("AP", "Avances et acomptes versés sur immobilisations", ad["AP"]),
        ("AR", "Titres de participation", ad["AR"]),
        ("AS", "Autres immobilisations financières", ad["AS"]),
    ]
    ac = liasse["actif_circulant_detail"]
    actif_circ_lines = [
        ("BH", "Fournisseurs, avances versées", ac["BH"]),
        ("BI", "Clients", ac["BI"]),
    ]

    pd_ = liasse["passif_detail"]
    passif_lines = [
        ("CA", "Capital", pd_["CA"]),
        ("CD", "Primes liées au capital social", pd_["CD"]),
        ("CF_CG", "Réserves", pd_["CF_CG"]),
        ("CH", "Report à nouveau (+ ou -)", pd_["CH"]),
        ("CJ", "Résultat net de l'exercice", bt["passif"]["Résultat net de l'exercice"]),
        ("CL", "Subventions d'investissement", pd_["CL"]),
        ("CM", "Provisions réglementées", pd_["CM"]),
        ("CP", "TOTAL CAPITAUX PROPRES ET RESSOURCES ASSIMILEES", None),
        ("DA", "Emprunts et dettes financières diverses", pd_["DA"]),
        ("DB", "Dettes de location-acquisition", pd_["DB"]),
        ("DC", "Provisions pour risques et charges", pd_["DC"]),
        ("DD", "TOTAL DETTES FINANCIERES ET RESSOURCES ASSIMILEES", None),
        ("DJ", "Fournisseurs d'exploitation", pd_["DJ"]),
        ("DH", "Clients, avances reçues / Fournisseurs avances (détail)", pd_["DH_avances"]),
        ("DK", "Dettes fiscales et sociales", pd_["DK"]),
        ("DM", "Autres dettes", pd_["DM"]),
    ]

    r = headers_row + 1
    for ref, label, val in actif_lines:
        ws.cell(row=r, column=1, value=ref)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=5, value=round(val.get("net", val) if isinstance(val, dict) else val))
        ws.cell(row=r, column=5).number_format = money_fmt
        r += 1
    ws.cell(row=r, column=1, value="AZ")
    ws.cell(row=r, column=2, value="TOTAL ACTIF IMMOBILISE").font = bold
    ws.cell(row=r, column=5, value=round(ad_net)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    r += 2
    ws.cell(row=r, column=1, value="BB")
    ws.cell(row=r, column=2, value="STOCKS ET ENCOURS")
    ws.cell(row=r, column=5, value=round(stocks_net)).number_format = money_fmt
    r += 1
    for ref, label, val in actif_circ_lines:
        ws.cell(row=r, column=1, value=ref)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=5, value=round(val)).number_format = money_fmt
        r += 1
    ws.cell(row=r, column=1, value="BK")
    ws.cell(row=r, column=2, value="TOTAL ACTIF CIRCULANT").font = bold
    ws.cell(row=r, column=5, value=round(stocks_net + creances_net)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    r += 2
    ws.cell(row=r, column=1, value="BT")
    ws.cell(row=r, column=2, value="TOTAL TRESORERIE-ACTIF").font = bold
    ws.cell(row=r, column=5, value=round(treso_actif_net)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    r += 2
    ws.cell(row=r, column=1, value="BZ")
    ws.cell(row=r, column=2, value="TOTAL GENERAL ACTIF").font = bold
    ws.cell(row=r, column=5, value=round(total_actif)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    last_actif_row = r

    r2 = headers_row + 1
    for ref, label, val in passif_lines:
        ws.cell(row=r2, column=7, value=ref)
        ws.cell(row=r2, column=8, value=label)
        if val is not None:
            ws.cell(row=r2, column=9, value=round(val)).number_format = money_fmt
        else:
            ws.cell(row=r2, column=8).font = bold
        r2 += 1
    total_passif = bt["total_passif"]
    ws.cell(row=r2, column=7, value="DZ")
    ws.cell(row=r2, column=8, value="TOTAL GENERAL PASSIF").font = bold
    ws.cell(row=r2, column=9, value=round(total_passif)).font = bold
    ws.cell(row=r2, column=9).number_format = money_fmt
    r2 += 2
    ws.cell(row=r2, column=7, value="Écart Actif - Passif :")
    ws.cell(row=r2, column=9, value=round(total_actif - total_passif)).number_format = money_fmt

    for col, w in zip("ABCDEFGHI", [6, 40, 14, 14, 14, 3, 6, 40, 16]):
        ws.column_dimensions[col].width = w

    # ---- RESULTAT ----
    cr = compute_liasse_resultat(conn)
    ws = wb.create_sheet("RESULTAT")
    ws["A1"] = "COMPTE DE RÉSULTAT — SYSTÈME NORMAL"
    ws["A1"].font = title_font
    company_row(ws, row=3)
    headers_row = 8
    for c, h in zip((1, 2, 5), ("REF", "LIBELLES", "EXERCICE N")):
        ws.cell(row=headers_row, column=c, value=h).font = bold
        ws.cell(row=headers_row, column=c).fill = header_fill

    resultat_lines = [
        ("TA", "Ventes de marchandises", cr["TA"]),
        ("RA", "Achats de marchandises", -cr["RA"]),
        ("XA", "MARGE COMMERCIALE", cr["XA"]),
        ("TB", "Ventes de produits fabriqués", cr["TB"]),
        ("TC", "Travaux, services vendus", cr["TC"]),
        ("TD", "Produits accessoires", cr["TD"]),
        ("XB", "CHIFFRE D'AFFAIRES", cr["XB"]),
        ("TE", "Production stockée (ou déstockage)", cr["TE"]),
        ("TG", "Subventions d'exploitation", cr["TG"]),
        ("TH", "Autres produits", cr["TH"]),
        ("RC", "Achats de matières premières et fournitures liées", -cr["RC"]),
        ("RE", "Autres achats", -cr["RE"]),
        ("RG", "Transports", -cr["RG"]),
        ("RH", "Services extérieurs", -cr["RH"]),
        ("RI", "Impôts et taxes", -cr["RI"]),
        ("RJ", "Autres charges", -cr["RJ"]),
        ("XC", "VALEUR AJOUTEE", cr["XC"]),
        ("RK", "Charges de personnel", -cr["RK"]),
        ("XD", "EXCEDENT BRUT D'EXPLOITATION", cr["XD"]),
        ("RL", "Dotations aux amortissements, provisions et dépréciations", -cr["RL"]),
        ("XE", "RESULTAT D'EXPLOITATION", cr["XE"]),
        ("TK", "Revenus financiers et assimilés", cr["TK"]),
        ("RM", "Frais financiers et charges assimilées", -cr["RM"]),
        ("XF", "RESULTAT FINANCIER", cr["XF"]),
        ("XG", "RESULTAT DES ACTIVITES ORDINAIRES", cr["XG"]),
        ("XH", "RESULTAT HORS ACTIVITES ORDINAIRES (non tracé)", cr["XH"]),
        ("RQ", "Participation des travailleurs (non tracée)", cr["RQ"]),
        ("RS", "Impôts sur le résultat (non tracé — IS à saisir séparément)", cr["RS"]),
        ("XI", "RESULTAT NET", cr["XI"]),
    ]
    bold_refs = {"XA", "XB", "XC", "XD", "XE", "XF", "XG", "XI"}
    r = headers_row + 1
    for ref, label, val in resultat_lines:
        ws.cell(row=r, column=1, value=ref)
        ws.cell(row=r, column=2, value=label)
        cell = ws.cell(row=r, column=5, value=round(val))
        cell.number_format = money_fmt
        if ref in bold_refs:
            ws.cell(row=r, column=2).font = bold
            cell.font = bold
        r += 1
    for col, w in zip("ABCDE", [6, 55, 3, 3, 16]):
        ws.column_dimensions[col].width = w

    # ---- TFT (simplifié, méthode directe) ----
    tft = compute_tft(conn, treso_ouverture=treso_ouverture)
    ws = wb.create_sheet("TFT")
    ws["A1"] = "TABLEAU DES FLUX DE TRÉSORERIE — méthode directe simplifiée"
    ws["A1"].font = title_font
    ws["A2"] = ("Cette version simplifiée (encaissements/décaissements de trésorerie classés EXP/INV/FIN) "
                "ne correspond PAS exactement au format officiel SYSCOHADA (méthode indirecte avec CAFG). "
                "Elle donne une image de la trésorerie mais doit être retravaillée avec un expert-comptable "
                "pour un dépôt officiel.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 45
    company_row(ws, row=4)
    tft_lines = [
        ("Trésorerie d'ouverture", tft["ouverture"]),
        ("Flux liés aux activités opérationnelles (EXP)", tft["exploitation"]),
        ("Flux liés aux activités d'investissement (INV)", tft["investissement"]),
        ("Flux liés aux activités de financement (FIN)", tft["financement"]),
        ("Flux non classés (à coder)", tft["non_classes"]),
        ("VARIATION NETTE DE TRESORERIE", tft["variation"]),
        ("TRESORERIE DE CLOTURE", tft["cloture"]),
    ]
    r = 10
    for label, val in tft_lines:
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=3, value=round(val))
        cell.number_format = money_fmt
        r += 1
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["C"].width = 16

    wb.save(path)
    return path


def export_liasse_fiscale_complete(conn, path, stock_initial=0.0):
    """Génère la liasse fiscale COMPLÈTE (mêmes 92 pages, mêmes dimensions que le
    modèle SYSCOHADA système normal fourni) : COUVERTURE/GARDE, BILAN et RESULTAT
    remplis automatiquement depuis vos écritures (soldes de clôture = solde
    d'ouverture + mouvements) ; TFT (officiel, vierge, + un onglet TFT simplifié
    calculé) ; toutes les autres pages (39 notes annexes, ~20 tableaux fiscaux DGI)
    sont conservées avec leur mise en page et leurs dimensions exactes, mais les
    montants qui provenaient du modèle sont effacés (ce ne sont pas vos chiffres)
    pour être complétées manuellement ou par votre expert-comptable."""
    import openpyxl
    from openpyxl.styles import Font

    template_path = os.path.join(_resource_dir(), "etats_financiers_template.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            "Le fichier modèle 'etats_financiers_template.xlsx' est introuvable dans "
            "l'exécutable. Cela signifie qu'il n'a pas été inclus lors de la compilation : "
            "vérifiez que ce fichier est bien présent à la racine du dépôt GitHub (à côté de "
            "main.py) et que .github/workflows/build.yml contient bien la ligne "
            "--add-data \"etats_financiers_template.xlsx;.\", puis relancez le build."
        )
    wb = openpyxl.load_workbook(template_path)
    green = Font(color="FF008000")

    # ---- Supprime les liens externes cassés (source du bandeau Excel
    #      « Impossible d'actualiser... valeurs depuis un classeur lié ») ----
    if getattr(wb, "_external_links", None):
        wb._external_links = []

    # ---- GARDE : identification de l'entité ----
    if "GARDE" in wb.sheetnames:
        g = wb["GARDE"]
        g["D22"] = get_company_value(conn, "societe_nom")
        g["C26"] = get_company_value(conn, "societe_sigle")
        g["C28"] = get_company_value(conn, "societe_adresse")
        g["D30"] = get_company_value(conn, "societe_ifu")
        g["D31"] = get_company_value(conn, "societe_teledeclarant")
        exdate = get_company_value(conn, "exercice_clos_le")
        if exdate:
            parsed = None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
                try:
                    parsed = datetime.strptime(exdate.strip(), fmt)
                    break
                except ValueError:
                    continue
            if parsed:
                g["E17"] = parsed
                g["E17"].number_format = "DD/MM/YYYY"
            else:
                g["E17"] = exdate

    # ---- BILAN ----
    liasse = compute_liasse_bilan(conn, stock_initial=stock_initial)
    bt = liasse["totaux"]
    ad = liasse["actif_detail"]
    ac = liasse["actif_circulant_detail"]
    pd_ = liasse["passif_detail"]

    actif_values = {
        "AE": ad["AE"]["net"], "AF": ad["AF"]["net"], "AG": ad["AG"]["net"], "AH": ad["AH"]["net"],
        "AD": ad["AE"]["net"] + ad["AF"]["net"] + ad["AG"]["net"] + ad["AH"]["net"],
        "AJ": ad["AJ"]["net"], "AK": ad["AK"]["net"], "AL": ad["AL"]["net"],
        "AM": ad["AM"]["net"], "AN": ad["AN"]["net"],
        "AI": ad["AJ"]["net"] + ad["AK"]["net"] + ad["AL"]["net"] + ad["AM"]["net"] + ad["AN"]["net"],
        "AP": ad["AP"]["net"], "AR": ad["AR"]["net"], "AS": ad["AS"]["net"],
        "AZ": bt["actif"]["Immobilisations nettes"],
        "BB": bt["actif"]["Stocks"],
        "BH": ac["BH"], "BI": ac["BI"],
        "BK": bt["actif"]["Stocks"] + bt["actif"]["Créances et emplois assimilés"],
        "BT": bt["actif"]["Trésorerie actif"],
        "BZ": bt["total_actif"],
    }
    passif_values = {
        "CA": pd_["CA"], "CD": pd_["CD"], "CF": pd_["CF_CG"], "CH": pd_["CH"],
        "CJ": bt["passif"]["Résultat net de l'exercice"],
        "CL": pd_["CL"], "CM": pd_["CM"],
        "CP": (pd_["CA"] + pd_["CD"] + pd_["CF_CG"] + pd_["CH"]
               + bt["passif"]["Résultat net de l'exercice"] + pd_["CL"] + pd_["CM"]),
        "DA": pd_["DA"], "DB": pd_["DB"], "DC": pd_["DC"],
        "DD": pd_["DA"] + pd_["DB"] + pd_["DC"],
        "DJ": pd_["DJ"], "DH": pd_["DH_avances"], "DK": pd_["DK"], "DM": pd_["DM"],
        "DP": pd_["DJ"] + pd_["DH_avances"] + pd_["DK"] + pd_["DM"],
        "DT": bt["passif"]["Trésorerie passif"],
        "DZ": bt["total_passif"],
    }

    if "BILAN" in wb.sheetnames:
        ws = wb["BILAN"]
        # Efface toutes les valeurs numériques préexistantes du modèle (Brut, Amort,
        # N-1) sur les lignes de données, pour n'y laisser QUE nos propres calculs.
        for row in range(11, 41):
            for col in (6, 7, 8, 9, 13, 14):  # F,G,H,I (actif) / M,N (passif)
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.value = None
        ws["C3"] = get_company_value(conn, "societe_nom")
        ws["C4"] = get_company_value(conn, "societe_adresse")
        ws["C5"] = get_company_value(conn, "societe_ifu")
        for ref, row in {
            "AD": 11, "AE": 12, "AF": 13, "AG": 14, "AH": 15, "AI": 16, "AJ": 17, "AK": 18,
            "AL": 19, "AM": 20, "AN": 21, "AP": 22, "AQ": 23, "AR": 24, "AS": 25, "AZ": 26,
            "BB": 28, "BH": 30, "BI": 31, "BK": 33, "BT": 37, "BZ": 39,
        }.items():
            if ref in actif_values:
                cell = ws.cell(row=row, column=8, value=round(actif_values[ref]))
                cell.font = green
        for ref, row in {
            "CA": 11, "CD": 13, "CF": 15, "CH": 17, "CJ": 18, "CL": 19, "CM": 20, "CP": 21,
            "DA": 22, "DB": 23, "DC": 24, "DD": 25, "DH": 27, "DJ": 29, "DK": 30, "DM": 31,
            "DP": 33, "DT": 36, "DZ": 39,
        }.items():
            if ref in passif_values:
                cell = ws.cell(row=row, column=13, value=round(passif_values[ref]))
                cell.font = green
        ws.cell(row=40, column=13, value="=H39-M39")  # écart de contrôle

    # ---- RESULTAT ----
    cr = compute_liasse_resultat(conn)
    if "RESULTAT" in wb.sheetnames:
        ws = wb["RESULTAT"]
        for row in range(11, 53):
            for col in (9, 10):  # I (exercice N), J (N-1)
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.value = None
        row_map = {"TA": 11, "RA": 12, "XA": 14, "TB": 15, "TC": 16, "TD": 17, "XB": 18,
                   "TE": 19, "TG": 21, "TH": 22, "RC": 24, "RE": 26, "RG": 28, "RH": 29,
                   "RI": 30, "RJ": 31, "XC": 32, "RK": 33, "XD": 34, "RL": 36, "XE": 37,
                   "TK": 38, "RM": 41, "XF": 43, "XG": 44, "XH": 49, "RQ": 50, "RS": 51, "XI": 52}
        sign_negative = {"RA", "RC", "RE", "RG", "RH", "RI", "RJ", "RK", "RL", "RM", "RQ", "RS"}
        for ref, row in row_map.items():
            val = cr.get(ref, 0.0)
            if ref in sign_negative:
                val = -abs(val)
            cell = ws.cell(row=row, column=9, value=round(val))
            cell.font = green

    # ---- TFT : remplit la vraie feuille officielle sur les lignes confirmées
    #      (ZA=10, FA=12, FB=13, FC=14, FD=15, FE=16), + un onglet
    #      supplémentaire avec le calcul complet (méthode indirecte — CAFG),
    #      les mêmes données que l'onglet TFT de l'application ----
    tft_off = compute_tft_officiel(conn)
    tft = compute_tft_indirect(conn)
    if "TFT" in wb.sheetnames:
        ws = wb["TFT"]
        for row in range(10, 42):
            for col in (9, 10):  # I (exercice N), J (N-1)
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.value = None
        for ref, row in (("ZA", 10), ("FA", 12), ("FB", 13), ("FC", 14), ("FD", 15), ("FE", 16)):
            c = ws.cell(row=row, column=9, value=round(tft_off[ref]))
            c.font = green
        ws["A44"] = (
            "Lignes ZA et FA à FE remplies automatiquement depuis vos écritures (identique à "
            "l'onglet TFT de l'application, section Flux opérationnels). Les lignes d'investissement "
            "et de financement (FF et suivantes) n'ont pas encore de position de cellule confirmée "
            "dans ce modèle — complétez-les manuellement, ou envoyez une capture des lignes "
            "suivantes pour qu'elles soient automatisées aussi. Voir l'onglet « TFT (méthode "
            "indirecte - CAFG) » pour le calcul complet (investissement et financement inclus)."
        )
    ws_tft = wb.create_sheet("TFT (méthode indirecte - CAFG)")
    ws_tft["A1"] = "TABLEAU DE FLUX DE TRÉSORERIE (méthode indirecte — CAFG)"
    ws_tft["A1"].font = Font(bold=True, size=12)
    tft_lignes = [
        ("A - Trésorerie nette au 1er janvier", tft["treso_ouverture"]),
        ("", None),
        ("DÉTERMINATION DE LA CAFG", None),
        ("Excédent Brut d'Exploitation (EBE)", tft["ebe"]),
        ("+ Revenus financiers", tft["revenus_financiers"]),
        ("- Frais financiers", tft["frais_financiers"]),
        ("CAPACITÉ D'AUTOFINANCEMENT GLOBALE (CAFG)", tft["cafg"]),
        ("- Variation des stocks", tft["variation_stocks"]),
        ("- Variation des créances", tft["variation_creances"]),
        ("+ Variation du passif circulant", tft["variation_dettes_circulantes"]),
        ("FLUX DES ACTIVITÉS OPÉRATIONNELLES (A)", tft["flux_operationnel"]),
        ("", None),
        ("FLUX DES ACTIVITÉS D'INVESTISSEMENT", None),
        ("- Acquisitions immobilisations incorporelles", tft["acquisitions_incorp"]),
        ("- Acquisitions immobilisations corporelles", tft["acquisitions_corp"]),
        ("- Acquisitions immobilisations financières", tft["acquisitions_fin"]),
        ("+ Cessions immobilisations incorporelles", tft["cessions_incorp"]),
        ("+ Cessions immobilisations corporelles", tft["cessions_corp"]),
        ("+ Cessions immobilisations financières", tft["cessions_fin"]),
        ("FLUX DES ACTIVITÉS D'INVESTISSEMENT (B)", tft["flux_investissement"]),
        ("", None),
        ("FLUX DES ACTIVITÉS DE FINANCEMENT", None),
        ("+ Augmentation de capital", tft["augmentation_capital"]),
        ("+ Subventions d'investissement reçues", tft["subventions_recues"]),
        ("- Prélèvements sur le capital", tft["prelevements_capital"]),
        ("- Dividendes versés", tft["dividendes_verses"]),
        ("+ Emprunts nouveaux", tft["emprunts_nouveaux"]),
        ("- Remboursements des emprunts", tft["remboursements_emprunts"]),
        ("FLUX DES ACTIVITÉS DE FINANCEMENT (C)", tft["flux_financement"]),
        ("", None),
        ("VARIATION DE LA TRÉSORERIE NETTE (A+B+C)", tft["variation_treso_nette"]),
        ("TRÉSORERIE NETTE CALCULÉE AU 31/12/N", tft["treso_cloture_calculee"]),
        ("CONTRÔLE — Trésorerie réelle (Balance, classe 5)", tft["treso_cloture_reelle"]),
        ("ÉCART", tft["ecart"]),
    ]
    for i, (label, val) in enumerate(tft_lignes):
        ws_tft.cell(row=3 + i, column=1, value=label)
        if val is not None:
            ws_tft.cell(row=3 + i, column=3, value=round(val))
    ws_tft.column_dimensions["A"].width = 55

    # ---- SITUATION FINANCIÈRE (FR-BFR-TN) : mêmes données que l'onglet
    #      correspondant de l'application ----
    sf = compute_situation_financiere(conn)
    ws_sf = wb.create_sheet("SITUATION FIN. (FR-BFR-TN)")
    ws_sf["A1"] = "SITUATION FINANCIÈRE (FR - BFR - TN)"
    ws_sf["A1"].font = Font(bold=True, size=12)
    sf_lignes = [
        ("Résultat net comptable", sf["resultat_net_comptable"]),
        ("EBE", sf["ebe"]), ("+ Revenus financiers", sf["revenus_financiers"]),
        ("- Frais financiers", sf["frais_financiers"]),
        ("CAFG", sf["cafg"]), ("- Dividendes versés", sf["dividendes_verses"]),
        ("AUTOFINANCEMENT", sf["autofinancement"]),
        ("Rentabilité économique (%)", sf["rentabilite_economique"]),
        ("Rentabilité financière (%)", sf["rentabilite_financiere"]),
        ("", None),
        ("Capitaux propres et ressources assimilées", sf["capitaux_propres_ressources"]),
        ("+ Dettes financières", sf["dettes_financieres"]),
        ("= RESSOURCES STABLES", sf["ressources_stables"]),
        ("- Actifs immobilisés", sf["actifs_immobilises"]),
        ("= FONDS DE ROULEMENT (FR)", sf["fonds_de_roulement"]),
        ("", None),
        ("+ Actif circulant d'exploitation", sf["actif_circulant_exploitation"]),
        ("- Passif circulant d'exploitation", sf["passif_circulant_exploitation"]),
        ("= Besoin de financement d'exploitation", sf["besoin_financement_exploitation"]),
        ("+ Actif circulant HAO", sf["actif_circulant_hao"]),
        ("- Passif circulant HAO", sf["passif_circulant_hao"]),
        ("= Besoin de financement HAO", sf["besoin_financement_hao"]),
        ("= BESOIN DE FINANCEMENT GLOBAL (BFR)", sf["besoin_financement_global"]),
        ("", None),
        ("TRÉSORERIE NETTE (FR - BFR)", sf["tresorerie_nette"]),
        ("Contrôle — Trésorerie réelle (Balance)", sf["controle_treso_reelle"]),
        ("Écart", sf["controle_ecart"]),
        ("", None),
        ("+ Flux activités opérationnelles", sf["flux_operationnel"]),
        ("- Flux activités d'investissement", sf["flux_investissement"]),
        ("+ Flux activités de financement", sf["flux_financement"]),
        ("VARIATION DE LA TRÉSORERIE NETTE DE LA PÉRIODE", sf["variation_treso_nette"]),
        ("", None),
        ("Endettement financier brut", sf["endettement_financier_brut"]),
        ("- Trésorerie actif", sf["treso_actif"]),
        ("= ENDETTEMENT FINANCIER NET", sf["endettement_financier_net"]),
    ]
    for i, (label, val) in enumerate(sf_lignes):
        ws_sf.cell(row=3 + i, column=1, value=label)
        if val is not None:
            ws_sf.cell(row=3 + i, column=3, value=round(val, 2))
    ws_sf.column_dimensions["A"].width = 55

    # ---- Toutes les autres pages : structure/dimensions conservées, valeurs
    #      chiffrées (issues du modèle) effacées pour éviter toute confusion ----
    skip = {"GARDE", "BILAN", "RESULTAT", "TFT", "TFT (méthode indirecte - CAFG)",
            "SITUATION FIN. (FR-BFR-TN)"}
    for name in wb.sheetnames:
        if name in skip:
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.value = None

    # ---- Filet de sécurité : efface tout texte résiduel qui ne serait pas le
    #      nom de VOTRE entité (au cas où le modèle contiendrait encore une
    #      dénomination sociale tierce, pour éviter toute confusion/litige) ----
    my_name = (get_company_value(conn, "societe_nom") or "").strip().upper()
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "GCM" in cell.value.upper():
                    if not my_name or my_name not in cell.value.upper():
                        cell.value = None

    # ---- NOTE 34 : Fiche de synthèse des principaux indicateurs financiers
    #      (SIG) — mêmes données que l'onglet Compte de résultat de l'app.
    #      Rempli APRÈS le nettoyage général ci-dessus (qui efface d'abord
    #      les anciennes valeurs littérales de l'entité précédente). ----
    if "NOTE 34" in wb.sheetnames:
        ws34 = wb["NOTE 34"]
        note34_map = {"XB": 11, "XA": 12, "XC": 13, "XD": 14, "XE": 15, "XF": 16,
                      "XG": 17, "XH": 18, "XI": 19}
        for ref, row in note34_map.items():
            c = ws34.cell(row=row, column=6, value=round(cr.get(ref, 0.0)))
            c.font = green
        try:
            exercice_n = get_current_exercice(conn)
            exercice_n1 = str(int(exercice_n) - 1)
            if any(e["exercice"] == exercice_n1 for e in list_exercices(conn)):
                cr_n1 = compute_liasse_resultat(conn, exercice=exercice_n1)
                for ref, row in note34_map.items():
                    ws34.cell(row=row, column=7, value=round(cr_n1.get(ref, 0.0)))
        except (ValueError, TypeError):
            pass

    # ---- Uniformise toutes les cellules de type date au format JJ/MM/AAAA ----
    date_format_markers = ("yy", "mm", "dd", "jj", "aaaa")
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                is_date_value = isinstance(cell.value, (datetime, date)) and not isinstance(cell.value, bool)
                fmt = (cell.number_format or "").lower()
                looks_like_date_format = fmt not in ("general", "@") and any(m in fmt for m in date_format_markers)
                if is_date_value or looks_like_date_format:
                    cell.number_format = "DD/MM/YYYY"

    wb.properties.creator = None
    wb.properties.lastModifiedBy = None

    wb.save(path)
    return path


if __name__ == "__main__":
    # Petit auto-test en ligne de commande (sans Tkinter).
    conn = get_connection(":memory:" if False else "test_core.db")
    print("Comptes chargés :", conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0])

    add_entry(conn, str(date.today()), "FA-0001", "AC", "601000", "", "Achat marchandises", 1000, 0)
    add_entry(conn, str(date.today()), "FA-0001", "AC", "445200", "", "TVA récupérable", 200, 0)
    add_entry(conn, str(date.today()), "FA-0001", "AC", "401000", "Ets Dupont", "Facture FA-0001", 0, 1200)
    add_entry(conn, str(date.today()), "FV-0001", "VE", "411000", "Société ABC", "Facture FV-0001", 1180, 0)
    add_entry(conn, str(date.today()), "FV-0001", "VE", "701000", "", "Vente marchandises", 0, 1180)

    d, c = totals_debit_credit(conn)
    print("Total débit / crédit :", d, c, "Équilibré :", d == c)

    print("\n--- Balance ---")
    for b in compute_balance(conn):
        print(b)

    print("\n--- Compte de résultat ---")
    cr = compute_compte_resultat(conn)
    print("Résultat net :", cr["resultat_net"])

    print("\n--- Bilan ---")
    bilan = compute_bilan(conn)
    print("Total actif :", bilan["total_actif"], "Total passif :", bilan["total_passif"], "Écart :", bilan["ecart"])

    print("\n--- TFT ---")
    print(compute_tft(conn))

    print("\n--- Grand livre (411000) ---")
    for r in compute_grand_livre(conn, "411000"):
        print(r)

    print("\n--- Stocks ---")
    for s in compute_stocks(conn):
        print(s)

    print("\n--- Production ---")
    print(compute_production(conn))

    conn.close()
    os.remove("test_core.db")
