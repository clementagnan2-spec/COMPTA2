# -*- coding: utf-8 -*-
"""
CLIENT — application de bureau qui se connecte à un SERVEUR
(voir server.py) par réseau local ou Internet, pour permettre à
PLUSIEURS UTILISATEURS de travailler EN MÊME TEMPS sur la même base de
données comptable.

Contrairement à main.py (qui ouvre directement un fichier SQLite local),
cette application n'ouvre AUCUN fichier local — toutes les opérations
(saisie, consultation) passent par le réseau via client_core.py.

Premier module entièrement fonctionnel de bout en bout : la SAISIE
COMPTABLE (multi-lignes, avec Bilan de contrôle en temps réel). Les
écrans Ventes / Achats / Stocks du circuit commercial suivent le même
principe (voir client_core.py + server.py RPC_WHITELIST à étendre) et
seront ajoutés en s'appuyant sur cette même architecture.
"""
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from datetime import date

import core  # fonctions PURES (sans accès base) réutilisées telles quelles : to_display_date, to_iso_date...
import client_core
from client_core import RemoteConnection, RemoteAuthError, RemoteCallError, RemoteConnectionError


def fmt_cfa(v):
    if v in (None, ""):
        return ""
    try:
        return f"{v:,.0f}".replace(",", " ")
    except (TypeError, ValueError):
        return str(v)


APPEL_ECHEC = object()  # sentinelle distincte de None (un appel reussi peut legitimement renvoyer None)


def appeler(widget, remote, fonction, *args, **kwargs):
    """Enveloppe tout appel réseau avec une gestion d'erreur unifiée
    (session expirée, serveur injoignable, erreur métier) — factorisé
    pour être réutilisé par tous les écrans du client (Saisie, GRH...).
    Renvoie APPEL_ECHEC (PAS None) en cas d'échec, pour ne jamais
    confondre un appel réussi qui renvoie légitimement None avec un échec."""
    try:
        return getattr(client_core, fonction)(remote, *args, **kwargs)
    except RemoteAuthError as exc:
        messagebox.showerror("Session expirée", str(exc), parent=widget)
        widget.winfo_toplevel().destroy()
    except RemoteConnectionError as exc:
        messagebox.showerror("Connexion perdue", str(exc), parent=widget)
    except RemoteCallError as exc:
        messagebox.showerror("Erreur", str(exc), parent=widget)
    return APPEL_ECHEC


class LoginWindow(tk.Tk):
    """Écran de connexion : adresse du serveur, port, identifiants —
    premier écran affiché au lancement du client."""

    def __init__(self):
        super().__init__()
        self.title("PLATEFORME INTEGREE DE GESTION — Client")
        self.geometry("460x420")
        self.resizable(False, False)
        try:
            icon_path = core.get_app_icon_path()
            self.iconbitmap(icon_path)
        except Exception:
            pass

        frame = ttk.Frame(self, padding=24)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Connexion au serveur", font=("Segoe UI", 15, "bold")).pack(anchor="w", pady=(0, 4))
        ttk.Label(frame, text=(
            "Renseignez l'adresse du poste serveur (sur le réseau local : son adresse IP, ex. "
            "192.168.1.10 — visible avec 'ipconfig' sur le poste serveur)."
        ), foreground="#595959", wraplength=400, justify="left").pack(anchor="w", pady=(0, 16))

        form = ttk.Frame(frame)
        form.pack(fill="x")
        ttk.Label(form, text="Adresse du serveur :").grid(row=0, column=0, sticky="w", pady=4)
        self.host_var = tk.StringVar(value="127.0.0.1")
        ttk.Entry(form, textvariable=self.host_var, width=28).grid(row=0, column=1, pady=4, sticky="w")

        ttk.Label(form, text="Port :").grid(row=1, column=0, sticky="w", pady=4)
        self.port_var = tk.StringVar(value="8765")
        ttk.Entry(form, textvariable=self.port_var, width=10).grid(row=1, column=1, pady=4, sticky="w")

        ttk.Label(form, text="Identifiant :").grid(row=2, column=0, sticky="w", pady=4)
        self.user_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.user_var, width=28).grid(row=2, column=1, pady=4, sticky="w")

        ttk.Label(form, text="Mot de passe :").grid(row=3, column=0, sticky="w", pady=4)
        self.pwd_var = tk.StringVar()
        pwd_entry = ttk.Entry(form, textvariable=self.pwd_var, width=28, show="•")
        pwd_entry.grid(row=3, column=1, pady=4, sticky="w")
        pwd_entry.bind("<Return>", lambda e: self.connecter())

        self.status_var = tk.StringVar()
        ttk.Label(frame, textvariable=self.status_var, foreground="#B00020", wraplength=400, justify="left").pack(
            anchor="w", pady=(12, 8))

        btns = ttk.Frame(frame)
        btns.pack(fill="x", pady=(8, 0))
        self.connect_btn = ttk.Button(btns, text="Se connecter", command=self.connecter)
        self.connect_btn.pack(side="left")
        ttk.Button(btns, text="Tester la connexion au serveur", command=self.tester).pack(side="left", padx=8)

        self.remote = None

    def tester(self):
        host = self.host_var.get().strip()
        try:
            port = int(self.port_var.get().strip())
        except ValueError:
            self.status_var.set("Le port doit être un nombre.")
            return
        remote = RemoteConnection(host, port, timeout=5)
        info = remote.ping()
        if info:
            version = info.get("version", "?")
            nb_fn = info.get("nb_fonctions_autorisees", "?")
            self.status_var.set(f"✓ Serveur joignable — version {version} ({nb_fn} fonctions autorisées).")
            self.status_var_color("#1F7A1F")
        else:
            self.status_var.set(f"✗ Serveur injoignable à {host}:{port} — vérifiez l'adresse, le port, et "
                                 f"que le serveur est bien démarré sur l'autre poste.")
            self.status_var_color("#B00020")

    def status_var_color(self, color):
        for w in self.winfo_children():
            pass  # simple — la couleur est déjà fixée à la création du Label ci-dessus

    def connecter(self):
        host = self.host_var.get().strip()
        nom_utilisateur = self.user_var.get().strip()
        mot_de_passe = self.pwd_var.get()
        if not host or not nom_utilisateur or not mot_de_passe:
            self.status_var.set("Adresse du serveur, identifiant et mot de passe sont obligatoires.")
            return
        try:
            port = int(self.port_var.get().strip())
        except ValueError:
            self.status_var.set("Le port doit être un nombre.")
            return

        self.connect_btn.configure(state="disabled")
        self.status_var.set("Connexion en cours…")
        self.update_idletasks()

        remote = RemoteConnection(host, port)
        try:
            remote.login(nom_utilisateur, mot_de_passe)
        except RemoteAuthError as exc:
            self.status_var.set(str(exc))
            self.connect_btn.configure(state="normal")
            return
        except RemoteConnectionError as exc:
            self.status_var.set(str(exc))
            self.connect_btn.configure(state="normal")
            return

        self.remote = remote
        self.destroy()


class ClientApp(tk.Tk):
    """Fenêtre principale du client, une fois connecté — barre de menu
    identique dans l'esprit à l'application de bureau (core.MENU_STRUCTURE),
    filtrée selon les sous-menus autorisés pour le niveau d'accès connecté
    (transmis par le serveur à la connexion). Seul l'écran Saisie est
    pleinement implémenté côté client pour l'instant — les autres
    sous-menus autorisés affichent un message clair plutôt que de planter,
    en attendant leur construction (même modèle à suivre)."""

    # Sous-menus du circuit commercial déjà pleinement fonctionnels côté client.
    IMPLEMENTED_SCREENS = {
        "saisie": lambda parent, remote: RemoteSaisieTab(parent, remote),
        "grh_personnel": lambda parent, remote: RemotePersonnelTab(parent, remote),
        "grh_time_sheet": lambda parent, remote: RemoteTimeSheetTab(parent, remote),
        "grh_kpi": lambda parent, remote: RemoteKpiTab(parent, remote),
        "grh_tableau_bord": lambda parent, remote: RemoteTableauBordGrhTab(parent, remote),
        "grh_hs": lambda parent, remote: RemoteHsTab(parent, remote),
        "grh_paie": lambda parent, remote: RemotePaieTab(parent, remote),
        "fournisseurs": lambda parent, remote: RemoteFournisseursTab(parent, remote),
        "reglements": lambda parent, remote: RemoteReglementsTab(parent, remote),
        "grand_livre": lambda parent, remote: RemoteGrandLivreTab(parent, remote),
        "balance": lambda parent, remote: RemoteBalanceTab(parent, remote),
        "bilan_syscohada": lambda parent, remote: RemoteBilanTab(parent, remote),
        "compte_resultat_sig": lambda parent, remote: RemoteEtatFormuleTab(
            parent, remote, "Compte de résultat (SIG)", "compute_cr"),
        "tft": lambda parent, remote: RemoteEtatFormuleTab(parent, remote, "TFT", "compute_tft_gabarit"),
        "situation_financiere": lambda parent, remote: RemoteEtatFormuleTab(
            parent, remote, "Situation financière", "compute_situation_fin"),
        "arrete_comptes": lambda parent, remote: RemoteArreteComptesTab(parent, remote),
        "clients": lambda parent, remote: RemoteClientsTab(parent, remote),
        "facturation": lambda parent, remote: RemoteFacturationTab(parent, remote),
        "stocks": lambda parent, remote: RemoteStocksTab(parent, remote),
        "machines": lambda parent, remote: RemoteMachinesTab(parent, remote),
        "tresorerie": lambda parent, remote: RemoteTresorerieTab(parent, remote),
        "rapprochement_bancaire": lambda parent, remote: RemoteRapprochementBancaireTab(parent, remote),
        "immobilisations": lambda parent, remote: RemoteImmobilisationsTab(parent, remote),
        "expression_besoin": lambda parent, remote: RemoteExpressionBesoinTab(parent, remote),
        "ep_bon_commande": lambda parent, remote: RemoteBonCommandeTab(parent, remote),
        "factures_frs": lambda parent, remote: RemoteFacturesFrsTab(parent, remote),
        "recouvrement": lambda parent, remote: RemoteRecouvrementTab(parent, remote),
        "commandes_client": lambda parent, remote: RemoteCommandeClientTab(parent, remote),
        "marges": lambda parent, remote: RemoteMargesTab(parent, remote),
        "contrats": lambda parent, remote: RemoteContratsTab(parent, remote),
        "bordereau_livraison": lambda parent, remote: RemoteBordereauLivraisonTab(parent, remote),
        "amortissements": lambda parent, remote: RemoteAmortissementsTab(parent, remote),
        "transport": lambda parent, remote: RemoteParcAutoTab(parent, remote),
        "missions": lambda parent, remote: RemoteMissionsTab(parent, remote),
        "pieces_rechange": lambda parent, remote: RemotePiecesRechangeTab(parent, remote),
        "reparations": lambda parent, remote: RemoteReparationsTab(parent, remote),
        "plan_analytique": lambda parent, remote: RemoteSimplePlanTab(
            parent, remote, "PLAN ANALYTIQUE", "list_analytic_codes", "add_analytic_code",
            "delete_analytic_code", extra_field="unite"),
        "plan_budgetaire": lambda parent, remote: RemoteSimplePlanTab(
            parent, remote, "PLAN BUDGÉTAIRE", "list_budget_codes", "add_budget_code",
            "delete_budget_code", extra_field="montant"),
        "plan_bailleur": lambda parent, remote: RemoteSimplePlanTab(
            parent, remote, "PLAN BAILLEURS DE FONDS", "list_donor_codes", "add_donor_code",
            "delete_donor_code"),
        "taux_tva": lambda parent, remote: RemoteSimplePlanTab(
            parent, remote, "TAUX DE TVA", "list_taux_tva", "add_taux_tva", "delete_taux_tva",
            extra_field="montant"),
        "taux_retenue": lambda parent, remote: RemoteSimplePlanTab(
            parent, remote, "TAUX DE RETENUE À LA SOURCE", "list_taux_retenue", "add_taux_retenue",
            "delete_taux_retenue", extra_field="montant"),
        "energie": lambda parent, remote: RemoteAnalytiquePeriodeTab(
            parent, remote, "Énergie",
            "Coûts d'énergie (eau, électricité, essence, gasoil, gaz...) par code analytique, sur l'exercice "
            "courant.", "ENERGIE-"),
        "maintenance": lambda parent, remote: RemoteAnalytiquePeriodeTab(
            parent, remote, "Maintenance",
            "Coûts de maintenance (véhicules, bâtiments, machines, informatique...) par code analytique, sur "
            "l'exercice courant.", "MAINT-"),
        "production": lambda parent, remote: RemoteProductionTab(parent, remote),
        "exercices": lambda parent, remote: RemoteExercicesTab(parent, remote),
        "synchronisation": lambda parent, remote: RemoteSynchronisationTab(parent, remote),
        "rapports_technique": lambda parent, remote: RemoteRapportTechniqueTab(parent, remote),
        "ouverture": lambda parent, remote: RemoteOuvertureTab(parent, remote),
        "plan_comptable": lambda parent, remote: RemotePlanComptableTab(parent, remote),
        "admin_factures": lambda parent, remote: RemotePlaceholderTab(
            parent, remote, "Modification des factures",
            "Réservé à l'application de bureau, par sécurité — modification de factures déjà validées, "
            "opération sensible non exposée à distance pour l'instant."),
        "admin_modele_bon_commande": lambda parent, remote: RemotePlaceholderTab(
            parent, remote, "Modèle de bon de commande",
            "Réservé à l'application de bureau — édition d'un modèle de document, opération locale au "
            "poste serveur."),
        "niveaux_acces": lambda parent, remote: RemotePlaceholderTab(
            parent, remote, "Niveaux d'accès",
            "Réservé à l'application de bureau, par sécurité — la gestion des niveaux d'accès et de leurs "
            "autorisations n'est volontairement pas exposée à distance (voir server.py RPC_WHITELIST)."),
        "utilisateurs": lambda parent, remote: RemotePlaceholderTab(
            parent, remote, "Utilisateurs",
            "Réservé à l'application de bureau, par sécurité — la création/suppression d'utilisateurs "
            "n'est volontairement pas exposée à distance."),
        "reinitialisation": lambda parent, remote: RemotePlaceholderTab(
            parent, remote, "Réinitialisation des données",
            "Réservé à l'application de bureau, par sécurité — opération destructrice et irréversible, "
            "volontairement non exposée à distance."),
    }

    def __init__(self, remote: RemoteConnection):
        super().__init__()
        self.remote = remote
        self.pages = {}
        self.report_callback_exception = self._report_callback_exception
        self.title(f"PLATEFORME INTEGREE DE GESTION — Client — {remote.nom_utilisateur} "
                    f"({remote.niveau_acces}) — {remote.host}:{remote.port}")
        self.geometry("1300x800")
        try:
            self.state("zoomed")
        except tk.TclError:
            pass
        try:
            self.iconbitmap(core.get_app_icon_path())
        except Exception:
            pass
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        top_bar = ttk.Frame(self, relief="raised", padding=4)
        top_bar.pack(fill="x")
        ttk.Label(top_bar, text=f"Connecté : {remote.nom_utilisateur} ({remote.niveau_acces})",
                  font=("Segoe UI", 9, "bold")).pack(side="left", padx=8)
        ttk.Label(top_bar, text=f"Serveur : {remote.host}:{remote.port}",
                  foreground="#595959").pack(side="left", padx=8)
        try:
            exercice_serveur = client_core.get_current_exercice(remote)
        except Exception:
            exercice_serveur = "?"
        ttk.Label(top_bar, text=f"Exercice comptable (serveur) : {exercice_serveur}",
                  font=("Segoe UI", 9, "bold"), foreground="#B00020").pack(side="left", padx=8)
        ttk.Label(top_bar, text=f"Version serveur : {getattr(remote, 'server_version', '?')}",
                  foreground="#595959").pack(side="left", padx=8)
        ttk.Button(top_bar, text="Se déconnecter", command=self._on_close).pack(side="right", padx=8)

        self._build_menu()

        self.content = ttk.Frame(self)
        self.content.pack(fill="both", expand=True)
        self.content.grid_rowconfigure(0, weight=1)
        self.content.grid_columnconfigure(0, weight=1)

        # Ouvre le premier sous-menu autorisé et implémenté par défaut
        # (généralement la Saisie), sinon un message d'accueil.
        premiere_cle = next((k for k in self.IMPLEMENTED_SCREENS if k in remote.menus_autorises), None)
        if premiere_cle:
            self.show(premiere_cle)
        else:
            self._show_accueil()

    def _build_menu(self):
        """Barre de menu — mêmes libellés et regroupements que
        l'application de bureau (core.MENU_STRUCTURE), mais un sous-menu
        n'apparaît que s'il est À LA FOIS autorisé pour ce niveau d'accès
        (remote.menus_autorises, transmis par le serveur) ET disponible
        côté client. Un menu de premier niveau sans aucun sous-menu
        correspondant est masqué entièrement — même logique que
        main.py:add_top_menu()."""
        menubar = tk.Menu(self)
        bold = ("Segoe UI", 9, "bold")

        for titre, items in core.MENU_STRUCTURE:
            items_visibles = [(label, key) for label, key in items if key in self.remote.menus_autorises]
            if not items_visibles:
                continue
            m = tk.Menu(menubar, tearoff=0)
            for label, key in items_visibles:
                suffix = "" if key in self.IMPLEMENTED_SCREENS else "  (bientôt disponible)"
                m.add_command(label=label + suffix, command=lambda k=key: self.show(k))
            menubar.add_cascade(label=titre, menu=m)
            menubar.entryconfig(menubar.index("end"), font=bold)

        self.config(menu=menubar)

    def show(self, key):
        if key not in self.IMPLEMENTED_SCREENS:
            messagebox.showinfo(
                "Bientôt disponible",
                f"Cet écran n'est pas encore disponible sur le client réseau — utilisez l'application de "
                f"bureau en attendant. Il suivra le même principe que l'écran Saisie une fois construit.",
                parent=self,
            )
            return
        if key not in self.pages:
            self.pages[key] = self.IMPLEMENTED_SCREENS[key](self.content, self.remote)
            self.pages[key].grid(row=0, column=0, sticky="nsew")
        for page_key, page in self.pages.items():
            if page_key == key:
                page.tkraise()

    def _show_accueil(self):
        frame = ttk.Frame(self.content)
        frame.grid(row=0, column=0, sticky="nsew")
        ttk.Label(frame, text="Aucun écran encore disponible pour votre niveau d'accès sur le client réseau.",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=16, pady=16)
        frame.tkraise()

    def _report_callback_exception(self, exc_type, exc_value, exc_traceback):
        """Gestionnaire d'erreurs global — SANS lui, une exception survenant
        dans un écran (ex. Immobilisations avec des données inhabituelles)
        serait silencieusement avalée par Tkinter (surtout en mode
        --windowed, sans console visible) : l'écran resterait vide, SANS
        AUCUN message, ce qui rend le diagnostic impossible pour
        l'utilisateur. Avec ce gestionnaire, toute erreur devient visible."""
        import traceback
        detail = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
        messagebox.showerror(
            "Erreur inattendue",
            f"Une erreur est survenue dans cet écran :\n\n{exc_type.__name__} : {exc_value}\n\n"
            f"Détail technique (à transmettre pour diagnostic) :\n{detail[-1500:]}",
        )

    def _on_close(self):
        self.remote.logout()
        self.destroy()


class RemoteSaisieTab(ttk.Frame):
    """Saisie comptable multi-lignes via le réseau — équivalent distant de
    SaisieTab (main.py), utilisant client_core au lieu de core directement."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.lignes = []  # [{"compte":..., "libelle":..., "debit":..., "credit":..., "quantite":..., "analytic_code":...}, ...]

        header = ttk.LabelFrame(self, text="En-tête de l'écriture")
        header.pack(fill="x", padx=8, pady=8)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(header, textvariable=self.date_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Pièce :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.piece_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.piece_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Journal :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.journal_var = tk.StringVar(value="OD")
        ttk.Combobox(header, textvariable=self.journal_var, width=6, values=("OD", "VE", "AC", "BQ", "CA"),
                     state="readonly").grid(row=0, column=5, padx=4)
        ttk.Label(header, text="Tiers (optionnel) :").grid(row=0, column=6, sticky="w", padx=(12, 4))
        self.tiers_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.tiers_var, width=20).grid(row=0, column=7, padx=4)

        ttk.Label(header, text="Fournisseur :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.fournisseur_var = tk.StringVar()
        self.fournisseur_combo = ttk.Combobox(header, textvariable=self.fournisseur_var, width=22)
        self.fournisseur_combo.grid(row=1, column=1, padx=4, pady=(4, 0))
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self.fournisseur_combo.bind("<FocusOut>", self._on_fournisseur_focusout)
        self._refresh_fournisseur_values()
        ttk.Label(header, text="Client :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.client_var = tk.StringVar()
        self.client_combo = ttk.Combobox(header, textvariable=self.client_var, width=22)
        self.client_combo.grid(row=1, column=3, padx=4, pady=(4, 0))
        self.client_combo.bind("<KeyRelease>", self._on_client_keyrelease)
        self.client_combo.bind("<FocusOut>", self._on_client_focusout)
        self._refresh_client_values()
        ttk.Label(header, text="Code budgétaire :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.budget_var = tk.StringVar()
        self.budget_combo = ttk.Combobox(header, textvariable=self.budget_var, width=16)
        self.budget_combo.grid(row=1, column=5, padx=4, pady=(4, 0))
        self.budget_combo.bind("<FocusOut>", self._on_budget_focusout)
        self._refresh_plan_values(self.budget_combo, "list_budget_codes")
        ttk.Label(header, text="Code bailleur :").grid(row=1, column=6, sticky="w", padx=(12, 4), pady=(4, 0))
        self.bailleur_var = tk.StringVar()
        self.bailleur_combo = ttk.Combobox(header, textvariable=self.bailleur_var, width=16)
        self.bailleur_combo.grid(row=1, column=7, padx=4, pady=(4, 0))
        self.bailleur_combo.bind("<FocusOut>", self._on_bailleur_focusout)
        self._refresh_plan_values(self.bailleur_combo, "list_donor_codes")
        ttk.Label(header, text=(
            "Fournisseur/Client ci-dessus : valeur par défaut pour toutes les lignes utilisant un compte "
            "40x (Fournisseurs) ou 41x (Clients)."
        ), foreground="#595959", wraplength=1000).grid(row=2, column=0, columnspan=8, sticky="w", padx=4, pady=(2, 0))

        stock_frame = ttk.LabelFrame(self, text=(
            "Compte stock (optionnel) — pour une facture globale d'achat (matière + transport/douane) "
            "ou une vente groupée à plusieurs clients : regroupe le mouvement de stock en un seul"))
        stock_frame.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Label(stock_frame, text="Compte stock :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.stock_compte_var = tk.StringVar()
        self.stock_compte_combo = ttk.Combobox(stock_frame, textvariable=self.stock_compte_var, width=28)
        self.stock_compte_combo.grid(row=0, column=1, padx=4)
        self.stock_compte_combo.bind("<KeyRelease>", self._on_stock_compte_keyrelease)
        self._refresh_stock_compte_values()
        ttk.Label(stock_frame, text="Sens :").grid(row=0, column=2, sticky="w", padx=(16, 4))
        self.stock_sens_var = tk.StringVar(value="Entrée (achat)")
        ttk.Combobox(stock_frame, textvariable=self.stock_sens_var, width=16, state="readonly",
                     values=["Entrée (achat)", "Sortie (vente)"]).grid(row=0, column=3, padx=4)
        ttk.Label(stock_frame, text="Quantité :").grid(row=0, column=4, sticky="w", padx=(16, 4))
        self.stock_qte_var = tk.StringVar()
        ttk.Entry(stock_frame, textvariable=self.stock_qte_var, width=12).grid(row=0, column=5, padx=4)
        ttk.Label(stock_frame, text=(
            "Entrée : le coût du stock = somme de TOUTES les lignes au débit ci-dessous (matière + "
            "frais accessoires : transport, douane, assurance...). Laissez « Compte stock » vide pour "
            "revenir au comportement ligne par ligne (une quantité indépendante par ligne)."
        ), foreground="#595959", wraplength=1000).grid(row=1, column=0, columnspan=6, sticky="w", padx=4, pady=(2, 4))

        ligne_frame = ttk.LabelFrame(self, text="Ajouter une ligne (compte au débit OU au crédit, pas les deux)")
        ligne_frame.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Label(ligne_frame, text="Compte :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.compte_var = tk.StringVar()
        self.compte_combo = ttk.Combobox(ligne_frame, textvariable=self.compte_var, width=26)
        self.compte_combo.grid(row=0, column=1, padx=4)
        self.compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
        ttk.Label(ligne_frame, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.libelle_var, width=28).grid(row=0, column=3, padx=4)
        ttk.Label(ligne_frame, text="Débit :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.debit_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.debit_var, width=12).grid(row=0, column=5, padx=4)
        ttk.Label(ligne_frame, text="Crédit :").grid(row=0, column=6, sticky="w", padx=(12, 4))
        self.credit_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.credit_var, width=12).grid(row=0, column=7, padx=4)
        ttk.Label(ligne_frame, text="Quantité :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.quantite_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.quantite_var, width=10).grid(
            row=1, column=1, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(ligne_frame, text="Code analytique :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.analytic_var = tk.StringVar()
        self.analytic_combo = ttk.Combobox(ligne_frame, textvariable=self.analytic_var, width=22)
        self.analytic_combo.grid(row=1, column=3, padx=4, pady=(4, 0))
        self._refresh_plan_values(self.analytic_combo, "list_analytic_codes")
        ttk.Label(ligne_frame, text=(
            "Quantité : requise pour générer automatiquement le mouvement de stock (achats 60x/comptes "
            "stock, ventes 70x). Code analytique : ex. AN-FAB, ENERGIE-EAU, MAINT-..."
        ), foreground="#595959", wraplength=1000).grid(row=2, column=0, columnspan=8, sticky="w", padx=4, pady=(2, 4))
        ttk.Button(ligne_frame, text="Ajouter la ligne", command=self.ajouter_ligne).grid(
            row=1, column=7, padx=12, pady=(4, 0))

        cols = ("compte", "libelle", "debit", "credit")
        self.tree_lignes = ttk.Treeview(self, columns=cols, show="headings", height=8)
        for c, h, w in zip(cols, ["Compte", "Libellé", "Débit", "Crédit"], [110, 340, 120, 120]):
            self.tree_lignes.heading(c, text=h)
            self.tree_lignes.column(c, width=w, anchor="w" if c in ("compte", "libelle") else "e")
        self.tree_lignes.pack(fill="x", padx=8, pady=(0, 4))
        ttk.Button(self, text="Supprimer la ligne sélectionnée", command=self.supprimer_ligne).pack(
            anchor="w", padx=8, pady=(0, 4))

        self.equilibre_var = tk.StringVar(value="Débit : 0   —   Crédit : 0")
        ttk.Label(self, textvariable=self.equilibre_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=8)
        ttk.Button(self, text="Enregistrer l'écriture (via le serveur)", command=self.enregistrer).pack(
            anchor="w", padx=8, pady=8)

        ttk.Separator(self).pack(fill="x", padx=8, pady=4)
        bottom = ttk.Frame(self)
        bottom.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        ttk.Label(bottom, text="Dernières écritures (exercice courant)", font=("Segoe UI", 10, "bold")).pack(
            anchor="w")
        ttk.Label(bottom, text=(
            "Cliquez une ligne pour la charger ci-dessous et la modifier, ou sélectionnez plusieurs "
            "lignes (Ctrl/Maj + clic, ou Ctrl+A) pour les supprimer."
        ), foreground="#595959").pack(anchor="w", pady=(0, 2))
        ttk.Button(bottom, text="Actualiser", command=self.refresh_entries).pack(anchor="w", pady=(2, 4))
        cols2 = ("id", "date", "piece", "journal", "compte", "libelle", "debit", "credit")
        self.tree_entries = ttk.Treeview(bottom, columns=cols2, show="headings", height=14,
                                          selectmode="extended")
        for c, h, w in zip(cols2, ["ID", "Date", "Pièce", "Journal", "Compte", "Libellé", "Débit", "Crédit"],
                           [40, 85, 90, 60, 90, 300, 110, 110]):
            self.tree_entries.heading(c, text=h)
            self.tree_entries.column(c, width=w, anchor="w" if c not in ("debit", "credit") else "e")
        self.tree_entries.pack(fill="both", expand=True, pady=(0, 4))
        self.tree_entries.bind("<<TreeviewSelect>>", self._on_entry_select)
        self.tree_entries.bind("<Control-a>", self._select_all_entries)
        self.tree_entries.bind("<Control-A>", self._select_all_entries)

        edit_frame = ttk.LabelFrame(self, text="Modifier l'écriture sélectionnée (une seule ligne à la fois)")
        edit_frame.pack(fill="x", padx=8, pady=(0, 8))
        self.edit_id_var = tk.StringVar(value="(aucune sélection)")
        ttk.Label(edit_frame, textvariable=self.edit_id_var, font=("Segoe UI", 9, "bold")).grid(
            row=0, column=0, columnspan=6, sticky="w", padx=4, pady=(4, 0))
        ttk.Label(edit_frame, text="Date (JJ/MM/AAAA) :").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.edit_date_var = tk.StringVar()
        ttk.Entry(edit_frame, textvariable=self.edit_date_var, width=12).grid(row=1, column=1, padx=4)
        ttk.Label(edit_frame, text="Pièce :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.edit_piece_var = tk.StringVar()
        ttk.Entry(edit_frame, textvariable=self.edit_piece_var, width=14).grid(row=1, column=3, padx=4)
        ttk.Label(edit_frame, text="Journal :").grid(row=1, column=4, sticky="w", padx=(12, 4))
        self.edit_journal_var = tk.StringVar()
        ttk.Combobox(edit_frame, textvariable=self.edit_journal_var, width=6,
                     values=("OD", "VE", "AC", "BQ", "CA"), state="readonly").grid(row=1, column=5, padx=4)
        ttk.Label(edit_frame, text="Compte :").grid(row=2, column=0, sticky="w", padx=4, pady=4)
        self.edit_compte_var = tk.StringVar()
        self.edit_compte_combo = ttk.Combobox(edit_frame, textvariable=self.edit_compte_var, width=22)
        self.edit_compte_combo.grid(row=2, column=1, padx=4)
        self.edit_compte_combo.bind("<KeyRelease>", self._on_edit_compte_keyrelease)
        ttk.Label(edit_frame, text="Libellé :").grid(row=2, column=2, sticky="w", padx=(12, 4))
        self.edit_libelle_var = tk.StringVar()
        ttk.Entry(edit_frame, textvariable=self.edit_libelle_var, width=28).grid(row=2, column=3, padx=4)
        ttk.Label(edit_frame, text="Débit :").grid(row=2, column=4, sticky="w", padx=(12, 4))
        self.edit_debit_var = tk.StringVar()
        ttk.Entry(edit_frame, textvariable=self.edit_debit_var, width=12).grid(row=2, column=5, padx=4)
        ttk.Label(edit_frame, text="Crédit :").grid(row=2, column=6, sticky="w", padx=(12, 4))
        self.edit_credit_var = tk.StringVar()
        ttk.Entry(edit_frame, textvariable=self.edit_credit_var, width=12).grid(row=2, column=7, padx=4)
        ttk.Label(edit_frame, text="Fournisseur (si compte 40x) :").grid(row=3, column=0, sticky="w", padx=4, pady=4)
        self.edit_fournisseur_var = tk.StringVar()
        self.edit_fournisseur_combo = ttk.Combobox(edit_frame, textvariable=self.edit_fournisseur_var, width=22)
        self.edit_fournisseur_combo.grid(row=3, column=1, padx=4)
        self.edit_fournisseur_combo.bind("<KeyRelease>", self._on_edit_fournisseur_keyrelease)
        ttk.Label(edit_frame, text="Client (si compte 41x) :").grid(row=3, column=2, sticky="w", padx=(12, 4))
        self.edit_client_var = tk.StringVar()
        self.edit_client_combo = ttk.Combobox(edit_frame, textvariable=self.edit_client_var, width=22)
        self.edit_client_combo.grid(row=3, column=3, padx=4)
        self.edit_client_combo.bind("<KeyRelease>", self._on_edit_client_keyrelease)
        edit_btns = ttk.Frame(edit_frame)
        edit_btns.grid(row=4, column=0, columnspan=8, sticky="w", padx=4, pady=6)
        ttk.Button(edit_btns, text="Enregistrer modification", command=self.update_entry_selection).pack(
            side="left", padx=2)
        ttk.Button(edit_btns, text="Supprimer (sélection multiple possible)",
                   command=self.delete_entry_selection).pack(side="left", padx=2)
        ttk.Button(edit_btns, text="Effacer le formulaire", command=self.clear_edit_form).pack(
            side="left", padx=2)

        self.refresh_entries()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_compte_keyrelease(self, event=None):
        query = self.compte_var.get().strip()
        items = self._appeler("search_accounts", query, limit=30)
        if items is not APPEL_ECHEC:
            self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_edit_compte_keyrelease(self, event=None):
        query = self._extract_code(self.edit_compte_var.get())
        items = self._appeler("search_accounts", query, limit=30)
        if items is not APPEL_ECHEC:
            self.edit_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_edit_fournisseur_keyrelease(self, event=None):
        query = self._extract_code(self.edit_fournisseur_var.get())
        items = self._appeler("list_fournisseurs", query or None)
        if items is not APPEL_ECHEC:
            self.edit_fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_edit_client_keyrelease(self, event=None):
        query = self._extract_code(self.edit_client_var.get())
        items = self._appeler("list_clients", query or None)
        if items is not APPEL_ECHEC:
            self.edit_client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _refresh_fournisseur_values(self):
        items = self._appeler("list_fournisseurs")
        if items is not APPEL_ECHEC:
            self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self._extract_code(self.fournisseur_var.get())
        if query:
            items = self._appeler("list_fournisseurs", query)
            if items is not APPEL_ECHEC:
                self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _refresh_client_values(self):
        items = self._appeler("list_clients")
        if items is not APPEL_ECHEC:
            self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _on_client_keyrelease(self, event=None):
        query = self._extract_code(self.client_var.get())
        if query:
            items = self._appeler("list_clients", query)
            if items is not APPEL_ECHEC:
                self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _refresh_plan_values(self, combo, list_fn):
        items = self._appeler(list_fn)
        if items is not APPEL_ECHEC:
            combo["values"] = [f"{c['code']} — {c['label']}" for c in items]

    def _refresh_stock_compte_values(self):
        stocks = self._appeler("compute_stocks_detail", prefixes=["31", "32", "33", "36"])
        if stocks is not APPEL_ECHEC:
            self.stock_compte_combo["values"] = [f"{s['code']} — {s['label']}" for s in stocks]

    def _on_stock_compte_keyrelease(self, event=None):
        query = self._extract_code(self.stock_compte_var.get())
        if query:
            items = self._appeler("search_accounts", query, limit=30)
            if items is not APPEL_ECHEC:
                self.stock_compte_combo["values"] = [
                    f"{a['code']} — {a['label']}" for a in items if a["classe"] == "3"]

    def _on_fournisseur_focusout(self, event=None):
        code = self._extract_code(self.fournisseur_var.get())
        if not code:
            return
        existe = self._appeler("fournisseur_exists", code)
        if existe is not APPEL_ECHEC and not existe:
            messagebox.showerror(
                "Fournisseur invalide",
                f"« {code} » n'est pas un code fournisseur existant. Tapez le nom ou le code pour "
                f"faire apparaître la liste, puis choisissez-y le fournisseur.", parent=self)
            self.fournisseur_var.set("")

    def _on_client_focusout(self, event=None):
        code = self._extract_code(self.client_var.get())
        if not code:
            return
        existe = self._appeler("client_exists", code)
        if existe is not APPEL_ECHEC and not existe:
            messagebox.showerror(
                "Client invalide",
                f"« {code} » n'est pas un code client existant. Tapez le nom ou le code pour faire "
                f"apparaître la liste, puis choisissez-y le client.", parent=self)
            self.client_var.set("")

    def _on_budget_focusout(self, event=None):
        code = self._extract_code(self.budget_var.get())
        if not code:
            return
        existe = self._appeler("budget_code_exists", code)
        if existe is not APPEL_ECHEC and not existe:
            messagebox.showerror(
                "Code budgétaire invalide",
                f"« {code} » n'existe pas dans le plan budgétaire. Choisissez-le dans la liste "
                f"déroulante plutôt que de le taper librement.", parent=self)
            self.budget_var.set("")

    def _on_bailleur_focusout(self, event=None):
        code = self._extract_code(self.bailleur_var.get())
        if not code:
            return
        existe = self._appeler("donor_code_exists", code)
        if existe is not APPEL_ECHEC and not existe:
            messagebox.showerror(
                "Code bailleur invalide",
                f"« {code} » n'existe pas dans le plan des bailleurs. Choisissez-le dans la liste "
                f"déroulante plutôt que de le taper librement.", parent=self)
            self.bailleur_var.set("")

    def _select_all_entries(self, event=None):
        self.tree_entries.selection_set(self.tree_entries.get_children())
        return "break"

    def _extract_code(self, raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def ajouter_ligne(self):
        compte = self._extract_code(self.compte_var.get())
        libelle = self.libelle_var.get().strip()
        if not compte or not libelle:
            messagebox.showwarning("Champ manquant", "Compte et libellé sont obligatoires.", parent=self)
            return
        existe = self._appeler("account_exists", compte)
        if existe is APPEL_ECHEC:
            return
        if not existe:
            messagebox.showerror(
                "Compte invalide",
                f"Le compte « {compte} » n'existe pas dans le plan comptable.\n\n"
                f"Tapez quelques chiffres ou lettres dans le champ Compte pour faire apparaître la liste "
                f"des comptes existants, puis choisissez-en un — ne tapez pas de numéro au hasard.",
                parent=self)
            return
        try:
            debit = float(self.debit_var.get() or 0)
            credit = float(self.credit_var.get() or 0)
            quantite = float(self.quantite_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Débit, Crédit et Quantité doivent être des nombres.", parent=self)
            return
        if debit and credit:
            messagebox.showwarning("Erreur", "Une ligne est soit au débit, soit au crédit — pas les deux.",
                                    parent=self)
            return
        if not debit and not credit:
            messagebox.showwarning("Erreur", "Renseignez un montant au débit ou au crédit.", parent=self)
            return
        analytic_code = self._extract_code(self.analytic_var.get())
        if analytic_code:
            existe = self._appeler("analytic_code_exists", analytic_code)
            if existe is APPEL_ECHEC:
                return
            if not existe:
                messagebox.showerror(
                    "Code analytique invalide",
                    f"Le code analytique « {analytic_code} » n'existe pas. Choisissez-le dans la liste "
                    f"déroulante plutôt que de le taper librement.", parent=self)
                return
        racine = compte[:2] if compte[:1] == "4" else compte[:1]
        fournisseur_defaut = self._extract_code(self.fournisseur_var.get())
        client_defaut = self._extract_code(self.client_var.get())
        if racine == "40" and not fournisseur_defaut:
            messagebox.showerror(
                "Fournisseur manquant",
                f"Le compte « {compte} » relève des Fournisseurs (racine 40) : choisissez le fournisseur "
                f"concerné dans le champ « Fournisseur » de l'en-tête avant d'ajouter cette ligne.",
                parent=self)
            return
        if racine == "41" and not client_defaut:
            messagebox.showerror(
                "Client manquant",
                f"Le compte « {compte} » relève des Clients (racine 41) : choisissez le client concerné "
                f"dans le champ « Client » de l'en-tête avant d'ajouter cette ligne.",
                parent=self)
            return
        self.lignes.append({
            "compte": compte, "libelle": libelle, "debit": debit, "credit": credit,
            "quantite": quantite, "analytic_code": analytic_code,
        })
        self.compte_var.set(""); self.libelle_var.set(""); self.debit_var.set(""); self.credit_var.set("")
        self.quantite_var.set(""); self.analytic_var.set("")
        self._refresh_lignes()

    def supprimer_ligne(self):
        sel = self.tree_lignes.selection()
        if not sel:
            return
        idx = self.tree_lignes.index(sel[0])
        del self.lignes[idx]
        self._refresh_lignes()

    def _refresh_lignes(self):
        for row in self.tree_lignes.get_children():
            self.tree_lignes.delete(row)
        total_debit = total_credit = 0.0
        for l in self.lignes:
            self.tree_lignes.insert("", "end", values=(
                l["compte"], l["libelle"], fmt_cfa(l["debit"]) if l["debit"] else "",
                fmt_cfa(l["credit"]) if l["credit"] else ""))
            total_debit += l["debit"]
            total_credit += l["credit"]
        etat = "✓ Équilibré" if abs(total_debit - total_credit) < 0.01 and self.lignes else ""
        self.equilibre_var.set(f"Débit : {fmt_cfa(total_debit)}   —   Crédit : {fmt_cfa(total_credit)}   {etat}")

    def enregistrer(self):
        if len(self.lignes) < 2:
            messagebox.showwarning("Écriture incomplète", "Ajoutez au moins deux lignes (au moins un débit et un "
                                                            "crédit).", parent=self)
            return
        total_debit = sum(l["debit"] for l in self.lignes)
        total_credit = sum(l["credit"] for l in self.lignes)
        if abs(total_debit - total_credit) >= 0.01:
            messagebox.showwarning("Écriture déséquilibrée",
                                    f"Débit ({fmt_cfa(total_debit)}) ≠ Crédit ({fmt_cfa(total_credit)}).",
                                    parent=self)
            return
        date_str = core.to_iso_date(self.date_var.get().strip())
        piece = self.piece_var.get().strip()
        if not date_str or not piece:
            messagebox.showwarning("Champ manquant", "Date et pièce sont obligatoires.", parent=self)
            return
        journal = self.journal_var.get().strip() or "OD"
        tiers = self.tiers_var.get().strip()
        fournisseur_code = self._extract_code(self.fournisseur_var.get())
        client_code = self._extract_code(self.client_var.get())
        budget_code = self._extract_code(self.budget_var.get())
        donor_code = self._extract_code(self.bailleur_var.get())

        compte_stock_global = self._extract_code(self.stock_compte_var.get()) or None
        quantite_stock_global = 0.0
        sens_stock_global = "entree"
        if compte_stock_global:
            existe = self._appeler("account_exists", compte_stock_global)
            if existe is APPEL_ECHEC:
                return
            if not existe:
                messagebox.showerror("Compte invalide",
                                      f"Le compte stock « {compte_stock_global} » n'existe pas.", parent=self)
                return
            if not self.stock_qte_var.get().strip():
                messagebox.showwarning(
                    "Champ manquant",
                    "La quantité est obligatoire quand un compte stock est choisi.", parent=self)
                return
            try:
                quantite_stock_global = float(self.stock_qte_var.get())
            except ValueError:
                messagebox.showerror("Erreur", "La quantité doit être un nombre.", parent=self)
                return
            if quantite_stock_global <= 0:
                messagebox.showerror("Erreur", "La quantité doit être strictement positive.", parent=self)
                return
            sens_stock_global = "sortie" if self.stock_sens_var.get().startswith("Sortie") else "entree"

        resultat = self._appeler(
            "add_ecriture_multi_lignes", date_str, piece, journal, self.lignes, tiers=tiers,
            fournisseur_code=fournisseur_code, client_code=client_code,
            budget_code=budget_code, donor_code=donor_code,
            compte_stock_global=compte_stock_global, quantite_stock_global=quantite_stock_global,
            sens_stock_global=sens_stock_global)
        if resultat is APPEL_ECHEC:
            return  # erreur déjà affichée par _appeler (session expirée, réseau, ou règle métier)
        messagebox.showinfo("Enregistré", f"Écriture « {piece} » enregistrée sur le serveur.", parent=self)
        self.lignes = []
        self._refresh_lignes()
        self.piece_var.set("")
        self.stock_compte_var.set("")
        self.stock_qte_var.set("")
        self.refresh_entries()

    def refresh_entries(self):
        exercice = self._appeler("get_current_exercice")
        if exercice is APPEL_ECHEC:
            return
        entries = self._appeler("list_entries", exercice=exercice)
        if entries is APPEL_ECHEC:
            return
        self._entries_by_id = {e["id"]: e for e in entries}
        for row in self.tree_entries.get_children():
            self.tree_entries.delete(row)
        for e in entries[-200:][::-1]:  # les 200 plus récentes, plus récentes en premier
            self.tree_entries.insert("", "end", iid=str(e["id"]), values=(
                e["id"], core.to_display_date(e["date"]), e["piece"], e["journal"], e["compte"], e["libelle"],
                fmt_cfa(e["debit"]) if e["debit"] else "", fmt_cfa(e["credit"]) if e["credit"] else ""))

    def _on_entry_select(self, event=None):
        sel = self.tree_entries.selection()
        if len(sel) != 1:
            # sélection multiple (ou vide) : uniquement utile pour la suppression groupée,
            # on ne charge pas le formulaire d'édition pour éviter toute ambiguïté.
            return
        entry = self._entries_by_id.get(int(sel[0]))
        if not entry:
            return
        self.edit_id_var.set(f"Modification de l'écriture ID {entry['id']}")
        self.edit_date_var.set(core.to_display_date(entry["date"]))
        self.edit_piece_var.set(entry["piece"] or "")
        self.edit_journal_var.set(entry["journal"] or "OD")
        self.edit_compte_var.set(entry["compte"])
        self.edit_libelle_var.set(entry["libelle"] or "")
        self.edit_debit_var.set(str(entry["debit"]) if entry["debit"] else "")
        self.edit_credit_var.set(str(entry["credit"]) if entry["credit"] else "")
        self.edit_fournisseur_var.set(entry.get("fournisseur_code") or "")
        self.edit_client_var.set(entry.get("client_code") or "")

    def clear_edit_form(self):
        self.tree_entries.selection_remove(self.tree_entries.selection())
        self.edit_id_var.set("(aucune sélection)")
        self.edit_date_var.set("")
        self.edit_piece_var.set("")
        self.edit_journal_var.set("")
        self.edit_compte_var.set("")
        self.edit_libelle_var.set("")
        self.edit_debit_var.set("")
        self.edit_credit_var.set("")
        self.edit_fournisseur_var.set("")
        self.edit_client_var.set("")

    def update_entry_selection(self):
        sel = self.tree_entries.selection()
        if len(sel) != 1:
            messagebox.showinfo(
                "Info", "Sélectionnez une seule écriture à modifier dans le tableau ci-dessus.", parent=self)
            return
        entry_id = int(sel[0])
        date_str = core.to_iso_date(self.edit_date_var.get().strip())
        compte = self._extract_code(self.edit_compte_var.get())
        libelle = self.edit_libelle_var.get().strip()
        if not date_str or not compte or not libelle:
            messagebox.showwarning("Champ manquant", "Date, compte et libellé sont obligatoires.", parent=self)
            return
        try:
            debit = float(self.edit_debit_var.get() or 0)
            credit = float(self.edit_credit_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Débit et Crédit doivent être des nombres.", parent=self)
            return
        if debit and credit:
            messagebox.showwarning(
                "Erreur", "Une ligne est soit au débit, soit au crédit — pas les deux.", parent=self)
            return
        fournisseur_code = self._extract_code(self.edit_fournisseur_var.get())
        client_code = self._extract_code(self.edit_client_var.get())
        if self._appeler(
            "update_entry", entry_id, date=date_str, piece=self.edit_piece_var.get().strip(),
            journal=self.edit_journal_var.get().strip() or "OD", compte=compte, libelle=libelle,
            debit=debit, credit=credit, fournisseur_code=fournisseur_code, client_code=client_code,
        ) is APPEL_ECHEC:
            return
        messagebox.showinfo("Modifié", f"Écriture ID {entry_id} modifiée sur le serveur.", parent=self)
        self.clear_edit_form()
        self.refresh_entries()

    def delete_entry_selection(self):
        sel = self.tree_entries.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ou plusieurs écritures à supprimer.",
                                 parent=self)
            return
        ids = [int(s) for s in sel]
        n = len(ids)
        if not messagebox.askyesno(
                "Confirmer", f"Supprimer {'cette écriture' if n == 1 else f'ces {n} écritures'} ? "
                             f"Cette action est irréversible.", parent=self):
            return
        if n == 1:
            if self._appeler("delete_entry", ids[0]) is APPEL_ECHEC:
                return
        else:
            resultat = self._appeler("delete_entries_bulk", ids)
            if resultat is APPEL_ECHEC:
                return
            deleted, errors = resultat
            if errors:
                messagebox.showwarning(
                    "Suppression partielle",
                    f"{deleted} écriture(s) supprimée(s). Erreurs :\n" + "\n".join(errors), parent=self)
        self.clear_edit_form()
        self.refresh_entries()


class RemotePersonnelTab(ttk.Frame):
    """Liste du personnel (GRH) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_id = None

        ttk.Label(self, text="LISTE DU PERSONNEL", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Employé")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Matricule :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.matricule_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.matricule_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Nom :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.nom_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.nom_var, width=16).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Prénom :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.prenom_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.prenom_var, width=16).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Poste :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.poste_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.poste_var, width=16).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Service :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.service_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.service_var, width=16).grid(row=1, column=3, padx=4, pady=(4, 0))
        ttk.Label(form, text="Statut :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.statut_var = tk.StringVar(value="actif")
        ttk.Combobox(form, textvariable=self.statut_var, width=13, state="readonly",
                     values=["actif", "congé", "suspendu", "parti"]).grid(row=1, column=5, padx=4, pady=(4, 0))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        cols = ("id", "matricule", "nom", "prenom", "poste", "service", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        for c, h, w in zip(cols, ["ID", "Matricule", "Nom", "Prénom", "Poste", "Service", "Statut"],
                           [40, 100, 130, 130, 150, 130, 90]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.matricule_var.set(v[1]); self.nom_var.set(v[2]); self.prenom_var.set(v[3])
        self.poste_var.set(v[4]); self.service_var.set(v[5]); self.statut_var.set(v[6])

    def clear_form(self):
        self.selected_id = None
        for var in (self.matricule_var, self.nom_var, self.prenom_var, self.poste_var, self.service_var):
            var.set("")
        self.statut_var.set("actif")

    def add(self):
        if not self.nom_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le nom est obligatoire.", parent=self)
            return
        r = self._appeler("add_personnel", self.nom_var.get(), matricule=self.matricule_var.get(),
                           prenom=self.prenom_var.get(), poste=self.poste_var.get(),
                           service=self.service_var.get(), statut=self.statut_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un employé.", parent=self)
            return
        r = self._appeler("update_personnel", self.selected_id, matricule=self.matricule_var.get().strip(),
                           nom=self.nom_var.get().strip(), prenom=self.prenom_var.get().strip(),
                           poste=self.poste_var.get().strip(), service=self.service_var.get().strip(),
                           statut=self.statut_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un employé.", parent=self)
            return
        if messagebox.askyesno("Confirmer", "Supprimer cet employé ?", parent=self):
            r = self._appeler("delete_personnel", self.selected_id)
            if r is APPEL_ECHEC:
                return
            self.clear_form()
            self.refresh()

    def refresh(self):
        personnel = self._appeler("list_personnel")
        if personnel is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for p in personnel:
            self.tree.insert("", "end", values=(
                p["id"], p["matricule"] or "", p["nom"], p["prenom"] or "", p["poste"] or "",
                p["service"] or "", p["statut"]))


class RemoteTimeSheetTab(ttk.Frame):
    """Time sheet (GRH) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote

        ttk.Label(self, text="TIME SHEET", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Nouveau pointage")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Employé :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.personnel_var = tk.StringVar()
        self.personnel_combo = ttk.Combobox(form, textvariable=self.personnel_var, width=26, state="readonly")
        self.personnel_combo.grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Heures :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.heures_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.heures_var, width=8).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Activité :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.activite_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.activite_var, width=40).grid(
            row=1, column=1, columnspan=3, padx=4, pady=(4, 0), sticky="we")
        ttk.Button(form, text="Ajouter le pointage", command=self.add).grid(row=1, column=5, padx=4, pady=(4, 0))

        cols = ("id", "employe", "date", "heures", "activite")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        for c, h, w in zip(cols, ["ID", "Employé", "Date", "Heures", "Activité"], [40, 180, 100, 80, 350]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=16, pady=8)
        ttk.Button(self, text="Supprimer la ligne sélectionnée", command=self.delete_sel).pack(
            anchor="w", padx=16, pady=(0, 12))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _refresh_personnel_values(self):
        personnel = self._appeler("list_personnel", actifs_only=True)
        if personnel is APPEL_ECHEC:
            return
        self.personnel_list = personnel
        self.personnel_combo["values"] = [f"{p['id']} — {p['prenom'] or ''} {p['nom']}".strip() for p in personnel]

    def add(self):
        raw = self.personnel_var.get()
        if not raw:
            messagebox.showwarning("Champ manquant", "Choisissez un employé.", parent=self)
            return
        personnel_id = int(raw.split(" — ", 1)[0])
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "La date est obligatoire.", parent=self)
            return
        try:
            heures = float(self.heures_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Les heures doivent être un nombre.", parent=self)
            return
        r = self._appeler("add_time_sheet", personnel_id, date_str, heures, activite=self.activite_var.get())
        if r is APPEL_ECHEC:
            return
        self.heures_var.set(""); self.activite_var.set("")
        self.refresh()

    def delete_sel(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.", parent=self)
            return
        ts_id = self.tree.item(sel[0], "values")[0]
        r = self._appeler("delete_time_sheet", ts_id)
        if r is APPEL_ECHEC:
            return
        self.refresh()

    def refresh(self):
        self._refresh_personnel_values()
        entries = self._appeler("list_time_sheet")
        if entries is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for t in entries:
            self.tree.insert("", "end", values=(
                t["id"], t["employe"], core.to_display_date(t["date_pointage"]), f"{t['heures']:g}",
                t["activite"] or ""))


class RemoteKpiTab(ttk.Frame):
    """KPI (GRH) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_id = None

        ttk.Label(self, text="KPI — INDICATEURS DE PERFORMANCE", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Indicateur")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Indicateur :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.indicateur_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.indicateur_var, width=28).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Service :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.service_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.service_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Période :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.periode_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.periode_var, width=14).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Valeur cible :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.cible_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.cible_var, width=10).grid(row=1, column=1, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Valeur réalisée :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.realisee_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.realisee_var, width=10).grid(row=1, column=3, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Statut :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.statut_var = tk.StringVar(value="en_cours")
        ttk.Combobox(form, textvariable=self.statut_var, width=13, state="readonly",
                     values=["en_cours", "atteint", "non_atteint"]).grid(row=1, column=5, padx=4, pady=(4, 0))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        cols = ("id", "indicateur", "service", "periode", "cible", "realisee", "taux", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        headers = ["ID", "Indicateur", "Service", "Période", "Cible", "Réalisée", "Taux %", "Statut"]
        widths = [40, 220, 100, 90, 80, 80, 70, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.indicateur_var.set(v[1]); self.service_var.set(v[2]); self.periode_var.set(v[3])
        self.cible_var.set(v[4]); self.realisee_var.set(v[5]); self.statut_var.set(v[7])

    def clear_form(self):
        self.selected_id = None
        for var in (self.indicateur_var, self.service_var, self.periode_var, self.cible_var, self.realisee_var):
            var.set("")
        self.statut_var.set("en_cours")

    def add(self):
        if not self.indicateur_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le nom de l'indicateur est obligatoire.", parent=self)
            return
        try:
            cible = float(self.cible_var.get() or 0)
            realisee = float(self.realisee_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Cible et Réalisée doivent être des nombres.", parent=self)
            return
        r = self._appeler("add_kpi", self.indicateur_var.get(), service=self.service_var.get(),
                           periode=self.periode_var.get(), valeur_cible=cible, valeur_realisee=realisee,
                           statut=self.statut_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un indicateur.", parent=self)
            return
        try:
            cible = float(self.cible_var.get() or 0)
            realisee = float(self.realisee_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Cible et Réalisée doivent être des nombres.", parent=self)
            return
        r = self._appeler("update_kpi", self.selected_id, indicateur=self.indicateur_var.get().strip(),
                           service=self.service_var.get().strip(), periode=self.periode_var.get().strip(),
                           valeur_cible=cible, valeur_realisee=realisee, statut=self.statut_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un indicateur.", parent=self)
            return
        if messagebox.askyesno("Confirmer", "Supprimer cet indicateur ?", parent=self):
            r = self._appeler("delete_kpi", self.selected_id)
            if r is APPEL_ECHEC:
                return
            self.clear_form()
            self.refresh()

    def refresh(self):
        kpis = self._appeler("list_kpi")
        if kpis is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for k in kpis:
            taux = f"{k['taux_realisation']:.0f}" if k["taux_realisation"] is not None else ""
            self.tree.insert("", "end", values=(
                k["id"], k["indicateur"], k["service"] or "", k["periode"] or "",
                f"{k['valeur_cible']:g}", f"{k['valeur_realisee']:g}", taux, k["statut"]))


class RemoteTableauBordGrhTab(ttk.Frame):
    """Tableau de bord GRH (synthèse en lecture seule) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="TABLEAU DE BORD GRH", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16)
        self.cards_frame = ttk.Frame(self)
        self.cards_frame.pack(fill="x", padx=16, pady=16)
        self.hs_frame = ttk.LabelFrame(self, text="Incidents HS ouverts, par gravité")
        self.hs_frame.pack(fill="x", padx=16, pady=8)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _card(self, parent, titre, valeur, col, couleur="#1F4E78"):
        f = ttk.Frame(parent, relief="solid", borderwidth=1)
        f.grid(row=0, column=col, padx=8, sticky="nsew")
        parent.columnconfigure(col, weight=1)
        tk.Label(f, text=titre, font=("Segoe UI", 9), bg="white", fg="#595959").pack(fill="x", padx=12, pady=(10, 0))
        tk.Label(f, text=str(valeur), font=("Segoe UI", 20, "bold"), bg="white", fg=couleur).pack(
            fill="x", padx=12, pady=(0, 10))

    def refresh(self):
        d = self._appeler("compute_tableau_bord_grh")
        if d is APPEL_ECHEC:
            return
        for w in self.cards_frame.winfo_children():
            w.destroy()
        for w in self.hs_frame.winfo_children():
            w.destroy()
        self._card(self.cards_frame, "Personnel actif", f"{d['nb_personnel_actif']} / {d['nb_personnel_total']}", 0)
        self._card(self.cards_frame, "Heures pointées (30j)", f"{d['total_heures_30j']:g} h", 1)
        self._card(self.cards_frame, "KPI en cours", d["nb_kpi_en_cours"], 2)
        self._card(self.cards_frame, "KPI atteints", d["nb_kpi_atteints"], 3, couleur="#1F7A1F")
        self._card(self.cards_frame, "KPI non atteints", d["nb_kpi_non_atteints"], 4, couleur="#B00020")
        self._card(self.cards_frame, "Incidents HS ouverts", d["nb_hs_ouverts"], 5,
                   couleur="#B00020" if d["nb_hs_ouverts"] else "#1F7A1F")
        if not d["hs_par_gravite"]:
            ttk.Label(self.hs_frame, text="Aucun incident ouvert.", foreground="#1F7A1F").pack(
                anchor="w", padx=12, pady=8)
        else:
            for gravite, nb in d["hs_par_gravite"].items():
                ttk.Label(self.hs_frame, text=f"• {gravite} : {nb}").pack(anchor="w", padx=12, pady=2)


class RemoteHsTab(ttk.Frame):
    """HS — Hygiène Santé (GRH) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_id = None

        ttk.Label(self, text="HS — HYGIÈNE SANTÉ", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Événement")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Date (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Type :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.type_var = tk.StringVar(value="incident")
        ttk.Combobox(form, textvariable=self.type_var, width=17, state="readonly",
                     values=["incident", "visite_medicale", "formation_securite", "distribution_epi"]).grid(
            row=0, column=3, padx=4)
        ttk.Label(form, text="Gravité :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.gravite_var = tk.StringVar()
        ttk.Combobox(form, textvariable=self.gravite_var, width=13, state="readonly",
                     values=["", "Mineure", "Modérée", "Grave"]).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Statut :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.statut_var = tk.StringVar(value="ouvert")
        ttk.Combobox(form, textvariable=self.statut_var, width=13, state="readonly",
                     values=["ouvert", "clos"]).grid(row=1, column=1, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Description :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.description_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.description_var, width=50).grid(
            row=1, column=3, columnspan=3, padx=4, pady=(4, 0), sticky="we")

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        cols = ("id", "date", "type", "gravite", "statut", "description")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        headers = ["ID", "Date", "Type", "Gravité", "Statut", "Description"]
        widths = [40, 90, 140, 90, 80, 380]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.date_var.set(v[1]); self.type_var.set(v[2])
        self.gravite_var.set(v[3]); self.statut_var.set(v[4]); self.description_var.set(v[5])

    def clear_form(self):
        self.selected_id = None
        self.date_var.set(date.today().strftime("%d/%m/%Y"))
        self.type_var.set("incident"); self.gravite_var.set(""); self.statut_var.set("ouvert")
        self.description_var.set("")

    def add(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "La date est obligatoire.", parent=self)
            return
        r = self._appeler("add_hs", date_str, type_evenement=self.type_var.get(),
                           description=self.description_var.get(), gravite=self.gravite_var.get(),
                           statut=self.statut_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un événement.", parent=self)
            return
        date_str = core.to_iso_date(self.date_var.get().strip())
        r = self._appeler("update_hs", self.selected_id, date_evenement=date_str, type_evenement=self.type_var.get(),
                           description=self.description_var.get().strip(), gravite=self.gravite_var.get(),
                           statut=self.statut_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un événement.", parent=self)
            return
        if messagebox.askyesno("Confirmer", "Supprimer cet événement ?", parent=self):
            r = self._appeler("delete_hs", self.selected_id)
            if r is APPEL_ECHEC:
                return
            self.clear_form()
            self.refresh()

    def refresh(self):
        events = self._appeler("list_hs")
        if events is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for h in events:
            self.tree.insert("", "end", values=(
                h["id"], core.to_display_date(h["date_evenement"]), h["type_evenement"], h["gravite"] or "",
                h["statut"], h["description"] or ""))


class RemotePaieBulletinsTab(ttk.Frame):
    """Saisie des éléments de gain de chaque employé pour une période de
    paie via le réseau — équivalent réseau complet de PaieBulletinsTab
    (bureau), réutilisant les employés déjà saisis dans GRH > Personnel."""

    CHAMPS = [
        ("classification", "Classification", "combo"),
        ("salaire_base", "Salaire de base", "num"),
        ("prime_anciennete", "Prime d'ancienneté", "num"),
        ("heures_sup", "Heures supplémentaires", "num"),
        ("sursalaire", "Sursalaire", "num"),
        ("gratification", "Gratification", "num"),
        ("indemnite_caisse", "Indemnité Caisse", "num"),
        ("indemnite_logement", "Indemnité Logement", "num"),
        ("indemnite_fonction", "Indemnité Fonction", "num"),
        ("indemnite_transport", "Indemnité Transport", "num"),
        ("personnes_a_charge", "Personnes à charge", "num"),
        ("retenue_pret", "Retenue prêt/avance", "num"),
    ]

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_bulletin_id = None
        self.selected_personnel_id = None
        self.personnel_by_label = {}

        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Label(top, text="Période (AAAA-MM) :").pack(side="left")
        self.periode_var = tk.StringVar(value=date.today().strftime("%Y-%m"))
        ttk.Entry(top, textvariable=self.periode_var, width=10).pack(side="left", padx=4)
        ttk.Button(top, text="Actualiser", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(top, text="Dupliquer vers une autre période...", command=self.dupliquer).pack(
            side="left", padx=8)

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=12, pady=(0, 4))
        ttk.Button(import_bar, text="Importer des bulletins (.xlsx)", command=self.import_xlsx).pack(
            side="left", padx=2)
        ttk.Button(import_bar, text="Télécharger un modèle (.xlsx)", command=self.download_template).pack(
            side="left", padx=2)

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        right = ttk.Frame(body, width=320)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        left = ttk.Frame(body)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))

        cols = ("matricule", "nom", "prenom", "classification", "salaire_base", "net_percu")
        self.tree = ttk.Treeview(left, columns=cols, show="headings")
        headers = ["Matricule", "Nom", "Prénom", "Classification", "Salaire base", "Net perçu (calculé)"]
        widths = [90, 150, 130, 110, 110, 130]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="x")
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        ttk.Label(right, text="Bulletin employé", font=("Segoe UI", 11, "bold")).pack(pady=(0, 8))
        ttk.Label(right, text="Employé :").pack(anchor="w")
        self.employe_var = tk.StringVar()
        self.employe_combo = ttk.Combobox(right, textvariable=self.employe_var, width=32, state="readonly")
        self.employe_combo.pack(anchor="w", pady=(0, 8))

        self.form_vars = {}
        form = ttk.Frame(right)
        form.pack(fill="x")
        for i, (key, label, kind) in enumerate(self.CHAMPS):
            ttk.Label(form, text=label).grid(row=i, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            if kind == "combo":
                w = ttk.Combobox(form, textvariable=var, values=["CADRE", "AUTRE"], state="readonly", width=16)
                var.set("AUTRE")
            else:
                w = ttk.Entry(form, textvariable=var, width=18)
                var.set("0")
            w.grid(row=i, column=1, pady=2, sticky="w")
            self.form_vars[key] = var

        retenues_frame = ttk.LabelFrame(right, text="Retenues et charges (calculées)")
        retenues_frame.pack(fill="x", pady=(12, 8))
        self.retenues_vars = {}
        retenues_labels = [
            ("cnss_salariale", "CNSS (salariale)"),
            ("iuts_net", "IUTS"),
            ("retenue_obligatoire", "Retenue obligatoire 1%"),
            ("retenue_pret", "Retenue prêt/avance"),
            ("net_percu", "NET PERÇU"),
            ("tpa_patronale", "TAP (charge patronale)"),
            ("cnss_patronale", "CNSS (charge patronale)"),
            ("cout_total_employeur", "Coût total employeur"),
        ]
        for i, (key, label) in enumerate(retenues_labels):
            ttk.Label(retenues_frame, text=label + " :").grid(row=i, column=0, sticky="w", padx=4, pady=1)
            var = tk.StringVar(value="—")
            bold = key in ("net_percu", "cout_total_employeur")
            ttk.Label(retenues_frame, textvariable=var,
                      font=("Segoe UI", 9, "bold" if bold else "normal")).grid(
                row=i, column=1, sticky="e", padx=4, pady=1)
            self.retenues_vars[key] = var
        ttk.Button(retenues_frame, text="Calculer (aperçu, sans enregistrer)",
                   command=self._calculer_retenues).grid(row=len(retenues_labels), column=0, columnspan=2,
                                                           pady=(6, 4))

        btns = ttk.Frame(right)
        btns.pack(pady=12)
        ttk.Button(btns, text="Enregistrer le bulletin", command=self.save_bulletin).grid(row=0, column=0, padx=2)
        ttk.Button(btns, text="Supprimer", command=self.delete_bulletin).grid(row=0, column=1, padx=2)
        ttk.Button(right, text="Vider le formulaire", command=self.clear_form).pack()
        ttk.Button(right, text="📄 Aperçu / Imprimer le bulletin (PDF)",
                   command=self.imprimer_bulletin).pack(pady=(8, 0))

        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _refresh_employes(self):
        items = self._appeler("list_personnel", actifs_only=True)
        if items is APPEL_ECHEC:
            return
        self.personnel_by_label = {}
        values = []
        for p in items:
            label = f"{p['matricule'] or p['id']} — {p['nom']} {p['prenom'] or ''}".strip()
            values.append(label)
            self.personnel_by_label[label] = p
        self.employe_combo["values"] = values

    def clear_form(self):
        self.selected_bulletin_id = None
        self.selected_personnel_id = None
        self.employe_var.set("")
        for key, _, kind in self.CHAMPS:
            self.form_vars[key].set("AUTRE" if kind == "combo" else "0")
        for var in self.retenues_vars.values():
            var.set("—")

    def _calculer_retenues(self):
        try:
            bulletin = {}
            for key, _, kind in self.CHAMPS:
                if kind == "combo":
                    bulletin[key] = self.form_vars[key].get()
                elif key == "personnes_a_charge":
                    bulletin[key] = int(float(self.form_vars[key].get() or 0))
                else:
                    bulletin[key] = float(self.form_vars[key].get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Merci de vérifier les valeurs numériques saisies.", parent=self)
            return
        params = self._appeler("get_paie_parametres")
        if params is APPEL_ECHEC:
            return
        resultat = self._appeler("compute_bulletin_paie", bulletin, params)
        if resultat is APPEL_ECHEC:
            return
        for key, var in self.retenues_vars.items():
            var.set(fmt_cfa(resultat.get(key, 0)))

    def imprimer_bulletin(self):
        if not self.selected_bulletin_id:
            messagebox.showinfo("Info", "Sélectionnez (ou enregistrez) d'abord un bulletin dans le tableau.",
                                 parent=self)
            return
        html = self._appeler("render_bulletin_paie_html", self.selected_bulletin_id)
        if html is APPEL_ECHEC:
            return
        import tempfile, webbrowser, os as _os
        path = _os.path.join(tempfile.gettempdir(), f"bulletin_paie_{self.selected_bulletin_id}.html")
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        webbrowser.open(f"file://{path}")

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        personnel_id = int(sel[0])
        self.selected_personnel_id = personnel_id
        periode = self.periode_var.get().strip()
        b = self._appeler("get_bulletin_paie", personnel_id, periode)
        if b is APPEL_ECHEC or not b:
            return
        self.selected_bulletin_id = b["id"]
        p = self._appeler("get_personnel", personnel_id)
        if p is APPEL_ECHEC:
            return
        label = f"{p['matricule'] or p['id']} — {p['nom']} {p['prenom'] or ''}".strip()
        self.employe_var.set(label)
        for key, _, _ in self.CHAMPS:
            self.form_vars[key].set(str(b.get(key, 0)))
        self._calculer_retenues()

    def save_bulletin(self):
        label = self.employe_var.get().strip()
        p = self.personnel_by_label.get(label)
        if not p:
            messagebox.showwarning("Champ manquant", "Choisissez un employé dans la liste.", parent=self)
            return
        periode = self.periode_var.get().strip()
        if len(periode) != 7 or periode[4] != "-":
            messagebox.showerror("Erreur", "Période invalide — format attendu : AAAA-MM (ex. 2026-08).",
                                  parent=self)
            return
        try:
            champs = {}
            for key, _, kind in self.CHAMPS:
                if kind == "combo":
                    champs[key] = self.form_vars[key].get()
                elif key == "personnes_a_charge":
                    champs[key] = int(float(self.form_vars[key].get() or 0))
                else:
                    champs[key] = float(self.form_vars[key].get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Merci de vérifier les valeurs numériques saisies.", parent=self)
            return
        if self._appeler("set_bulletin_paie", p["id"], periode, **champs) is APPEL_ECHEC:
            return
        b = self._appeler("get_bulletin_paie", p["id"], periode)
        if b is not APPEL_ECHEC and b:
            self.selected_bulletin_id = b["id"]
        self.refresh()
        self._calculer_retenues()
        messagebox.showinfo("Enregistré", "Bulletin de paie enregistré.", parent=self)

    def delete_bulletin(self):
        if not self.selected_bulletin_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un bulletin dans le tableau.", parent=self)
            return
        if messagebox.askyesno("Confirmer", "Supprimer ce bulletin de paie ?", parent=self):
            if self._appeler("delete_bulletin_paie", self.selected_bulletin_id) is APPEL_ECHEC:
                return
            self.clear_form()
            self.refresh()

    def dupliquer(self):
        cible = simpledialog.askstring(
            "Dupliquer vers une autre période",
            "Copier les bulletins de la période affichée vers quelle période (AAAA-MM) ?",
            initialvalue=self.periode_var.get().strip(), parent=self)
        if not cible:
            return
        n = self._appeler("dupliquer_bulletins_periode", self.periode_var.get().strip(), cible.strip())
        if n is APPEL_ECHEC:
            return
        messagebox.showinfo("Terminé", f"{n} bulletin(s) dupliqué(s) vers {cible.strip()}.", parent=self)

    def download_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_bulletins_paie.xlsx", title="Enregistrer le modèle", parent=self)
        if not path:
            return
        try:
            core.export_paie_bulletins_template(path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de la création du modèle : {exc}", parent=self)
            return
        messagebox.showinfo("Modèle créé", f"Modèle enregistré :\n{path}", parent=self)

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer des bulletins de paie", parent=self)
        if not path:
            return
        try:
            rows = core.parse_paie_bulletins_xlsx(path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de la lecture du fichier : {exc}", parent=self)
            return
        resultat = self._appeler("apply_paie_bulletins_rows", rows)
        if resultat is APPEL_ECHEC:
            return
        imported, warnings = resultat
        self.refresh()
        msg = f"{imported} bulletin(s) importé(s)/mis à jour sur le serveur."
        if warnings:
            msg += "\n\nAvertissements :\n" + "\n".join(warnings[:20])
        messagebox.showinfo("Import terminé", msg, parent=self)

    def refresh(self):
        self._refresh_employes()
        for row in self.tree.get_children():
            self.tree.delete(row)
        periode = self.periode_var.get().strip()
        if len(periode) != 7:
            return
        etat = self._appeler("compute_paie_periode", periode)
        if etat is APPEL_ECHEC:
            return
        for l in etat["lignes"]:
            self.tree.insert("", "end", iid=str(l["personnel_id"]), values=(
                l["matricule"], l["nom"], l["prenom"] or "", l["classification"],
                fmt_cfa(l["salaire_base"]), fmt_cfa(l["net_percu"])))


class RemotePaieEtatTab(ttk.Frame):
    """État de paie calculé pour une période via le réseau — équivalent
    réseau complet de PaieEtatTab (bureau)."""

    RESULT_COLS = [
        ("matricule", "Matricule", 80), ("nom", "Nom", 120), ("prenom", "Prénom", 100),
        ("remuneration_totale", "Rém. Totale", 100), ("cnss_salariale", "CNSS", 90),
        ("salaire_brut", "Sal. Brut", 100), ("base_imposable", "Base Imp.", 100),
        ("iuts_net", "IUTS", 90), ("salaire_net", "Salaire Net", 100),
        ("retenue_obligatoire", "Ret. Oblig.", 90), ("retenue_pret", "Ret. Prêt", 90),
        ("net_percu", "Net Perçu", 110), ("cout_total_employeur", "Coût Employeur", 120),
    ]

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.last_etat = None

        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Label(top, text="Période (AAAA-MM) :").pack(side="left")
        self.periode_var = tk.StringVar(value=date.today().strftime("%Y-%m"))
        ttk.Entry(top, textvariable=self.periode_var, width=10).pack(side="left", padx=4)
        ttk.Button(top, text="Calculer la paie", command=self.calculer).pack(side="left", padx=12)
        ttk.Button(top, text="Exporter vers Excel", command=self.export_excel).pack(side="left", padx=4)
        ttk.Button(top, text="Valider la paie (comptabiliser)", command=self.valider).pack(side="left", padx=12)

        self.statut_var = tk.StringVar()
        ttk.Label(self, textvariable=self.statut_var, foreground="#B00020", font=("Segoe UI", 9, "bold")).pack(
            anchor="w", padx=12)

        cols = [c[0] for c in self.RESULT_COLS]
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        for key, label, width in self.RESULT_COLS:
            self.tree.heading(key, text=label)
            self.tree.column(key, width=width, anchor="w" if key in ("matricule", "nom", "prenom") else "e")
        self.tree.pack(fill="x", padx=12, pady=(0, 4))
        self.tree.bind("<Double-1>", self._on_double_click)
        ttk.Label(self, text="Double-cliquez une ligne pour l'aperçu avant impression du bulletin.",
                  foreground="#595959").pack(anchor="w", padx=12)

        self.totaux_var = tk.StringVar()
        ttk.Label(self, textvariable=self.totaux_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=12, pady=8)

        self.calculer()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def calculer(self):
        periode = self.periode_var.get().strip()
        if len(periode) != 7 or periode[4] != "-":
            messagebox.showerror("Erreur", "Période invalide — format attendu : AAAA-MM (ex. 2026-08).",
                                  parent=self)
            return
        etat = self._appeler("compute_paie_periode", periode)
        if etat is APPEL_ECHEC:
            return
        self.last_etat = etat
        for row in self.tree.get_children():
            self.tree.delete(row)
        for l in etat["lignes"]:
            values = [fmt_cfa(l[k]) if k not in ("matricule", "nom", "prenom") else l[k]
                      for k, _, _ in self.RESULT_COLS]
            self.tree.insert("", "end", iid=str(l["bulletin_id"]), values=values)
        t = etat["totaux"]
        self.totaux_var.set(
            f"Total Net Perçu : {fmt_cfa(t['net_percu'])}   Total CNSS : {fmt_cfa(t['cnss_total'])}   "
            f"Total IUTS : {fmt_cfa(t['iuts_net'])}   Total Ret. Oblig. : {fmt_cfa(t['retenue_obligatoire'])}   "
            f"Coût total employeur : {fmt_cfa(t['cout_total_employeur'])} F CFA")
        validee = self._appeler("est_periode_paie_validee", periode)
        if validee is not APPEL_ECHEC and validee:
            self.statut_var.set(
                f"✓ Paie de {periode} déjà VALIDÉE (comptabilisée) — les bulletins ne sont plus modifiables.")
        else:
            self.statut_var.set("")
        if not etat["lignes"]:
            messagebox.showinfo("Info", "Aucun bulletin saisi pour cette période — utilisez l'onglet Bulletins.",
                                 parent=self)

    def valider(self):
        periode = self.periode_var.get().strip()
        if len(periode) != 7 or periode[4] != "-":
            messagebox.showerror("Erreur", "Période invalide — format attendu : AAAA-MM (ex. 2026-08).",
                                  parent=self)
            return
        validee = self._appeler("est_periode_paie_validee", periode)
        if validee is APPEL_ECHEC:
            return
        if validee:
            messagebox.showinfo("Info", f"La paie de {periode} est déjà validée.", parent=self)
            return
        etat = self._appeler("compute_paie_periode", periode)
        if etat is APPEL_ECHEC:
            return
        if not etat["lignes"]:
            messagebox.showwarning("Rien à valider", "Aucun bulletin saisi pour cette période.", parent=self)
            return
        t = etat["totaux"]
        if not messagebox.askyesno(
            "Confirmer la validation de la paie",
            f"Valider la paie de {periode} pour {len(etat['lignes'])} employé(s) ?\n\n"
            f"Total Net à payer : {fmt_cfa(t['net_percu'])} F CFA\n"
            f"Total CNSS (salariale + patronale) : {fmt_cfa(t['cnss_total'])} F CFA\n"
            f"Total IUTS : {fmt_cfa(t['iuts_net'])} F CFA\n"
            f"Coût total employeur : {fmt_cfa(t['cout_total_employeur'])} F CFA\n\n"
            f"Cette action envoie les écritures comptables sur le serveur (débit charges de personnel, "
            f"crédit CNSS/IUTS/rémunérations dues) et VERROUILLE les bulletins de cette période — ils ne "
            f"pourront plus être modifiés. Cette action est définitive.", parent=self
        ):
            return
        resultat = self._appeler("valider_paie_periode", periode)
        if resultat is APPEL_ECHEC:
            return
        _, piece = resultat
        messagebox.showinfo("Validation terminée",
                             f"Paie de {periode} comptabilisée (pièce {piece}). Les écritures sont visibles "
                             f"dans le menu SAISIE.", parent=self)
        self.calculer()

    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        bulletin_id = int(sel[0])
        html = self._appeler("render_bulletin_paie_html", bulletin_id)
        if html is APPEL_ECHEC:
            return
        import tempfile, webbrowser, os
        path = os.path.join(tempfile.gettempdir(), f"bulletin_paie_{bulletin_id}.html")
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        webbrowser.open(f"file://{path}")

    def export_excel(self):
        if not self.last_etat or not self.last_etat["lignes"]:
            messagebox.showinfo("Info", "Rien à exporter — calculez d'abord la paie de la période.", parent=self)
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile=f"Etat_paie_{self.periode_var.get().strip()}.xlsx", title="Exporter l'état de paie",
            parent=self)
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messagebox.showerror(
                "Module manquant",
                "L'export Excel nécessite le module openpyxl, absent de cette version du client.",
                parent=self)
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Etat de paie"
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for i, (_, label, _) in enumerate(self.RESULT_COLS, start=1):
            c = ws.cell(row=1, column=i, value=label)
            c.font = header_font
            c.fill = header_fill
        for r, l in enumerate(self.last_etat["lignes"], start=2):
            for i, (key, _, _) in enumerate(self.RESULT_COLS, start=1):
                ws.cell(row=r, column=i, value=l[key])
        wb.save(path)
        messagebox.showinfo("Export terminé", f"État de paie exporté :\n{path}", parent=self)


class RemotePaieParametresTab(ttk.Frame):
    """Paramètres de paie via le réseau — équivalent réseau complet de
    PaieParametresTab (bureau)."""

    CHAMPS = [
        ("taux_cnss_salarie", "Taux CNSS salariale", True),
        ("plafond_cnss", "Plafond rémunération CNSS", False),
        ("cnss_salariale_plafonnee", "CNSS salariale plafonnée", False),
        ("taux_cnss_patronale", "Taux CNSS patronale", True),
        ("taux_tpa", "Taux TPA (patronale)", True),
        ("taux_retenue_obligatoire", "Taux retenue obligatoire", True),
        ("abattement_cadre", "Abattement CADRE", True),
        ("abattement_autre", "Abattement AUTRE", True),
        ("taux_plafond_fiscal", "Taux plafond fiscal", True),
    ]

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="PARAMÈTRES DE PAIE", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Taux et plafonds utilisés pour tous les calculs de paie (CNSS, TPA, abattements). "
            "Les indemnités exonérées (Logement/Fonction/Transport) et le barème IUTS restent aux valeurs "
            "réglementaires par défaut."
        ), foreground="#595959", wraplength=1000).pack(anchor="w", padx=16, pady=(0, 12))

        form = ttk.Frame(self)
        form.pack(anchor="w", padx=16)
        self.vars = {}
        for i, (key, label, is_pct) in enumerate(self.CHAMPS):
            ttk.Label(form, text=label + (" (%)" if is_pct else "") + " :").grid(
                row=i, column=0, sticky="w", padx=4, pady=4)
            var = tk.StringVar()
            ttk.Entry(form, textvariable=var, width=14).grid(row=i, column=1, padx=4, pady=4)
            self.vars[key] = (var, is_pct)
        ttk.Button(self, text="Enregistrer les paramètres", command=self.save).pack(
            anchor="w", padx=16, pady=12)

        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        params = self._appeler("get_paie_parametres")
        if params is APPEL_ECHEC:
            return
        for key, (var, is_pct) in self.vars.items():
            val = params.get(key, 0)
            var.set(str(val * 100 if is_pct else val))

    def save(self):
        params = self._appeler("get_paie_parametres")
        if params is APPEL_ECHEC:
            return
        try:
            for key, (var, is_pct) in self.vars.items():
                val = float(var.get())
                params[key] = val / 100 if is_pct else val
        except ValueError:
            messagebox.showerror("Erreur", "Toutes les valeurs doivent être des nombres.", parent=self)
            return
        if self._appeler("set_paie_parametres", params) is APPEL_ECHEC:
            return
        messagebox.showinfo("Enregistré", "Paramètres de paie enregistrés.", parent=self)


class RemotePaieTab(ttk.Frame):
    """Regroupe la saisie des bulletins, l'état de paie calculé et les
    paramètres via le réseau — équivalent réseau complet de PaieTab
    (bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        inner = ttk.Notebook(self)
        inner.pack(fill="both", expand=True)
        self.bulletins_tab = RemotePaieBulletinsTab(inner, remote)
        self.etat_tab = RemotePaieEtatTab(inner, remote)
        self.params_tab = RemotePaieParametresTab(inner, remote)
        inner.add(self.bulletins_tab, text="Bulletins")
        inner.add(self.etat_tab, text="État de paie")
        inner.add(self.params_tab, text="Paramètres de paie")

    def refresh(self):
        self.bulletins_tab.refresh()
        self.etat_tab.calculer()
        self.params_tab.refresh()


class RemoteRapprochementCompteDialog(tk.Toplevel):
    """Rapprochement bancaire d'UN SEUL compte via le réseau — équivalent
    réseau complet de RapprochementCompteDialog (bureau)."""

    def __init__(self, parent, remote: RemoteConnection, compte, libelle, on_saved=None):
        super().__init__(parent)
        self.remote = remote
        self.compte = compte
        self.on_saved = on_saved
        self.title(f"Rapprochement bancaire — {compte} {libelle}")
        self.geometry("820x600")
        self.transient(parent)
        self.grab_set()
        self._row_entry_ids = {}

        ttk.Label(self, text=f"{compte} — {libelle}", font=("Segoe UI", 13, "bold")).pack(
            anchor="w", padx=12, pady=(12, 4))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=12, pady=4)
        ttk.Label(filt, text="Du (JJ/MM/AAAA) :").pack(side="left")
        self.date_from_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_from_var, width=12).pack(side="left", padx=4)
        ttk.Label(filt, text="Au (JJ/MM/AAAA) :").pack(side="left", padx=(12, 0))
        self.date_to_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_to_var, width=12).pack(side="left", padx=4)
        ttk.Button(filt, text="Afficher", command=self.refresh).pack(side="left", padx=8)
        ttk.Label(filt, text="Solde du relevé bancaire :").pack(side="left", padx=(20, 4))
        self.solde_releve_var = tk.StringVar()
        releve_entry = ttk.Entry(filt, textvariable=self.solde_releve_var, width=16)
        releve_entry.pack(side="left")
        releve_entry.bind("<KeyRelease>", lambda e: self._maj_ecart())
        ttk.Button(filt, text="📄 État de rapprochement (aperçu avant impression, PDF)",
                   command=self.afficher_etat).pack(side="left", padx=(16, 0))

        cols = ("pointe", "date", "piece", "libelle", "debit", "credit", "solde")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        headers = ["Pointé", "Date", "Pièce", "Libellé", "Débit", "Crédit", "Solde cumulé"]
        widths = [55, 85, 70, 260, 100, 100, 120]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            anchor = "center" if c in ("pointe", "debit", "credit", "solde") else "w"
            self.tree.column(c, width=w, anchor=anchor)
        self.tree.tag_configure("pointe", background="#D9EAD3")
        self.tree.pack(fill="both", expand=True, padx=12, pady=8)
        self.tree.bind("<Button-1>", self._on_click)

        self.synthese_var = tk.StringVar()
        ttk.Label(self, textvariable=self.synthese_var, font=("Segoe UI", 11, "bold")).pack(
            anchor="w", padx=12, pady=(0, 4))
        self.ecart_var = tk.StringVar()
        self.ecart_label = ttk.Label(self, textvariable=self.ecart_var, font=("Segoe UI", 12, "bold"))
        self.ecart_label.pack(anchor="w", padx=12, pady=(0, 12))

        self._solde_debut = 0.0
        self._total_pointe = 0.0
        self._solde_fin = 0.0
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._row_entry_ids = {}
        date_from = core.to_iso_date(self.date_from_var.get()) if self.date_from_var.get().strip() else None
        date_to = core.to_iso_date(self.date_to_var.get()) if self.date_to_var.get().strip() else None
        comptes = self._appeler("compute_mouvements_prefixe_periode", self.compte, date_from=date_from,
                                 date_to=date_to)
        if comptes is APPEL_ECHEC:
            return
        if not comptes:
            self.synthese_var.set("Aucun mouvement sur cette période pour ce compte.")
            self.ecart_var.set("")
            return
        c = comptes[0]
        for m in c["mouvements"]:
            iid = self.tree.insert("", "end", tags=("pointe",) if m["pointe"] else (), values=(
                "☑" if m["pointe"] else "☐", core.to_display_date(m["date"]), m["piece"] or "",
                m["libelle"] or "", f"{fmt_cfa(m['debit'])}" if m["debit"] else "",
                f"{fmt_cfa(m['credit'])}" if m["credit"] else "", f"{fmt_cfa(m['solde_cumule'])}",
            ))
            self._row_entry_ids[iid] = m["id"]
        self._solde_debut = c["solde_debut_periode"]
        self._total_pointe = c["total_pointe"]
        self._solde_fin = c["solde_fin_periode"]
        self.synthese_var.set(
            f"Solde comptable début de période : {fmt_cfa(self._solde_debut)}    "
            f"Solde comptable fin de période : {fmt_cfa(self._solde_fin)}")
        self._maj_ecart()

    def _maj_ecart(self):
        solde_pointe = self._solde_debut + self._total_pointe
        if self.solde_releve_var.get().strip():
            try:
                solde_releve = float(self.solde_releve_var.get().replace(" ", "").replace(",", "."))
                ecart = solde_pointe - solde_releve
                if abs(ecart) < 1:
                    self.ecart_var.set(f"Solde pointé : {fmt_cfa(solde_pointe)}  —  ✓ Rapprochement OK "
                                        f"(écart : {fmt_cfa(ecart)})")
                    self.ecart_label.configure(foreground="#1F7A1F")
                else:
                    self.ecart_var.set(f"Solde pointé : {fmt_cfa(solde_pointe)}  —  Écart avec le relevé : "
                                        f"{fmt_cfa(ecart)}")
                    self.ecart_label.configure(foreground="#B00020")
                return
            except ValueError:
                pass
        self.ecart_var.set(f"Solde pointé (mouvements cochés) : {fmt_cfa(solde_pointe)} — "
                            f"saisissez le solde du relevé bancaire pour voir l'écart.")
        self.ecart_label.configure(foreground="black")

    def afficher_etat(self):
        solde_releve = None
        if self.solde_releve_var.get().strip():
            try:
                solde_releve = float(self.solde_releve_var.get().replace(" ", "").replace(",", "."))
            except ValueError:
                pass
        date_from = core.to_iso_date(self.date_from_var.get()) if self.date_from_var.get().strip() else None
        date_to = core.to_iso_date(self.date_to_var.get()) if self.date_to_var.get().strip() else None
        html = self._appeler("render_etat_rapprochement_html", self.compte, date_from=date_from,
                              date_to=date_to, solde_releve=solde_releve)
        if html is APPEL_ECHEC:
            return
        import tempfile, webbrowser, os
        path = os.path.join(tempfile.gettempdir(), f"etat_rapprochement_{self.compte}.html")
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        webbrowser.open(f"file://{path}")

    def _on_click(self, event):
        if self.tree.identify_region(event.x, event.y) != "cell":
            return
        if self.tree.identify_column(event.x) != "#1":
            return
        row = self.tree.identify_row(event.y)
        if row not in self._row_entry_ids:
            return
        entry_id = self._row_entry_ids[row]
        deja_pointe = self.tree.set(row, "pointe") == "☑"
        if self._appeler("set_pointage_bancaire", entry_id, not deja_pointe) is APPEL_ECHEC:
            return
        self.refresh()
        if self.on_saved:
            self.on_saved()


class RemoteArreteComptesTab(ttk.Frame):
    """Tableau de vérification avant clôture (« arrêté de comptes ») via
    le réseau — équivalent réseau complet de ArreteComptesTab (bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.last_resultat = None

        top = ttk.Frame(self)
        top.pack(fill="x", padx=16, pady=(16, 4))
        ttk.Label(top, text="ARRÊTÉ DE COMPTES", font=("Segoe UI", 14, "bold")).pack(side="left")
        ttk.Label(top, text="   Date d'arrêté (JJ/MM/AAAA) :").pack(side="left", padx=(20, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(top, textvariable=self.date_var, width=12).pack(side="left")
        ttk.Button(top, text="Calculer", command=self.calculer).pack(side="left", padx=12)
        ttk.Label(self, text=(
            "Photographie des points à vérifier avant de clôturer une période — ne modifie rien. "
            "Cliquez sur un onglet pour le détail de chaque contrôle."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        self.tab_fournisseurs = self._make_tab_tree(
            "Fournisseurs",
            "Comptes 40x avec mouvement sur l'exercice. En rouge : solde DÉBITEUR (anomalie possible — "
            "un fournisseur ne devrait normalement pas nous devoir de l'argent, sauf avance versée).",
            ("compte", "libelle", "solde"), ["Compte", "Libellé", "Solde"], [90, 320, 130])

        self.tab_fnp = self._make_tab_tree(
            "Factures non parvenues",
            "Commandes fournisseurs LIVRÉES mais pour lesquelles aucune facture d'achat n'a été saisie "
            "depuis — à vérifier manuellement (détection approximative, pas de lien direct commande / "
            "facture dans le logiciel).",
            ("piece", "fournisseur", "montant", "date_commande", "date_livraison"),
            ["Pièce", "Fournisseur", "Montant", "Date commande", "Date livraison"], [90, 220, 110, 110, 110])

        self.tab_clients = self._make_tab_tree(
            "Clients",
            "Balance âgée des impayés clients (voir COMMERCIAL > Paiement/Recouvrement pour le détail par tranche).",
            ("client", "total"), ["Client", "Total impayé"], [300, 150])

        self.tab_banques = self._make_tab_tree(
            "Rapprochements bancaires",
            "Solde comptable de chaque compte banque à la date d'arrêté — cliquez une ligne pour pointer "
            "ses mouvements par rapport au relevé bancaire.",
            ("compte", "libelle", "solde"), ["Compte", "Libellé", "Solde comptable"], [90, 320, 150])
        self.tab_banques.bind("<<TreeviewSelect>>", self._on_select_banque)

        self.tab_impots = self._make_tab_tree(
            "Impôts & charges sociales",
            "Soldes des comptes TVA (443/444), IUTS/retenues (447) et CNSS (43) à la date d'arrêté.",
            ("compte", "libelle", "solde"), ["Compte", "Libellé", "Solde"], [90, 320, 130])

        self.tab_engagements = self._make_tab_tree(
            "Engagements en retard",
            "Commandes fournisseurs et factures clients en dépassement de délai (livraison ou paiement).",
            ("type", "piece", "tiers", "montant", "statut"),
            ["Type", "Pièce", "Tiers", "Montant", "Statut"], [90, 90, 220, 110, 180])

        self.tab_paie = self._make_tab_tree(
            "Paie",
            "Statut des 3 dernières périodes de paie saisies (validée = déjà comptabilisée et verrouillée).",
            ("periode", "statut"), ["Période", "Statut"], [120, 250])

        self.calculer()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _make_tab_tree(self, titre, description, cols, headers, widths):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text=titre)
        ttk.Label(frame, text=description, foreground="#595959", wraplength=1050).pack(
            anchor="w", padx=8, pady=8)
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        for c, h, w in zip(cols, headers, widths):
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor="w")
        tree.tag_configure("alerte", foreground="#B00020")
        tree.pack(fill="x", padx=8, pady=(0, 8))
        return tree

    def _on_select_banque(self, event=None):
        sel = self.tab_banques.selection()
        if not sel:
            return
        v = self.tab_banques.item(sel[0], "values")
        compte, libelle = v[0], v[1]
        RemoteRapprochementCompteDialog(self, self.remote, compte, libelle, on_saved=self.calculer)

    def calculer(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showerror("Erreur", "Date invalide — format attendu : JJ/MM/AAAA.", parent=self)
            return
        resultat = self._appeler("compute_arrete_comptes", date_arrete=date_str)
        if resultat is APPEL_ECHEC:
            return
        self.last_resultat = resultat

        for tree in (self.tab_fournisseurs, self.tab_fnp, self.tab_clients, self.tab_banques,
                     self.tab_impots, self.tab_engagements, self.tab_paie):
            for row in tree.get_children():
                tree.delete(row)

        anomalies_comptes = {b["code"] for b in resultat["fournisseurs"]["anomalies"]}
        for b in resultat["fournisseurs"]["comptes"]:
            tag = ("alerte",) if b["code"] in anomalies_comptes else ()
            self.tab_fournisseurs.insert("", "end", tags=tag, values=(
                b["code"], b["label"], fmt_cfa(b["solde_cloture"])))

        for c in resultat["factures_non_parvenues"]:
            self.tab_fnp.insert("", "end", tags=("alerte",), values=(
                c["piece"] or "", c["raison_sociale"], fmt_cfa(c["montant"]),
                core.to_display_date(c["date_commande"]), core.to_display_date(c["date_livraison_reelle"])))

        for cl in resultat["clients_balance_agee"]:
            self.tab_clients.insert("", "end", tags=("alerte",), values=(
                cl["raison_sociale"], fmt_cfa(cl["total"])))

        for b in resultat["rapprochements_bancaires"]:
            self.tab_banques.insert("", "end", values=(b["compte"], b["libelle"], fmt_cfa(b["solde_comptable"])))

        for c in resultat["impots"] + resultat["charges_sociales"]:
            if c["solde_fin_periode"]:
                self.tab_impots.insert("", "end", values=(c["code"], c["label"], fmt_cfa(c["solde_fin_periode"])))

        for c in resultat["engagements_fournisseurs_retard"]:
            self.tab_engagements.insert("", "end", tags=("alerte",), values=(
                "Fournisseur", c["piece"] or "", c["raison_sociale"], fmt_cfa(c["montant"]),
                c["statut_livraison"] if c["depassement_livraison"] else c["statut_paiement"]))
        for f in resultat["factures_clients_retard"]:
            self.tab_engagements.insert("", "end", tags=("alerte",), values=(
                "Client", f["piece"] or "", f["raison_sociale"], fmt_cfa(f["montant"]), f["statut_paiement"]))

        for p in resultat["paie_statuts"]:
            tag = () if p["validee"] else ("alerte",)
            self.tab_paie.insert("", "end", tags=tag, values=(
                p["periode"], "✓ Validée (comptabilisée)" if p["validee"] else "En attente de validation"))

    def refresh(self):
        self.calculer()


class RemoteFournisseursTab(ttk.Frame):
    """Fournisseurs (ENGAGEMENTS-PROJETS) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_code = None

        ttk.Label(self, text="FOURNISSEURS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Fournisseur")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Code :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.code_var = tk.StringVar()
        self.code_entry = ttk.Entry(form, textvariable=self.code_var, width=14)
        self.code_entry.grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Raison sociale :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.raison_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.raison_var, width=30).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Contact :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.contact_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.contact_var, width=20).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Téléphone :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.telephone_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.telephone_var, width=18).grid(row=1, column=3, padx=4, pady=(4, 0))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        recherche = ttk.Frame(self)
        recherche.pack(fill="x", padx=16, pady=(0, 4))
        ttk.Label(recherche, text="Rechercher :").pack(side="left")
        self.recherche_var = tk.StringVar()
        recherche_entry = ttk.Entry(recherche, textvariable=self.recherche_var, width=30)
        recherche_entry.pack(side="left", padx=4)
        recherche_entry.bind("<KeyRelease>", lambda e: self.refresh())

        web_bar = ttk.LabelFrame(self, text="Trouver de nouveaux fournisseurs sur Internet")
        web_bar.pack(fill="x", padx=16, pady=(4, 4))
        ttk.Label(web_bar, text="Produit recherché :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.web_produit_var = tk.StringVar()
        ttk.Entry(web_bar, textvariable=self.web_produit_var, width=28).grid(row=0, column=1, padx=4)
        ttk.Label(web_bar, text="Ville :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.web_ville_var = tk.StringVar(value="Ouagadougou")
        ttk.Entry(web_bar, textvariable=self.web_ville_var, width=18).grid(row=0, column=3, padx=4)
        ttk.Button(web_bar, text="Rechercher sur Internet", command=self.rechercher_internet).grid(
            row=0, column=4, padx=12)
        ttk.Label(web_bar, text=(
            "Ouvre votre navigateur avec une recherche Google déjà remplie. Les résultats s'affichent "
            "dans le navigateur, pas ici — copiez ensuite les coordonnées du fournisseur choisi dans le "
            "formulaire ci-dessus pour l'enregistrer."
        ), foreground="#595959", wraplength=1050).grid(row=1, column=0, columnspan=5, sticky="w", padx=4, pady=(2, 4))

        cols = ("code", "raison_sociale", "contact", "telephone")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        for c, h, w in zip(cols, ["Code", "Raison sociale", "Contact", "Téléphone"], [100, 280, 180, 140]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def rechercher_internet(self):
        produit = self.web_produit_var.get().strip()
        if not produit:
            messagebox.showwarning("Champ manquant", "Indiquez le produit recherché.", parent=self)
            return
        ville = self.web_ville_var.get().strip()
        requete = f"fournisseur {produit} {ville}".strip()
        import webbrowser
        from urllib.parse import quote_plus
        webbrowser.open(f"https://www.google.com/search?q={quote_plus(requete)}")

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_code = v[0]
        self.code_var.set(v[0]); self.raison_var.set(v[1])
        self.contact_var.set(v[2]); self.telephone_var.set(v[3])
        self.code_entry.configure(state="disabled")

    def clear_form(self):
        self.selected_code = None
        self.code_var.set(""); self.raison_var.set(""); self.contact_var.set(""); self.telephone_var.set("")
        self.code_entry.configure(state="normal")

    def add(self):
        if not self.code_var.get().strip() or not self.raison_var.get().strip():
            messagebox.showwarning("Champ manquant", "Code et raison sociale sont obligatoires.", parent=self)
            return
        r = self._appeler("add_fournisseur", self.code_var.get(), self.raison_var.get(),
                           contact=self.contact_var.get(), telephone=self.telephone_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un fournisseur.", parent=self)
            return
        r = self._appeler("add_fournisseur", self.selected_code, self.raison_var.get(),
                           contact=self.contact_var.get(), telephone=self.telephone_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un fournisseur.", parent=self)
            return
        if messagebox.askyesno("Confirmer", f"Supprimer le fournisseur « {self.selected_code} » ?", parent=self):
            r = self._appeler("delete_fournisseur", self.selected_code)
            if r is APPEL_ECHEC:
                return
            self.clear_form()
            self.refresh()

    def refresh(self):
        fournisseurs = self._appeler("list_fournisseurs", self.recherche_var.get().strip() or None)
        if fournisseurs is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for f in fournisseurs:
            self.tree.insert("", "end", values=(
                f["code"], f["raison_sociale"], f["contact"] or "", f["telephone"] or ""))


class RemoteReglementsTab(ttk.Frame):
    """Règlements fournisseurs (ENGAGEMENTS-PROJETS) via le réseau —
    reçoit automatiquement les factures d'achat validées (déjà
    comptabilisées, pas de nouvelle écriture ici), permet aussi d'en
    créer directement. Choisir la banque/caisse et enregistrer le
    paiement comptabilise Débit fournisseur (401000) / Crédit banque."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.reglement_id_selectionne = None
        self.lignes = []

        ttk.Label(self, text="RÈGLEMENTS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))
        ttk.Label(self, text=(
            "Les factures d'achat validées (menu ENGAGEMENTS-PROJETS > Factures frs) apparaissent ici "
            "automatiquement, déjà « validée » — la charge est déjà comptabilisée. Sélectionnez-en une, "
            "choisissez la banque/caisse et cliquez « Enregistrer le paiement »."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        header = ttk.LabelFrame(self, text="Nouveau règlement (créé directement, hors circuit Factures frs)")
        header.pack(fill="x", padx=16, pady=4)
        ttk.Label(header, text="Numéro :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.numero_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.numero_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(header, textvariable=self.date_var, width=12).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Fournisseur :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.fournisseur_var = tk.StringVar()
        self.fournisseur_combo = ttk.Combobox(header, textvariable=self.fournisseur_var, width=26)
        self.fournisseur_combo.grid(row=0, column=5, padx=4)
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self._refresh_fournisseur_values()
        ttk.Button(header, text="Créer le règlement", command=self.creer).grid(row=0, column=6, padx=12)

        ligne_frame = ttk.LabelFrame(self, text="Lignes (une fois le règlement créé, sélectionné dans la liste)")
        ligne_frame.pack(fill="x", padx=16, pady=(0, 8))
        ttk.Label(ligne_frame, text="Compte de charge :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.compte_var = tk.StringVar()
        self.compte_combo = ttk.Combobox(ligne_frame, textvariable=self.compte_var, width=26)
        self.compte_combo.grid(row=0, column=1, padx=4)
        self.compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
        ttk.Label(ligne_frame, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.libelle_var, width=26).grid(row=0, column=3, padx=4)
        ttk.Label(ligne_frame, text="Quantité :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.quantite_var = tk.StringVar(value="1")
        ttk.Entry(ligne_frame, textvariable=self.quantite_var, width=8).grid(row=0, column=5, padx=4)
        ttk.Label(ligne_frame, text="Prix unitaire :").grid(row=0, column=6, sticky="w", padx=(12, 4))
        self.prix_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.prix_var, width=12).grid(row=0, column=7, padx=4)
        ttk.Button(ligne_frame, text="Ajouter la ligne", command=self.ajouter_ligne).grid(row=0, column=8, padx=12)

        self.tree_lignes = ttk.Treeview(self, columns=("compte", "libelle", "qte", "prix"), show="headings",
                                         height=5)
        for c, h, w in zip(("compte", "libelle", "qte", "prix"), ["Compte", "Libellé", "Qté", "Prix unit."],
                           [100, 300, 60, 110]):
            self.tree_lignes.heading(c, text=h)
            self.tree_lignes.column(c, width=w, anchor="w")
        self.tree_lignes.pack(fill="x", padx=16, pady=(0, 4))
        ttk.Button(self, text="Valider le règlement (comptabilise l'achat sur le serveur — règlement "
                               "créé directement uniquement)", command=self.valider).pack(
            anchor="w", padx=16, pady=(4, 8))

        echeancier_frame = ttk.LabelFrame(self, text="Échéancier de paiement du règlement sélectionné")
        echeancier_frame.pack(fill="x", padx=16, pady=(0, 8))
        ttk.Label(echeancier_frame, text=(
            "Une ou plusieurs tranches — pour un engagement payable en plusieurs fois sur plusieurs "
            "mois plutôt qu'en un seul versement."
        ), foreground="#595959", wraplength=1050).grid(row=0, column=0, columnspan=6, sticky="w", padx=4, pady=(4, 2))
        ttk.Button(echeancier_frame, text="Modifier l'échéancier (plusieurs tranches possibles)",
                   command=self.modifier_echeancier).grid(row=1, column=0, padx=4, pady=4, sticky="w")
        cols_ech = ("id", "tranche", "date", "montant", "statut")
        self.tree_echeances = ttk.Treeview(echeancier_frame, columns=cols_ech, show="headings", height=4)
        headers_ech = ["ID", "Tranche", "Échéance", "Montant", "Statut"]
        for c, h, w in zip(cols_ech, headers_ech, [0, 60, 100, 120, 100]):
            self.tree_echeances.heading(c, text=h)
            self.tree_echeances.column(c, width=w, anchor="w", stretch=(c != "id"))
        self.tree_echeances.column("id", width=0, stretch=False)
        self.tree_echeances["displaycolumns"] = ("tranche", "date", "montant", "statut")
        self.tree_echeances.grid(row=2, column=0, columnspan=6, sticky="we", padx=4, pady=4)
        self.tree_echeances.tag_configure("retard", foreground="#B00020")
        self.tree_echeances.tag_configure("payee", foreground="#1F7A1F")

        ttk.Label(echeancier_frame, text="Payer la tranche sélectionnée — Date :").grid(
            row=3, column=0, sticky="w", padx=4, pady=(4, 4))
        self.date_paiement_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(echeancier_frame, textvariable=self.date_paiement_var, width=12).grid(
            row=3, column=1, padx=4, sticky="w")
        ttk.Label(echeancier_frame, text="Compte banque/caisse :").grid(row=3, column=2, sticky="w", padx=(12, 4))
        self.compte_paiement_var = tk.StringVar()
        self.compte_paiement_combo = ttk.Combobox(echeancier_frame, textvariable=self.compte_paiement_var, width=28)
        self.compte_paiement_combo.grid(row=3, column=3, padx=4, sticky="w")
        self.compte_paiement_combo.bind("<KeyRelease>", self._on_compte_paiement_keyrelease)
        self._refresh_compte_paiement_values()
        ttk.Button(echeancier_frame, text="Enregistrer le paiement de cette tranche",
                   command=self.enregistrer_paiement).grid(row=3, column=4, padx=12)
        self.paiement_statut_var = tk.StringVar()
        ttk.Label(echeancier_frame, textvariable=self.paiement_statut_var, foreground="#1F7A1F").grid(
            row=4, column=0, columnspan=6, sticky="w", padx=4, pady=(2, 4))

        ttk.Separator(self).pack(fill="x", padx=16, pady=4)
        ttk.Label(self, text="Factures à régler", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16)
        filtre_bar = ttk.Frame(self)
        filtre_bar.pack(fill="x", padx=16, pady=(2, 4))
        ttk.Button(filtre_bar, text="Actualiser", command=self.refresh).pack(side="left", padx=(0, 12))
        self.filtre_var = tk.StringVar(value="Toutes")
        for f in ["Toutes", "Non soldées", "En retard", "Partiellement payées", "Soldées", "Brouillons"]:
            ttk.Radiobutton(filtre_bar, text=f, value=f, variable=self.filtre_var, command=self.refresh).pack(
                side="left", padx=4)
        cols = ("id", "numero", "date", "fournisseur", "montant", "statut", "paiement")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=10)
        headers = ["ID", "Numéro", "Date", "Fournisseur", "Montant", "Ventilation", "Paiement"]
        for c, h, w in zip(cols, headers, [0, 100, 90, 220, 110, 150, 200]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w", stretch=(c != "id"))
        self.tree.column("id", width=0, stretch=False)
        self.tree["displaycolumns"] = ("numero", "date", "fournisseur", "montant", "statut", "paiement")
        self.tree.tag_configure("retard", foreground="#B00020")
        self.tree.tag_configure("soldee", foreground="#1F7A1F")
        self.tree.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.tree.bind("<<TreeviewSelect>>", self._on_select_reglement)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _refresh_fournisseur_values(self):
        items = self._appeler("list_fournisseurs")
        if items is not APPEL_ECHEC:
            self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self._extract_code(self.fournisseur_var.get())
        if query:
            items = self._appeler("list_fournisseurs", query)
            if items is not APPEL_ECHEC:
                self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _refresh_compte_values(self):
        items = self._appeler("search_accounts", "6", limit=100)
        if items is not APPEL_ECHEC:
            self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "6"]

    def _on_compte_keyrelease(self, event=None):
        query = self._extract_code(self.compte_var.get())
        if query:
            items = self._appeler("search_accounts", query, limit=50)
            if items is not APPEL_ECHEC:
                self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "6"]

    def _refresh_compte_paiement_values(self):
        items = self._appeler("search_accounts", "5", limit=100)
        if items is not APPEL_ECHEC:
            self.compte_paiement_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "5"]

    def _on_compte_paiement_keyrelease(self, event=None):
        query = self._extract_code(self.compte_paiement_var.get())
        if query:
            items = self._appeler("search_accounts", query, limit=50)
            if items is not APPEL_ECHEC:
                self.compte_paiement_combo["values"] = [
                    f"{a['code']} — {a['label']}" for a in items if a["classe"] == "5"]

    def creer(self):
        if not self.numero_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le numéro est obligatoire.", parent=self)
            return
        date_str = core.to_iso_date(self.date_var.get().strip())
        rid = self._appeler("create_reglement", self.numero_var.get(), date_str,
                             fournisseur_code=self._extract_code(self.fournisseur_var.get()))
        if rid is APPEL_ECHEC:
            return
        self.reglement_id_selectionne = rid
        messagebox.showinfo("Créé", f"Règlement « {self.numero_var.get()} » créé (brouillon) — ajoutez des "
                                     f"lignes puis validez.", parent=self)
        self.numero_var.set("")
        self.refresh()

    def ajouter_ligne(self):
        if not self.reglement_id_selectionne:
            messagebox.showinfo("Info", "Créez ou sélectionnez d'abord un règlement dans la liste.", parent=self)
            return
        compte = self._extract_code(self.compte_var.get())
        libelle = self.libelle_var.get().strip()
        if not compte or not libelle:
            messagebox.showwarning("Champ manquant", "Compte de charge et libellé sont obligatoires.", parent=self)
            return
        try:
            qte = float(self.quantite_var.get() or 0)
            prix = float(self.prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        r = self._appeler("add_ligne_reglement", self.reglement_id_selectionne, compte, libelle, qte,
                           prix_unitaire=prix)
        if r is APPEL_ECHEC:
            return
        self.compte_var.set(""); self.libelle_var.set(""); self.quantite_var.set("1"); self.prix_var.set("")
        self._refresh_lignes()

    def valider(self):
        if not self.reglement_id_selectionne:
            messagebox.showinfo("Info", "Sélectionnez d'abord un règlement dans la liste.", parent=self)
            return
        if not messagebox.askyesno("Valider ce règlement",
                                    "Le règlement va être comptabilisé sur le serveur (débit des comptes de "
                                    "charge, crédit fournisseur). Continuer ?", parent=self):
            return
        r = self._appeler("valider_reglement", self.reglement_id_selectionne)
        if r is APPEL_ECHEC:
            return
        messagebox.showinfo("Validé", "Règlement comptabilisé sur le serveur.", parent=self)
        self.refresh()

    def refresh_echeances(self):
        for row in self.tree_echeances.get_children():
            self.tree_echeances.delete(row)
        if not self.reglement_id_selectionne:
            return
        echeances = self._appeler("list_echeances_reglement", self.reglement_id_selectionne)
        if echeances is APPEL_ECHEC:
            return
        for e in echeances:
            tag = "payee" if e["statut"] == "Payée" else ("retard" if e["statut"] == "En retard" else "")
            self.tree_echeances.insert("", "end", iid=str(e["id"]), tags=(tag,) if tag else (), values=(
                e["id"], e["numero_tranche"], core.to_display_date(e["date_echeance"]),
                fmt_cfa(e["montant"]), e["statut"]))
        restant = sum(e["montant"] for e in echeances if e["statut"] != "Payée")
        if restant <= 0 and echeances:
            self.paiement_statut_var.set("✓ Toutes les tranches sont payées.")
        else:
            self.paiement_statut_var.set(f"Reste à payer : {fmt_cfa(restant)} F CFA." if echeances else "")

    def modifier_echeancier(self):
        if not self.reglement_id_selectionne:
            messagebox.showinfo("Info", "Sélectionnez d'abord un règlement dans la liste.", parent=self)
            return
        totals = self._appeler("compute_reglement_totals", self.reglement_id_selectionne)
        if totals is APPEL_ECHEC:
            return
        RemoteEcheancierDialog(self, self.remote, self.reglement_id_selectionne, totals["net_a_payer"],
                                self.refresh_echeances)

    def enregistrer_paiement(self):
        sel = self.tree_echeances.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une tranche dans l'échéancier ci-dessus.",
                                 parent=self)
            return
        echeance_id = int(sel[0])
        date_paiement = core.to_iso_date(self.date_paiement_var.get().strip())
        if not date_paiement:
            messagebox.showwarning("Champ manquant", "Saisissez la date de paiement.", parent=self)
            return
        compte = self._extract_code(self.compte_paiement_var.get())
        if not compte:
            messagebox.showwarning("Champ manquant", "Choisissez le compte banque ou caisse.", parent=self)
            return
        montant = self._appeler("enregistrer_paiement_echeance", echeance_id, date_paiement, compte)
        if montant is APPEL_ECHEC:
            return
        self.refresh_echeances()
        messagebox.showinfo("Paiement comptabilisé",
                             f"Paiement de {fmt_cfa(montant)} comptabilisé sur le serveur (Débit fournisseur, "
                             f"Crédit banque/caisse).", parent=self)
        self.refresh()

    def _on_select_reglement(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.reglement_id_selectionne = int(v[0])
        self._refresh_lignes()
        self.refresh_echeances()

    def _refresh_lignes(self):
        for row in self.tree_lignes.get_children():
            self.tree_lignes.delete(row)
        if not self.reglement_id_selectionne:
            return
        lignes = self._appeler("list_lignes_reglement", self.reglement_id_selectionne)
        if lignes is APPEL_ECHEC:
            return
        for l in lignes:
            self.tree_lignes.insert("", "end", values=(
                l["compte_charge"] or "⚠ à choisir", l["libelle"], f"{l['quantite']:g}",
                fmt_cfa(l["prix_unitaire"])))

    def refresh(self):
        self._refresh_compte_values()
        reglements = self._appeler("list_reglements_avec_statut_paiement")
        if reglements is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        filtre = self.filtre_var.get()
        for r in reglements:
            if filtre == "Non soldées" and r["statut_paiement"] not in ("Non soldée", "En retard") and \
                    "Partiellement" not in r["statut_paiement"]:
                continue
            if filtre == "En retard" and not r.get("en_retard"):
                continue
            if filtre == "Partiellement payées" and "Partiellement" not in r["statut_paiement"]:
                continue
            if filtre == "Soldées" and r["statut_paiement"] != "✓ Soldée":
                continue
            if filtre == "Brouillons" and r["statut"] == "validee":
                continue
            statut_validation = "Validé" if r["statut"] == "validee" else "Brouillon — à compléter"
            tag = "retard" if r.get("en_retard") else ("soldee" if r["statut_paiement"] == "✓ Soldée" else "")
            self.tree.insert("", "end", iid=str(r["id"]), tags=(tag,) if tag else (), values=(
                r["id"], r["numero"], core.to_display_date(r["date_reglement"]), r["raison_sociale"],
                fmt_cfa(r["montant_total"]), statut_validation, r["statut_paiement"]))


class RemoteEcheancierDialog(tk.Toplevel):
    """Définit l'échéancier de paiement d'un Règlement via le réseau —
    équivalent réseau complet de EcheancierDialog (bureau)."""

    def __init__(self, parent, remote: RemoteConnection, reglement_id, net_a_payer, on_saved):
        super().__init__(parent)
        self.remote = remote
        self.reglement_id = reglement_id
        self.net_a_payer = net_a_payer
        self.on_saved = on_saved
        self.title("Échéancier de paiement")
        self.geometry("560x420")
        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text=(
            f"Net à payer : {fmt_cfa(net_a_payer)} F CFA — répartissez ce montant sur une ou plusieurs "
            f"échéances (la somme des tranches doit correspondre exactement)."
        ), wraplength=520, foreground="#595959").pack(anchor="w", padx=10, pady=(10, 6))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=10)
        ttk.Label(form, text="Échéance (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Montant :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Button(form, text="Ajouter la tranche", command=self.ajouter_tranche).grid(row=0, column=4, padx=12)

        self.tree = ttk.Treeview(self, columns=("date", "montant"), show="headings", height=8)
        self.tree.heading("date", text="Échéance")
        self.tree.heading("montant", text="Montant")
        self.tree.column("date", width=140, anchor="w")
        self.tree.column("montant", width=160, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=10, pady=8)
        ttk.Button(self, text="Retirer la tranche sélectionnée", command=self.retirer_tranche).pack(
            anchor="w", padx=10)

        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=10, pady=6)

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Enregistrer l'échéancier", command=self.enregistrer).pack(side="left")
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=8)

        self.tranches = []
        existantes = appeler(self, remote, "list_echeances_reglement", reglement_id)
        if existantes is not APPEL_ECHEC:
            for e in existantes:
                if e["statut"] != "Payée":
                    self.tranches.append({"date_echeance": e["date_echeance"], "montant": e["montant"]})
        self._refresh_tree()

    def _refresh_tree(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        total = 0.0
        for t in self.tranches:
            self.tree.insert("", "end", values=(core.to_display_date(t["date_echeance"]), fmt_cfa(t["montant"])))
            total += t["montant"]
        ecart = total - self.net_a_payer
        etat = "✓ correspond au net à payer" if abs(ecart) < 1 else f"⚠ écart de {fmt_cfa(ecart)}"
        self.total_var.set(f"Total des tranches : {fmt_cfa(total)} F CFA — {etat}")

    def ajouter_tranche(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "Date d'échéance invalide.", parent=self)
            return
        try:
            montant = float(self.montant_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.", parent=self)
            return
        if montant <= 0:
            messagebox.showerror("Erreur", "Le montant doit être strictement positif.", parent=self)
            return
        self.tranches.append({"date_echeance": date_str, "montant": montant})
        self.montant_var.set("")
        self._refresh_tree()

    def retirer_tranche(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.tranches[idx]
        self._refresh_tree()

    def enregistrer(self):
        if not self.tranches:
            messagebox.showwarning("Vide", "Ajoutez au moins une tranche.", parent=self)
            return
        if appeler(self, self.remote, "set_echeancier_reglement", self.reglement_id,
                   self.tranches) is APPEL_ECHEC:
            return
        messagebox.showinfo("Enregistré", "Échéancier mis à jour sur le serveur.", parent=self)
        self.on_saved()
        self.destroy()


class RemoteEtatFormuleTab(ttk.Frame):
    """Écran générique en lecture seule pour les états basés sur
    compute_etat_formule_generique() côté serveur (Compte de résultat,
    TFT, Situation financière) — même principe que EtatFormuleTab dans
    l'application de bureau, réutilisé pour les 3 rapports qui partagent
    la même structure « Rubrique | N (| N-1 | %) »."""

    def __init__(self, parent, remote: RemoteConnection, titre, fonction):
        super().__init__(parent)
        self.remote = remote
        self.titre = titre
        self.fonction = fonction

        ttk.Label(self, text=titre, font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 8))

        self.tree = ttk.Treeview(self, columns=("libelle", "n", "n1", "pct"), show="headings", height=32)
        self.tree.heading("libelle", text="Rubrique")
        self.tree.column("libelle", width=460, anchor="w", stretch=True)
        for c in ("n", "n1", "pct"):
            self.tree.column(c, width=150, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        d = self._appeler(self.fonction)
        if d is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        headers = {"N": "Exercice N", "N-1": "Exercice N-1", "%": "%"}
        cols = ("n", "n1", "pct")
        for col_key, label in zip(cols, d["colonnes"] + [""] * 3):
            self.tree.heading(col_key, text=headers.get(label, label) if label else "")
        for l in d["lignes"]:
            valeurs = [l.get(c, None) for c in d["colonnes"]]
            while len(valeurs) < 3:
                valeurs.append(None)
            self.tree.insert("", "end", values=(l["libelle"], fmt_cfa(valeurs[0]), fmt_cfa(valeurs[1]),
                                                 fmt_cfa(valeurs[2])))
        if d.get("errors"):
            detail = "\n".join(f"• Cellule {coord} : {msg}" for coord, _formula, msg in d["errors"][:10])
            messagebox.showwarning(
                "Formules en erreur",
                f"{len(d['errors'])} formule(s) du gabarit n'ont pas pu être évaluées. Les autres lignes "
                f"restent correctes.\n\n{detail}", parent=self,
            )


class RemoteBilanTab(ttk.Frame):
    """Bilan SYSCOHADA en lecture seule via le réseau — même moteur
    (compute_bilan_detaille) que l'application de bureau, présentation
    simplifiée (colonnes Net et Net N-1 uniquement, sans le détail
    Brut/Amortissements, pour une vue rapide)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="BILAN SYSCOHADA", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=16, pady=(0, 4))
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left")
        self.ecart_var = tk.StringVar()
        self.ecart_label = ttk.Label(btn_bar, textvariable=self.ecart_var, font=("Segoe UI", 10, "bold"))
        self.ecart_label.pack(side="left", padx=16)

        columns_frame = ttk.Frame(self)
        columns_frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        columns_frame.columnconfigure(0, weight=1)
        columns_frame.columnconfigure(1, weight=1)
        self.tree_actif = ttk.Treeview(columns_frame, columns=("libelle", "net", "net_n1"), show="headings", height=30)
        for c, h, w in zip(("libelle", "net", "net_n1"), ["ACTIF", "Net", "Net N-1"], [260, 130, 130]):
            self.tree_actif.heading(c, text=h)
            self.tree_actif.column(c, width=w, anchor="w" if c == "libelle" else "e")
        self.tree_actif.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        self.tree_passif = ttk.Treeview(columns_frame, columns=("libelle", "montant", "montant_n1"),
                                         show="headings", height=30)
        for c, h, w in zip(("libelle", "montant", "montant_n1"), ["PASSIF", "Montant", "Montant N-1"],
                           [280, 140, 140]):
            self.tree_passif.heading(c, text=h)
            self.tree_passif.column(c, width=w, anchor="w" if c == "libelle" else "e")
        self.tree_passif.grid(row=0, column=1, sticky="nsew", padx=(4, 0))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _add_actif(self, titre, lignes, total_label, total_val, total_val_n1, montant_field="net"):
        self.tree_actif.insert("", "end", values=(titre, "", ""))
        for l in lignes:
            montant = l.get(montant_field, 0)
            montant_n1 = l.get(f"{montant_field}_n1", 0)
            if montant or montant_n1:
                self.tree_actif.insert("", "end", values=(f"  {l['label']}", fmt_cfa(montant), fmt_cfa(montant_n1)))
        self.tree_actif.insert("", "end", values=(f"  {total_label}", fmt_cfa(total_val), fmt_cfa(total_val_n1)))

    def _add_passif(self, titre, lignes, total_label, total_val, total_val_n1):
        self.tree_passif.insert("", "end", values=(titre, "", ""))
        for l in lignes:
            montant = l.get("sous_total", 0)
            montant_n1 = l.get("sous_total_n1", 0)
            if montant or montant_n1:
                self.tree_passif.insert("", "end", values=(f"  {l['label']}", fmt_cfa(montant), fmt_cfa(montant_n1)))
        self.tree_passif.insert("", "end", values=(f"  {total_label}", fmt_cfa(total_val), fmt_cfa(total_val_n1)))

    def refresh(self):
        d = self._appeler("compute_bilan_detaille")
        if d is APPEL_ECHEC:
            return
        for tree in (self.tree_actif, self.tree_passif):
            for row in tree.get_children():
                tree.delete(row)
        a, p = d["actif"], d["passif"]
        self._add_actif("IMMOBILISATIONS", a["immobilisations"], "Total immobilisations nettes",
                         a["total_immo_net"], a["total_immo_net_n1"])
        self._add_actif("STOCKS", a["stocks"], "Total stocks", a["total_stocks"], a["total_stocks_n1"],
                         montant_field="sous_total")
        self._add_actif("CRÉANCES", a["creances"], "Total créances", a["total_creances"], a["total_creances_n1"],
                         montant_field="sous_total")
        self._add_actif("TRÉSORERIE ACTIF", a["tresorerie"], "Total trésorerie actif", a["total_tresorerie"],
                         a["total_tresorerie_n1"], montant_field="sous_total")
        self.tree_actif.insert("", "end", values=("TOTAL ACTIF", fmt_cfa(d["total_actif"]), fmt_cfa(d["total_actif_n1"])))
        self._add_passif("CAPITAUX PROPRES", p["capitaux_propres"], "Total capitaux propres",
                          p["total_capitaux_propres"], p["total_capitaux_propres_n1"])
        self._add_passif("DETTES CIRCULANTES", p["dettes"], "Total dettes circulantes", p["total_dettes"],
                          p["total_dettes_n1"])
        self._add_passif("TRÉSORERIE PASSIF", p["tresorerie"], "Total trésorerie passif", p["total_tresorerie"],
                          p["total_tresorerie_n1"])
        self.tree_passif.insert("", "end", values=("TOTAL PASSIF", fmt_cfa(d["total_passif"]), fmt_cfa(d["total_passif_n1"])))

        ecart = d["ecart"]
        if abs(ecart) < 1:
            self.ecart_var.set(f"✓ Actif = Passif ({fmt_cfa(d['total_actif'])})")
            self.ecart_label.configure(foreground="#1F7A1F")
        else:
            self.ecart_var.set(f"⚠ Écart Actif - Passif : {fmt_cfa(ecart)}")
            self.ecart_label.configure(foreground="#B00020")


class RemoteBalanceTab(ttk.Frame):
    """Balance générale en lecture seule via le réseau — 6 colonnes
    (Ouverture/Mouvement/Clôture Débit/Crédit), même moteur que
    l'application de bureau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="BALANCE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 4))
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        cols = ("compte", "libelle", "ouv_debit", "ouv_credit", "mvt_debit", "mvt_credit", "sold_debit", "sold_credit")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        headers = ["N° Compte", "Libellé", "Ouv. Débit", "Ouv. Crédit", "Mvt Débit", "Mvt Crédit",
                   "Clôt. Débit", "Clôt. Crédit"]
        widths = [90, 220, 100, 100, 100, 100, 100, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w" if c in ("compte", "libelle") else "e")
        self.tree.tag_configure("classe_total", background="#DCE6F1", font=("Segoe UI", 9, "bold"))
        self.tree.tag_configure("grand_total", background="#1F4E78", foreground="white", font=("Segoe UI", 10, "bold"))
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        d = self._appeler("compute_balance_detaillee")
        if d is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for c in d["classes"]:
            for l in c["lignes"]:
                self.tree.insert("", "end", values=(
                    l["code"], l["label"], fmt_cfa(l["ouverture_debit"]), fmt_cfa(l["ouverture_credit"]),
                    fmt_cfa(l["cumul_debit"]), fmt_cfa(l["cumul_credit"]), fmt_cfa(l["solde_debit"]),
                    fmt_cfa(l["solde_credit"])))
            st = c["sous_total"]
            self.tree.insert("", "end", tags=("classe_total",), values=(
                "", f"TOTAL CLASSE {c['classe']}", fmt_cfa(st["ouverture_debit"]), fmt_cfa(st["ouverture_credit"]),
                fmt_cfa(st["cumul_debit"]), fmt_cfa(st["cumul_credit"]), fmt_cfa(st["solde_debit"]),
                fmt_cfa(st["solde_credit"])))
        gt = d["grand_total"]
        self.tree.insert("", "end", tags=("grand_total",), values=(
            "", "TOTAL BALANCE", fmt_cfa(gt["ouverture_debit"]), fmt_cfa(gt["ouverture_credit"]),
            fmt_cfa(gt["cumul_debit"]), fmt_cfa(gt["cumul_credit"]), fmt_cfa(gt["solde_debit"]),
            fmt_cfa(gt["solde_credit"])))


class RemoteGrandLivreTab(ttk.Frame):
    """Grand livre en lecture seule via le réseau — détail écriture par
    écriture, groupé par compte puis par classe."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="GRAND LIVRE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 4))
        cols = ("date", "piece", "journal", "libelle", "debit", "credit")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=30)
        for c, h, w in zip(cols, ["Date", "Pièce", "Journal", "Libellé", "Débit", "Crédit"],
                           [85, 90, 60, 400, 110, 110]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w" if c not in ("debit", "credit") else "e")
        self.tree.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        classes = self._appeler("compute_grand_livre_complet")
        if classes is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for c in classes:
            for compte in c["comptes"]:
                self.tree.insert("", "end", values=("", "", "", f"{compte['code']} — {compte['label']}", "", ""))
                for l in compte["lignes"]:
                    self.tree.insert("", "end", values=(
                        core.to_display_date(l["date"]), l["piece"] or "", l["journal"] or "", l["libelle"] or "",
                        fmt_cfa(l["debit"]) if l["debit"] else "", fmt_cfa(l["credit"]) if l["credit"] else ""))
                self.tree.insert("", "end", values=(
                    "", "", "", f"TOTAL COMPTE {compte['code']} — Solde {compte['sens']}",
                    fmt_cfa(compte["total_debit"]), fmt_cfa(compte["total_credit"])))


class RemoteClientsTab(ttk.Frame):
    """Clients (COMMERCIAL) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_code = None

        ttk.Label(self, text="CLIENTS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Client")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Code :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.code_var = tk.StringVar()
        self.code_entry = ttk.Entry(form, textvariable=self.code_var, width=14)
        self.code_entry.grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Raison sociale :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.raison_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.raison_var, width=30).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Contact :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.contact_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.contact_var, width=20).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Téléphone :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.telephone_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.telephone_var, width=18).grid(row=1, column=3, padx=4, pady=(4, 0))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        web_bar = ttk.LabelFrame(self, text="Trouver de nouveaux clients / prospects sur Internet")
        web_bar.pack(fill="x", padx=16, pady=(0, 4))
        ttk.Label(web_bar, text="Produit / service vendu :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.web_produit_var = tk.StringVar()
        ttk.Entry(web_bar, textvariable=self.web_produit_var, width=28).grid(row=0, column=1, padx=4)
        ttk.Label(web_bar, text="Ville :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.web_ville_var = tk.StringVar(value="Ouagadougou")
        ttk.Entry(web_bar, textvariable=self.web_ville_var, width=18).grid(row=0, column=3, padx=4)
        ttk.Button(web_bar, text="Rechercher sur Internet", command=self.rechercher_internet).grid(
            row=0, column=4, padx=12)
        ttk.Label(web_bar, text=(
            "Ouvre votre navigateur avec une recherche Google déjà remplie (entreprises susceptibles "
            "d'acheter ce produit/service dans cette ville). Copiez ensuite les coordonnées du prospect "
            "choisi dans le formulaire ci-dessus pour l'enregistrer."
        ), foreground="#595959", wraplength=1050).grid(row=1, column=0, columnspan=5, sticky="w", padx=4, pady=(2, 4))

        cols = ("code", "raison_sociale", "contact", "telephone")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        for c, h, w in zip(cols, ["Code", "Raison sociale", "Contact", "Téléphone"], [100, 280, 180, 140]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def rechercher_internet(self):
        produit = self.web_produit_var.get().strip()
        if not produit:
            messagebox.showwarning("Champ manquant", "Indiquez le produit ou service vendu.", parent=self)
            return
        ville = self.web_ville_var.get().strip()
        requete = f"entreprises acheteurs {produit} {ville}".strip()
        import webbrowser
        from urllib.parse import quote_plus
        webbrowser.open(f"https://www.google.com/search?q={quote_plus(requete)}")

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_code = v[0]
        self.code_var.set(v[0]); self.raison_var.set(v[1])
        self.contact_var.set(v[2]); self.telephone_var.set(v[3])
        self.code_entry.configure(state="disabled")

    def clear_form(self):
        self.selected_code = None
        self.code_var.set(""); self.raison_var.set(""); self.contact_var.set(""); self.telephone_var.set("")
        self.code_entry.configure(state="normal")

    def add(self):
        if not self.code_var.get().strip() or not self.raison_var.get().strip():
            messagebox.showwarning("Champ manquant", "Code et raison sociale sont obligatoires.", parent=self)
            return
        r = self._appeler("add_client", self.code_var.get(), self.raison_var.get(),
                           contact=self.contact_var.get(), telephone=self.telephone_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un client.", parent=self)
            return
        r = self._appeler("add_client", self.selected_code, self.raison_var.get(),
                           contact=self.contact_var.get(), telephone=self.telephone_var.get())
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un client.", parent=self)
            return
        if messagebox.askyesno("Confirmer", f"Supprimer le client « {self.selected_code} » ?", parent=self):
            r = self._appeler("delete_client", self.selected_code)
            if r is APPEL_ECHEC:
                return
            self.clear_form()
            self.refresh()

    def refresh(self):
        clients = self._appeler("list_clients")
        if clients is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for c in clients:
            self.tree.insert("", "end", values=(c["code"], c["raison_sociale"], c["contact"] or "",
                                                 c["telephone"] or ""))


class RemoteCommandeClientEcheancierDialog(tk.Toplevel):
    """Planifie l'échéancier de règlement PRÉVU d'une Commande client via
    le réseau — équivalent réseau complet de CommandeClientEcheancierDialog
    (bureau)."""

    def __init__(self, parent, remote: RemoteConnection, commande_id, montant_ttc, on_saved):
        super().__init__(parent)
        self.remote = remote
        self.commande_id = commande_id
        self.montant_ttc = montant_ttc
        self.on_saved = on_saved
        self.title("Échéancier de règlement prévu")
        self.geometry("560x420")
        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text=(
            f"Montant TTC : {fmt_cfa(montant_ttc)} F CFA — répartissez ce montant sur une ou plusieurs "
            f"échéances prévues (purement planifié, rien n'est encore comptabilisé)."
        ), wraplength=520, foreground="#595959").pack(anchor="w", padx=10, pady=(10, 6))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=10)
        ttk.Label(form, text="Échéance (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Montant :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Button(form, text="Ajouter la tranche", command=self.ajouter_tranche).grid(row=0, column=4, padx=12)

        self.tree = ttk.Treeview(self, columns=("date", "montant"), show="headings", height=8)
        self.tree.heading("date", text="Échéance")
        self.tree.heading("montant", text="Montant")
        self.tree.column("date", width=140, anchor="w")
        self.tree.column("montant", width=160, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=10, pady=8)
        ttk.Button(self, text="Retirer la tranche sélectionnée", command=self.retirer_tranche).pack(
            anchor="w", padx=10)

        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=10, pady=6)

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Enregistrer l'échéancier", command=self.enregistrer).pack(side="left")
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=8)

        self.tranches = []
        existantes = appeler(self, remote, "list_echeances_commande_client", commande_id)
        if existantes is not APPEL_ECHEC:
            self.tranches = [{"date_echeance": t["date_echeance"], "montant": t["montant"]} for t in existantes]
        self._refresh_tree()

    def _refresh_tree(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        total = 0.0
        for t in self.tranches:
            self.tree.insert("", "end", values=(core.to_display_date(t["date_echeance"]), fmt_cfa(t["montant"])))
            total += t["montant"]
        ecart = total - self.montant_ttc
        etat = "✓ correspond au montant TTC" if abs(ecart) < 1 else f"⚠ écart de {fmt_cfa(ecart)}"
        self.total_var.set(f"Total des tranches : {fmt_cfa(total)} F CFA — {etat}")

    def ajouter_tranche(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "Date d'échéance invalide.", parent=self)
            return
        try:
            montant = float(self.montant_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.", parent=self)
            return
        if montant <= 0:
            messagebox.showerror("Erreur", "Le montant doit être strictement positif.", parent=self)
            return
        self.tranches.append({"date_echeance": date_str, "montant": montant})
        self.montant_var.set("")
        self._refresh_tree()

    def retirer_tranche(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.tranches[idx]
        self._refresh_tree()

    def enregistrer(self):
        if not self.tranches:
            messagebox.showwarning("Vide", "Ajoutez au moins une tranche.", parent=self)
            return
        if appeler(self, self.remote, "set_echeancier_commande_client", self.commande_id,
                   self.tranches) is APPEL_ECHEC:
            return
        messagebox.showinfo("Enregistré", "Échéancier planifié sur le serveur.", parent=self)
        self.on_saved()
        self.destroy()


class RemoteCommandeClientDialog(tk.Toplevel):
    """Détail d'une Commande client via le réseau — équivalent réseau
    complet de CommandeClientDialog (bureau). La validation fait basculer
    la commande en FACTURE DE VENTE (brouillon, non comptabilisée) avec
    les lignes recopiées."""

    def __init__(self, parent, remote: RemoteConnection, commande_id, on_saved):
        super().__init__(parent)
        self.remote = remote
        self.commande_id = commande_id
        self.on_saved = on_saved
        self.selected_ligne_id = None
        self.title("Commande client")
        self.geometry("1080x680")
        self.minsize(850, 500)
        self.transient(parent)
        self.grab_set()

        commande = appeler(self, remote, "get_commande_client", commande_id)
        if commande is APPEL_ECHEC or not commande:
            self.destroy()
            return
        self.validee = commande["statut"] == "validee"

        header = ttk.LabelFrame(self, text="Informations")
        header.pack(fill="x", padx=10, pady=8)
        ttk.Label(header, text="N° :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.numero_var = tk.StringVar(value=commande["numero"])
        ttk.Entry(header, textvariable=self.numero_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=core.to_display_date(commande["date_commande"]))
        ttk.Entry(header, textvariable=self.date_var, width=12).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Client :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.client_var = tk.StringVar(value=commande["client_code"] or "")
        self.client_combo = ttk.Combobox(header, textvariable=self.client_var, width=24)
        self.client_combo.grid(row=0, column=5, padx=4)
        self.client_combo.bind("<KeyRelease>", self._on_client_keyrelease)
        self._refresh_client_values()

        ttk.Label(header, text="Paiement attendu :").grid(row=1, column=0, sticky="w", padx=4, pady=(6, 0))
        self.date_paiement_var = tk.StringVar(value=core.to_display_date(commande.get("date_paiement_attendu") or ""))
        ttk.Entry(header, textvariable=self.date_paiement_var, width=12).grid(
            row=1, column=1, padx=4, pady=(6, 0), sticky="w")
        ttk.Label(header, text="TVA % :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(6, 0))
        self.tva_var = tk.StringVar(value=str(commande.get("tva_taux") or 0))
        ttk.Entry(header, textvariable=self.tva_var, width=8).grid(row=1, column=3, padx=4, pady=(6, 0), sticky="w")
        ttk.Label(header, text="Compte TVA :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(6, 0))
        self.tva_compte_var = tk.StringVar(value=commande.get("tva_compte") or core.COMPTE_TVA_VENTES)
        ttk.Entry(header, textvariable=self.tva_compte_var, width=14).grid(
            row=1, column=5, padx=4, pady=(6, 0), sticky="w")

        echeancier_frame = ttk.Frame(header)
        echeancier_frame.grid(row=2, column=0, columnspan=6, sticky="we", padx=4, pady=(6, 0))
        ttk.Button(echeancier_frame, text="Planifier un échéancier (plusieurs tranches, optionnel)",
                   command=self.modifier_echeancier).pack(side="left")
        self.echeancier_apercu_var = tk.StringVar()
        ttk.Label(echeancier_frame, textvariable=self.echeancier_apercu_var, foreground="#595959").pack(
            side="left", padx=12)
        self._refresh_echeancier_apercu()

        origine = "" if not commande.get("facture_vente_id") else \
            f"Facture générée : n° interne {commande['facture_vente_id']}"
        self.statut_var = tk.StringVar(
            value=f"Statut : {'VALIDÉE (facture générée)' if self.validee else 'Brouillon'}   {origine}")
        ttk.Label(header, textvariable=self.statut_var, font=("Segoe UI", 10, "bold")).grid(
            row=3, column=0, columnspan=4, sticky="w", padx=4, pady=(4, 0))
        self.retard_var = tk.StringVar()
        self.retard_label = ttk.Label(header, textvariable=self.retard_var, font=("Segoe UI", 10, "bold"))
        self.retard_label.grid(row=3, column=4, columnspan=2, sticky="w", padx=4, pady=(4, 0))
        self._refresh_retard()

        lignes_frame = ttk.LabelFrame(self, text=(
            "Lignes — un compte de produit (classe 70x) est OBLIGATOIRE sur chaque ligne pour pouvoir valider"))
        lignes_frame.pack(fill="both", padx=10, pady=6)
        form = ttk.Frame(lignes_frame)
        form.pack(fill="x", padx=6, pady=4)
        ttk.Label(form, text="Compte de produit (classe 70x) :").grid(row=0, column=0, sticky="w")
        self.ligne_compte_var = tk.StringVar()
        self.ligne_compte_combo = ttk.Combobox(form, textvariable=self.ligne_compte_var, width=26)
        self.ligne_compte_combo.grid(row=0, column=1, padx=4)
        self.ligne_compte_combo.bind("<KeyRelease>", self._on_ligne_compte_keyrelease)
        self._refresh_ligne_compte_values()
        ttk.Label(form, text="Code analytique :").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.ligne_analytic_var = tk.StringVar()
        self.ligne_analytic_combo = ttk.Combobox(form, textvariable=self.ligne_analytic_var, width=18)
        self.ligne_analytic_combo.grid(row=0, column=3, padx=4)
        self._refresh_ligne_analytic_values()

        ttk.Label(form, text="Libellé :").grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.ligne_libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_libelle_var, width=26).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Quantité :").grid(row=1, column=2, sticky="w", padx=(12, 0), pady=(4, 0))
        self.ligne_qte_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=8).grid(row=1, column=3, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Prix unitaire :").grid(row=1, column=4, sticky="w", padx=(12, 0), pady=(4, 0))
        self.ligne_prix_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_prix_var, width=10).grid(row=1, column=5, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Unité :").grid(row=1, column=6, sticky="w", padx=(12, 0), pady=(4, 0))
        self.ligne_unite_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_unite_var, width=8).grid(row=1, column=7, padx=4, pady=(4, 0))

        btn_row = ttk.Frame(lignes_frame)
        btn_row.pack(fill="x", padx=6)
        self.update_ligne_btn = ttk.Button(btn_row, text="Mettre à jour la ligne sélectionnée",
                                            command=self.update_ligne)
        self.update_ligne_btn.pack(side="left")
        self.add_ligne_btn = ttk.Button(btn_row, text="Ajouter une nouvelle ligne", command=self.add_ligne)
        self.add_ligne_btn.pack(side="left", padx=8)
        self.delete_ligne_btn = ttk.Button(btn_row, text="Supprimer la ligne sélectionnée", command=self.delete_ligne)
        self.delete_ligne_btn.pack(side="left", padx=8)

        cols = ("id", "compte", "libelle", "quantite", "prix", "montant", "unite", "analytique")
        self.tree = ttk.Treeview(lignes_frame, columns=cols, show="headings", height=10)
        headers = ["ID", "Compte", "Libellé", "Qté", "Prix unit.", "Montant HT", "Unité", "Analytique"]
        widths = [40, 90, 220, 70, 90, 110, 70, 130]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=6, pady=6)
        self.tree.bind("<<TreeviewSelect>>", self._on_select_ligne)
        self.total_var = tk.StringVar()
        ttk.Label(lignes_frame, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=6, pady=(0, 6))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        self.save_btn = ttk.Button(btns, text="Enregistrer", command=self.save)
        self.save_btn.pack(side="left", padx=4)
        self.valider_btn = ttk.Button(btns, text="Valider (génère la facture)", command=self.valider)
        self.valider_btn.pack(side="left", padx=4)
        self.corriger_btn = ttk.Button(btns, text="Corriger (repasser en brouillon)", command=self.corriger)
        self.corriger_btn.pack(side="left", padx=4)
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=4)

        self.refresh_lignes()
        self._apply_lock()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _refresh_client_values(self):
        items = self._appeler("list_clients")
        if items is not APPEL_ECHEC:
            self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _on_client_keyrelease(self, event=None):
        query = self._extract_code(self.client_var.get())
        if query:
            items = self._appeler("list_clients", query)
            if items is not APPEL_ECHEC:
                self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _refresh_ligne_compte_values(self):
        items = self._appeler("search_accounts", "7", limit=100)
        if items is not APPEL_ECHEC:
            self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "7"]

    def _on_ligne_compte_keyrelease(self, event=None):
        query = self._extract_code(self.ligne_compte_var.get())
        if query:
            items = self._appeler("search_accounts", query, limit=50)
            if items is not APPEL_ECHEC:
                self.ligne_compte_combo["values"] = [
                    f"{a['code']} — {a['label']}" for a in items if a["classe"] == "7"]

    def _refresh_ligne_analytic_values(self):
        codes = self._appeler("list_analytic_codes")
        if codes is not APPEL_ECHEC:
            self.ligne_analytic_combo["values"] = [f"{c['code']} — {c['label']}" for c in codes]

    def _apply_lock(self):
        state = "disabled" if self.validee else "normal"
        for w in (self.update_ligne_btn, self.add_ligne_btn, self.delete_ligne_btn, self.save_btn, self.valider_btn):
            w.configure(state=state)
        self.corriger_btn.configure(state="normal" if self.validee else "disabled")

    def _on_select_ligne(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_ligne_id = v[0]
        self.ligne_compte_var.set(v[1])
        self.ligne_libelle_var.set(v[2])
        self.ligne_qte_var.set(v[3])
        self.ligne_prix_var.set(v[4])
        self.ligne_unite_var.set(v[6])
        self.ligne_analytic_var.set(v[7])

    def _clear_ligne_form(self):
        self.selected_ligne_id = None
        self.ligne_compte_var.set(""); self.ligne_libelle_var.set("")
        self.ligne_qte_var.set(""); self.ligne_prix_var.set("")
        self.ligne_unite_var.set(""); self.ligne_analytic_var.set("")

    def refresh_lignes(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        lignes = self._appeler("list_lignes_commande_client", self.commande_id)
        if lignes is APPEL_ECHEC:
            return
        total = 0.0
        for l in lignes:
            total += l["montant_ht"]
            self.tree.insert("", "end", values=(
                l["id"], l["compte_vente"] or "⚠ à choisir", l["libelle"], f"{l['quantite']:g}",
                f"{fmt_cfa(l['prix_unitaire'])}", f"{fmt_cfa(l['montant_ht'])}", l["unite"] or "",
                l["analytic_code"] or ""))
        self.total_var.set(f"Total HT : {fmt_cfa(total)}")

    def add_ligne(self):
        libelle = self.ligne_libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé est obligatoire.", parent=self)
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
            prix = float(self.ligne_prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        compte = self._extract_code(self.ligne_compte_var.get()) or None
        analytic = self._extract_code(self.ligne_analytic_var.get()) or None
        if self._appeler("add_ligne_commande_client", self.commande_id, libelle, qte, prix_unitaire=prix,
                          unite=self.ligne_unite_var.get().strip() or None,
                          compte_vente=compte, analytic_code=analytic) is APPEL_ECHEC:
            return
        self._clear_ligne_form()
        self.refresh_lignes()

    def update_ligne(self):
        if not self.selected_ligne_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne dans le tableau.", parent=self)
            return
        libelle = self.ligne_libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé est obligatoire.", parent=self)
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
            prix = float(self.ligne_prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        compte = self._extract_code(self.ligne_compte_var.get()) or None
        analytic = self._extract_code(self.ligne_analytic_var.get()) or None
        if self._appeler(
            "update_ligne_commande_client", self.selected_ligne_id, libelle=libelle, quantite=qte,
            prix_unitaire=prix, unite=self.ligne_unite_var.get().strip() or None,
            compte_vente=compte, analytic_code=analytic
        ) is APPEL_ECHEC:
            return
        self._clear_ligne_form()
        self.refresh_lignes()

    def delete_ligne(self):
        if not self.selected_ligne_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.", parent=self)
            return
        if self._appeler("delete_ligne_commande_client", self.selected_ligne_id) is APPEL_ECHEC:
            return
        self._clear_ligne_form()
        self.refresh_lignes()

    def _refresh_retard(self):
        commandes = self._appeler("list_commandes_client")
        if commandes is APPEL_ECHEC:
            return
        match = next((c for c in commandes if c["id"] == self.commande_id), None)
        if not match or not match.get("statut_paiement"):
            self.retard_var.set("")
            return
        self.retard_var.set(f"Paiement : {match['statut_paiement']}")
        self.retard_label.configure(foreground="#B00020" if match["depassement_paiement"] else "#1F7A1F")

    def _refresh_echeancier_apercu(self):
        tranches = self._appeler("list_echeances_commande_client", self.commande_id)
        if tranches is APPEL_ECHEC:
            return
        if not tranches:
            self.echeancier_apercu_var.set("Aucun échéancier planifié (paiement en une fois à la date ci-dessus).")
        else:
            self.echeancier_apercu_var.set(
                f"{len(tranches)} tranche(s) planifiée(s) : " +
                ", ".join(f"{core.to_display_date(t['date_echeance'])} ({fmt_cfa(t['montant'])})" for t in tranches))

    def modifier_echeancier(self):
        totals = self._appeler("compute_commande_client_totals", self.commande_id)
        if totals is APPEL_ECHEC:
            return
        if totals["total_ttc"] <= 0:
            messagebox.showinfo("Info", "Ajoutez d'abord des lignes avec un montant avant de planifier "
                                         "un échéancier.", parent=self)
            return
        RemoteCommandeClientEcheancierDialog(self, self.remote, self.commande_id, totals["total_ttc"],
                                              self._refresh_echeancier_apercu)

    def save(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        try:
            tva = float(self.tva_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le taux de TVA doit être un nombre.", parent=self)
            return
        if self._appeler(
            "update_commande_client", self.commande_id,
            numero=self.numero_var.get().strip(), date_commande=date_str,
            client_code=self._extract_code(self.client_var.get()),
            date_paiement_attendu=core.to_iso_date(self.date_paiement_var.get().strip()) or None,
            tva_taux=tva, tva_compte=self.tva_compte_var.get().strip() or core.COMPTE_TVA_VENTES,
        ) is APPEL_ECHEC:
            return
        self._refresh_retard()
        messagebox.showinfo("Enregistré", "Commande enregistrée sur le serveur.", parent=self)
        self.on_saved()

    def valider(self):
        if not messagebox.askyesno(
            "Valider cette commande",
            "Cette commande va être verrouillée et va générer une FACTURE DE VENTE (brouillon, non "
            "comptabilisée) avec les lignes recopiées. Aucune écriture comptable n'est envoyée à ce "
            "stade — c'est la validation de cette facture, avec sa date de règlement prévu (menu "
            "COMMERCIAL > Facturation), qui comptabilisera la vente et l'enverra dans Paiement/Recouvrement."
            "\n\nContinuer ?",
            parent=self,
        ):
            return
        self.save()
        r = self._appeler("valider_commande_client", self.commande_id)
        if r is APPEL_ECHEC:
            return
        messagebox.showinfo(
            "Validée",
            "Commande validée — une facture de vente brouillon a été créée dans COMMERCIAL > Facturation. "
            "Complétez-la si besoin puis validez-la avec sa date de règlement prévu pour comptabiliser "
            "la vente et l'envoyer dans Paiement/Recouvrement.",
            parent=self,
        )
        self.on_saved()
        self.destroy()

    def corriger(self):
        if not messagebox.askyesno(
            "Corriger cette commande",
            "Cette commande est déjà validée : elle va repasser en brouillon modifiable (impossible si "
            "la facture générée est elle-même déjà validée — corrigez-la d'abord).\n\nContinuer ?",
            parent=self,
        ):
            return
        r = self._appeler("devalider_commande_client", self.commande_id)
        if r is APPEL_ECHEC:
            return
        self.validee = False
        self.statut_var.set("Statut : Brouillon")
        self._apply_lock()
        self.refresh_lignes()
        messagebox.showinfo("Repassée en brouillon",
                             "La commande est de nouveau modifiable. Corrigez les lignes puis cliquez "
                             "sur « Valider ».", parent=self)
        self.on_saved()


class RemoteCommandeClientTab(ttk.Frame):
    """Liste des Commandes client via le réseau (menu COMMERCIAL) —
    équivalent réseau complet de CommandeClientTab (bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="COMMANDES CLIENT", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Enregistrez ici les commandes reçues d'un client. Une commande validée génère une facture "
            "de vente (brouillon) — c'est la validation de CETTE facture (COMMERCIAL > Facturation) qui "
            "comptabilise la vente et l'envoie dans Paiement/Recouvrement. Cliquez sur une ligne pour l'ouvrir."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=16, pady=4)
        ttk.Button(btn_bar, text="Nouvelle commande", command=self.new_commande).pack(side="left")
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left", padx=8)

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=16, pady=8)
        cols = ("numero", "date", "client", "nb_lignes", "statut", "paiement_attendu", "statut_paiement")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        headers = ["N°", "Date", "Client", "Lignes", "Statut", "Paiement attendu", "Statut paiement"]
        widths = [110, 100, 220, 60, 130, 130, 160]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("retard", foreground="#B00020", font=("Segoe UI", 9, "bold"))
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.tree.bind("<Button-1>", self._on_click)
        self._by_iid = {}
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def new_commande(self):
        numero = simpledialog.askstring("Nouvelle commande client", "N° :", parent=self)
        if not numero:
            return
        cid = self._appeler("create_commande_client", numero, date.today().strftime("%Y-%m-%d"))
        if cid is APPEL_ECHEC:
            return
        self.refresh()
        RemoteCommandeClientDialog(self, self.remote, cid, on_saved=self.refresh)

    def _on_click(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        self.tree.selection_set(iid)
        cid = self._by_iid.get(iid)
        if cid:
            RemoteCommandeClientDialog(self, self.remote, cid, on_saved=self.refresh)

    def refresh(self):
        commandes = self._appeler("list_commandes_client")
        if commandes is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._by_iid = {}
        for commande in commandes:
            nb_result = self._appeler("list_lignes_commande_client", commande["id"])
            nb = len(nb_result) if nb_result is not APPEL_ECHEC else 0
            statut = "Validée" if commande["statut"] == "validee" else "Brouillon"
            tags = ("retard",) if commande.get("depassement_paiement") else ()
            iid = self.tree.insert("", "end", tags=tags, values=(
                commande["numero"], core.to_display_date(commande["date_commande"]),
                commande.get("raison_sociale") or commande.get("client_code") or "", nb, statut,
                core.to_display_date(commande.get("date_paiement_attendu") or ""),
                commande.get("statut_paiement") or "",
            ))
            self._by_iid[iid] = commande["id"]


class RemoteFactureVenteEcheancierDialog(tk.Toplevel):
    """Planifie l'échéancier de règlement PRÉVU d'une Facture de vente en
    brouillon via le réseau — équivalent réseau complet de
    FactureVenteEcheancierDialog (bureau)."""

    def __init__(self, parent, remote: RemoteConnection, facture_id, montant_ttc, on_saved):
        super().__init__(parent)
        self.remote = remote
        self.facture_id = facture_id
        self.montant_ttc = montant_ttc
        self.on_saved = on_saved
        self.title("Échéancier de règlement prévu")
        self.geometry("560x420")
        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text=(
            f"Montant TTC : {fmt_cfa(montant_ttc)} F CFA — répartissez ce montant sur une ou plusieurs "
            f"échéances prévues (purement planifié, rien n'est encore comptabilisé)."
        ), wraplength=520, foreground="#595959").pack(anchor="w", padx=10, pady=(10, 6))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=10)
        ttk.Label(form, text="Échéance (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Montant :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Button(form, text="Ajouter la tranche", command=self.ajouter_tranche).grid(row=0, column=4, padx=12)

        self.tree = ttk.Treeview(self, columns=("date", "montant"), show="headings", height=8)
        self.tree.heading("date", text="Échéance")
        self.tree.heading("montant", text="Montant")
        self.tree.column("date", width=140, anchor="w")
        self.tree.column("montant", width=160, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=10, pady=8)
        ttk.Button(self, text="Retirer la tranche sélectionnée", command=self.retirer_tranche).pack(
            anchor="w", padx=10)

        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=10, pady=6)

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Enregistrer l'échéancier", command=self.enregistrer).pack(side="left")
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=8)

        self.tranches = []
        existantes = appeler(self, remote, "list_echeances_facture_vente", facture_id)
        if existantes is not APPEL_ECHEC:
            self.tranches = [{"date_echeance": t["date_echeance"], "montant": t["montant"]} for t in existantes]
        self._refresh_tree()

    def _refresh_tree(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        total = 0.0
        for t in self.tranches:
            self.tree.insert("", "end", values=(core.to_display_date(t["date_echeance"]), fmt_cfa(t["montant"])))
            total += t["montant"]
        ecart = total - self.montant_ttc
        etat = "✓ correspond au montant TTC" if abs(ecart) < 1 else f"⚠ écart de {fmt_cfa(ecart)}"
        self.total_var.set(f"Total des tranches : {fmt_cfa(total)} F CFA — {etat}")

    def ajouter_tranche(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "Date d'échéance invalide.", parent=self)
            return
        try:
            montant = float(self.montant_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.", parent=self)
            return
        if montant <= 0:
            messagebox.showerror("Erreur", "Le montant doit être strictement positif.", parent=self)
            return
        self.tranches.append({"date_echeance": date_str, "montant": montant})
        self.montant_var.set("")
        self._refresh_tree()

    def retirer_tranche(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.tranches[idx]
        self._refresh_tree()

    def enregistrer(self):
        if not self.tranches:
            messagebox.showwarning("Vide", "Ajoutez au moins une tranche.", parent=self)
            return
        if appeler(self, self.remote, "set_echeancier_facture_vente", self.facture_id,
                   self.tranches) is APPEL_ECHEC:
            return
        messagebox.showinfo("Enregistré", "Échéancier planifié sur le serveur.", parent=self)
        self.on_saved()
        self.destroy()


class RemoteFacturationTab(ttk.Frame):
    """Facturation clients (COMMERCIAL) via le réseau — une facture
    validée comptabilise directement la vente (débit client, crédit
    compte de vente + TVA), même moteur que l'application de bureau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.facture_id_selectionnee = None

        ttk.Label(self, text="FACTURATION", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        header = ttk.LabelFrame(self, text="Nouvelle facture")
        header.pack(fill="x", padx=16, pady=4)
        ttk.Label(header, text="Numéro :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.numero_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.numero_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(header, textvariable=self.date_var, width=12).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Client :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.client_var = tk.StringVar()
        self.client_combo = ttk.Combobox(header, textvariable=self.client_var, width=26)
        self.client_combo.grid(row=0, column=5, padx=4)
        self.client_combo.bind("<KeyRelease>", self._on_client_keyrelease)
        self._refresh_client_values()
        ttk.Button(header, text="Créer la facture", command=self.creer).grid(row=0, column=6, padx=12)
        ttk.Button(header, text="✓ Valider et envoyer en Saisie", command=self.valider).grid(
            row=0, column=7, padx=4)
        ttk.Label(header, text="Date de règlement prévu (JJ/MM/AAAA) :").grid(
            row=1, column=0, columnspan=2, sticky="w", padx=4, pady=(6, 0))
        self.date_paiement_prevu_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.date_paiement_prevu_var, width=14).grid(
            row=1, column=2, sticky="w", padx=4, pady=(6, 0))
        ttk.Label(header, text=(
            "Obligatoire pour valider — dès qu'elle est renseignée, la facture est comptabilisée et "
            "devient visible dans COMMERCIAL > Paiement/Recouvrement."
        ), foreground="#595959", wraplength=650).grid(row=1, column=3, columnspan=4, sticky="w", padx=4, pady=(6, 0))
        echeancier_frame = ttk.Frame(header)
        echeancier_frame.grid(row=2, column=0, columnspan=7, sticky="we", padx=4, pady=(6, 0))
        ttk.Button(echeancier_frame, text="Planifier un échéancier (plusieurs tranches, optionnel)",
                   command=self.modifier_echeancier_facture).pack(side="left")
        self.echeancier_apercu_var = tk.StringVar()
        ttk.Label(echeancier_frame, textvariable=self.echeancier_apercu_var, foreground="#595959").pack(
            side="left", padx=12)

        self.sync_frame = ttk.Frame(header)
        self.sync_frame.grid(row=3, column=0, columnspan=8, sticky="we", padx=4, pady=(6, 0))
        ttk.Label(self.sync_frame, text=(
            "⚠ Cette facture est validée mais n'a jamais été envoyée dans Paiement/Recouvrement "
            "(validée avant l'existence de cette liaison automatique) :"
        ), foreground="#B00020").pack(side="left")
        ttk.Button(self.sync_frame, text="Envoyer vers Paiement/Recouvrement (sans toucher à la comptabilité)",
                   command=self.synchroniser_recouvrement).pack(side="left", padx=8)
        self.sync_frame.grid_remove()

        ligne_frame = ttk.LabelFrame(self, text="Lignes (une fois la facture créée, sélectionnée dans la liste)")
        ligne_frame.pack(fill="x", padx=16, pady=(0, 8))
        ttk.Label(ligne_frame, text="Compte de vente :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.compte_var = tk.StringVar()
        self.compte_combo = ttk.Combobox(ligne_frame, textvariable=self.compte_var, width=26)
        self.compte_combo.grid(row=0, column=1, padx=4)
        self.compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
        self._refresh_compte_values()
        ttk.Label(ligne_frame, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.libelle_var, width=26).grid(row=0, column=3, padx=4)
        ttk.Label(ligne_frame, text="Quantité :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.quantite_var = tk.StringVar(value="1")
        ttk.Entry(ligne_frame, textvariable=self.quantite_var, width=8).grid(row=0, column=5, padx=4)
        ttk.Label(ligne_frame, text="Prix unitaire :").grid(row=0, column=6, sticky="w", padx=(12, 4))
        self.prix_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.prix_var, width=12).grid(row=0, column=7, padx=4)
        ttk.Button(ligne_frame, text="Ajouter la ligne", command=self.ajouter_ligne).grid(row=0, column=8, padx=12)

        self.tree_lignes = ttk.Treeview(self, columns=("compte", "libelle", "qte", "prix"), show="headings",
                                         height=6)
        for c, h, w in zip(("compte", "libelle", "qte", "prix"), ["Compte", "Libellé", "Qté", "Prix unit."],
                           [100, 300, 60, 110]):
            self.tree_lignes.heading(c, text=h)
            self.tree_lignes.column(c, width=w, anchor="w")
        self.tree_lignes.pack(fill="x", padx=16, pady=(0, 4))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=8)
        ttk.Button(btns, text="Valider la facture (comptabilise la vente sur le serveur)",
                   command=self.valider).pack(side="left")
        ttk.Button(btns, text="Supprimer la facture sélectionnée (brouillon uniquement)",
                   command=self.supprimer_facture).pack(side="left", padx=8)
        ttk.Button(btns, text="Aperçu avant impression",
                   command=self.imprimer_facture).pack(side="left", padx=8)

        ttk.Separator(self).pack(fill="x", padx=16, pady=4)
        ttk.Label(self, text="Factures existantes", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(2, 4))
        cols = ("id", "numero", "date", "client", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=10)
        for c, h, w in zip(cols, ["ID", "Numéro", "Date", "Client", "Statut"], [40, 100, 90, 260, 100]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.tree.bind("<<TreeviewSelect>>", self._on_select_facture)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _refresh_client_values(self):
        items = self._appeler("list_clients")
        if items is not APPEL_ECHEC:
            self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _on_client_keyrelease(self, event=None):
        query = self._extract_code(self.client_var.get())
        if query:
            items = self._appeler("list_clients", query)
            if items is not APPEL_ECHEC:
                self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _refresh_compte_values(self):
        items = self._appeler("search_accounts", "7", limit=50)
        if items is not APPEL_ECHEC:
            self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "7"]

    def _on_compte_keyrelease(self, event=None):
        query = self._extract_code(self.compte_var.get())
        items = self._appeler("search_accounts", query, limit=30)
        if items is not APPEL_ECHEC:
            self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def creer(self):
        if not self.numero_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le numéro est obligatoire.", parent=self)
            return
        client_code = self._extract_code(self.client_var.get())
        if not client_code:
            messagebox.showwarning("Champ manquant", "Choisissez un client.", parent=self)
            return
        date_str = core.to_iso_date(self.date_var.get().strip())
        fid = self._appeler("create_facture_vente", self.numero_var.get().strip(), date_str, client_code)
        if fid is APPEL_ECHEC:
            return
        self.facture_id_selectionnee = fid
        messagebox.showinfo("Créée", f"Facture « {self.numero_var.get()} » créée (brouillon) — ajoutez des "
                                      f"lignes puis validez.", parent=self)
        self.numero_var.set("")
        self.refresh()

    def ajouter_ligne(self):
        if not self.facture_id_selectionnee:
            messagebox.showinfo("Info", "Créez ou sélectionnez d'abord une facture dans la liste.", parent=self)
            return
        compte = self._extract_code(self.compte_var.get())
        libelle = self.libelle_var.get().strip()
        if not compte or not libelle:
            messagebox.showwarning("Champ manquant", "Compte de vente et libellé sont obligatoires.", parent=self)
            return
        try:
            qte = float(self.quantite_var.get() or 0)
            prix = float(self.prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        r = self._appeler("add_ligne_facture_vente", self.facture_id_selectionnee, compte, libelle, qte, prix)
        if r is APPEL_ECHEC:
            return
        self.compte_var.set(""); self.libelle_var.set(""); self.quantite_var.set("1"); self.prix_var.set("")
        self._refresh_lignes()

    def valider(self):
        if not self.facture_id_selectionnee:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture dans la liste.", parent=self)
            return
        date_paiement_prevu = core.to_iso_date(self.date_paiement_prevu_var.get().strip())
        if not date_paiement_prevu:
            messagebox.showwarning(
                "Date manquante",
                "Renseignez la date de règlement prévu avant de valider — c'est elle qui déclenche la "
                "comptabilisation et la visibilité dans Paiement/Recouvrement.", parent=self)
            return
        if not messagebox.askyesno("Valider cette facture",
                                    "La facture va être comptabilisée sur le serveur (débit client, crédit "
                                    "vente + TVA), et apparaîtra dans COMMERCIAL > Paiement/Recouvrement avec son "
                                    "échéancier. Continuer ?", parent=self):
            return
        r = self._appeler("valider_facture_vente", self.facture_id_selectionnee,
                           date_paiement_prevu=date_paiement_prevu)
        if r is APPEL_ECHEC:
            return
        messagebox.showinfo("Validée", "Facture comptabilisée sur le serveur.", parent=self)
        self.refresh()

    def imprimer_facture(self):
        if not self.facture_id_selectionnee:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture dans la liste.", parent=self)
            return
        html = self._appeler("render_facture_vente_html", self.facture_id_selectionnee)
        if html is APPEL_ECHEC:
            return
        import tempfile
        import webbrowser
        import os
        path = os.path.join(tempfile.gettempdir(), f"facture_vente_{self.facture_id_selectionnee}.html")
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        webbrowser.open(f"file://{path}")

    def supprimer_facture(self):
        if not self.facture_id_selectionnee:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture dans la liste.", parent=self)
            return
        if not messagebox.askyesno("Confirmer",
                                    "Supprimer cette facture et toutes ses lignes ? Impossible si elle est "
                                    "déjà validée.", parent=self):
            return
        r = self._appeler("delete_facture_vente", self.facture_id_selectionnee)
        if r is APPEL_ECHEC:
            return
        self.facture_id_selectionnee = None
        self._refresh_lignes()
        self.refresh()

    def _on_select_facture(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.facture_id_selectionnee = int(v[0])
        self._refresh_lignes()
        self._refresh_echeancier_apercu()
        self._refresh_sync_warning()

    def _refresh_sync_warning(self):
        if not self.facture_id_selectionnee:
            self.sync_frame.grid_remove()
            return
        f = self._appeler("get_facture_vente", self.facture_id_selectionnee)
        if f is APPEL_ECHEC or not f:
            self.sync_frame.grid_remove()
            return
        if f["statut"] == "validee" and not f.get("facture_client_id"):
            self.sync_frame.grid()
        else:
            self.sync_frame.grid_remove()

    def synchroniser_recouvrement(self):
        date_str = core.to_iso_date(self.date_paiement_prevu_var.get().strip())
        if not date_str:
            messagebox.showwarning(
                "Date manquante",
                "Renseignez la date de règlement prévu ci-dessus avant d'envoyer cette facture vers "
                "Paiement/Recouvrement.", parent=self)
            return
        if self._appeler("synchroniser_facture_vente_recouvrement", self.facture_id_selectionnee,
                          date_str) is APPEL_ECHEC:
            return
        messagebox.showinfo(
            "Envoyée",
            "Facture envoyée vers Paiement/Recouvrement, avec son échéancier — aucune écriture "
            "comptable n'a été modifiée.", parent=self)
        self._refresh_sync_warning()

    def _refresh_echeancier_apercu(self):
        if not self.facture_id_selectionnee:
            self.echeancier_apercu_var.set("")
            return
        tranches = self._appeler("list_echeances_facture_vente", self.facture_id_selectionnee)
        if tranches is APPEL_ECHEC:
            return
        if not tranches:
            self.echeancier_apercu_var.set("Aucun échéancier planifié (paiement en une fois à la date ci-dessus).")
        else:
            self.echeancier_apercu_var.set(
                f"{len(tranches)} tranche(s) planifiée(s) : " +
                ", ".join(f"{core.to_display_date(t['date_echeance'])} ({fmt_cfa(t['montant'])})" for t in tranches))

    def modifier_echeancier_facture(self):
        if not self.facture_id_selectionnee:
            messagebox.showinfo("Info", "Créez ou sélectionnez d'abord une facture.", parent=self)
            return
        totals = self._appeler("compute_facture_totals", self.facture_id_selectionnee)
        if totals is APPEL_ECHEC:
            return
        if totals["total_ttc"] <= 0:
            messagebox.showinfo("Info", "Ajoutez d'abord des lignes avant de planifier un échéancier.", parent=self)
            return
        RemoteFactureVenteEcheancierDialog(self, self.remote, self.facture_id_selectionnee, totals["total_ttc"],
                                            self._refresh_echeancier_apercu)

    def _refresh_lignes(self):
        for row in self.tree_lignes.get_children():
            self.tree_lignes.delete(row)
        if not self.facture_id_selectionnee:
            return
        lignes = self._appeler("list_lignes_facture_vente", self.facture_id_selectionnee)
        if lignes is APPEL_ECHEC:
            return
        for l in lignes:
            self.tree_lignes.insert("", "end", values=(
                l["compte_vente"], l["libelle"], f"{l['quantite']:g}", fmt_cfa(l["prix_unitaire"])))

    def refresh(self):
        factures = self._appeler("list_factures_vente")
        if factures is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for f in factures:
            self.tree.insert("", "end", values=(
                f["id"], f["numero"], core.to_display_date(f["date_facture"]), f["raison_sociale"], f["statut"]))


class RemoteStocksSyntheseTab(ttk.Frame):
    """Détail réel de chaque compte de stock, avec édition du stock initial
    (valeur + quantité) — équivalent réseau complet de StocksSyntheseTab."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_code = None
        ttk.Label(self, text=(
            "Détail RÉEL de chaque compte de stock utilisé (pas seulement les comptes centralisateurs "
            "310000/320000/331000/360000). Cliquez une ligne, modifiez la valeur puis « Enregistrer »."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=8, pady=(8, 0))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=8, pady=4)
        ttk.Label(filt, text="Catégorie :").pack(side="left")
        self.categorie_var = tk.StringVar(value="Toutes")
        ttk.Combobox(filt, textvariable=self.categorie_var, width=28, state="readonly", values=[
            "Toutes", "31 — Marchandises", "32 — Matières premières",
            "33 — Autres approvisionnements", "36 — Produits finis",
        ]).pack(side="left", padx=4)
        ttk.Button(filt, text="Filtrer", command=self.refresh).pack(side="left", padx=8)

        ttk.Label(filt, text="Marge de valorisation des produits finis par défaut (%) :").pack(
            side="left", padx=(24, 4))
        self.marge_defaut_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.marge_defaut_var, width=6).pack(side="left", padx=2)
        ttk.Button(filt, text="Enregistrer la marge", command=self.save_marge_defaut).pack(side="left", padx=4)

        edit_bar = ttk.Frame(self)
        edit_bar.pack(fill="x", padx=8, pady=4)
        ttk.Label(edit_bar, text="Stock initial (valeur) du compte sélectionné :").pack(side="left")
        self.initial_var = tk.StringVar()
        ttk.Entry(edit_bar, textvariable=self.initial_var, width=14).pack(side="left", padx=4)
        ttk.Button(edit_bar, text="Enregistrer la valeur", command=self.save_initial).pack(side="left", padx=4)
        ttk.Label(edit_bar, text="Quantité initiale :").pack(side="left", padx=(16, 0))
        self.qte_initial_var = tk.StringVar()
        ttk.Entry(edit_bar, textvariable=self.qte_initial_var, width=14).pack(side="left", padx=4)
        ttk.Button(edit_bar, text="Enregistrer la quantité", command=self.save_qte_initial).pack(
            side="left", padx=4)
        ttk.Label(edit_bar, text="(pour un nouveau compte : tapez son n° ci-dessus, puis enregistrez)",
                  foreground="#595959").pack(side="left", padx=8)

        cols = ("code", "label", "initial", "entrees", "sorties", "final",
                "qte_initiale", "qte_entrees", "qte_sorties", "qte_finale", "cump")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["N° Compte", "Libellé", "Stock initial", "Entrées (Débit)", "Sorties (Crédit)", "Stock final",
                   "Qté initiale", "Qté entrées", "Qté sorties", "Qté finale", "Coût unit. moyen"]
        widths = [90, 190, 100, 100, 100, 100, 80, 80, 80, 80, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=8, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.selected_code = values[0]
        self.initial_var.set(values[2])
        self.qte_initial_var.set(values[6])

    def save_marge_defaut(self):
        try:
            value = float(self.marge_defaut_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La marge doit être un nombre.", parent=self)
            return
        if self._appeler("set_setting", "marge_production_defaut", value) is APPEL_ECHEC:
            return
        messagebox.showinfo("Enregistré", "Marge de valorisation par défaut enregistrée.", parent=self)

    def save_initial(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un compte de stock dans le tableau "
                                         "(ou saisissez son code dans le champ ci-dessus).", parent=self)
            return
        try:
            value = float(self.initial_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le stock initial doit être un nombre.", parent=self)
            return
        if self._appeler("set_stock_initial", self.selected_code, value) is APPEL_ECHEC:
            return
        self.refresh()

    def save_qte_initial(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un compte de stock dans le tableau.", parent=self)
            return
        try:
            value = float(self.qte_initial_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La quantité initiale doit être un nombre.", parent=self)
            return
        if self._appeler("set_stock_qte_initiale", self.selected_code, value) is APPEL_ECHEC:
            return
        self.refresh()

    def refresh(self):
        marge = self._appeler("get_setting", "marge_production_defaut", 30.0)
        if marge is not APPEL_ECHEC:
            self.marge_defaut_var.set(str(marge))
        cat = self.categorie_var.get()
        prefixes = [cat.split(" — ")[0].strip()] if cat != "Toutes" else None
        stocks = self._appeler("compute_stocks_detail", prefixes=prefixes)
        if stocks is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for s in stocks:
            cump = f"{fmt_cfa(s['cout_unitaire_moyen'])}" if s["cout_unitaire_moyen"] is not None else "—"
            self.tree.insert("", "end", values=(
                s["code"], s["label"], f"{fmt_cfa(s['stock_initial'])}",
                f"{fmt_cfa(s['entrees'])}", f"{fmt_cfa(s['sorties'])}", f"{fmt_cfa(s['stock_final'])}",
                f"{s['qte_initiale']:g}", f"{s['qte_entrees']:g}", f"{s['qte_sorties']:g}",
                f"{s['qte_finale']:g}", cump,
            ))


class RemoteStocksMouvementsTab(ttk.Frame):
    """Détail de tous les mouvements comptables sur les comptes de stock
    (classe 3), avec leur origine — équivalent réseau de StocksMouvementsTab."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text=(
            "Tous les mouvements comptables des comptes de stock de l'exercice en cours, y compris "
            "ceux générés automatiquement par la validation d'une facture de vente ou d'achat."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=8, pady=(8, 4))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=8, pady=4)
        ttk.Label(filt, text="Filtrer par origine :").pack(side="left")
        self.origine_var = tk.StringVar(value="Toutes")
        ttk.Combobox(filt, textvariable=self.origine_var, width=18, state="readonly",
                     values=["Toutes", "Facturation", "Facture frs", "Saisie directe (auto)",
                             "Saisie manuelle"]).pack(side="left", padx=4)
        ttk.Button(filt, text="Filtrer", command=self.refresh).pack(side="left", padx=8)

        cols = ("date", "piece", "compte", "compte_label", "libelle", "debit", "credit", "quantite",
                "qte_cumulee", "valeur_cumulee", "origine")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["Date", "Pièce", "Compte", "Libellé du compte", "Libellé écriture",
                   "Débit (valeur)", "Crédit (valeur)", "Qté mvt", "Qté cumulée", "Valeur cumulée", "Origine"]
        widths = [90, 80, 80, 150, 190, 90, 90, 70, 90, 100, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("auto", foreground="#1F4E78")
        self.tree.pack(fill="both", padx=8, pady=8)

        self.totals_var = tk.StringVar()
        ttk.Label(self, textvariable=self.totals_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=8, pady=(0, 8))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        mouvements = self._appeler("compute_mouvements_stocks")
        if mouvements is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        filtre = self.origine_var.get()
        total_d = total_c = 0.0
        for m in mouvements:
            if filtre != "Toutes" and m["origine"] != filtre:
                continue
            tags = ("auto",) if m["origine"] != "Saisie manuelle" else ()
            self.tree.insert("", "end", tags=tags, values=(
                core.to_display_date(m["date"]), m["piece"] or "", m["compte"], m["compte_label"],
                m["libelle"] or "", f"{fmt_cfa(m['debit'])}" if m["debit"] else "",
                f"{fmt_cfa(m['credit'])}" if m["credit"] else "", f"{m['quantite']:g}" if m["quantite"] else "",
                f"{m['qte_cumulee']:g}", f"{fmt_cfa(m['valeur_cumulee'])}",
                m["origine"],
            ))
            total_d += m["debit"]
            total_c += m["credit"]
        self.totals_var.set(f"TOTAL — Débit : {fmt_cfa(total_d)}   Crédit : {fmt_cfa(total_c)}")


class RemoteStocksTab(ttk.Frame):
    """Regroupe la synthèse par compte (éditable) et le détail des
    mouvements comptables — équivalent réseau complet de StocksTab (bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        inner = ttk.Notebook(self)
        inner.pack(fill="both", expand=True)
        self.synthese_tab = RemoteStocksSyntheseTab(inner, remote)
        self.mouvements_tab = RemoteStocksMouvementsTab(inner, remote)
        inner.add(self.synthese_tab, text="Synthèse par compte")
        inner.add(self.mouvements_tab, text="Mouvements comptables (classe 3)")

    def refresh(self):
        self.synthese_tab.refresh()
        self.mouvements_tab.refresh()


class RemoteRapprochementBancaireTab(ttk.Frame):
    """Rapprochement bancaire via le réseau (menu TRESORERIE) —
    équivalent réseau complet de RapprochementBancaireTab (bureau) :
    liste des comptes de banque, clic pour ouvrir le rapprochement
    détaillé de ce compte."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="RAPPROCHEMENT BANCAIRE", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Tous les comptes de banque (52xxxx) avec leur solde comptable actuel. Cliquez sur une "
            "ligne pour pointer ses mouvements par rapport au relevé bancaire papier."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 8))

        cols = ("compte", "libelle", "solde")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=20)
        for c, h, w in zip(cols, ["Compte", "Libellé", "Solde comptable"], [100, 400, 180]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w" if c != "solde" else "e")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        compte, libelle = v[0], v[1]
        RemoteRapprochementCompteDialog(self, self.remote, compte, libelle, on_saved=self.refresh)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        comptes = self._appeler("compute_comptes_prefixe_periode", "52")
        if comptes is APPEL_ECHEC:
            return
        for c in comptes:
            self.tree.insert("", "end", values=(c["code"], c["label"], fmt_cfa(c["solde_fin_periode"])))


class RemoteTresorerieTab(ttk.Frame):
    """Trésorerie via le réseau — 3 onglets, comme le bureau : Banques
    (Entrées/Sorties), Engagements à payer, Échéances (prévisionnel)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="TRÉSORERIE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=16, pady=8)
        tab_banques = ttk.Frame(notebook)
        tab_engagements = ttk.Frame(notebook)
        tab_echeances = ttk.Frame(notebook)
        notebook.add(tab_banques, text="Banques (Entrées / Sorties)")
        notebook.add(tab_engagements, text="Engagements à payer")
        notebook.add(tab_echeances, text="Échéances (prévisionnel)")

        self._build_banques(tab_banques)
        self._build_engagements(tab_engagements)
        self._build_echeances(tab_echeances)
        self.refresh()

    def _build_banques(self, parent):
        ttk.Button(parent, text="Actualiser", command=self.refresh).pack(anchor="w", padx=8, pady=(8, 4))
        cols1 = ("compte", "libelle", "debut", "entrees", "sorties", "fin")
        self.tree_banques = ttk.Treeview(parent, columns=cols1, show="headings", height=14)
        for c, h, w in zip(cols1, ["Compte", "Libellé", "Solde début", "Entrées", "Sorties", "Solde fin"],
                           [90, 220, 130, 130, 130, 140]):
            self.tree_banques.heading(c, text=h)
            self.tree_banques.column(c, width=w, anchor="w" if c in ("compte", "libelle") else "e")
        self.tree_banques.tag_configure("total", background="#1F4E78", foreground="white",
                                         font=("Segoe UI", 9, "bold"))
        self.tree_banques.pack(fill="both", expand=True, padx=8, pady=8)

    def _build_engagements(self, parent):
        ttk.Label(parent, text=(
            "Règlements déjà validés (charge comptabilisée) dont le paiement bancaire n'a pas encore été "
            "enregistré — ce que l'entreprise doit encore décaisser, comparé à la trésorerie disponible."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=8, pady=(8, 8))
        ttk.Button(parent, text="Actualiser", command=self.refresh).pack(anchor="w", padx=8)
        self.synthese_var = tk.StringVar()
        ttk.Label(parent, textvariable=self.synthese_var, font=("Segoe UI", 10, "bold"), wraplength=1200).pack(
            anchor="w", padx=8, pady=(8, 8))
        cols2 = ("numero", "date", "fournisseur", "montant", "statut")
        self.tree_engagements = ttk.Treeview(parent, columns=cols2, show="headings", height=14)
        for c, h, w in zip(cols2, ["N° Règlement", "Date", "Fournisseur", "Montant restant à payer", "Statut"],
                           [130, 100, 220, 170, 220]):
            self.tree_engagements.heading(c, text=h)
            self.tree_engagements.column(c, width=w, anchor="w" if c != "montant" else "e")
        self.tree_engagements.tag_configure("total", background="#1F4E78", foreground="white",
                                             font=("Segoe UI", 9, "bold"))
        self.tree_engagements.tag_configure("retard", foreground="#B00020")
        self.tree_engagements.pack(fill="both", expand=True, padx=8, pady=8)

    def _build_echeances(self, parent):
        ttk.Label(parent, text=(
            "Tous les engagements fournisseurs (bons de commande — même pas encore facturés — et "
            "factures validées) et clients ayant une échéance de paiement pas encore réglée, répartis "
            "par mois avec le(s) compte(s) de charge concerné(s)."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=8, pady=(8, 8))
        ech_top = ttk.Frame(parent)
        ech_top.pack(fill="x", padx=8)
        ttk.Label(ech_top, text="À partir du (JJ/MM/AAAA) :").pack(side="left")
        self.echeance_date_from_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(ech_top, textvariable=self.echeance_date_from_var, width=12).pack(side="left", padx=4)
        ttk.Label(ech_top, text="Nombre de mois :").pack(side="left", padx=(12, 4))
        self.echeance_nb_mois_var = tk.StringVar(value="6")
        ttk.Entry(ech_top, textvariable=self.echeance_nb_mois_var, width=4).pack(side="left")
        ttk.Button(ech_top, text="Actualiser", command=self.refresh).pack(side="left", padx=8)
        self.echeances_synthese_var = tk.StringVar()
        ttk.Label(parent, textvariable=self.echeances_synthese_var, font=("Segoe UI", 11, "bold")).pack(
            anchor="w", padx=8, pady=8)
        self.tree_echeances = ttk.Treeview(parent, show="headings", height=14)
        self.tree_echeances.tag_configure("sous_total", background="#DCE6F1", font=("Segoe UI", 9, "bold"))
        self.tree_echeances.tag_configure("total", background="#1F4E78", foreground="white",
                                           font=("Segoe UI", 10, "bold"))
        self.tree_echeances.pack(fill="both", expand=True, padx=8, pady=8)

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        r1 = self._appeler("compute_tresorerie_banques_horizontal")
        if r1 is APPEL_ECHEC:
            return
        lignes, total = r1
        for row in self.tree_banques.get_children():
            self.tree_banques.delete(row)
        for l in lignes:
            self.tree_banques.insert("", "end", values=(
                l["code"], l["label"], fmt_cfa(l["solde_debut_periode"]), fmt_cfa(l["debit_periode"]),
                fmt_cfa(l["credit_periode"]), fmt_cfa(l["solde_fin_periode"])))
        self.tree_banques.insert("", "end", tags=("total",), values=(
            "TOTAL", "", fmt_cfa(total["solde_debut_periode"]), fmt_cfa(total["debit_periode"]),
            fmt_cfa(total["credit_periode"]), fmt_cfa(total["solde_fin_periode"])))

        d = self._appeler("compute_engagements_a_payer")
        if d is APPEL_ECHEC:
            return
        for row in self.tree_engagements.get_children():
            self.tree_engagements.delete(row)
        for e in d["engagements"]:
            tag = ("retard",) if e.get("en_retard") else ()
            self.tree_engagements.insert("", "end", tags=tag, values=(
                e["numero"], core.to_display_date(e["date_reglement"]), e["raison_sociale"],
                fmt_cfa(e["net_a_payer"]), e.get("statut_paiement", "")))
        self.tree_engagements.insert("", "end", tags=("total",), values=(
            "TOTAL ENGAGEMENTS", "", "", fmt_cfa(d["total_engagements"]), ""))
        etat = "✓ peut faire face" if d["peut_faire_face"] else "⚠ insuffisant"
        self.synthese_var.set(f"Trésorerie disponible : {fmt_cfa(d['treso_disponible'])}   —   Engagements : "
                               f"{fmt_cfa(d['total_engagements'])}   —   {etat}")

        date_from = core.to_iso_date(self.echeance_date_from_var.get().strip())
        try:
            nb_mois = max(1, min(12, int(self.echeance_nb_mois_var.get() or 6)))
        except ValueError:
            nb_mois = 6
        p = self._appeler("compute_echeances_tresorerie_pivot", date_from=date_from, nb_mois=nb_mois)
        if p is APPEL_ECHEC:
            return
        for row in self.tree_echeances.get_children():
            self.tree_echeances.delete(row)
        cols = ["type", "tiers", "piece", "compte"] + [f"mois{i}" for i in range(len(p["colonnes"]))] + ["total"]
        self.tree_echeances["columns"] = cols
        headers = ["Type", "Tiers", "Pièce", "Compte(s) de charge"] + p["colonnes"] + ["Total"]
        widths = [140, 200, 90, 130] + [110] * len(p["colonnes"]) + [120]
        for c, h, w in zip(cols, headers, widths):
            self.tree_echeances.heading(c, text=h)
            self.tree_echeances.column(c, width=w, anchor="w" if c in ("type", "tiers", "piece", "compte") else "e")
        for l in p["lignes"]:
            self.tree_echeances.insert("", "end", values=(
                l["type"], l["tiers"], l["piece"], l["compte_charge"],
                *[fmt_cfa(m) if m else "" for m in l["montants"]], fmt_cfa(l["total"])))
        self.tree_echeances.insert("", "end", tags=("sous_total",), values=(
            "SOUS-TOTAL PRODUITS (CLIENTS)", "", "", "",
            *[fmt_cfa(t) for t in p["sous_total_produits"]], fmt_cfa(p["total_produits"])))
        self.tree_echeances.insert("", "end", tags=("sous_total",), values=(
            "SOUS-TOTAL CHARGES (FOURNISSEURS)", "", "", "",
            *[fmt_cfa(t) for t in p["sous_total_charges"]], fmt_cfa(p["total_charges"])))
        self.tree_echeances.insert("", "end", tags=("total",), values=(
            "TOTAL", "", "", "", *[fmt_cfa(t) for t in p["totaux_colonnes"]], fmt_cfa(p["total_general"])))
        self.echeances_synthese_var.set(
            f"Sur les {nb_mois} prochains mois — impact net sur la trésorerie : "
            f"{fmt_cfa(p['total_general'])} F CFA")


class RemoteImmobilisationsTab(ttk.Frame):
    """Immobilisations via le réseau — équivalent réseau complet de
    ImmobilisationsTab (bureau) : sélection d'un compte, puis
    fournisseur / prix d'achat / date d'acquisition éditables."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_compte = None
        ttk.Label(self, text="IMMOBILISATIONS", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Comptes de classe 2 ayant un solde dans la Balance. Sélectionnez une ligne pour renseigner "
            "son fournisseur et son prix d'achat."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=16, pady=(0, 4))
        ttk.Button(import_bar, text="Importer des immobilisations (.xlsx)", command=self.import_xlsx).pack(
            side="left", padx=2)
        ttk.Button(import_bar, text="Télécharger un modèle (.xlsx)", command=self.download_template).pack(
            side="left", padx=2)

        form = ttk.LabelFrame(self, text="Fiche du compte sélectionné")
        form.pack(fill="x", padx=16, pady=4)
        self.compte_label_var = tk.StringVar(value="(sélectionnez une ligne dans le tableau ci-dessous)")
        ttk.Label(form, textvariable=self.compte_label_var, font=("Segoe UI", 9, "bold")).grid(
            row=0, column=0, columnspan=4, sticky="w", padx=4, pady=4)
        ttk.Label(form, text="Fournisseur :").grid(row=1, column=0, sticky="w", padx=4)
        self.fournisseur_var = tk.StringVar()
        self.fournisseur_combo = ttk.Combobox(form, textvariable=self.fournisseur_var, width=28)
        self.fournisseur_combo.grid(row=1, column=1, padx=4, pady=4)
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self._refresh_fournisseur_values()
        ttk.Label(form, text="Prix d'achat :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.prix_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.prix_var, width=16).grid(row=1, column=3, padx=4)
        ttk.Label(form, text="Date d'acquisition :").grid(row=1, column=4, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=1, column=5, padx=4)
        ttk.Label(form, text="Base de répartition (quantité annuelle) :").grid(
            row=2, column=0, sticky="w", padx=4, pady=(4, 0))
        self.base_qte_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.base_qte_var, width=12).grid(row=2, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Unité (tonnes, heures...) :").grid(row=2, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.base_unite_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.base_unite_var, width=16).grid(row=2, column=3, padx=4, pady=(4, 0))
        ttk.Label(form, text="Amortissement annuel (si pas comptabilisé) :").grid(
            row=2, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.amort_manuel_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.amort_manuel_var, width=16).grid(row=2, column=5, padx=4, pady=(4, 0))
        ttk.Label(form, text=(
            "Pour utiliser cet équipement dans une recette de fabrication (composant « Amortissement "
            "d'équipement ») : indiquez sa capacité annuelle normale (ex. 5000 tonnes/an ou 2000 heures/an). "
            "Le coût unitaire = amortissement RÉELLEMENT comptabilisé (dotations 68x/28x déjà saisies) ÷ "
            "cette capacité ; si aucune dotation n'est encore comptabilisée pour cet équipement, le montant "
            "« Amortissement annuel » saisi ci-dessus est utilisé à la place, en attendant."
        ), foreground="#595959", wraplength=1050).grid(row=3, column=0, columnspan=6, sticky="w", padx=4, pady=(2, 0))
        ttk.Button(form, text="Enregistrer la fiche", command=self.save_fiche).grid(row=1, column=6, padx=12)
        ttk.Button(form, text="Modifier la fiche", command=self.save_fiche).grid(row=2, column=6, padx=12, pady=(4, 0))

        cols = ("compte", "libelle", "categorie", "fournisseur", "prix_achat", "taux", "brut", "amort", "net")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        headers = ["Compte", "Libellé", "Catégorie", "Fournisseur", "Prix d'achat", "Taux %",
                   "Valeur brute", "Amortissement", "Valeur nette"]
        widths = [80, 200, 220, 160, 110, 60, 110, 110, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 12))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _refresh_fournisseur_values(self):
        items = self._appeler("list_fournisseurs")
        if items is APPEL_ECHEC:
            return
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self.fournisseur_var.get().strip()
        items = self._appeler("list_fournisseurs", query)
        if items is APPEL_ECHEC:
            return
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_compte = v[0]
        self.compte_label_var.set(f"{v[0]} — {v[1]}")
        fiche = self._appeler("get_immobilisation_fiche", v[0])
        if fiche is APPEL_ECHEC:
            return
        fournisseur = None
        if fiche["fournisseur_code"]:
            fournisseur = self._appeler("get_fournisseur", fiche["fournisseur_code"])
            if fournisseur is APPEL_ECHEC:
                return
        self.fournisseur_var.set(
            f"{fiche['fournisseur_code']} — {fournisseur['raison_sociale']}"
            if fournisseur else (fiche["fournisseur_code"] or ""))
        self.prix_var.set(str(fiche["prix_achat"]) if fiche["prix_achat"] else "")
        self.date_var.set(core.to_display_date(fiche["date_acquisition"] or ""))
        self.base_qte_var.set(str(fiche["base_repartition_quantite"]) if fiche.get("base_repartition_quantite") else "")
        self.base_unite_var.set(fiche.get("base_repartition_unite") or "")
        self.amort_manuel_var.set(
            str(fiche["amortissement_annuel_manuel"]) if fiche.get("amortissement_annuel_manuel") else "")

    def save_fiche(self):
        if not self.selected_compte:
            messagebox.showinfo("Info", "Sélectionnez d'abord un compte dans le tableau.", parent=self)
            return
        raw = self.fournisseur_var.get().strip()
        fournisseur_code = raw.split(" — ", 1)[0].strip() if " — " in raw else raw
        try:
            prix = float(self.prix_var.get()) if self.prix_var.get().strip() else 0
        except ValueError:
            messagebox.showerror("Erreur", "Le prix d'achat doit être un nombre.", parent=self)
            return
        base_qte = None
        if self.base_qte_var.get().strip():
            try:
                base_qte = float(self.base_qte_var.get())
            except ValueError:
                messagebox.showerror("Erreur", "La base de répartition doit être un nombre.", parent=self)
                return
        amort_manuel = None
        if self.amort_manuel_var.get().strip():
            try:
                amort_manuel = float(self.amort_manuel_var.get())
            except ValueError:
                messagebox.showerror("Erreur", "L'amortissement annuel doit être un nombre.", parent=self)
                return
        if self._appeler("set_immobilisation_fiche", self.selected_compte, fournisseur_code=fournisseur_code or None,
                          prix_achat=prix, date_acquisition=core.to_iso_date(self.date_var.get().strip()),
                          base_repartition_quantite=base_qte,
                          base_repartition_unite=self.base_unite_var.get().strip() or None,
                          amortissement_annuel_manuel=amort_manuel) is APPEL_ECHEC:
            return
        self.refresh()
        messagebox.showinfo("Enregistré", "Fiche d'immobilisation enregistrée.", parent=self)

    def download_template(self):
        # Ne nécessite aucun accès au serveur : le modèle est un fichier vierge,
        # généré directement sur ce poste avec le module core embarqué dans le client.
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_immobilisations.xlsx", title="Enregistrer le modèle", parent=self,
        )
        if not path:
            return
        try:
            core.export_immobilisations_template(path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de la création du modèle : {exc}", parent=self)
            return
        messagebox.showinfo("Modèle créé", f"Modèle enregistré :\n{path}", parent=self)

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer des immobilisations", parent=self)
        if not path:
            return
        try:
            rows = core.parse_immobilisations_xlsx(path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de la lecture du fichier : {exc}", parent=self)
            return
        resultat = self._appeler("apply_immobilisations_rows", rows)
        if resultat is APPEL_ECHEC:
            return
        imported, warnings = resultat
        self.refresh()
        msg = f"{imported} fiche(s) d'immobilisation importée(s)/mise(s) à jour sur le serveur."
        if warnings:
            msg += "\n\nAvertissements :\n" + "\n".join(warnings[:20])
        messagebox.showinfo("Import terminé", msg, parent=self)

    def refresh(self):
        immos = self._appeler("compute_immobilisations_liste")
        if immos is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for l in immos:
            self.tree.insert("", "end", values=(
                l["compte"], l["libelle"], l.get("categorie") or "",
                l.get("fournisseur_nom") or l.get("fournisseur_code") or "",
                fmt_cfa(l["prix_achat"]) if l["prix_achat"] else "", f"{l['taux_pct']:g}",
                fmt_cfa(l["valeur_brute"]), fmt_cfa(l["amortissement"]), fmt_cfa(l["valeur_nette"])))


class RemoteExpressionBesoinTab(ttk.Frame):
    """Expression de besoin (ENGAGEMENTS-PROJETS) via le réseau — la
    validation fait automatiquement basculer vers un Bon de commande
    (même moteur que l'application de bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.expression_id_selectionnee = None

        ttk.Label(self, text="EXPRESSION DE BESOIN", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))

        header = ttk.LabelFrame(self, text="Nouvelle expression de besoin")
        header.pack(fill="x", padx=16, pady=4)
        ttk.Label(header, text="Numéro :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.numero_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.numero_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(header, textvariable=self.date_var, width=12).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Demandeur :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.demandeur_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.demandeur_var, width=18).grid(row=0, column=5, padx=4)
        ttk.Button(header, text="Créer", command=self.creer).grid(row=0, column=6, padx=12)

        ligne_frame = ttk.LabelFrame(self, text="Lignes (une fois créée, sélectionnée dans la liste)")
        ligne_frame.pack(fill="x", padx=16, pady=(0, 8))
        ttk.Label(ligne_frame, text="Libellé :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.libelle_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.libelle_var, width=30).grid(row=0, column=1, padx=4)
        ttk.Label(ligne_frame, text="Quantité :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.quantite_var = tk.StringVar(value="1")
        ttk.Entry(ligne_frame, textvariable=self.quantite_var, width=10).grid(row=0, column=3, padx=4)
        ttk.Label(ligne_frame, text="Unité :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.unite_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.unite_var, width=10).grid(row=0, column=5, padx=4)
        ttk.Button(ligne_frame, text="Ajouter la ligne", command=self.ajouter_ligne).grid(row=0, column=6, padx=12)

        ttk.Button(self, text="Valider (bascule en Bon de commande sur le serveur)",
                   command=self.valider).pack(anchor="w", padx=16, pady=8)

        ttk.Separator(self).pack(fill="x", padx=16, pady=4)
        ttk.Label(self, text="Expressions existantes", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(2, 4))
        cols = ("id", "numero", "date", "demandeur", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=12)
        for c, h, w in zip(cols, ["ID", "Numéro", "Date", "Demandeur", "Statut"], [40, 100, 90, 200, 100]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def creer(self):
        if not self.numero_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le numéro est obligatoire.", parent=self)
            return
        date_str = core.to_iso_date(self.date_var.get().strip())
        eid = self._appeler("create_expression_besoin", self.numero_var.get(), date_str,
                             demandeur=self.demandeur_var.get().strip())
        if eid is APPEL_ECHEC:
            return
        self.expression_id_selectionnee = eid
        self.numero_var.set("")
        self.refresh()

    def ajouter_ligne(self):
        if not self.expression_id_selectionnee:
            messagebox.showinfo("Info", "Créez ou sélectionnez d'abord une expression dans la liste.", parent=self)
            return
        if not self.libelle_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le libellé est obligatoire.", parent=self)
            return
        try:
            qte = float(self.quantite_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La quantité doit être un nombre.", parent=self)
            return
        r = self._appeler("add_ligne_expression_besoin", self.expression_id_selectionnee,
                           self.libelle_var.get(), qte, unite=self.unite_var.get() or None)
        if r is APPEL_ECHEC:
            return
        self.libelle_var.set(""); self.quantite_var.set("1"); self.unite_var.set("")
        messagebox.showinfo("Ajouté", "Ligne ajoutée.", parent=self)

    def valider(self):
        if not self.expression_id_selectionnee:
            messagebox.showinfo("Info", "Sélectionnez d'abord une expression dans la liste.", parent=self)
            return
        if not messagebox.askyesno("Valider", "Cette expression va basculer en Bon de commande. Continuer ?",
                                    parent=self):
            return
        r = self._appeler("valider_expression_besoin", self.expression_id_selectionnee)
        if r is APPEL_ECHEC:
            return
        messagebox.showinfo("Validée", "Bon de commande créé sur le serveur (menu Bon de commande).", parent=self)
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        self.expression_id_selectionnee = int(self.tree.item(sel[0], "values")[0])

    def refresh(self):
        expressions = self._appeler("list_expressions_besoin")
        if expressions is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for e in expressions:
            self.tree.insert("", "end", values=(
                e["id"], e["numero"], core.to_display_date(e["date_demande"]), e["demandeur"] or "", e["statut"]))


class RemoteBonCommandeEcheancierDialog(tk.Toplevel):
    """Planifie l'échéancier de paiement PRÉVU d'un Bon de commande via le
    réseau — équivalent réseau complet de BonCommandeEcheancierDialog
    (bureau)."""

    def __init__(self, parent, remote: RemoteConnection, bon_id, net_a_payer, on_saved):
        super().__init__(parent)
        self.remote = remote
        self.bon_id = bon_id
        self.net_a_payer = net_a_payer
        self.on_saved = on_saved
        self.title("Échéancier de paiement prévu")
        self.geometry("560x420")
        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text=(
            f"Net à payer : {fmt_cfa(net_a_payer)} F CFA — répartissez ce montant sur une ou plusieurs "
            f"échéances prévues (purement planifié, rien n'est encore comptabilisé)."
        ), wraplength=520, foreground="#595959").pack(anchor="w", padx=10, pady=(10, 6))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=10)
        ttk.Label(form, text="Échéance (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Montant :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Button(form, text="Ajouter la tranche", command=self.ajouter_tranche).grid(row=0, column=4, padx=12)

        self.tree = ttk.Treeview(self, columns=("date", "montant"), show="headings", height=8)
        self.tree.heading("date", text="Échéance")
        self.tree.heading("montant", text="Montant")
        self.tree.column("date", width=140, anchor="w")
        self.tree.column("montant", width=160, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=10, pady=8)
        ttk.Button(self, text="Retirer la tranche sélectionnée", command=self.retirer_tranche).pack(
            anchor="w", padx=10)

        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=10, pady=6)

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Enregistrer l'échéancier", command=self.enregistrer).pack(side="left")
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=8)

        self.tranches = []
        existantes = appeler(self, remote, "list_echeances_bon_commande", bon_id)
        if existantes is not APPEL_ECHEC:
            self.tranches = [{"date_echeance": t["date_echeance"], "montant": t["montant"]} for t in existantes]
        self._refresh_tree()

    def _refresh_tree(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        total = 0.0
        for t in self.tranches:
            self.tree.insert("", "end", values=(core.to_display_date(t["date_echeance"]), fmt_cfa(t["montant"])))
            total += t["montant"]
        ecart = total - self.net_a_payer
        etat = "✓ correspond au net à payer" if abs(ecart) < 1 else f"⚠ écart de {fmt_cfa(ecart)}"
        self.total_var.set(f"Total des tranches : {fmt_cfa(total)} F CFA — {etat}")

    def ajouter_tranche(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "Date d'échéance invalide.", parent=self)
            return
        try:
            montant = float(self.montant_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.", parent=self)
            return
        if montant <= 0:
            messagebox.showerror("Erreur", "Le montant doit être strictement positif.", parent=self)
            return
        self.tranches.append({"date_echeance": date_str, "montant": montant})
        self.montant_var.set("")
        self._refresh_tree()

    def retirer_tranche(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.tranches[idx]
        self._refresh_tree()

    def enregistrer(self):
        if not self.tranches:
            messagebox.showwarning("Vide", "Ajoutez au moins une tranche.", parent=self)
            return
        if appeler(self, self.remote, "set_echeancier_bon_commande", self.bon_id,
                   self.tranches) is APPEL_ECHEC:
            return
        messagebox.showinfo("Enregistré", "Échéancier planifié sur le serveur.", parent=self)
        self.on_saved()
        self.destroy()


class RemoteBonCommandeTab(ttk.Frame):
    """Bon de commande (ENGAGEMENTS-PROJETS) via le réseau — la validation
    comptabilise directement l'achat, comme sur l'application de bureau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.bon_id_selectionne = None

        ttk.Label(self, text="BON DE COMMANDE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        header = ttk.LabelFrame(self, text="Nouveau bon de commande")
        header.pack(fill="x", padx=16, pady=4)
        ttk.Label(header, text="Numéro :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.numero_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.numero_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(header, textvariable=self.date_var, width=12).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Fournisseur (code) :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.fournisseur_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.fournisseur_var, width=14).grid(row=0, column=5, padx=4)
        ttk.Button(header, text="Créer", command=self.creer).grid(row=0, column=6, padx=12)

        ligne_frame = ttk.LabelFrame(self, text="Lignes — un compte débiteur (charge ou immobilisation) est "
                                                  "OBLIGATOIRE pour pouvoir valider")
        ligne_frame.pack(fill="x", padx=16, pady=(0, 8))
        ttk.Label(ligne_frame, text="Compte débiteur :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.compte_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.compte_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Label(ligne_frame, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.libelle_var, width=26).grid(row=0, column=3, padx=4)
        ttk.Label(ligne_frame, text="Quantité :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.quantite_var = tk.StringVar(value="1")
        ttk.Entry(ligne_frame, textvariable=self.quantite_var, width=8).grid(row=0, column=5, padx=4)
        ttk.Label(ligne_frame, text="Prix unitaire :").grid(row=0, column=6, sticky="w", padx=(12, 4))
        self.prix_var = tk.StringVar()
        ttk.Entry(ligne_frame, textvariable=self.prix_var, width=12).grid(row=0, column=7, padx=4)
        ttk.Button(ligne_frame, text="Ajouter la ligne", command=self.ajouter_ligne).grid(row=0, column=8, padx=12)

        ttk.Button(self, text="Valider (comptabilise + crée le Bordereau sur le serveur)",
                   command=self.valider).pack(anchor="w", padx=16, pady=8)
        ttk.Button(self, text="Planifier un échéancier (plusieurs tranches, optionnel)",
                   command=self.modifier_echeancier).pack(anchor="w", padx=16, pady=(0, 8))

        ttk.Separator(self).pack(fill="x", padx=16, pady=4)
        ttk.Label(self, text="Bons de commande existants", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(2, 4))
        cols = ("id", "numero", "date", "fournisseur", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=12)
        for c, h, w in zip(cols, ["ID", "Numéro", "Date", "Fournisseur", "Statut"], [40, 100, 90, 200, 100]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def creer(self):
        if not self.numero_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le numéro est obligatoire.", parent=self)
            return
        date_str = core.to_iso_date(self.date_var.get().strip())
        bid = self._appeler("create_ep_bon_commande", self.numero_var.get(), date_str,
                             fournisseur_code=self.fournisseur_var.get().strip())
        if bid is APPEL_ECHEC:
            return
        self.bon_id_selectionne = bid
        self.numero_var.set("")
        self.refresh()

    def ajouter_ligne(self):
        if not self.bon_id_selectionne:
            messagebox.showinfo("Info", "Créez ou sélectionnez d'abord un bon dans la liste.", parent=self)
            return
        compte = self.compte_var.get().strip()
        libelle = self.libelle_var.get().strip()
        if not compte or not libelle:
            messagebox.showwarning("Champ manquant", "Compte débiteur et libellé sont obligatoires.", parent=self)
            return
        try:
            qte = float(self.quantite_var.get() or 0)
            prix = float(self.prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        r = self._appeler("add_ligne_ep_bon_commande", self.bon_id_selectionne, libelle, qte,
                           prix_unitaire=prix, compte_charge=compte)
        if r is APPEL_ECHEC:
            return
        self.compte_var.set(""); self.libelle_var.set(""); self.quantite_var.set("1"); self.prix_var.set("")
        messagebox.showinfo("Ajoutée", "Ligne ajoutée.", parent=self)

    def modifier_echeancier(self):
        if not self.bon_id_selectionne:
            messagebox.showinfo("Info", "Sélectionnez d'abord un bon dans la liste.", parent=self)
            return
        totals = self._appeler("compute_ep_bon_commande_totals", self.bon_id_selectionne)
        if totals is APPEL_ECHEC:
            return
        if totals["net_a_payer"] <= 0:
            messagebox.showinfo("Info", "Ajoutez d'abord des lignes avec un montant avant de planifier "
                                         "un échéancier.", parent=self)
            return
        RemoteBonCommandeEcheancierDialog(self, self.remote, self.bon_id_selectionne, totals["net_a_payer"],
                                           lambda: None)
        messagebox.showinfo("Ajoutée", "Ligne ajoutée.", parent=self)

    def valider(self):
        if not self.bon_id_selectionne:
            messagebox.showinfo("Info", "Sélectionnez d'abord un bon dans la liste.", parent=self)
            return
        if not messagebox.askyesno(
                "Valider ce bon de commande",
                "Le bon va générer une FACTURE D'ACHAT (brouillon, non comptabilisée) avec les lignes "
                "recopiées. Aucune écriture comptable n'est envoyée à ce stade — c'est la validation de "
                "cette facture, avec sa date de règlement prévu (menu ENGAGEMENTS-PROJETS > Factures "
                "frs), qui comptabilisera l'achat et créera le Bordereau de livraison. Continuer ?",
                parent=self):
            return
        r = self._appeler("valider_ep_bon_commande", self.bon_id_selectionne)
        if r is APPEL_ECHEC:
            return
        messagebox.showinfo(
            "Validé",
            "Bon de commande validé — une facture d'achat brouillon a été créée dans ENGAGEMENTS-PROJETS "
            "> Factures frs. Complétez-la si besoin puis validez-la avec sa date de règlement prévu.",
            parent=self)
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        self.bon_id_selectionne = int(self.tree.item(sel[0], "values")[0])

    def refresh(self):
        bons = self._appeler("list_ep_bons_commande")
        if bons is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for b in bons:
            self.tree.insert("", "end", values=(
                b["id"], b["numero"], core.to_display_date(b["date_commande"]), b.get("fournisseur_code") or "",
                b["statut"]))


class RemoteFacturesRecouvrementTab(ttk.Frame):
    """Journal des factures clients (suivi des retards de paiement) —
    équivalent réseau complet de la sous-partie « Factures » de
    RecouvrementTab (bureau) : création de facture + enregistrement du
    paiement (comptabilisé automatiquement)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_id = None

        ttk.Label(self, text="PAIEMENT/RECOUVREMENT — SUIVI DES RETARDS DE PAIEMENT CLIENTS",
                  font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Enregistrez ici chaque facture émise à un client. L'échéance de paiement est calculée "
            "automatiquement à partir du délai par défaut du client, à la date de facture. Renseignez "
            "ensuite la date réelle de paiement au fur et à mesure des encaissements."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 8))

        form = ttk.LabelFrame(self, text="Nouvelle facture")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Client :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.client_var = tk.StringVar()
        self.client_combo = ttk.Combobox(form, textvariable=self.client_var, width=28)
        self.client_combo.grid(row=0, column=1, padx=4)
        self.client_combo.bind("<KeyRelease>", self._on_client_keyrelease)
        self._refresh_client_values()

        ttk.Label(form, text="N° Pièce :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.piece_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.piece_var, width=14).grid(row=0, column=3, padx=4)

        ttk.Label(form, text="Libellé :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.libelle_var, width=26).grid(row=0, column=5, padx=4)

        ttk.Label(form, text="Montant :").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=16).grid(row=1, column=1, padx=4)

        ttk.Label(form, text="Date facture (JJ/MM/AAAA) :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.date_facture_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_facture_var, width=14).grid(row=1, column=3, padx=4)

        ttk.Button(form, text="Créer la facture (échéance auto)", command=self.add_facture).grid(
            row=1, column=4, columnspan=2, sticky="w", padx=12, pady=4)

        ttk.Button(form, text="Créer la facture (échéance auto)", command=self.add_facture).grid(
            row=1, column=4, columnspan=2, sticky="w", padx=12, pady=4)
        ttk.Label(form, text=(
            "Les factures créées via COMMERCIAL > Facturation apparaissent ici automatiquement, "
            "avec leur échéancier — ce formulaire sert uniquement pour une facture saisie directement."
        ), foreground="#595959", wraplength=1000).grid(row=2, column=0, columnspan=6, sticky="w", padx=4, pady=(4, 0))

        ttk.Button(self, text="Supprimer la facture sélectionnée", command=self.delete_facture).pack(
            anchor="w", padx=16, pady=(8, 4))
        ttk.Label(self, text="Cliquez sur une ligne pour ouvrir le paiement selon l'échéancier.",
                  foreground="#595959").pack(anchor="w", padx=16, pady=(0, 4))

        cols = ("id", "client", "piece", "libelle", "montant", "date_facture",
                "echeance_paiement", "statut_paiement")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["ID", "Client", "Pièce", "Libellé", "Montant", "Date facture",
                   "Échéance paiement", "Statut paiement"]
        widths = [40, 180, 90, 200, 110, 110, 130, 220]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("depasse", foreground="#B00020")
        self.tree.tag_configure("soldee", foreground="#1F7A1F")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<Button-1>", self._on_click)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _refresh_client_values(self):
        items = self._appeler("list_clients")
        if items is APPEL_ECHEC:
            return
        self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _on_client_keyrelease(self, event=None):
        query = self._extract_code(self.client_var.get())
        if query:
            items = self._appeler("list_clients", query)
            if items is APPEL_ECHEC:
                return
            self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _on_click(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        self.tree.selection_set(iid)
        self.selected_id = int(iid)
        RemoteRecouvrementPaiementDialog(self, self.remote, self.selected_id, on_saved=self.refresh)

    def add_facture(self):
        code = self._extract_code(self.client_var.get())
        if not code:
            messagebox.showwarning("Champ manquant", "Choisissez un client.", parent=self)
            return
        existe = self._appeler("client_exists", code)
        if existe is APPEL_ECHEC:
            return
        if not existe:
            messagebox.showerror(
                "Client invalide",
                f"Le client « {code} » n'existe pas. Créez-le d'abord dans l'onglet Clients.", parent=self)
            return
        if not self.montant_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le montant est obligatoire.", parent=self)
            return
        try:
            montant = float(self.montant_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.", parent=self)
            return
        if montant <= 0:
            messagebox.showerror("Erreur", "Le montant doit être strictement positif.", parent=self)
            return
        date_facture = core.to_iso_date(self.date_facture_var.get().strip())
        if not date_facture:
            messagebox.showwarning("Champ manquant", "La date de facture est obligatoire.", parent=self)
            return
        if self._appeler("add_facture", code, self.piece_var.get().strip(), self.libelle_var.get().strip(),
                          montant, date_facture) is APPEL_ECHEC:
            return
        self.piece_var.set("")
        self.libelle_var.set("")
        self.montant_var.set("")
        self.refresh()

    def delete_facture(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture.", parent=self)
            return
        if messagebox.askyesno("Confirmer", "Supprimer cette facture ?", parent=self):
            if self._appeler("delete_facture", self.selected_id) is APPEL_ECHEC:
                return
            self.selected_id = None
            self.refresh()

    def refresh(self):
        self._refresh_client_values()
        factures = self._appeler("list_factures_avec_statut_paiement")
        if factures is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for f in factures:
            tag = "depasse" if f.get("en_retard") else ("soldee" if f["statut_paiement_detail"] == "✓ Soldée" else "")
            self.tree.insert("", "end", iid=str(f["id"]), tags=(tag,) if tag else (), values=(
                f["id"], f["raison_sociale"], f["piece"] or "", f["libelle"] or "",
                fmt_cfa(f["montant"]), core.to_display_date(f["date_facture"]),
                core.to_display_date(f["date_echeance_paiement"]), f["statut_paiement_detail"],
            ))


class RemoteRecouvrementPaiementDialog(tk.Toplevel):
    """Paiement d'une facture client selon son échéancier via le réseau —
    équivalent réseau complet de RecouvrementPaiementDialog (bureau)."""

    def __init__(self, parent, remote: RemoteConnection, facture_client_id, on_saved=None):
        super().__init__(parent)
        self.remote = remote
        self.facture_client_id = facture_client_id
        self.on_saved = on_saved
        factures = appeler(self, remote, "list_factures")
        facture = next((f for f in factures if f["id"] == facture_client_id), None) if factures is not APPEL_ECHEC else None
        titre = f"{facture['piece']} — {facture['raison_sociale']}" if facture else ""
        self.title(f"Paiement — {titre}")
        self.geometry("720x480")
        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text=f"{titre} — {fmt_cfa(facture['montant']) if facture else ''} F CFA",
                  font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=12, pady=(12, 8))

        ttk.Label(self, text="Échéancier de règlement :").pack(anchor="w", padx=12)
        cols = ("id", "tranche", "date", "montant", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=8)
        for c, h, w in zip(cols, ["ID", "Tranche", "Échéance", "Montant", "Statut"], [0, 60, 100, 130, 120]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w", stretch=(c != "id"))
        self.tree.column("id", width=0, stretch=False)
        self.tree["displaycolumns"] = ("tranche", "date", "montant", "statut")
        self.tree.pack(fill="both", expand=True, padx=12, pady=6)
        self.tree.tag_configure("retard", foreground="#B00020")
        self.tree.tag_configure("payee", foreground="#1F7A1F")

        ttk.Button(self, text="Modifier l'échéancier (plusieurs tranches possibles)",
                   command=self.modifier_echeancier).pack(anchor="w", padx=12, pady=(0, 8))

        paiement_frame = ttk.LabelFrame(self, text="Payer la tranche sélectionnée")
        paiement_frame.pack(fill="x", padx=12, pady=(0, 8))
        ttk.Label(paiement_frame, text="Date de paiement :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.date_paiement_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(paiement_frame, textvariable=self.date_paiement_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(paiement_frame, text="Compte banque/caisse :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.compte_paiement_var = tk.StringVar()
        self.compte_paiement_combo = ttk.Combobox(paiement_frame, textvariable=self.compte_paiement_var, width=26)
        self.compte_paiement_combo.grid(row=0, column=3, padx=4)
        self.compte_paiement_combo.bind("<KeyRelease>", self._on_compte_paiement_keyrelease)
        self._refresh_compte_paiement_values()
        ttk.Button(paiement_frame, text="Enregistrer le paiement de cette tranche",
                   command=self.enregistrer_paiement).grid(row=0, column=4, padx=12)

        self.statut_var = tk.StringVar()
        ttk.Label(self, textvariable=self.statut_var, foreground="#1F7A1F").pack(anchor="w", padx=12, pady=(0, 12))

        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _refresh_compte_paiement_values(self):
        items = self._appeler("search_accounts", "", limit=200)
        if items is not APPEL_ECHEC:
            self.compte_paiement_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "5"]

    def _on_compte_paiement_keyrelease(self, event=None):
        query = self._extract_code(self.compte_paiement_var.get())
        items = self._appeler("search_accounts", query, limit=50)
        if items is not APPEL_ECHEC:
            self.compte_paiement_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "5"]

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        echeances = self._appeler("list_echeances_client", self.facture_client_id)
        if echeances is APPEL_ECHEC:
            return
        for e in echeances:
            tag = "payee" if e["statut"] == "Payée" else ("retard" if e["statut"] == "En retard" else "")
            self.tree.insert("", "end", iid=str(e["id"]), tags=(tag,) if tag else (), values=(
                e["id"], e["numero_tranche"], core.to_display_date(e["date_echeance"]),
                fmt_cfa(e["montant"]), e["statut"]))
        restant = sum(e["montant"] for e in echeances if e["statut"] != "Payée")
        self.statut_var.set("✓ Toutes les tranches sont payées." if restant <= 0 and echeances
                             else f"Reste à payer : {fmt_cfa(restant)} F CFA.")

    def modifier_echeancier(self):
        echeances = self._appeler("list_echeances_client", self.facture_client_id)
        if echeances is APPEL_ECHEC:
            return
        totals_ttc = sum(e["montant"] for e in echeances)
        RemoteEcheancierClientDialog(self, self.remote, self.facture_client_id, totals_ttc, self.refresh)

    def enregistrer_paiement(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une tranche dans l'échéancier ci-dessus.",
                                 parent=self)
            return
        echeance_id = int(sel[0])
        date_paiement = core.to_iso_date(self.date_paiement_var.get().strip())
        if not date_paiement:
            messagebox.showwarning("Champ manquant", "Saisissez la date de paiement.", parent=self)
            return
        compte = self._extract_code(self.compte_paiement_var.get())
        if not compte:
            messagebox.showwarning("Champ manquant", "Choisissez le compte banque ou caisse.", parent=self)
            return
        montant = self._appeler("enregistrer_paiement_echeance_client", echeance_id, date_paiement, compte)
        if montant is APPEL_ECHEC:
            return
        self.refresh()
        messagebox.showinfo("Paiement comptabilisé",
                             f"Paiement de {fmt_cfa(montant)} comptabilisé sur le serveur (Débit banque/caisse, "
                             f"Crédit client).", parent=self)
        if self.on_saved:
            self.on_saved()


class RemoteEcheancierClientDialog(tk.Toplevel):
    """Modifie l'échéancier de règlement d'une facture client déjà
    comptabilisée via le réseau — équivalent réseau complet de
    EcheancierClientDialog (bureau)."""

    def __init__(self, parent, remote: RemoteConnection, facture_client_id, montant_total, on_saved):
        super().__init__(parent)
        self.remote = remote
        self.facture_client_id = facture_client_id
        self.montant_total = montant_total
        self.on_saved = on_saved
        self.title("Échéancier de règlement")
        self.geometry("560x420")
        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text=(
            f"Montant total : {fmt_cfa(montant_total)} F CFA — répartissez ce montant sur une ou plusieurs "
            f"échéances (la somme des tranches doit correspondre exactement)."
        ), wraplength=520, foreground="#595959").pack(anchor="w", padx=10, pady=(10, 6))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=10)
        ttk.Label(form, text="Échéance (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Montant :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Button(form, text="Ajouter la tranche", command=self.ajouter_tranche).grid(row=0, column=4, padx=12)

        self.tree = ttk.Treeview(self, columns=("date", "montant"), show="headings", height=8)
        self.tree.heading("date", text="Échéance")
        self.tree.heading("montant", text="Montant")
        self.tree.column("date", width=140, anchor="w")
        self.tree.column("montant", width=160, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=10, pady=8)
        ttk.Button(self, text="Retirer la tranche sélectionnée", command=self.retirer_tranche).pack(
            anchor="w", padx=10)

        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=10, pady=6)

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Enregistrer l'échéancier", command=self.enregistrer).pack(side="left")
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=8)

        self.tranches = []
        existantes = appeler(self, remote, "list_echeances_client", facture_client_id)
        if existantes is not APPEL_ECHEC:
            for e in existantes:
                if e["statut"] != "Payée":
                    self.tranches.append({"date_echeance": e["date_echeance"], "montant": e["montant"]})
        self._refresh_tree()

    def _refresh_tree(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        total = 0.0
        for t in self.tranches:
            self.tree.insert("", "end", values=(core.to_display_date(t["date_echeance"]), fmt_cfa(t["montant"])))
            total += t["montant"]
        ecart = total - self.montant_total
        etat = "✓ correspond au montant total" if abs(ecart) < 1 else f"⚠ écart de {fmt_cfa(ecart)}"
        self.total_var.set(f"Total des tranches : {fmt_cfa(total)} F CFA — {etat}")

    def ajouter_tranche(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "Date d'échéance invalide.", parent=self)
            return
        try:
            montant = float(self.montant_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.", parent=self)
            return
        if montant <= 0:
            messagebox.showerror("Erreur", "Le montant doit être strictement positif.", parent=self)
            return
        self.tranches.append({"date_echeance": date_str, "montant": montant})
        self.montant_var.set("")
        self._refresh_tree()

    def retirer_tranche(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.tranches[idx]
        self._refresh_tree()

    def enregistrer(self):
        if not self.tranches:
            messagebox.showwarning("Vide", "Ajoutez au moins une tranche.", parent=self)
            return
        if appeler(self, self.remote, "set_echeancier_client", self.facture_client_id,
                   self.tranches) is APPEL_ECHEC:
            return
        messagebox.showinfo("Enregistré", "Échéancier mis à jour sur le serveur.", parent=self)
        self.on_saved()
        self.destroy()


class RemoteBalanceAgeeTab(ttk.Frame):
    """Balance âgée des créances clients via le réseau, avec seuils de
    tranches réglables — équivalent réseau complet de la sous-partie
    « Balance âgée » de RecouvrementTab (bureau).

    compute_balance_agee() renvoie « tranches » comme une LISTE indexée
    [0..len(seuils)] (et non un dict par libellé de tranche) — l'affichage
    doit donc utiliser les mêmes seuils que ceux effectivement envoyés au
    serveur pour construire les en-têtes de colonnes."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="BALANCE ÂGÉE DES CRÉANCES CLIENTS", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Répartit le montant des factures NON PAYÉES de chaque client par ancienneté (jours "
            "écoulés depuis la date de facture). Choisissez les seuils des tranches ci-dessous."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 8))

        seuils_bar = ttk.Frame(self)
        seuils_bar.pack(fill="x", padx=16, pady=4)
        ttk.Label(seuils_bar, text="Seuils des tranches (jours) :").pack(side="left")
        self.seuil1_var = tk.StringVar(value="30")
        self.seuil2_var = tk.StringVar(value="60")
        self.seuil3_var = tk.StringVar(value="90")
        ttk.Entry(seuils_bar, textvariable=self.seuil1_var, width=6).pack(side="left", padx=4)
        ttk.Label(seuils_bar, text="/").pack(side="left")
        ttk.Entry(seuils_bar, textvariable=self.seuil2_var, width=6).pack(side="left", padx=4)
        ttk.Label(seuils_bar, text="/").pack(side="left")
        ttk.Entry(seuils_bar, textvariable=self.seuil3_var, width=6).pack(side="left", padx=4)
        ttk.Button(seuils_bar, text="Appliquer", command=self.refresh).pack(side="left", padx=12)
        ttk.Label(seuils_bar, text="Préréglages :").pack(side="left", padx=(20, 4))
        ttk.Button(seuils_bar, text="30/60/90", command=lambda: self._preset_seuils(30, 60, 90)).pack(
            side="left", padx=2)
        ttk.Button(seuils_bar, text="15/30/60", command=lambda: self._preset_seuils(15, 30, 60)).pack(
            side="left", padx=2)
        ttk.Button(seuils_bar, text="30/60/120", command=lambda: self._preset_seuils(30, 60, 120)).pack(
            side="left", padx=2)

        cols = ("client", "t0", "t1", "t2", "t3", "total")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        self.tree.column("client", width=260, anchor="w")
        self.tree.tag_configure("total", background="#1F4E78", foreground="white", font=("Segoe UI", 9, "bold"))
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _preset_seuils(self, s1, s2, s3):
        self.seuil1_var.set(str(s1))
        self.seuil2_var.set(str(s2))
        self.seuil3_var.set(str(s3))
        self.refresh()

    def _get_seuils(self):
        try:
            s1 = int(self.seuil1_var.get())
            s2 = int(self.seuil2_var.get())
            s3 = int(self.seuil3_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Les seuils doivent être des nombres entiers (jours).", parent=self)
            return None
        if not (s1 < s2 < s3):
            messagebox.showerror("Erreur", "Les seuils doivent être croissants (ex. 30 < 60 < 90).", parent=self)
            return None
        return (s1, s2, s3)

    def refresh(self):
        seuils = self._get_seuils()
        if not seuils:
            return
        s1, s2, s3 = seuils
        headers = ["Client", f"0-{s1} j", f"{s1+1}-{s2} j", f"{s2+1}-{s3} j", f">{s3} j", "Total"]
        widths = [260, 110, 110, 110, 110, 120]
        cols = ("client", "t0", "t1", "t2", "t3", "total")
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w" if c == "client" else "e")

        clients = self._appeler("compute_balance_agee", seuils=list(seuils))
        if clients is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        totaux = [0.0, 0.0, 0.0, 0.0]
        for c in clients:
            tranches = c["tranches"]
            self.tree.insert("", "end", values=(
                c["raison_sociale"], fmt_cfa(tranches[0]), fmt_cfa(tranches[1]),
                fmt_cfa(tranches[2]), fmt_cfa(tranches[3]), fmt_cfa(c["total"])))
            for i in range(4):
                totaux[i] += tranches[i]
        self.tree.insert("", "end", tags=("total",), values=(
            "TOTAL", fmt_cfa(totaux[0]), fmt_cfa(totaux[1]), fmt_cfa(totaux[2]), fmt_cfa(totaux[3]),
            fmt_cfa(sum(totaux))))


class RemoteRecouvrementTab(ttk.Frame):
    """Regroupe le journal des factures clients et la balance âgée —
    équivalent réseau complet de RecouvrementTab (bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True)
        self.factures_tab = RemoteFacturesRecouvrementTab(notebook, remote)
        self.agee_tab = RemoteBalanceAgeeTab(notebook, remote)
        notebook.add(self.factures_tab, text="Factures")
        notebook.add(self.agee_tab, text="Balance âgée")

    def refresh(self):
        self.factures_tab.refresh()
        self.agee_tab.refresh()


class RemoteMargesTab(ttk.Frame):
    """Marges bénéficiaires (COMMERCIAL) via le réseau — mêmes indicateurs
    que la Liasse fiscale (marge commerciale, valeur ajoutée, résultat)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.text = tk.Text(self, font=("Consolas", 11), wrap="none")
        self.text.pack(fill="both", expand=True, padx=16, pady=16)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        cr = self._appeler("compute_liasse_resultat")
        if cr is APPEL_ECHEC:
            return
        label_ca = "Chiffre d'affaires (XB)"
        label_re = "Résultat d'exploitation (XE)"
        lines = [
            "MARGES BÉNÉFICIAIRES", "=" * 60, "",
            f"  {'Ventes de marchandises (TA)':<45} {cr['TA']:>14,.2f}",
            f"  {'Achats de marchandises (RA)':<45} {-cr['RA']:>14,.2f}",
            f"  {'MARGE COMMERCIALE (XA)':<45} {cr['XA']:>14,.2f}", "",
            f"  {label_ca:<45} {cr['XB']:>14,.2f}",
            f"  {'VALEUR AJOUTÉE (XC)':<45} {cr['XC']:>14,.2f}",
            f"  {label_re:<45} {cr['XE']:>14,.2f}",
            f"  {'RÉSULTAT NET (XI)':<45} {cr['XI']:>14,.2f}",
        ]
        self.text.delete("1.0", "end")
        self.text.insert("1.0", "\n".join(lines))


class RemoteContratsTab(ttk.Frame):
    """Contrats fournisseurs (ENGAGEMENTS-PROJETS) via le réseau — suivi
    complet des délais de livraison et paiement, équivalent réseau de
    ContratsTab (bureau) : création, mise à jour des dates réelles de
    livraison/paiement, et suppression."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_id = None

        ttk.Label(self, text="CONTRATS FOURNISSEURS — SUIVI DES DÉLAIS",
                  font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Enregistrez ici chaque commande/contrat passé avec un fournisseur. Les échéances de "
            "livraison et de paiement sont calculées automatiquement à partir des délais par défaut "
            "du fournisseur, à la date de commande. Renseignez ensuite les dates réelles au fur et à "
            "mesure — les dépassements sont signalés automatiquement."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 8))

        form = ttk.LabelFrame(self, text="Nouvelle commande / contrat")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Fournisseur :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.fournisseur_var = tk.StringVar()
        self.fournisseur_combo = ttk.Combobox(form, textvariable=self.fournisseur_var, width=28)
        self.fournisseur_combo.grid(row=0, column=1, padx=4)
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self._refresh_fournisseur_values()

        ttk.Label(form, text="N° Pièce :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.piece_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.piece_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Libellé :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.libelle_var, width=26).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Montant :").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=16).grid(row=1, column=1, padx=4)
        ttk.Label(form, text="Date commande (JJ/MM/AAAA) :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.date_commande_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_commande_var, width=14).grid(row=1, column=3, padx=4)
        ttk.Button(form, text="Créer la commande (échéances auto)", command=self.add_commande).grid(
            row=1, column=4, columnspan=2, sticky="w", padx=12, pady=4)

        update_frame = ttk.LabelFrame(self, text="Mettre à jour la commande sélectionnée (dates réelles)")
        update_frame.pack(fill="x", padx=16, pady=(8, 4))
        ttk.Label(update_frame, text="Date livraison réelle (JJ/MM/AAAA) :").grid(
            row=0, column=0, sticky="w", padx=4, pady=4)
        self.livraison_reelle_var = tk.StringVar()
        ttk.Entry(update_frame, textvariable=self.livraison_reelle_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Button(update_frame, text="Enregistrer la livraison", command=self.save_livraison).grid(
            row=0, column=2, padx=8)
        ttk.Label(update_frame, text="Date paiement réel (JJ/MM/AAAA) :").grid(
            row=0, column=3, sticky="w", padx=(20, 4))
        self.paiement_reel_var = tk.StringVar()
        ttk.Entry(update_frame, textvariable=self.paiement_reel_var, width=14).grid(row=0, column=4, padx=4)
        ttk.Button(update_frame, text="Enregistrer le paiement", command=self.save_paiement).grid(
            row=0, column=5, padx=8)
        ttk.Button(update_frame, text="Supprimer la commande sélectionnée", command=self.delete_commande).grid(
            row=0, column=6, padx=20)

        cols = ("id", "fournisseur", "piece", "libelle", "montant", "date_commande",
                "livraison_prevue", "statut_livraison", "echeance_paiement", "statut_paiement")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["ID", "Fournisseur", "Pièce", "Libellé", "Montant", "Date commande",
                   "Livraison prévue", "Statut livraison", "Échéance paiement", "Statut paiement"]
        widths = [40, 160, 80, 160, 100, 100, 100, 140, 110, 140]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("depasse", foreground="#B00020")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _refresh_fournisseur_values(self):
        items = self._appeler("list_fournisseurs")
        if items is APPEL_ECHEC:
            return
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self._extract_code(self.fournisseur_var.get())
        if query:
            items = self._appeler("list_fournisseurs", query)
            if items is APPEL_ECHEC:
                return
            self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        self.selected_id = int(self.tree.item(sel[0], "values")[0])

    def add_commande(self):
        code = self._extract_code(self.fournisseur_var.get())
        if not code:
            messagebox.showwarning("Champ manquant", "Choisissez un fournisseur.", parent=self)
            return
        existe = self._appeler("fournisseur_exists", code)
        if existe is APPEL_ECHEC:
            return
        if not existe:
            messagebox.showerror(
                "Fournisseur invalide",
                f"Le fournisseur « {code} » n'existe pas. Créez-le d'abord dans l'onglet Fournisseurs.",
                parent=self)
            return
        try:
            montant = float(self.montant_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.", parent=self)
            return
        date_commande = core.to_iso_date(self.date_commande_var.get().strip())
        if not date_commande:
            messagebox.showwarning("Champ manquant", "La date de commande est obligatoire.", parent=self)
            return
        if self._appeler("add_commande", code, self.piece_var.get().strip(), self.libelle_var.get().strip(),
                          montant, date_commande) is APPEL_ECHEC:
            return
        self.piece_var.set("")
        self.libelle_var.set("")
        self.montant_var.set("")
        self.refresh()

    def save_livraison(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une commande dans le tableau.", parent=self)
            return
        d = core.to_iso_date(self.livraison_reelle_var.get().strip())
        if not d:
            messagebox.showwarning("Champ manquant", "Saisissez la date de livraison réelle.", parent=self)
            return
        if self._appeler("update_commande", self.selected_id, date_livraison_reelle=d) is APPEL_ECHEC:
            return
        self.livraison_reelle_var.set("")
        self.refresh()

    def save_paiement(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une commande dans le tableau.", parent=self)
            return
        d = core.to_iso_date(self.paiement_reel_var.get().strip())
        if not d:
            messagebox.showwarning("Champ manquant", "Saisissez la date de paiement réel.", parent=self)
            return
        if self._appeler("update_commande", self.selected_id, date_paiement_reel=d) is APPEL_ECHEC:
            return
        self.paiement_reel_var.set("")
        self.refresh()

    def delete_commande(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une commande.", parent=self)
            return
        if messagebox.askyesno("Confirmer", "Supprimer cette commande ?", parent=self):
            if self._appeler("delete_commande", self.selected_id) is APPEL_ECHEC:
                return
            self.selected_id = None
            self.refresh()

    def refresh(self):
        self._refresh_fournisseur_values()
        commandes = self._appeler("list_commandes")
        if commandes is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for c in commandes:
            tags = ("depasse",) if (c["depassement_livraison"] or c["depassement_paiement"]) else ()
            self.tree.insert("", "end", tags=tags, values=(
                c["id"], c["raison_sociale"], c["piece"] or "", c["libelle"] or "",
                fmt_cfa(c["montant"]), core.to_display_date(c["date_commande"]),
                core.to_display_date(c["date_livraison_prevue"]), c["statut_livraison"],
                core.to_display_date(c["date_echeance_paiement"]), c["statut_paiement"],
            ))


class RemoteFactureEcheancierDialog(tk.Toplevel):
    """Planifie l'échéancier de paiement PRÉVU d'une Facture d'achat en
    brouillon via le réseau — équivalent réseau complet de
    FactureEcheancierDialog (bureau)."""

    def __init__(self, parent, remote: RemoteConnection, facture_id, net_a_payer, on_saved):
        super().__init__(parent)
        self.remote = remote
        self.facture_id = facture_id
        self.net_a_payer = net_a_payer
        self.on_saved = on_saved
        self.title("Échéancier de paiement prévu")
        self.geometry("560x420")
        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text=(
            f"Net à payer : {fmt_cfa(net_a_payer)} F CFA — répartissez ce montant sur une ou plusieurs "
            f"échéances prévues (purement planifié, rien n'est encore comptabilisé)."
        ), wraplength=520, foreground="#595959").pack(anchor="w", padx=10, pady=(10, 6))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=10)
        ttk.Label(form, text="Échéance (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Montant :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Button(form, text="Ajouter la tranche", command=self.ajouter_tranche).grid(row=0, column=4, padx=12)

        self.tree = ttk.Treeview(self, columns=("date", "montant"), show="headings", height=8)
        self.tree.heading("date", text="Échéance")
        self.tree.heading("montant", text="Montant")
        self.tree.column("date", width=140, anchor="w")
        self.tree.column("montant", width=160, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=10, pady=8)
        ttk.Button(self, text="Retirer la tranche sélectionnée", command=self.retirer_tranche).pack(
            anchor="w", padx=10)

        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=10, pady=6)

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Enregistrer l'échéancier", command=self.enregistrer).pack(side="left")
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=8)

        self.tranches = []
        existantes = appeler(self, remote, "list_echeances_facture_achat", facture_id)
        if existantes is not APPEL_ECHEC:
            self.tranches = [{"date_echeance": t["date_echeance"], "montant": t["montant"]} for t in existantes]
        self._refresh_tree()

    def _refresh_tree(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        total = 0.0
        for t in self.tranches:
            self.tree.insert("", "end", values=(core.to_display_date(t["date_echeance"]), fmt_cfa(t["montant"])))
            total += t["montant"]
        ecart = total - self.net_a_payer
        etat = "✓ correspond au net à payer" if abs(ecart) < 1 else f"⚠ écart de {fmt_cfa(ecart)}"
        self.total_var.set(f"Total des tranches : {fmt_cfa(total)} F CFA — {etat}")

    def ajouter_tranche(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "Date d'échéance invalide.", parent=self)
            return
        try:
            montant = float(self.montant_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.", parent=self)
            return
        if montant <= 0:
            messagebox.showerror("Erreur", "Le montant doit être strictement positif.", parent=self)
            return
        self.tranches.append({"date_echeance": date_str, "montant": montant})
        self.montant_var.set("")
        self._refresh_tree()

    def retirer_tranche(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        del self.tranches[idx]
        self._refresh_tree()

    def enregistrer(self):
        if not self.tranches:
            messagebox.showwarning("Vide", "Ajoutez au moins une tranche.", parent=self)
            return
        if appeler(self, self.remote, "set_echeancier_facture_achat", self.facture_id,
                   self.tranches) is APPEL_ECHEC:
            return
        messagebox.showinfo("Enregistré", "Échéancier planifié sur le serveur.", parent=self)
        self.on_saved()
        self.destroy()


class RemoteFacturesFrsTab(ttk.Frame):
    """Factures fournisseurs (achats) via le réseau — équivalent réseau
    complet de FacturesFrsTab (bureau) : reçoit les factures générées
    automatiquement par la validation d'un Bon de commande (ENGAGEMENTS-
    PROJETS > Expression de besoin), permet d'en créer directement, et
    sa validation (avec la date de règlement prévu, obligatoire) envoie
    les écritures comptables sur le serveur — avec entrée de stock
    automatique pour les lignes marchandises/matières premières."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.current_facture_id = None
        self._factures_cache = []

        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Button(top, text="Nouvelle facture", command=self.new_facture).pack(side="left", padx=2)
        ttk.Button(top, text="Supprimer cette facture", command=self.delete_facture).pack(side="left", padx=2)
        self.corriger_btn = ttk.Button(top, text="Corriger cette facture (erreur sur les chiffres)",
                                        command=self.corriger_facture)
        self.corriger_btn.pack(side="left", padx=2)
        ttk.Button(top, text="✓ Valider et envoyer en Saisie", command=self.valider).pack(side="left", padx=8)
        self.statut_var = tk.StringVar()
        ttk.Label(top, textvariable=self.statut_var, font=("Segoe UI", 10, "bold")).pack(side="left", padx=16)

        ttk.Label(self, text="Factures existantes — cliquez une ligne pour la charger ci-dessous :").pack(
            anchor="w", padx=12)
        cols_liste = ("id", "numero", "fournisseur", "date", "statut", "net_a_payer")
        self.tree_factures = ttk.Treeview(self, columns=cols_liste, show="headings", height=5)
        headers_liste = ["ID", "Numéro", "Fournisseur", "Date", "Statut", "Net à payer"]
        for c, h, w in zip(cols_liste, headers_liste, [0, 100, 240, 90, 100, 120]):
            self.tree_factures.heading(c, text=h)
            self.tree_factures.column(c, width=w, anchor="w", stretch=(c != "id"))
        self.tree_factures.column("id", width=0, stretch=False)
        self.tree_factures["displaycolumns"] = ("numero", "fournisseur", "date", "statut", "net_a_payer")
        self.tree_factures.pack(fill="x", padx=12, pady=(2, 8))
        self.tree_factures.bind("<<TreeviewSelect>>", self._on_facture_selected)

        info = ttk.Frame(self)
        info.pack(fill="x", padx=12, pady=4)
        ttk.Label(info, text="N° Facture :").grid(row=0, column=0, sticky="w", padx=4)
        self.numero_var = tk.StringVar()
        ttk.Entry(info, textvariable=self.numero_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(info, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(info, textvariable=self.date_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Label(info, text="Fournisseur (compte 40) :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.fournisseur_var = tk.StringVar()
        self.fournisseur_combo = ttk.Combobox(info, textvariable=self.fournisseur_var, width=26)
        self.fournisseur_combo.grid(row=0, column=5, padx=4)
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self._refresh_fournisseur_values()

        ttk.Label(info, text="Retenue % :").grid(row=1, column=0, sticky="w", padx=4, pady=(6, 0))
        self.retenue_taux_var = tk.StringVar()
        ttk.Entry(info, textvariable=self.retenue_taux_var, width=6).grid(
            row=1, column=1, sticky="w", padx=4, pady=(6, 0))
        ttk.Label(info, text="Compte retenue (classe 44) :").grid(
            row=1, column=2, sticky="w", padx=(12, 4), pady=(6, 0))
        self.retenue_compte_var = tk.StringVar()
        self.retenue_compte_combo = ttk.Combobox(info, textvariable=self.retenue_compte_var, width=30)
        self.retenue_compte_combo.grid(row=1, column=3, columnspan=2, sticky="w", padx=4, pady=(6, 0))
        self.retenue_compte_combo.bind("<KeyRelease>", self._on_retenue_compte_keyrelease)
        self._refresh_retenue_compte_values()
        self._init_retenue_defaults()

        ttk.Label(info, text="Date de règlement prévu (JJ/MM/AAAA) :").grid(
            row=2, column=0, columnspan=2, sticky="w", padx=4, pady=(6, 0))
        self.date_paiement_prevu_var = tk.StringVar()
        ttk.Entry(info, textvariable=self.date_paiement_prevu_var, width=14).grid(
            row=2, column=2, sticky="w", padx=4, pady=(6, 0))
        ttk.Label(info, text=(
            "Obligatoire pour valider — dès qu'elle est renseignée, la facture est comptabilisée sur le "
            "serveur et devient visible dans ENGAGEMENTS-PROJETS > Contrats et dans TRÉSORERIE."
        ), foreground="#595959", wraplength=650).grid(row=2, column=3, columnspan=3, sticky="w", padx=4, pady=(6, 0))

        echeancier_frame = ttk.Frame(info)
        echeancier_frame.grid(row=3, column=0, columnspan=6, sticky="we", padx=4, pady=(6, 0))
        ttk.Button(echeancier_frame, text="Planifier un échéancier (plusieurs tranches, optionnel)",
                   command=self.modifier_echeancier_facture).pack(side="left")
        self.echeancier_apercu_var = tk.StringVar()
        ttk.Label(echeancier_frame, textvariable=self.echeancier_apercu_var, foreground="#595959").pack(
            side="left", padx=12)

        form = ttk.LabelFrame(self, text="Ajouter une ligne (produit/service acheté — compte 6x)")
        form.pack(fill="x", padx=12, pady=6)
        ttk.Label(form, text="Compte d'achat :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.ligne_compte_var = tk.StringVar()
        self.ligne_compte_combo = ttk.Combobox(form, textvariable=self.ligne_compte_var, width=34)
        self.ligne_compte_combo.grid(row=0, column=1, padx=4)
        self.ligne_compte_combo.bind("<KeyRelease>", self._on_ligne_compte_keyrelease)
        self._refresh_ligne_compte_values()
        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.ligne_libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_libelle_var, width=26).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Quantité :").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.ligne_qte_var = tk.StringVar(value="1")
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=10).grid(row=1, column=1, sticky="w", padx=4)
        ttk.Label(form, text="Prix unitaire :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.ligne_prix_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_prix_var, width=14).grid(row=1, column=3, sticky="w", padx=4)
        ttk.Label(form, text="Code analytique :").grid(row=1, column=4, sticky="w", padx=(12, 4))
        self.ligne_analytic_var = tk.StringVar()
        self.ligne_analytic_combo = ttk.Combobox(form, textvariable=self.ligne_analytic_var, width=20)
        self.ligne_analytic_combo.grid(row=1, column=5, padx=4, sticky="w")
        self._refresh_ligne_analytic_values()
        ttk.Button(form, text="Ajouter la ligne", command=self.add_ligne).grid(row=1, column=6, padx=12)

        cols = ("id", "compte", "libelle", "type_stock", "qte", "prix", "analytique", "montant")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=6)
        headers = ["ID", "Compte", "Libellé", "Impact stock", "Qté", "Prix unit.", "Code analytique", "Montant HT"]
        widths = [40, 90, 220, 110, 60, 100, 110, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="x", padx=12, pady=6)
        ttk.Button(self, text="Supprimer la ligne sélectionnée", command=self.delete_ligne).pack(
            anchor="w", padx=12)

        self.totals_var = tk.StringVar()
        ttk.Label(self, textvariable=self.totals_var, font=("Segoe UI", 11, "bold")).pack(
            anchor="w", padx=12, pady=(8, 0))

        docs_frame = ttk.Frame(self)
        docs_frame.pack(fill="x", padx=12, pady=(8, 0))
        entete_col = ttk.Frame(docs_frame)
        entete_col.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ttk.Label(entete_col, text="En-tête de la facture (modifiable, pour l'impression) :").pack(anchor="w")
        self.entete_text = tk.Text(entete_col, height=2, font=("Segoe UI", 10))
        self.entete_text.pack(fill="x", pady=(0, 8))
        pied_col = ttk.Frame(docs_frame)
        pied_col.pack(side="left", fill="x", expand=True, padx=(8, 0))
        ttk.Label(pied_col, text="Pied de page de la facture (modifiable) :").pack(anchor="w")
        self.pied_text = tk.Text(pied_col, height=2, font=("Segoe UI", 10))
        self.pied_text.pack(fill="x", pady=(0, 8))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=12, pady=8)
        ttk.Button(btns, text="Enregistrer la facture", command=self.save_facture).pack(side="left", padx=2)
        ttk.Button(btns, text="Aperçu avant impression", command=self.imprimer_facture).pack(side="left", padx=2)
        ttk.Button(btns, text="Valider et envoyer en Saisie", command=self.valider).pack(side="left", padx=2)

        self.refresh_factures_list()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _refresh_fournisseur_values(self):
        items = self._appeler("list_fournisseurs")
        if items is not APPEL_ECHEC:
            self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self._extract_code(self.fournisseur_var.get())
        if query:
            items = self._appeler("list_fournisseurs", query)
            if items is not APPEL_ECHEC:
                self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _refresh_retenue_compte_values(self):
        items = self._appeler("search_accounts", "44", limit=100)
        if items is not APPEL_ECHEC:
            self.retenue_compte_combo["values"] = [
                f"{a['code']} — {a['label']}" for a in items if a["code"][:2] == "44"]

    def _on_retenue_compte_keyrelease(self, event=None):
        query = self._extract_code(self.retenue_compte_var.get())
        if query:
            items = self._appeler("search_accounts", query, limit=50)
            if items is not APPEL_ECHEC:
                self.retenue_compte_combo["values"] = [
                    f"{a['code']} — {a['label']}" for a in items if a["code"][:2] == "44"]

    def _init_retenue_defaults(self):
        taux = self._appeler("get_setting", "retenue_taux_defaut", core.RETENUE_TAUX_DEFAUT)
        if taux is not APPEL_ECHEC:
            self.retenue_taux_var.set(str(taux))
        compte = self._appeler("get_text_setting", "retenue_compte_defaut", core.COMPTE_RETENUE_DEFAUT)
        if compte is not APPEL_ECHEC:
            self.retenue_compte_var.set(compte)

    def _refresh_ligne_compte_values(self):
        items = self._appeler("search_accounts", "6", limit=100)
        if items is not APPEL_ECHEC:
            self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "6"]

    def _on_ligne_compte_keyrelease(self, event=None):
        query = self._extract_code(self.ligne_compte_var.get())
        if query:
            items = self._appeler("search_accounts", query, limit=50)
            if items is not APPEL_ECHEC:
                self.ligne_compte_combo["values"] = [
                    f"{a['code']} — {a['label']}" for a in items if a["classe"] == "6"]

    def _refresh_ligne_analytic_values(self):
        codes = self._appeler("list_analytic_codes")
        if codes is not APPEL_ECHEC:
            self.ligne_analytic_combo["values"] = [f"{c['code']} — {c['label']}" for c in codes]

    def refresh_factures_list(self):
        factures = self._appeler("list_factures_achat")
        if factures is APPEL_ECHEC:
            return
        self._factures_cache = factures
        for row in self.tree_factures.get_children():
            self.tree_factures.delete(row)
        for f in factures:
            totals = self._appeler("compute_facture_achat_totals", f["id"])
            net = fmt_cfa(totals["net_a_payer"]) if totals is not APPEL_ECHEC else ""
            self.tree_factures.insert("", "end", iid=str(f["id"]), values=(
                f["id"], f["numero"], f["raison_sociale"], core.to_display_date(f["date_facture"]),
                f["statut"], net))
        if self.current_facture_id is None and factures:
            self.current_facture_id = factures[0]["id"]
            self.tree_factures.selection_set(str(self.current_facture_id))
        elif self.current_facture_id is not None:
            try:
                self.tree_factures.selection_set(str(self.current_facture_id))
            except tk.TclError:
                pass
        self.load_facture()

    def _on_facture_selected(self, event=None):
        sel = self.tree_factures.selection()
        if not sel:
            return
        self.current_facture_id = int(sel[0])
        self.load_facture()

    def load_facture(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.entete_text.delete("1.0", "end")
        self.pied_text.delete("1.0", "end")
        if not self.current_facture_id:
            self.statut_var.set("Aucune facture — créez-en une nouvelle.")
            self.totals_var.set("")
            self.corriger_btn.configure(state="disabled")
            return
        f = self._appeler("get_facture_achat", self.current_facture_id)
        if f is APPEL_ECHEC or not f:
            self.current_facture_id = None
            self.statut_var.set("")
            self.corriger_btn.configure(state="disabled")
            return
        self.numero_var.set(f["numero"])
        self.date_var.set(core.to_display_date(f["date_facture"]))
        fournisseur = self._appeler("get_fournisseur", f["fournisseur_code"])
        if fournisseur is APPEL_ECHEC:
            return
        self.fournisseur_var.set(
            f"{f['fournisseur_code']} — {fournisseur['raison_sociale']}" if fournisseur else f["fournisseur_code"])
        self.retenue_taux_var.set(str(f["retenue_taux"]))
        self.retenue_compte_var.set(f["retenue_compte"])
        self.date_paiement_prevu_var.set(core.to_display_date(f.get("date_paiement_prevu") or ""))
        self.entete_text.insert("1.0", f["entete"] or "")
        self.pied_text.insert("1.0", f["pied_page"] or "")
        statut_label = "VALIDÉE (écritures envoyées en Saisie)" if f["statut"] == "validee" else "Brouillon"
        if f.get("bon_commande_id") and f["statut"] != "validee":
            statut_label += " — générée depuis un Bon de commande"
        self.statut_var.set(f"Statut : {statut_label}")
        self.corriger_btn.configure(state="normal" if f["statut"] == "validee" else "disabled")

        lignes = self._appeler("list_lignes_facture_achat", self.current_facture_id)
        if lignes is APPEL_ECHEC:
            return
        for l in lignes:
            impact = {"marchandise": "Stock marchandises (31)", "matiere_premiere": "Stock matières (32)"}.get(
                l["type_stock"], "Aucun (service)")
            self.tree.insert("", "end", values=(
                l["id"], l["compte_achat"], l["libelle"], impact,
                f"{l['quantite']:g}", fmt_cfa(l["prix_unitaire"]), l.get("analytic_code") or "",
                fmt_cfa(l["montant_ht"])))
        totals = self._appeler("compute_facture_achat_totals", self.current_facture_id)
        if totals is APPEL_ECHEC:
            return
        self.totals_var.set(
            f"TOTAL HT : {fmt_cfa(totals['total_ht'])}    Retenue ({totals['retenue_taux']:g}%) : "
            f"{fmt_cfa(totals['retenue_montant'])}    NET À PAYER : {fmt_cfa(totals['net_a_payer'])}")
        self._refresh_echeancier_apercu()

    def _refresh_echeancier_apercu(self):
        if not self.current_facture_id:
            self.echeancier_apercu_var.set("")
            return
        tranches = self._appeler("list_echeances_facture_achat", self.current_facture_id)
        if tranches is APPEL_ECHEC:
            return
        if not tranches:
            self.echeancier_apercu_var.set("Aucun échéancier planifié (paiement en une fois à la date ci-dessus).")
        else:
            self.echeancier_apercu_var.set(
                f"{len(tranches)} tranche(s) planifiée(s) : " +
                ", ".join(f"{core.to_display_date(t['date_echeance'])} ({fmt_cfa(t['montant'])})" for t in tranches))

    def modifier_echeancier_facture(self):
        if not self.current_facture_id:
            messagebox.showinfo("Info", "Créez ou sélectionnez d'abord une facture.", parent=self)
            return
        totals = self._appeler("compute_facture_achat_totals", self.current_facture_id)
        if totals is APPEL_ECHEC:
            return
        if totals["net_a_payer"] <= 0:
            messagebox.showinfo("Info", "Ajoutez d'abord des lignes avec un montant avant de planifier "
                                         "un échéancier.", parent=self)
            return
        RemoteFactureEcheancierDialog(self, self.remote, self.current_facture_id, totals["net_a_payer"],
                                       self._refresh_echeancier_apercu)

    def new_facture(self):
        numero = simpledialog.askstring("Nouvelle facture", "N° de facture :", parent=self)
        if not numero:
            return
        fournisseur_code = self._extract_code(self.fournisseur_var.get())
        if not fournisseur_code:
            messagebox.showinfo("Fournisseur requis",
                                 "Choisissez d'abord un fournisseur existant dans le champ Fournisseur.",
                                 parent=self)
            return
        date_str = core.to_iso_date(self.date_var.get().strip()) or date.today().strftime("%Y-%m-%d")
        try:
            retenue_taux = float(self.retenue_taux_var.get() or 0)
        except ValueError:
            retenue_taux = 0
        retenue_compte = self._extract_code(self.retenue_compte_var.get()) or core.COMPTE_RETENUE_DEFAUT
        fid = self._appeler("create_facture_achat", numero, date_str, fournisseur_code,
                             retenue_taux=retenue_taux, retenue_compte=retenue_compte)
        if fid is APPEL_ECHEC:
            return
        self.current_facture_id = fid
        self.refresh_factures_list()

    def save_facture(self):
        if not self.current_facture_id:
            messagebox.showinfo("Info", "Créez d'abord une facture.", parent=self)
            return
        try:
            retenue_taux = float(self.retenue_taux_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le taux de retenue doit être un nombre.", parent=self)
            return
        if self._appeler(
            "update_facture_achat", self.current_facture_id,
            numero=self.numero_var.get().strip(), date_facture=core.to_iso_date(self.date_var.get().strip()),
            fournisseur_code=self._extract_code(self.fournisseur_var.get()),
            entete=self.entete_text.get("1.0", "end").strip(), pied_page=self.pied_text.get("1.0", "end").strip(),
            retenue_taux=retenue_taux, retenue_compte=self._extract_code(self.retenue_compte_var.get()) or "447800",
            date_paiement_prevu=core.to_iso_date(self.date_paiement_prevu_var.get().strip()),
        ) is APPEL_ECHEC:
            return
        self._appeler("set_setting", "retenue_taux_defaut", retenue_taux)
        self._appeler("set_text_setting", "retenue_compte_defaut", self._extract_code(self.retenue_compte_var.get()))
        messagebox.showinfo("Enregistré", "Facture enregistrée (brouillon).", parent=self)
        self.refresh_factures_list()

    def add_ligne(self):
        if not self.current_facture_id:
            messagebox.showinfo("Info", "Créez d'abord une facture.", parent=self)
            return
        compte = self._extract_code(self.ligne_compte_var.get())
        libelle = self.ligne_libelle_var.get().strip()
        if not compte or not libelle:
            messagebox.showwarning("Champ manquant", "Compte d'achat et libellé sont obligatoires.", parent=self)
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
            prix = float(self.ligne_prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        analytic_code = self._extract_code(self.ligne_analytic_var.get()) or None
        if self._appeler("add_ligne_facture_achat", self.current_facture_id, compte, libelle, qte, prix,
                          analytic_code=analytic_code) is APPEL_ECHEC:
            return
        self.ligne_compte_var.set(""); self.ligne_libelle_var.set("")
        self.ligne_qte_var.set("1"); self.ligne_prix_var.set(""); self.ligne_analytic_var.set("")
        self.load_facture()

    def delete_ligne(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.", parent=self)
            return
        ligne_id = int(self.tree.item(sel[0], "values")[0])
        if self._appeler("delete_ligne_facture_achat", ligne_id) is APPEL_ECHEC:
            return
        self.load_facture()

    def valider(self):
        if not self.current_facture_id:
            messagebox.showinfo("Info", "Créez d'abord une facture.", parent=self)
            return
        self.save_facture()
        date_paiement_prevu = core.to_iso_date(self.date_paiement_prevu_var.get().strip())
        if not date_paiement_prevu:
            messagebox.showwarning(
                "Date manquante",
                "Renseignez la date de règlement prévu avant de valider — c'est elle qui déclenche la "
                "comptabilisation et la visibilité dans Contrats/Trésorerie.", parent=self)
            return
        if not messagebox.askyesno(
            "Confirmer la validation",
            "Valider cette facture ? Les écritures comptables seront envoyées sur le serveur (débit "
            "achats, crédit fournisseur, retenue à la source, et entrée de stock automatique pour les "
            "lignes marchandises/matières premières). Cette action est définitive.", parent=self
        ):
            return
        resultat = self._appeler("valider_facture_achat", self.current_facture_id,
                                  date_paiement_prevu=date_paiement_prevu)
        if resultat is APPEL_ECHEC:
            return
        msg = "Facture validée et écritures envoyées en Saisie."
        if resultat:
            msg += "\n\nAvertissements :\n" + "\n".join(resultat)
        messagebox.showinfo("Validation terminée", msg, parent=self)
        self.refresh_factures_list()

    def delete_facture(self):
        if not self.current_facture_id:
            return
        if messagebox.askyesno("Confirmer", "Supprimer cette facture ?", parent=self):
            if self._appeler("delete_facture_achat", self.current_facture_id) is APPEL_ECHEC:
                return
            self.current_facture_id = None
            self.refresh_factures_list()

    def corriger_facture(self):
        if not self.current_facture_id:
            return
        if not messagebox.askyesno(
            "Corriger cette facture",
            "Cette facture est déjà validée : ses écritures comptables (débit achats, crédit "
            "fournisseur, retenue à la source, entrée de stock) vont être RETIRÉES de la Saisie et la "
            "facture repassera en brouillon modifiable.\n\nVous pourrez alors corriger les chiffres puis "
            "la revalider.\n\nContinuer ?", parent=self
        ):
            return
        if self._appeler("devalider_facture_achat", self.current_facture_id) is APPEL_ECHEC:
            return
        messagebox.showinfo("Facture repassée en brouillon",
                             "La facture est de nouveau modifiable. Corrigez les chiffres puis cliquez "
                             "sur « Valider et envoyer en Saisie ».", parent=self)
        self.refresh_factures_list()

    def imprimer_facture(self):
        if not self.current_facture_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture.", parent=self)
            return
        f = self._appeler("get_facture_achat", self.current_facture_id)
        if f is APPEL_ECHEC or not f:
            return
        html = self._appeler("render_facture_achat_html" if f["statut"] == "validee" else "render_bon_commande_html",
                              self.current_facture_id)
        if html is APPEL_ECHEC:
            return
        import tempfile, webbrowser, os
        path = os.path.join(tempfile.gettempdir(), f"facture_achat_{self.current_facture_id}.html")
        with open(path, "w", encoding="utf-8") as f_out:
            f_out.write(html)
        webbrowser.open(f"file://{path}")


class RemoteBordereauLivraisonTab(ttk.Frame):
    """Bordereau de livraison (ENGAGEMENTS-PROJETS) via le réseau —
    consultation et confirmation de réception (quantités livrées)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.bordereau_id_selectionne = None

        ttk.Label(self, text="BORDEREAU DE LIVRAISON", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 4))

        cols = ("id", "numero", "date", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=10)
        for c, h, w in zip(cols, ["ID", "Numéro", "Date", "Statut"], [40, 120, 100, 100]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="x", padx=16, pady=(0, 8))
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        ttk.Label(self, text="Lignes (quantités commandées / livrées)", font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=16)
        self.tree_lignes = ttk.Treeview(self, columns=("libelle", "qte_cmd", "qte_liv", "unite"),
                                         show="headings", height=10)
        for c, h, w in zip(("libelle", "qte_cmd", "qte_liv", "unite"),
                           ["Libellé", "Qté commandée", "Qté livrée", "Unité"], [300, 120, 120, 80]):
            self.tree_lignes.heading(c, text=h)
            self.tree_lignes.column(c, width=w, anchor="w")
        self.tree_lignes.pack(fill="both", expand=True, padx=16, pady=(4, 8))
        ttk.Button(self, text="Valider la réception (confirme les quantités livrées)",
                   command=self.valider).pack(anchor="w", padx=16, pady=(0, 16))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        self.bordereau_id_selectionne = int(self.tree.item(sel[0], "values")[0])
        self._refresh_lignes()

    def _refresh_lignes(self):
        for row in self.tree_lignes.get_children():
            self.tree_lignes.delete(row)
        if not self.bordereau_id_selectionne:
            return
        lignes = self._appeler("list_lignes_bordereau_livraison", self.bordereau_id_selectionne)
        if lignes is APPEL_ECHEC:
            return
        for l in lignes:
            self.tree_lignes.insert("", "end", values=(
                l["libelle"], f"{l['quantite_commandee']:g}", f"{l['quantite_livree']:g}", l["unite"] or ""))

    def valider(self):
        if not self.bordereau_id_selectionne:
            messagebox.showinfo("Info", "Sélectionnez d'abord un bordereau.", parent=self)
            return
        r = self._appeler("valider_bordereau_livraison", self.bordereau_id_selectionne)
        if r is APPEL_ECHEC:
            return
        messagebox.showinfo("Validé", "Réception confirmée sur le serveur.", parent=self)
        self.refresh()

    def refresh(self):
        bordereaux = self._appeler("list_bordereaux_livraison")
        if bordereaux is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for b in bordereaux:
            self.tree.insert("", "end", values=(
                b["id"], b["numero"], core.to_display_date(b["date_livraison"]), b["statut"]))


class RemoteAmortissementsTab(ttk.Frame):
    """Taux d'amortissement par catégorie (IMMOBILISATIONS) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="AMORTISSEMENTS — TAUX PAR CATÉGORIE", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))
        self.tree = ttk.Treeview(self, columns=("categorie", "taux"), show="headings", height=12)
        self.tree.heading("categorie", text="Catégorie")
        self.tree.heading("taux", text="Taux (%)")
        self.tree.column("categorie", width=350, anchor="w")
        self.tree.column("taux", width=100, anchor="e")
        self.tree.pack(fill="x", padx=16, pady=(0, 8))
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        form = ttk.Frame(self)
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Nouveau taux (%) pour la catégorie sélectionnée :").pack(side="left")
        self.taux_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.taux_var, width=8).pack(side="left", padx=8)
        ttk.Button(form, text="Enregistrer", command=self.enregistrer).pack(side="left")
        self.selected_categorie = None
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_categorie = v[0]
        self.taux_var.set(v[1])

    def enregistrer(self):
        if not self.selected_categorie:
            messagebox.showinfo("Info", "Sélectionnez d'abord une catégorie.", parent=self)
            return
        try:
            taux = float(self.taux_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le taux doit être un nombre.", parent=self)
            return
        r = self._appeler("set_taux_amortissement", self.selected_categorie, taux)
        if r is APPEL_ECHEC:
            return
        self.refresh()

    def refresh(self):
        taux = self._appeler("list_taux_amortissement")
        if taux is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for t in taux:
            self.tree.insert("", "end", values=(t["categorie"], f"{t['taux_pct']:g}"))


class RemoteParcAutoTab(ttk.Frame):
    """Parc auto (TRANSPORT) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_id = None
        ttk.Label(self, text="PARC AUTO", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))
        form = ttk.LabelFrame(self, text="Véhicule")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Immatriculation :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.immat_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.immat_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Marque :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.marque_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.marque_var, width=16).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Modèle :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.modele_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.modele_var, width=16).grid(row=0, column=5, padx=4)
        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left", padx=8)
        cols = ("id", "immat", "marque", "modele")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        for c, h, w in zip(cols, ["ID", "Immatriculation", "Marque", "Modèle"], [40, 140, 160, 160]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if sel:
            self.selected_id = self.tree.item(sel[0], "values")[0]

    def add(self):
        if not self.immat_var.get().strip():
            messagebox.showwarning("Champ manquant", "L'immatriculation est obligatoire.", parent=self)
            return
        r = self._appeler("add_vehicule", self.immat_var.get(), marque=self.marque_var.get(),
                           modele=self.modele_var.get())
        if r is APPEL_ECHEC:
            return
        self.immat_var.set(""); self.marque_var.set(""); self.modele_var.set("")
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un véhicule.", parent=self)
            return
        r = self._appeler("delete_vehicule", self.selected_id)
        if r is APPEL_ECHEC:
            return
        self.refresh()

    def refresh(self):
        vehicules = self._appeler("list_vehicules")
        if vehicules is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for v in vehicules:
            self.tree.insert("", "end", values=(v["id"], v["immatriculation"], v["marque"] or "", v["modele"] or ""))


class RemoteMissionsTab(ttk.Frame):
    """Missions (TRANSPORT) via le réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_id = None
        ttk.Label(self, text="MISSIONS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))
        form = ttk.LabelFrame(self, text="Mission")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Destination :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.destination_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.destination_var, width=20).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Chauffeur :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.chauffeur_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.chauffeur_var, width=18).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Motif :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.motif_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.motif_var, width=20).grid(row=0, column=5, padx=4)
        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left", padx=8)
        cols = ("id", "destination", "chauffeur", "motif")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        for c, h, w in zip(cols, ["ID", "Destination", "Chauffeur", "Motif"], [40, 200, 180, 220]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if sel:
            self.selected_id = self.tree.item(sel[0], "values")[0]

    def add(self):
        if not self.destination_var.get().strip():
            messagebox.showwarning("Champ manquant", "La destination est obligatoire.", parent=self)
            return
        r = self._appeler("add_mission", self.destination_var.get(), chauffeur=self.chauffeur_var.get(),
                           motif=self.motif_var.get())
        if r is APPEL_ECHEC:
            return
        self.destination_var.set(""); self.chauffeur_var.set(""); self.motif_var.set("")
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une mission.", parent=self)
            return
        r = self._appeler("delete_mission", self.selected_id)
        if r is APPEL_ECHEC:
            return
        self.refresh()

    def refresh(self):
        missions = self._appeler("list_missions")
        if missions is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for m in missions:
            self.tree.insert("", "end", values=(m["id"], m["destination"], m["chauffeur"] or "", m["motif"] or ""))


class RemotePiecesRechangeTab(ttk.Frame):
    """Pièces de rechange (TRANSPORT/MAINTENANCE-QUALITÉ, partagé) via le
    réseau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_id = None
        ttk.Label(self, text="PIÈCES DE RECHANGE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))
        form = ttk.LabelFrame(self, text="Pièce")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Désignation :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.designation_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.designation_var, width=24).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Quantité stock :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.qte_var = tk.StringVar(value="0")
        ttk.Entry(form, textvariable=self.qte_var, width=10).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Coût unitaire :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.cout_var = tk.StringVar(value="0")
        ttk.Entry(form, textvariable=self.cout_var, width=12).grid(row=0, column=5, padx=4)
        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left", padx=8)
        cols = ("id", "designation", "qte", "cout")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        for c, h, w in zip(cols, ["ID", "Désignation", "Qté stock", "Coût unitaire"], [40, 300, 100, 120]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if sel:
            self.selected_id = self.tree.item(sel[0], "values")[0]

    def add(self):
        if not self.designation_var.get().strip():
            messagebox.showwarning("Champ manquant", "La désignation est obligatoire.", parent=self)
            return
        try:
            qte = float(self.qte_var.get() or 0)
            cout = float(self.cout_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et coût doivent être des nombres.", parent=self)
            return
        r = self._appeler("add_piece_rechange", self.designation_var.get(), quantite_stock=qte, cout_unitaire=cout)
        if r is APPEL_ECHEC:
            return
        self.designation_var.set(""); self.qte_var.set("0"); self.cout_var.set("0")
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une pièce.", parent=self)
            return
        r = self._appeler("delete_piece_rechange", self.selected_id)
        if r is APPEL_ECHEC:
            return
        self.refresh()

    def refresh(self):
        pieces = self._appeler("list_pieces_rechange")
        if pieces is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for p in pieces:
            self.tree.insert("", "end", values=(
                p["id"], p["designation"], f"{p['quantite_stock']:g}", fmt_cfa(p["cout_unitaire"])))


class RemoteReparationDialog(tk.Toplevel):
    """Détail d'une réparation (double-clic) via le réseau — pièces
    utilisées (décrémente le stock côté serveur) + main d'œuvre —
    équivalent réseau de ReparationDialog (bureau)."""

    def __init__(self, parent, remote: RemoteConnection, reparation_id, on_saved):
        super().__init__(parent)
        self.remote = remote
        self.reparation_id = reparation_id
        self.on_saved = on_saved
        self.title("Réparation")
        self.geometry("850x560")
        self.transient(parent)
        self.grab_set()

        rep = appeler(self, remote, "get_reparation", reparation_id)
        if rep is APPEL_ECHEC or not rep:
            self.destroy()
            return

        header = ttk.LabelFrame(self, text="Informations")
        header.pack(fill="x", padx=10, pady=8)
        ttk.Label(header, text="Description :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.description_var = tk.StringVar(value=rep["description"])
        ttk.Entry(header, textvariable=self.description_var, width=40).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Garage :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.garage_var = tk.StringVar(value=rep["garage"] or "")
        ttk.Entry(header, textvariable=self.garage_var, width=20).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Main d'œuvre :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.mo_var = tk.StringVar(value=str(rep["cout_main_oeuvre"]))
        ttk.Entry(header, textvariable=self.mo_var, width=14).grid(row=1, column=1, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(header, text="Statut :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.statut_var = tk.StringVar(value=rep["statut"])
        ttk.Combobox(header, textvariable=self.statut_var, width=17, state="readonly",
                     values=["en_cours", "terminee"]).grid(row=1, column=3, padx=4, pady=(4, 0))

        lignes_frame = ttk.LabelFrame(self, text="Pièces utilisées (décrémente le stock)")
        lignes_frame.pack(fill="both", padx=10, pady=6)
        form = ttk.Frame(lignes_frame)
        form.pack(fill="x", padx=6, pady=4)
        ttk.Label(form, text="Pièce :").grid(row=0, column=0, sticky="w")
        self.piece_var = tk.StringVar()
        self.piece_combo = ttk.Combobox(form, textvariable=self.piece_var, width=32, state="readonly")
        self.piece_combo.grid(row=0, column=1, padx=4)
        self._refresh_pieces()
        ttk.Label(form, text="Quantité :").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.qte_var = tk.StringVar(value="1")
        ttk.Entry(form, textvariable=self.qte_var, width=8).grid(row=0, column=3, padx=4)
        ttk.Button(form, text="Ajouter", command=self.add_ligne).grid(row=0, column=4, padx=12)

        cols = ("id", "designation", "qte", "cout_unit", "montant")
        self.tree = ttk.Treeview(lignes_frame, columns=cols, show="headings", height=10)
        for c, h, w in zip(cols, ["ID", "Pièce", "Quantité", "Coût unit.", "Montant"], [40, 300, 90, 100, 110]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=6, pady=6)
        ttk.Button(lignes_frame, text="Supprimer la ligne sélectionnée (restitue le stock)",
                   command=self.delete_ligne).pack(anchor="w", padx=6, pady=(0, 6))
        self.total_var = tk.StringVar()
        ttk.Label(lignes_frame, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=6, pady=(0, 6))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Enregistrer", command=self.save).pack(side="left", padx=4)
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=4)

        self.refresh_lignes()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _refresh_pieces(self):
        pieces = self._appeler("list_pieces_rechange")
        if pieces is APPEL_ECHEC:
            return
        self.pieces = pieces
        self.piece_combo["values"] = [f"{p['id']} — {p['designation']} ({p['quantite_stock']:g} en stock)"
                                       for p in pieces]

    def add_ligne(self):
        raw = self.piece_var.get()
        if not raw:
            messagebox.showwarning("Champ manquant", "Choisissez une pièce.", parent=self)
            return
        piece_id = int(raw.split(" — ", 1)[0])
        try:
            qte = float(self.qte_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "La quantité doit être un nombre.", parent=self)
            return
        if self._appeler("add_ligne_reparation", self.reparation_id, piece_id, quantite=qte) is APPEL_ECHEC:
            return
        self.qte_var.set("1")
        self._refresh_pieces()
        self.refresh_lignes()

    def delete_ligne(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.", parent=self)
            return
        ligne_id = self.tree.item(sel[0], "values")[0]
        if self._appeler("delete_ligne_reparation", ligne_id) is APPEL_ECHEC:
            return
        self._refresh_pieces()
        self.refresh_lignes()

    def refresh_lignes(self):
        lignes = self._appeler("list_lignes_reparation", self.reparation_id)
        if lignes is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for l in lignes:
            self.tree.insert("", "end", values=(
                l["id"], l["designation"], f"{l['quantite']:g}",
                fmt_cfa(l["cout_unitaire"]), fmt_cfa(l["montant"])))
        cout_total = self._appeler("compute_cout_total_reparation", self.reparation_id)
        if cout_total is APPEL_ECHEC:
            return
        self.total_var.set(f"Coût total (pièces + main d'œuvre) : {fmt_cfa(cout_total)}")

    def save(self):
        try:
            mo = float(self.mo_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La main d'œuvre doit être un nombre.", parent=self)
            return
        if self._appeler("update_reparation", self.reparation_id, description=self.description_var.get().strip(),
                          garage=self.garage_var.get().strip(), cout_main_oeuvre=mo,
                          statut=self.statut_var.get()) is APPEL_ECHEC:
            return
        self.refresh_lignes()
        messagebox.showinfo("Enregistré", "Réparation enregistrée.", parent=self)
        self.on_saved()


class RemoteMachinesTab(ttk.Frame):
    """Machines industrielles via le réseau — équivalent réseau complet de
    MachinesTab (bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_id = None
        ttk.Label(self, text="MACHINES INDUSTRIELLES", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Machine")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Nom / désignation :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.nom_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.nom_var, width=26).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Catégorie :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.categorie_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.categorie_var, width=18).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Compte d'immobilisation (optionnel) :").grid(
            row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.compte_var = tk.StringVar()
        self.compte_combo = ttk.Combobox(form, textvariable=self.compte_var, width=24)
        self.compte_combo.grid(row=1, column=1, padx=4, pady=(4, 0))
        self.compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
        self._refresh_compte_values()
        ttk.Label(form, text="Statut :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.statut_var = tk.StringVar(value="en fonctionnement")
        ttk.Combobox(form, textvariable=self.statut_var, width=18, state="readonly",
                     values=["en fonctionnement", "en panne", "en maintenance", "à l'arrêt"]).grid(
            row=1, column=3, padx=4, pady=(4, 0))
        ttk.Label(form, text="Notes :").grid(row=2, column=0, sticky="w", padx=4, pady=(4, 0))
        self.notes_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.notes_var, width=60).grid(
            row=2, column=1, columnspan=3, sticky="we", padx=4, pady=(4, 0))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=16, pady=8)
        cols = ("id", "nom", "categorie", "compte", "statut", "notes")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        for c, h, w in zip(cols, ["ID", "Nom", "Catégorie", "Compte", "Statut", "Notes"],
                           [40, 220, 140, 100, 140, 260]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("panne", foreground="#B00020")
        self.tree.tag_configure("maintenance", foreground="#B8860B")
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _refresh_compte_values(self):
        items = self._appeler("search_accounts", "2", limit=200)
        if items is not APPEL_ECHEC:
            self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "2"]

    def _on_compte_keyrelease(self, event=None):
        query = self.compte_var.get().strip()
        if query:
            items = self._appeler("search_accounts", query, limit=50)
            if items is not APPEL_ECHEC:
                self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "2"]

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.nom_var.set(v[1]); self.categorie_var.set(v[2]); self.compte_var.set(v[3])
        self.statut_var.set(v[4]); self.notes_var.set(v[5])

    def clear_form(self):
        self.selected_id = None
        for var in (self.nom_var, self.categorie_var, self.compte_var, self.notes_var):
            var.set("")
        self.statut_var.set("en fonctionnement")

    def add(self):
        if not self.nom_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le nom de la machine est obligatoire.", parent=self)
            return
        if self._appeler("add_machine", self.nom_var.get().strip(), categorie=self.categorie_var.get().strip(),
                          compte_immobilisation=self._extract_code(self.compte_var.get()),
                          statut=self.statut_var.get(), notes=self.notes_var.get().strip()) is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une machine.", parent=self)
            return
        if self._appeler("update_machine", self.selected_id, nom=self.nom_var.get().strip(),
                          categorie=self.categorie_var.get().strip(),
                          compte_immobilisation=self._extract_code(self.compte_var.get()),
                          statut=self.statut_var.get(), notes=self.notes_var.get().strip()) is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une machine.", parent=self)
            return
        if messagebox.askyesno("Confirmer", "Supprimer cette machine ?", parent=self):
            if self._appeler("delete_machine", self.selected_id) is APPEL_ECHEC:
                return
            self.clear_form()
            self.refresh()

    def refresh(self):
        machines = self._appeler("list_machines")
        if machines is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for m in machines:
            tag = "panne" if m["statut"] == "en panne" else ("maintenance" if m["statut"] == "en maintenance" else "")
            self.tree.insert("", "end", tags=(tag,) if tag else (), values=(
                m["id"], m["nom"], m["categorie"] or "", m["compte_immobilisation"] or "",
                m["statut"], m["notes"] or ""))


class RemoteRapportTechniqueTab(ttk.Frame):
    """Tableau de bord technique combiné via le réseau — équivalent
    réseau complet de RapportTechniqueTab (bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote

        canvas = tk.Canvas(self, highlightthickness=0)
        vscroll = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas_window = canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(canvas_window, width=e.width))
        canvas.configure(yscrollcommand=vscroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")
        canvas.bind("<Enter>", lambda e: canvas.bind_all(
            "<MouseWheel>", lambda ev: canvas.yview_scroll(int(-ev.delta / 120), "units")))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        ttk.Label(scrollable, text="RAPPORTS TECHNIQUES", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(scrollable, text=(
            "Vue combinée : production (stocks critiques), machines industrielles, parc auto "
            "(disponibilité pour le ravitaillement) et maintenance."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        btn_bar = ttk.Frame(scrollable)
        btn_bar.pack(fill="x", padx=16, pady=(0, 8))
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left")
        ttk.Button(btn_bar, text="📊 Voir le rapport complet avec graphiques (aperçu avant impression, PDF)",
                   command=self.afficher_rapport_complet).pack(side="left", padx=12)

        self.cartes_var = tk.StringVar()
        ttk.Label(scrollable, textvariable=self.cartes_var, font=("Segoe UI", 11, "bold"),
                  wraplength=1100).pack(anchor="w", padx=16, pady=(0, 12))

        self.tab_stocks = self._make_section(scrollable, "1. Production — Stocks critiques",
            ("compte", "libelle", "qte", "seuil"), ["Compte", "Désignation", "Quantité actuelle", "Seuil d'alerte"],
            [90, 260, 130, 130])
        self.tab_machines = self._make_section(scrollable, "2. Machines industrielles — État de fonctionnement",
            ("nom", "categorie", "compte", "statut"), ["Machine", "Catégorie", "Compte", "Statut"],
            [220, 150, 100, 150])
        self.tab_vehicules = self._make_section(scrollable, "3. Parc automobile — Disponibilité",
            ("immat", "marque", "chauffeur", "statut"), ["Immatriculation", "Marque/Modèle", "Chauffeur", "Statut"],
            [130, 200, 150, 150])
        self.tab_maintenance = self._make_section(scrollable, "4. Maintenance — Réparations en cours",
            ("date", "cible", "description", "statut", "cout"), ["Date", "Machine/Véhicule", "Description",
                                                                    "Statut", "Coût M.O."], [90, 180, 260, 110, 110])

        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _make_section(self, parent, titre, cols, headers, widths):
        ttk.Label(parent, text=titre, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=16, pady=(12, 2))
        tree = ttk.Treeview(parent, columns=cols, show="headings", height=6)
        for c, h, w in zip(cols, headers, widths):
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor="w")
        tree.tag_configure("alerte", foreground="#B00020")
        tree.pack(fill="x", padx=16, pady=(0, 8))
        return tree

    def afficher_rapport_complet(self):
        html = self._appeler("render_rapport_technique_html")
        if html is APPEL_ECHEC:
            return
        import tempfile, webbrowser, os
        path = os.path.join(tempfile.gettempdir(), "rapport_technique.html")
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        webbrowser.open(f"file://{path}")

    def refresh(self):
        d = self._appeler("compute_rapport_technique")
        if d is APPEL_ECHEC:
            return

        nb_machines_hs = len(d["machines_par_statut"].get("en panne", [])) + \
            len(d["machines_par_statut"].get("en maintenance", []))
        self.cartes_var.set(
            f"⚠ Stocks sous seuil : {len(d['stocks_critiques'])}     "
            f"⚠ Machines hors service : {nb_machines_hs}     "
            f"🚚 Véhicules disponibles : {d['nb_disponibles']}/{d['nb_vehicules']} "
            f"({d['taux_disponibilite']:.0f}%)     "
            f"🔧 Réparations en cours : {len(d['reparations_en_cours'])}"
        )

        for row in self.tab_stocks.get_children():
            self.tab_stocks.delete(row)
        for s in d["stocks_critiques"]:
            self.tab_stocks.insert("", "end", tags=("alerte",), values=(
                s["code"], s["label"], f"{s['qte_finale']:g}", f"{s['seuil_alerte']:g}"))
        if not d["stocks_critiques"]:
            self.tab_stocks.insert("", "end", values=("", "✓ Aucun stock sous son seuil d'alerte", "", ""))

        for row in self.tab_machines.get_children():
            self.tab_machines.delete(row)
        for m in sorted(d["machines"], key=lambda m: m["statut"] != "en fonctionnement"):
            tag = "alerte" if m["statut"] in ("en panne", "en maintenance") else ()
            self.tab_machines.insert("", "end", tags=(tag,) if tag else (), values=(
                m["nom"], m["categorie"] or "", m["compte_immobilisation"] or "", m["statut"]))

        for row in self.tab_vehicules.get_children():
            self.tab_vehicules.delete(row)
        for v in sorted(d["vehicules"], key=lambda v: v["statut"] != "actif"):
            tag = "alerte" if v["statut"] != "actif" else ()
            self.tab_vehicules.insert("", "end", tags=(tag,) if tag else (), values=(
                v["immatriculation"], f"{v['marque'] or ''} {v['modele'] or ''}".strip(),
                v["chauffeur_affecte"] or "", v["statut"]))

        for row in self.tab_maintenance.get_children():
            self.tab_maintenance.delete(row)
        for r in d["reparations_en_cours"]:
            self.tab_maintenance.insert("", "end", tags=("alerte",), values=(
                core.to_display_date(r["date_reparation"]), r["machine_nom"] or r["immatriculation"] or "—",
                r["description"], r["statut"], fmt_cfa(r["cout_main_oeuvre"])))
        if not d["reparations_en_cours"]:
            self.tab_maintenance.insert("", "end", values=("", "✓ Aucune réparation en cours", "", "", ""))


class RemoteReparationsTab(ttk.Frame):
    """Réparations (TRANSPORT) via le réseau — équivalent réseau complet
    de ReparationsTab (bureau) : création + détail (double-clic) avec
    pièces utilisées et main d'œuvre, pour un véhicule OU une machine."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self._by_iid = {}
        ttk.Label(self, text="RÉPARATIONS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Véhicule :").grid(row=0, column=0, sticky="w")
        self.vehicule_var = tk.StringVar()
        self.vehicule_combo = ttk.Combobox(form, textvariable=self.vehicule_var, width=20, state="readonly")
        self.vehicule_combo.grid(row=0, column=1, padx=4)
        ttk.Label(form, text="ou Machine :").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.machine_var = tk.StringVar()
        self.machine_combo = ttk.Combobox(form, textvariable=self.machine_var, width=20, state="readonly")
        self.machine_combo.grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Description :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.description_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.description_var, width=30).grid(row=0, column=5, padx=4)
        ttk.Button(form, text="Nouvelle réparation", command=self.new_reparation).grid(row=0, column=6, padx=12)
        ttk.Label(form, text="Choisissez soit un véhicule, soit une machine (pas les deux).",
                  foreground="#595959").grid(row=1, column=0, columnspan=7, sticky="w", pady=(2, 0))

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=16, pady=8)
        cols = ("id", "cible", "date", "description", "garage", "cout_mo", "cout_total", "statut")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        headers = ["ID", "Véhicule / Machine", "Date", "Description", "Garage", "Main d'œuvre", "Coût total", "Statut"]
        widths = [40, 160, 90, 220, 130, 100, 100, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self._on_double_click)
        ttk.Label(self, text="Double-cliquez sur une réparation pour gérer ses pièces utilisées.",
                  foreground="#595959").pack(anchor="w", padx=16, pady=(4, 8))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _refresh_vehicule_values(self):
        vehicules = self._appeler("list_vehicules")
        if vehicules is APPEL_ECHEC:
            return
        self.vehicules = vehicules
        self.vehicule_combo["values"] = [f"{v['id']} — {v['immatriculation']}" for v in vehicules]

    def _refresh_machine_values(self):
        machines = self._appeler("list_machines")
        if machines is APPEL_ECHEC:
            return
        self.machines = machines
        self.machine_combo["values"] = [f"{m['id']} — {m['nom']}" for m in machines]

    def new_reparation(self):
        if not self.description_var.get().strip():
            messagebox.showwarning("Champ manquant", "La description est obligatoire.", parent=self)
            return
        vehicule_id = None
        raw = self.vehicule_var.get()
        if raw:
            vehicule_id = int(raw.split(" — ", 1)[0])
        machine_id = None
        raw_m = self.machine_var.get()
        if raw_m:
            machine_id = int(raw_m.split(" — ", 1)[0])
        rid = self._appeler("create_reparation", self.description_var.get().strip(), vehicule_id=vehicule_id,
                             machine_id=machine_id)
        if rid is APPEL_ECHEC:
            return
        self.description_var.set("")
        self.refresh()
        RemoteReparationDialog(self, self.remote, rid, on_saved=self.refresh)

    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        rid = self._by_iid.get(sel[0])
        if rid:
            RemoteReparationDialog(self, self.remote, rid, on_saved=self.refresh)

    def refresh(self):
        self._refresh_vehicule_values()
        self._refresh_machine_values()
        reparations = self._appeler("list_reparations")
        if reparations is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._by_iid = {}
        for r in reparations:
            cout_total = self._appeler("compute_cout_total_reparation", r["id"])
            if cout_total is APPEL_ECHEC:
                return
            cible = r.get("machine_nom") or r.get("immatriculation") or ""
            iid = self.tree.insert("", "end", values=(
                r["id"], cible, core.to_display_date(r.get("date_reparation") or ""),
                r["description"], r.get("garage") or "", fmt_cfa(r["cout_main_oeuvre"]), fmt_cfa(cout_total),
                r["statut"]))
            self._by_iid[iid] = r["id"]


class RemoteSimplePlanTab(ttk.Frame):
    """Écran générique code/libellé (+ montant/unité optionnels) via le
    réseau — réutilisé pour Plan analytique, Plan budgétaire, Plan
    bailleurs de fonds, Taux TVA, Taux retenue — même principe que
    _SimplePlanTab dans l'application de bureau."""

    def __init__(self, parent, remote: RemoteConnection, titre, list_fn, add_fn, delete_fn,
                 code_label="Code", extra_field=None):
        super().__init__(parent)
        self.remote = remote
        self.titre = titre
        self.list_fn_name = list_fn
        self.add_fn_name = add_fn
        self.delete_fn_name = delete_fn
        self.extra_field = extra_field  # None, "unite", ou "montant"
        self.selected_code = None

        ttk.Label(self, text=titre, font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))
        form = ttk.LabelFrame(self, text="Élément")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text=f"{code_label} :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.code_var = tk.StringVar()
        self.code_entry = ttk.Entry(form, textvariable=self.code_var, width=16)
        self.code_entry.grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.label_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.label_var, width=40).grid(row=0, column=3, padx=4)
        if extra_field:
            libelle_champ = "Unité (L, Kw, H...)" if extra_field == "unite" else "Montant / Taux (%)"
            ttk.Label(form, text=f"{libelle_champ} :").grid(row=0, column=4, sticky="w", padx=(12, 4))
            self.extra_var = tk.StringVar()
            ttk.Entry(form, textvariable=self.extra_var, width=12).grid(row=0, column=5, padx=4)
        else:
            self.extra_var = None

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        cols = ("code", "label", "extra") if extra_field else ("code", "label")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        self.tree.heading("code", text=code_label)
        self.tree.heading("label", text="Libellé")
        self.tree.column("code", width=140, anchor="w")
        self.tree.column("label", width=400, anchor="w")
        if extra_field:
            self.tree.heading("extra", text="Unité" if extra_field == "unite" else "Montant/Taux")
            self.tree.column("extra", width=120, anchor="e" if extra_field == "montant" else "w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_code = v[0]
        self.code_var.set(v[0]); self.label_var.set(v[1])
        if self.extra_var is not None and len(v) > 2:
            self.extra_var.set(v[2])
        self.code_entry.configure(state="disabled")

    def clear_form(self):
        self.selected_code = None
        self.code_var.set(""); self.label_var.set("")
        if self.extra_var is not None:
            self.extra_var.set("")
        self.code_entry.configure(state="normal")

    def _extra_kwargs(self):
        if not self.extra_field:
            return {}
        raw = self.extra_var.get().strip()
        if self.extra_field == "unite":
            return {"unite": raw or None}
        try:
            return {"montant": float(raw) if raw else 0}
        except ValueError:
            messagebox.showerror("Erreur", "Le montant/taux doit être un nombre.", parent=self)
            return None

    def add(self):
        if not self.code_var.get().strip() or not self.label_var.get().strip():
            messagebox.showwarning("Champ manquant", "Code et libellé sont obligatoires.", parent=self)
            return
        kwargs = self._extra_kwargs()
        if kwargs is None:
            return
        r = self._appeler(self.add_fn_name, self.code_var.get(), self.label_var.get(), **kwargs)
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un élément.", parent=self)
            return
        kwargs = self._extra_kwargs()
        if kwargs is None:
            return
        r = self._appeler(self.add_fn_name, self.selected_code, self.label_var.get(), **kwargs)
        if r is APPEL_ECHEC:
            return
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un élément.", parent=self)
            return
        if messagebox.askyesno("Confirmer", f"Supprimer « {self.selected_code} » ?", parent=self):
            r = self._appeler(self.delete_fn_name, self.selected_code)
            if r is APPEL_ECHEC:
                return
            self.clear_form()
            self.refresh()

    def refresh(self):
        items = self._appeler(self.list_fn_name)
        if items is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for it in items:
            if self.extra_field == "unite":
                self.tree.insert("", "end", values=(it["code"], it["label"], it.get("unite") or ""))
            elif self.extra_field == "montant":
                self.tree.insert("", "end", values=(it["code"], it["label"], fmt_cfa(it.get("montant"))))
            else:
                self.tree.insert("", "end", values=(it["code"], it["label"]))


class RemoteAnalytiquePeriodeTab(ttk.Frame):
    """Coûts analytiques par catégorie (Énergie, Maintenance) via le
    réseau — même principe que AnalytiquePeriodeTab dans l'application de
    bureau."""

    def __init__(self, parent, remote: RemoteConnection, titre, description, prefix):
        super().__init__(parent)
        self.remote = remote
        self.prefix = prefix
        ttk.Label(self, text=titre.upper(), font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=description, foreground="#595959", wraplength=1200, justify="left").pack(
            anchor="w", padx=16, pady=(0, 8))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 4))
        cols = ("code", "label", "debut", "periode", "fin")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=20)
        for c, h, w in zip(cols, ["Code", "Libellé", "Début période", "Charge période", "Cumul fin"],
                           [110, 260, 130, 130, 130]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w" if c in ("code", "label") else "e")
        self.tree.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        items = self._appeler("compute_couts_analytiques_categorie", self.prefix)
        if items is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for i in items:
            self.tree.insert("", "end", values=(
                i["code"], i["label"], fmt_cfa(i["solde_debut_periode"]),
                fmt_cfa(i["debit_periode"] - i["credit_periode"]), fmt_cfa(i["solde_fin_periode"])))


class RemoteCoutsFabricationPeriodeTab(ttk.Frame):
    """Coûts de fabrication réels de la période (écritures taguées AN-FAB) —
    lecture seule, équivalent réseau de CoutsFabricationPeriodeTab (bureau).

    compute_production() renvoie un RÉSUMÉ (dict) — ventes, production
    stockée, postes de coûts, marge — et non une liste de lignes ; l'affichage
    reprend donc le même format texte que côté bureau plutôt qu'un tableau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text=(
            "Pour qu'une charge remonte ici, saisissez le code analytique « AN-FAB » "
            "sur la ligne correspondante dans l'onglet Saisie."
        ), foreground="#595959").pack(anchor="w", padx=8, pady=(8, 0))
        self.text = tk.Text(self, font=("Consolas", 11), wrap="none")
        self.text.pack(fill="both", padx=8, pady=8)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(pady=(0, 8))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        p = self._appeler("compute_production")
        if p is APPEL_ECHEC:
            return
        lines = ["PRODUCTION DE L'EXERCICE", "=" * 60,
                 f"  {'Ventes (produits finis, travaux, services)':<50} {p['ventes']:>12,.2f}",
                 f"  {'Production stockée (variation stock 360000)':<50} {p['production_stockee']:>12,.2f}",
                 f"  {'VALEUR DE LA PRODUCTION':<50} {p['valeur_production']:>12,.2f}",
                 "", "COÛTS DE FABRICATION (axe AN-FAB)", "=" * 60]
        for poste in p["postes_cout"]:
            lines.append(f"  {poste['label']:<50} {poste['montant']:>12,.2f}")
        lines += [f"  {'COÛT DE PRODUCTION':<50} {p['cout_production']:>12,.2f}", "",
                  f"MARGE SUR COÛT DE PRODUCTION{'':<34}{p['marge']:>12,.2f}"]
        self.text.delete("1.0", "end")
        self.text.insert("1.0", "\n".join(lines))


class RemoteRecetteFabricationTab(ttk.Frame):
    """Nomenclature de fabrication (BOM) via le réseau — équivalent complet
    de RecetteFabricationTab (bureau) : créer/supprimer un produit fini,
    composer sa recette, calculer le coût de production et VALIDER la
    fabrication (écritures envoyées comme depuis le bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        self.selected_produit = None

        canvas = tk.Canvas(self, highlightthickness=0)
        vscroll = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas_window = canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(canvas_window, width=e.width))
        canvas.configure(yscrollcommand=vscroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        vscroll.pack(side="right", fill="y")
        canvas.bind("<Enter>", lambda e: canvas.bind_all(
            "<MouseWheel>", lambda ev: canvas.yview_scroll(int(-ev.delta / 120), "units")))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        top = ttk.Frame(scrollable)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Label(top, text="Produit fini :").pack(side="left")
        self.produit_var = tk.StringVar()
        self.produit_combo = ttk.Combobox(top, textvariable=self.produit_var, width=30, state="readonly")
        self.produit_combo.pack(side="left", padx=4)
        self.produit_combo.bind("<<ComboboxSelected>>", self._on_produit_selected)
        ttk.Button(top, text="Nouveau produit fini", command=self._new_produit).pack(side="left", padx=8)
        ttk.Button(top, text="Supprimer ce produit", command=self._delete_produit).pack(side="left", padx=2)

        params = ttk.Frame(scrollable)
        params.pack(fill="x", padx=12, pady=(0, 8))
        ttk.Label(params, text="Quantité produite par recette :").pack(side="left")
        self.qte_produite_var = tk.StringVar()
        ttk.Entry(params, textvariable=self.qte_produite_var, width=8).pack(side="left", padx=4)
        ttk.Label(params, text="Marge (%) :").pack(side="left", padx=(16, 0))
        self.marge_var = tk.StringVar()
        ttk.Entry(params, textvariable=self.marge_var, width=8).pack(side="left", padx=4)
        ttk.Label(params, text="Compte stock produit fini (classe 36) :").pack(side="left", padx=(16, 0))
        self.compte_stock_pf_var = tk.StringVar()
        self.compte_stock_pf_combo = ttk.Combobox(params, textvariable=self.compte_stock_pf_var, width=26)
        self.compte_stock_pf_combo.pack(side="left", padx=4)
        self.compte_stock_pf_combo.bind("<KeyRelease>", self._on_compte_pf_keyrelease)
        self._refresh_compte_pf_values()
        ttk.Button(params, text="Enregistrer ces paramètres", command=self._save_params).pack(side="left", padx=8)

        form = ttk.LabelFrame(scrollable, text="Ajouter un composant à la recette")
        form.pack(fill="x", padx=12, pady=4)
        ttk.Label(form, text="Type :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.type_var = tk.StringVar(value="matiere")
        type_combo = ttk.Combobox(form, textvariable=self.type_var, width=22, state="readonly",
                                   values=list(core.LIGNE_TYPES.values()))
        type_combo.set(core.LIGNE_TYPES["matiere"])
        type_combo.grid(row=0, column=1, padx=4)
        type_combo.bind("<<ComboboxSelected>>", self._on_type_changed)
        self.type_combo = type_combo

        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.libelle_var, width=22).grid(row=0, column=3, padx=4)

        self.compte_label = ttk.Label(form, text="Compte de stock :")
        self.compte_label.grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.compte_var = tk.StringVar()
        self.compte_combo = ttk.Combobox(form, textvariable=self.compte_var, width=26)
        self.compte_combo.grid(row=0, column=5, padx=4)
        self.compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
        self.compte_combo.bind("<<ComboboxSelected>>", self._on_compte_changed)
        self.compte_combo.bind("<FocusOut>", self._on_compte_changed)
        self._stocks_cache = {}
        self._refresh_stock_accounts()

        self.ligne_qte_label = ttk.Label(form, text="Quantité :")
        self.ligne_qte_label.grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.ligne_qte_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=10).grid(row=1, column=1, padx=4, sticky="w")

        self.cout_label = ttk.Label(form, text="Coût unitaire (si pas de compte de stock) :")
        self.cout_label.grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.ligne_cout_var = tk.StringVar()
        self.cout_entry = ttk.Entry(form, textvariable=self.ligne_cout_var, width=12)
        self.cout_entry.grid(row=1, column=3, padx=4, sticky="w")

        self.compte_apercu_var = tk.StringVar()
        ttk.Label(form, textvariable=self.compte_apercu_var, foreground="#1F7A1F", wraplength=480,
                  justify="left").grid(row=2, column=4, columnspan=2, sticky="w", padx=(12, 4))

        self.analytic_label = ttk.Label(form, text="Code analytique (Énergie/Maintenance...) :")
        self.analytic_label.grid(row=1, column=4, sticky="w", padx=(12, 4))
        self.ligne_analytic_var = tk.StringVar()
        self.analytic_combo = ttk.Combobox(form, textvariable=self.ligne_analytic_var, width=26)
        self.analytic_combo.grid(row=1, column=5, padx=4, sticky="w")
        self.analytic_combo.bind("<<ComboboxSelected>>", self._on_analytic_changed)
        self.analytic_combo.bind("<FocusOut>", self._on_analytic_changed)
        self._refresh_analytic_values()

        self.analytic_apercu_var = tk.StringVar()
        ttk.Label(form, textvariable=self.analytic_apercu_var, foreground="#1F7A1F", wraplength=480,
                  justify="left").grid(row=3, column=4, columnspan=2, sticky="w", padx=(12, 4))

        ttk.Button(form, text="Ajouter le composant", command=self.add_ligne).grid(row=4, column=5, padx=4, pady=4)

        cols = ("id", "type", "libelle", "compte", "quantite", "cout_unitaire", "analytique", "source", "montant")
        self.tree = ttk.Treeview(scrollable, columns=cols, show="headings", height=8)
        headers = ["ID", "Type", "Libellé", "Compte", "Quantité", "Coût unitaire", "Code analytique",
                   "Origine du coût", "Montant"]
        widths = [40, 90, 180, 90, 80, 110, 130, 170, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=12, pady=8)
        ttk.Button(scrollable, text="Supprimer le composant sélectionné", command=self.delete_ligne).pack(
            anchor="w", padx=12)

        self.result_text = tk.Text(scrollable, font=("Consolas", 11), height=8, wrap="none")
        self.result_text.pack(fill="x", padx=12, pady=8)

        ttk.Button(scrollable, text="Valider la fabrication (comptabiliser)", command=self.valider_fabrication).pack(
            anchor="w", padx=12, pady=(0, 8))

        self._on_type_changed()
        self.refresh_produits()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _refresh_stock_accounts(self):
        if self.type_combo.get() == core.LIGNE_TYPES["amortissement"]:
            immos = self._appeler("compute_immobilisations_liste")
            if immos is APPEL_ECHEC:
                return
            self._stocks_cache = {}
            self.compte_combo["values"] = [f"{i['compte']} — {i['libelle']}" for i in immos]
        else:
            stocks = self._appeler("compute_stocks_detail", prefixes=["31", "32", "33", "36"])
            if stocks is APPEL_ECHEC:
                return
            self._stocks_cache = {s["code"]: s for s in stocks}
            self.compte_combo["values"] = [f"{s['code']} — {s['label']}" for s in stocks]

    def _on_compte_keyrelease(self, event=None):
        query = self._extract_code(self.compte_var.get())
        if not query:
            return
        items = self._appeler("search_accounts", query, limit=30)
        if items is APPEL_ECHEC:
            return
        if self.type_combo.get() == core.LIGNE_TYPES["amortissement"]:
            self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items if a["classe"] == "2"]
        else:
            self.compte_combo["values"] = [
                f"{a['code']} — {a['label']}" for a in items if a["code"][:1] in ("3",)]

    def _on_compte_changed(self, event=None):
        code = self._extract_code(self.compte_var.get())
        if not code:
            self.compte_apercu_var.set("")
            return
        if self.type_combo.get() == core.LIGNE_TYPES["amortissement"]:
            fiche = self._appeler("get_immobilisation_fiche", code)
            if fiche is APPEL_ECHEC:
                return
            base = fiche.get("base_repartition_quantite")
            unite = fiche.get("base_repartition_unite") or "unité"
            if not base:
                self.compte_apercu_var.set(
                    f"Base de répartition non renseignée pour ce compte — allez dans IMMOBILISATIONS, "
                    f"sélectionnez « {code} » et indiquez sa quantité annuelle de référence (ex. "
                    f"5000 tonnes/an ou 2000 heures/an).")
                return
            cu = self._appeler("compute_cout_amortissement_unitaire", code)
            if cu is APPEL_ECHEC:
                return
            if cu is not None:
                self.compte_apercu_var.set(
                    f"Coût d'amortissement : {fmt_cfa(cu)} F CFA / {unite} "
                    f"(amortissement de la période ÷ {base:g} {unite}/an — sera utilisé automatiquement)")
            else:
                self.compte_apercu_var.set(
                    "Aucun amortissement comptabilisé pour cet équipement pour l'instant — saisissez un "
                    "coût unitaire manuel en attendant.")
            return
        stock = self._stocks_cache.get(code)
        if stock is None:
            # Compte pas dans le cache initial (trouvé via la recherche) — on va le chercher.
            detail = self._appeler("compute_stocks_detail", prefixes=[code[:2]])
            if detail is APPEL_ECHEC:
                return
            self._stocks_cache.update({s["code"]: s for s in detail})
            stock = self._stocks_cache.get(code)
        if stock is None:
            self.compte_apercu_var.set(f"Compte « {code} » introuvable parmi les comptes de stock.")
            return
        cu = stock.get("cout_unitaire_moyen")
        if cu is not None:
            self.compte_apercu_var.set(
                f"Coût unitaire moyen en stock : {fmt_cfa(cu)} F CFA — "
                f"{stock['qte_finale']:g} unité(s) disponible(s) (sera utilisé automatiquement)")
        else:
            self.compte_apercu_var.set(
                "Aucune quantité en stock pour ce compte pour l'instant — saisissez un coût unitaire "
                "manuel en attendant, ou renseignez d'abord un stock initial (onglet Stocks).")

    def _refresh_analytic_values(self):
        codes = self._appeler("list_analytic_codes")
        if codes is APPEL_ECHEC:
            return
        self.analytic_combo["values"] = [f"{c['code']} — {c['label']}" for c in codes]

    def _on_analytic_changed(self, event=None):
        code = self._extract_code(self.ligne_analytic_var.get())
        if not code:
            self.analytic_apercu_var.set("")
            self.ligne_qte_label.configure(text="Quantité :")
            return
        unite = self._appeler("get_analytic_code_unite", code)
        if unite is APPEL_ECHEC:
            return
        cu = self._appeler("compute_cout_unitaire_moyen_analytique", code, toutes_dates=True)
        if cu is APPEL_ECHEC:
            return
        if cu is not None:
            self.analytic_apercu_var.set(
                f"Coût moyen pondéré constaté : {fmt_cfa(cu)} F CFA / {unite or 'unité'} "
                f"(sera utilisé automatiquement)")
        else:
            self.analytic_apercu_var.set(
                f"Aucune quantité comptabilisée sous ce code pour l'instant — "
                f"saisissez un coût unitaire manuel en attendant.")
        self.ligne_qte_label.configure(text=f"Quantité ({unite}) :" if unite else "Quantité :")

    def _refresh_compte_pf_values(self):
        stocks = self._appeler("compute_stocks_detail", prefixes=["36"])
        if stocks is APPEL_ECHEC:
            return
        values = [f"{s['code']} — {s['label']}" for s in stocks]
        if "360000 — PRODUITS FINIS" not in values:
            existe = self._appeler("account_exists", "360000")
            if existe is APPEL_ECHEC:
                return
            if existe:
                label = self._appeler("get_account_label", "360000")
                if label is APPEL_ECHEC:
                    return
                values.insert(0, f"360000 — {label}")
        self.compte_stock_pf_combo["values"] = values

    def _on_compte_pf_keyrelease(self, event=None):
        query = self._extract_code(self.compte_stock_pf_var.get())
        if query:
            items = self._appeler("search_accounts", query, limit=50)
            if items is APPEL_ECHEC:
                return
            self.compte_stock_pf_combo["values"] = [
                f"{a['code']} — {a['label']}" for a in items if a["code"].startswith("36")]

    def _on_type_changed(self, event=None):
        type_label = self.type_combo.get()
        actif = type_label in (core.LIGNE_TYPES["matiere"], core.LIGNE_TYPES["amortissement"])
        self.compte_combo.configure(state="normal" if actif else "disabled")
        if actif:
            self.compte_label.configure(
                text="Compte d'immobilisation :" if type_label == core.LIGNE_TYPES["amortissement"]
                else "Compte de stock :")
            self._refresh_stock_accounts()
        else:
            self.compte_var.set("")
            self.compte_apercu_var.set("")

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _type_key(self):
        label = self.type_combo.get()
        for key, val in core.LIGNE_TYPES.items():
            if val == label:
                return key
        return "autre"

    def refresh_produits(self):
        produits = self._appeler("list_produits_finis")
        if produits is APPEL_ECHEC:
            return
        self.produit_combo["values"] = [f"{p['code']} — {p['nom']}" for p in produits]
        if produits and not self.selected_produit:
            self.selected_produit = produits[0]["code"]
            self.produit_var.set(f"{produits[0]['code']} — {produits[0]['nom']}")
        self.refresh()

    def _on_produit_selected(self, event=None):
        self.selected_produit = self._extract_code(self.produit_var.get())
        self.refresh()

    def _new_produit(self):
        code = simpledialog.askstring("Nouveau produit fini", "Code du produit :", parent=self)
        if not code:
            return
        nom = simpledialog.askstring("Nouveau produit fini", "Nom du produit :", parent=self)
        if not nom:
            return
        marge_defaut = self._appeler("get_setting", "marge_production_defaut", 30.0)
        if marge_defaut is APPEL_ECHEC:
            return
        if self._appeler("add_produit_fini", code.strip(), nom.strip(),
                          marge_pourcentage=marge_defaut) is APPEL_ECHEC:
            return
        self.selected_produit = code.strip()
        self.refresh_produits()

    def _delete_produit(self):
        if not self.selected_produit:
            return
        if messagebox.askyesno(
                "Confirmer", f"Supprimer le produit « {self.selected_produit} » et sa recette ?", parent=self):
            if self._appeler("delete_produit_fini", self.selected_produit) is APPEL_ECHEC:
                return
            self.selected_produit = None
            self.refresh_produits()

    def _save_params(self):
        if not self.selected_produit:
            return
        produit = self._appeler("get_produit_fini", self.selected_produit)
        if produit is APPEL_ECHEC:
            return
        try:
            qte = float(self.qte_produite_var.get() or 1)
            marge = float(self.marge_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité produite et marge doivent être des nombres.", parent=self)
            return
        compte_stock = self._extract_code(self.compte_stock_pf_var.get()) or "360000"
        existe = self._appeler("account_exists", compte_stock)
        if existe is APPEL_ECHEC:
            return
        if not existe:
            messagebox.showerror("Compte invalide", f"Le compte « {compte_stock} » n'existe pas.", parent=self)
            return
        if self._appeler("add_produit_fini", self.selected_produit, produit["nom"],
                          produit["description"] or "", qte, marge, compte_stock) is APPEL_ECHEC:
            return
        self.refresh()

    def add_ligne(self):
        if not self.selected_produit:
            messagebox.showinfo("Info", "Créez ou sélectionnez d'abord un produit fini.", parent=self)
            return
        libelle = self.libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé du composant est obligatoire.", parent=self)
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La quantité doit être un nombre.", parent=self)
            return
        type_key = self._type_key()
        compte = self._extract_code(self.compte_var.get()) if type_key in ("matiere", "amortissement") else None
        cout_unitaire = None
        if self.ligne_cout_var.get().strip():
            try:
                cout_unitaire = float(self.ligne_cout_var.get())
            except ValueError:
                messagebox.showerror("Erreur", "Le coût unitaire doit être un nombre.", parent=self)
                return
        if type_key in ("matiere", "amortissement") and not compte and cout_unitaire is None:
            messagebox.showwarning(
                "Champ manquant",
                "Choisissez un compte (stock ou immobilisation) ou saisissez un coût unitaire manuel.",
                parent=self)
            return
        analytic_code = self._extract_code(self.ligne_analytic_var.get()) or None
        if self._appeler("add_recette_ligne", self.selected_produit, type_key, libelle, qte, compte,
                          cout_unitaire, analytic_code=analytic_code) is APPEL_ECHEC:
            return
        self.libelle_var.set("")
        self.ligne_qte_var.set("")
        self.ligne_cout_var.set("")
        self.ligne_analytic_var.set("")
        self.refresh()

    def delete_ligne(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord un composant dans le tableau.", parent=self)
            return
        ligne_id = int(self.tree.item(sel[0], "values")[0])
        if self._appeler("delete_recette_ligne", ligne_id) is APPEL_ECHEC:
            return
        self.refresh()

    def refresh(self):
        self._refresh_stock_accounts()
        self._refresh_analytic_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.result_text.delete("1.0", "end")
        if not self.selected_produit:
            return
        produit = self._appeler("get_produit_fini", self.selected_produit)
        if produit is APPEL_ECHEC or not produit:
            return
        self.qte_produite_var.set(str(produit["quantite_produite"]))
        self.marge_var.set(str(produit["marge_pourcentage"]))
        self._refresh_compte_pf_values()
        label_pf = self._appeler("get_account_label", produit["compte_stock"])
        if label_pf is APPEL_ECHEC:
            return
        self.compte_stock_pf_var.set(f"{produit['compte_stock']} — {label_pf}")

        resultat = self._appeler("compute_cout_production", self.selected_produit)
        if resultat is APPEL_ECHEC:
            return
        for l in resultat["lignes"]:
            self.tree.insert("", "end", values=(
                l["id"], core.LIGNE_TYPES.get(l["type_ligne"], l["type_ligne"]), l["libelle"],
                l["compte"] or "", f"{l['quantite']:g}", f"{fmt_cfa(l['cout_unitaire_utilise'])}",
                l.get("analytic_code") or "",
                l["source_cout"], f"{fmt_cfa(l['montant'])}",
            ))
        lines = [
            f"COÛT DE PRODUCTION — {produit['nom']} ({self.selected_produit})", "=" * 70,
            f"  {'Coût de production total (recette)':<45} {resultat['cout_production_total']:>15,.2f}",
            f"  {'Quantité produite':<45} {resultat['quantite_produite']:>15,g}",
            f"  {'COÛT DE PRODUCTION UNITAIRE':<45} {resultat['cout_unitaire_produit']:>15,.2f}", "",
            f"  {'Marge appliquée':<45} {resultat['marge_pourcentage']:>14,g} %",
            f"  {'PRIX DE VENTE UNITAIRE SUGGÉRÉ':<45} {resultat['prix_vente_unitaire']:>15,.2f}",
            f"  {'dont marge unitaire':<45} {resultat['marge_unitaire']:>15,.2f}",
        ]
        self.result_text.insert("1.0", "\n".join(lines))

    def valider_fabrication(self):
        if not self.selected_produit:
            messagebox.showinfo("Info", "Sélectionnez d'abord un produit fini.", parent=self)
            return
        self._save_params()
        resultat = self._appeler("compute_cout_production", self.selected_produit)
        if resultat is APPEL_ECHEC:
            return
        if not resultat["lignes"]:
            messagebox.showwarning(
                "Recette vide", "Ajoutez au moins un composant à la recette avant de valider.", parent=self)
            return
        if not messagebox.askyesno(
            "Confirmer la validation de la fabrication",
            f"Valider la fabrication de « {resultat['produit']['nom']} » ?\n\n"
            f"Coût de production : {fmt_cfa(resultat['cout_production_total'])}\n"
            f"Quantité produite : {resultat['quantite_produite']:g}\n"
            f"Valeur du produit fini mis en stock (coût + marge {resultat['marge_pourcentage']:g}%) : "
            f"{fmt_cfa(resultat['prix_vente_total'])}\n\n"
            f"Cette action va DIMINUER les matières premières consommées (quantité et valeur) et "
            f"AUGMENTER le stock de produit fini, avec envoi des écritures dans le menu SAISIE. "
            f"Cette action est définitive.", parent=self
        ):
            return
        resultat2 = self._appeler("valider_fabrication", self.selected_produit)
        if resultat2 is APPEL_ECHEC:
            return
        _, warnings = resultat2
        msg = "Fabrication validée. Les matières premières ont été décrémentées et le produit fini mis en stock."
        if warnings:
            msg += "\n\nAvertissements :\n" + "\n".join(warnings)
        messagebox.showinfo("Validation terminée", msg, parent=self)
        self.refresh()


class RemoteProductionTab(ttk.Frame):
    """Regroupe la nomenclature de fabrication (coût de production, prix de
    vente) et le suivi des coûts réels de fabrication de la période —
    équivalent réseau complet de ProductionTab (bureau)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        inner = ttk.Notebook(self)
        inner.pack(fill="both", expand=True)
        self.recette_tab = RemoteRecetteFabricationTab(inner, remote)
        self.periode_tab = RemoteCoutsFabricationPeriodeTab(inner, remote)
        inner.add(self.recette_tab, text="Recettes / Coût de production")
        inner.add(self.periode_tab, text="Coûts de fabrication (période)")

    def refresh(self):
        self.recette_tab.refresh_produits()
        self.periode_tab.refresh()


class RemoteExercicesTab(ttk.Frame):
    """Exercices comptables (clôture) via le réseau — consultation et
    clôture (report des soldes vers l'exercice suivant)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="EXERCICES COMPTABLES", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))
        ttk.Label(self, text=(
            "La clôture calcule les soldes de clôture de tous les comptes de bilan et les reporte comme "
            "soldes d'ouverture de l'exercice suivant. Cette action est IRRÉVERSIBLE."
        ), foreground="#B00020", wraplength=1100, justify="left").pack(anchor="w", padx=16, pady=(0, 8))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 4))
        cols = ("exercice", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        self.tree.heading("exercice", text="Exercice")
        self.tree.heading("statut", text="Statut")
        self.tree.column("exercice", width=140, anchor="w")
        self.tree.column("statut", width=140, anchor="w")
        self.tree.pack(fill="x", padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.selected_exercice = None
        ttk.Button(self, text="Clôturer l'exercice sélectionné (reporte les soldes)",
                   command=self.cloturer).pack(anchor="w", padx=16, pady=(0, 16))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if sel:
            self.selected_exercice = self.tree.item(sel[0], "values")[0]

    def cloturer(self):
        if not self.selected_exercice:
            messagebox.showinfo("Info", "Sélectionnez d'abord un exercice.", parent=self)
            return
        if not messagebox.askyesno(
            "Clôturer cet exercice",
            f"Clôturer définitivement l'exercice {self.selected_exercice} ? Cette action est IRRÉVERSIBLE.",
            parent=self,
        ):
            return
        r = self._appeler("close_exercice", self.selected_exercice)
        if r is APPEL_ECHEC:
            return
        messagebox.showinfo("Clôturé", f"Exercice {self.selected_exercice} clôturé sur le serveur.", parent=self)
        self.refresh()

    def refresh(self):
        exercices = self._appeler("list_exercices")
        if exercices is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for e in exercices:
            self.tree.insert("", "end", values=(e["exercice"], "Clôturé" if e["cloture"] else "Ouvert"))


class RemotePlaceholderTab(ttk.Frame):
    """Écran non encore défini — même principe que PlaceholderTab dans
    l'application de bureau."""

    def __init__(self, parent, remote: RemoteConnection, titre, message):
        super().__init__(parent)
        ttk.Label(self, text=titre.upper(), font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))
        ttk.Label(self, text=message, foreground="#595959", wraplength=1100, justify="left").pack(
            anchor="w", padx=16)


class RemoteSynchronisationTab(ttk.Frame):
    """Synchronisation — explique pourquoi cette opération de maintenance
    du schéma n'est pas exposée à distance (volontairement, pour la
    sécurité — voir server.py RPC_WHITELIST)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        ttk.Label(self, text="SYNCHRONISATION", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))
        ttk.Label(self, text=(
            "Cette opération de maintenance (mise à jour du schéma de la base de données) n'est "
            "volontairement pas exposée à distance, par sécurité — elle reste réservée à l'application "
            "de bureau ou au poste serveur directement. Le serveur applique déjà automatiquement toute "
            "mise à jour de schéma nécessaire à son propre démarrage."
        ), foreground="#595959", wraplength=1100, justify="left").pack(anchor="w", padx=16)


class RemoteOuvertureTab(ttk.Frame):
    """Soldes d'ouverture via le réseau — Débit/Crédit avec totaux et
    contrôle d'équilibre, même principe que l'application de bureau."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote

        ttk.Label(self, text="SOLDES D'OUVERTURE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 4))

        form = ttk.LabelFrame(self, text="Saisir / modifier un solde")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Compte :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.compte_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.compte_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Débit :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.debit_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.debit_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Crédit :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.credit_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.credit_var, width=14).grid(row=0, column=5, padx=4)
        ttk.Button(form, text="Enregistrer", command=self.enregistrer).grid(row=0, column=6, padx=12)

        cols = ("code", "label", "debit", "credit")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=22)
        headers = ["N° Compte", "Libellé", "Débit", "Crédit"]
        widths = [90, 380, 130, 130]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w" if c in ("code", "label") else "e")
        self.tree.tag_configure("total", background="#1F4E78", foreground="white", font=("Segoe UI", 10, "bold"))
        self.tree.pack(fill="both", padx=16, pady=8)

        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=16, pady=(0, 16))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def enregistrer(self):
        compte = self.compte_var.get().strip()
        if not compte:
            messagebox.showwarning("Champ manquant", "Le compte est obligatoire.", parent=self)
            return
        try:
            debit = float(self.debit_var.get() or 0)
            credit = float(self.credit_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Débit et Crédit doivent être des nombres.", parent=self)
            return
        if debit and credit:
            messagebox.showwarning("Erreur", "Un compte est soit au débit, soit au crédit — pas les deux.",
                                    parent=self)
            return
        solde = debit - credit
        r = self._appeler("set_opening_balance", compte, solde)
        if r is APPEL_ECHEC:
            return
        self.compte_var.set(""); self.debit_var.set(""); self.credit_var.set("")
        self.refresh()

    def refresh(self):
        soldes = self._appeler("list_opening_balances")
        if soldes is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        total_debit = total_credit = 0.0
        for s in soldes:
            solde = s["solde"]
            debit = solde if solde > 0 else 0.0
            credit = -solde if solde < 0 else 0.0
            self.tree.insert("", "end", values=(
                s["code"], s["label"], fmt_cfa(debit) if debit else "", fmt_cfa(credit) if credit else ""))
            total_debit += debit
            total_credit += credit
        self.tree.insert("", "end", tags=("total",), values=(
            "", "TOTAL", fmt_cfa(total_debit), fmt_cfa(total_credit)))
        ecart = total_debit - total_credit
        etat = "Équilibré ✓" if abs(ecart) < 0.01 else "NON ÉQUILIBRÉ ✗"
        self.total_var.set(f"Total Débit : {fmt_cfa(total_debit)}   —   Total Crédit : {fmt_cfa(total_credit)}   "
                            f"—   {etat}")


class RemotePlanComptableTab(ttk.Frame):
    """Plan comptable en lecture seule via le réseau — recherche par code
    ou libellé (la création/modification de comptes reste réservée à
    l'application de bureau, opération structurante rarement nécessaire
    à distance)."""

    def __init__(self, parent, remote: RemoteConnection):
        super().__init__(parent)
        self.remote = remote
        ttk.Label(self, text="PLAN COMPTABLE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        search_bar = ttk.Frame(self)
        search_bar.pack(fill="x", padx=16, pady=4)
        ttk.Label(search_bar, text="Rechercher (code ou libellé) :").pack(side="left")
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_bar, textvariable=self.search_var, width=30)
        search_entry.pack(side="left", padx=6)
        search_entry.bind("<KeyRelease>", lambda e: self.refresh())
        cols = ("code", "label", "classe")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=26)
        for c, h, w in zip(cols, ["Code", "Libellé", "Classe"], [110, 420, 80]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.refresh()

    def _appeler(self, fonction, *args, **kwargs):
        return appeler(self, self.remote, fonction, *args, **kwargs)

    def refresh(self):
        comptes = self._appeler("search_accounts", self.search_var.get().strip(), limit=300)
        if comptes is APPEL_ECHEC:
            return
        for row in self.tree.get_children():
            self.tree.delete(row)
        for c in comptes:
            self.tree.insert("", "end", values=(c["code"], c["label"], c.get("classe", "")))


def main():
    login = LoginWindow()
    login.mainloop()
    if login.remote is None:
        return  # fenêtre fermée sans connexion réussie
    app = ClientApp(login.remote)
    app.mainloop()


if __name__ == "__main__":
    main()
